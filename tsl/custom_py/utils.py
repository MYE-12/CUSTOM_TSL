import frappe
import json
from frappe.utils import (
	add_days,
	add_months,
	cint,
	date_diff,
	flt,
	get_first_day,
	get_last_day,
	get_link_to_form,
	getdate,
	rounded,
	today,
)
from frappe import _
from frappe.model.mapper import get_mapped_doc
from frappe.utils.file_manager import save_file
from frappe.utils.file_manager import get_file
from frappe.utils import add_to_date
import requests
from datetime import datetime
from erpnext.setup.utils import get_exchange_rate
from frappe.utils.csvutils import read_csv_content
from tsl.tsl.doctype.leave_application_form.leave_application_form import validate_balance_leaves
# from hrms.payroll.doctype.salary_slip.salary_slip import generate_password_for_pdf
from frappe.utils.background_jobs import enqueue
from frappe.core.doctype.communication.email import make

#from frappe.permissions import has_role
# @frappe.whitelist()
# def currency():

# 	url = "https://api.exchangerate.host/USD"
# 	payload = {}
# 	headers = {}
# 	response = requests.request("GET", url, headers=headers, data=payload)
# 	data = response.json()
# 	print((data['rates']['KWD']))

# @frappe.whitelist()
# def send_sales_reminder(): # Reflects in Sales Track
	
#     sales_person = ["maaz@tsl-me.com","vazeem@tsl-me.com"]
#     sales_report = frappe.db.sql("""select date,sales_user from `tabSales Track` where date = CURDATE()""",as_dict=1)

#     if sales_report == []:
#         frappe.sendmail(
#                         recipients= sales_person,
#                        cc = ["yousuf@tsl-me.com"],
#                         subject="Daily Sales Report Reminder",
#                         message = """Dear Team,<br><br>Kindly fill the Daily Report by EOD<br><br>Thanks & Regards,<br><br>
#                         <b>​TSL Company</b><br>
#                         Bldg: 1473, Unit: 13, Street: 24, Block: 1, Al Rai Industrial Area, Kuwait.<br>
#                         Tel: +965 24741313 | Fax: +965 24741311 | <br>
#                         Email: info@tsl-me.com | Web: www.tsl-me.com"""
#                         )
#     for sr in sales_report:
#         print(sr.sales_person)
#         if sales_person[0] not in sr.sales_user:
#             print(sales_person[0])
#             frappe.sendmail(
#                     recipients= [sales_person[0]],
#                     cc = ["yousuf@tsl-me.com"],
#                     subject="Daily Sales Report Reminder",
#                     message = """Dear Team,<br><br>Kindly fill the Daily Report by EOD<br><br>Thanks & Regards,<br><br>
#                             <b>​TSL Company</b><br>
#                             Bldg: 1473, Unit: 13, Street: 24, Block: 1, Al Rai Industrial Area, Kuwait.<br>
#                             Tel: +965 24741313 | Fax: +965 24741311 | <br>
#                             Email: info@tsl-me.com | Web: www.tsl-me.com"""
#                             )
#         if sales_person[1] not in sr.sales_user:
#             print(sales_person[1])
#             frappe.sendmail(
#                 recipients= sales_person[1],
#                 cc = ["yousuf@tsl-me.com"],
#                             subject="Daily Sales Report Reminder",
#                             message = """Dear Team,<br><br>Kindly fill the Daily Report by EOD<br><br>Thanks & Regards,<br><br>
#                             <b>​TSL Company</b><br>
#                             Bldg: 1473, Unit: 13, Street: 24, Block: 1, Al Rai Industrial Area, Kuwait.<br>
#                             Tel: +965 24741313 | Fax: +965 24741311 | <br>
#                             Email: info@tsl-me.com | Web: www.tsl-me.com"""
#                             )

@frappe.whitelist()
def create_rfq_int(ps): # Request For Quotation
	doc = frappe.get_doc("Initial Evaluation",ps)
	new_doc = frappe.new_doc("Request for Quotation")
	new_doc.company = doc.company
	new_doc.branch = frappe.db.get_value("Work Order Data",doc.work_order_data,"branch")
	new_doc.initial_evaluation = ps
	new_doc.work_order_data = doc.work_order_data
	new_doc.department = frappe.db.get_value("Work Order Data",doc.work_order_data,"department")
	new_doc.items=[]
	psn = doc.items[-1].part_sheet_no
	for i in doc.get("items"):
		invent = [i[0] for i in frappe.db.get_list("Warehouse",{"company":doc.company,"is_branch":1},"name",as_list=1)]
		actual = frappe.db.get_value("Bin",{"item_code":i.part,"warehouse":["in",invent]},"actual_qty")
		needed = i.qty - (actual or 0)
		if needed > 0:
			qty = needed
		if i.parts_availability == "No" and psn == i.part_sheet_no:
			new_doc.append("items",{
				"item_code":i.part,
				"item_name":i.part_name,
				"description":i.part_name,
				'model':i.model,
				"category":i.category,
				"sub_category":i.sub_category,
				"mfg":i.manufacturer,
				'serial_no':i.serial_no,
				"uom":"Nos",
				"stock_uom":"Nos",
				"conversion_factor":1,
				"stock_qty":1,
				"qty":qty or 0,
				"schedule_date":add_to_date(new_doc.transaction_date,days = 2),
				"warehouse":new_doc.branch,
				"branch":new_doc.branch,
				"work_order_data":doc.work_order_data,
				"department":frappe.db.get_value("Work Order Data",doc.work_order_data,"department")
			})
	return new_doc

@frappe.whitelist() # Stock against the Parts requested in initial Evaluation
def getstock_detail(item_details,company):
	item_details = json.loads(item_details)
	data = ''
	if item_details:
		data = ''
		data += '<h4><center><b>STOCK DETAILS</b></center></h4>'
		data += '<h6><B>Note:</h6>'
		data += '<table class="table table-bordered" style = font-size:12px width=100% ><tr><td>REPAIR KUWAIT - <b>TSL</b></td><td>REPAIR DUBAI - <b>TSL</b></td><td>RIYADH MAIN -<b>TSL SA</b></td><td>JEDDAH -<b>TSL SA</b></td><td>DAMMAM- <b>TSL SA</b></tr>'
		data += '</table>'

		data += '<table class="table table-bordered">'
		data += '<tr>'
		data += '<td style="width:07%;padding:1px;font-size:14px;font-size:12px;background-color:#3333ff;color:white;"><center><b>SKU</b><center></td>'
		data += '<td style="width:07%;padding:1px;font-size:14px;font-size:12px;background-color:#3333ff;color:white;"><center><b>PART NUMBER</b><center></td>'
		if company == "TSL COMPANY - Kuwait":
			data += '<td style="width:07%;padding:1px;font-size:14px;font-size:12px;background-color:#3333ff;color:white;"><center><b>AVAILABLE STOCK - ( KW )</b><center></td>'
		if company == "TSL COMPANY - UAE":
			data += '<td style="width:07%;padding:1px;font-size:14px;font-size:12px;background-color:#3333ff;color:white;"><center><b>AVAILABLE STOCK-( DUBAI )</b><center></td>'
		if company == "TSL COMPANY - KSA":
			data += '<td style="width:07%;padding:1px;font-size:14px;font-size:12px;background-color:#3333ff;color:white;"><center><b>STOCK - ( RIYADH-SA )</b><center></td>'
			data += '<td style="width:07%;padding:1px;font-size:14px;font-size:12px;background-color:#3333ff;color:white;"><center><b>STOCK - ( JEDDAH-SA )</b><center></td>'
			data += '<td style="width:07%;padding:1px;font-size:14px;font-size:12px;background-color:#3333ff;color:white;"><center><b>STOCK - ( DAMMAM-SA )</b><center></td>'
		data += '</tr>'
		
		for j in item_details:
			country = frappe.get_value("Company",{"name":company},["country"])
			warehouse_stock = frappe.db.sql("""
			select sum(b.actual_qty) as qty from `tabBin` b join `tabWarehouse` wh on wh.name = b.warehouse join `tabCompany` c on c.name = wh.company where c.country = '%s' and b.item_code = '%s'
			""" % (country,j["sku"]),as_dict=True)[0]
			if not warehouse_stock["qty"]:
				warehouse_stock["qty"] = 0		
			new_po = frappe.db.sql("""select sum(`tabPurchase Order Item`.qty) as qty,sum(`tabPurchase Order Item`.received_qty) as d_qty from `tabPurchase Order` 
			left join `tabPurchase Order Item` on `tabPurchase Order`.name = `tabPurchase Order Item`.parent
			where `tabPurchase Order Item`.item_code = '%s' and `tabPurchase Order`.docstatus = 1 """ % (j["sku"]), as_dict=True)[0]
			if not new_po['qty']:
				new_po['qty'] = 0
			if not new_po['d_qty']:
				new_po['d_qty'] = 0
			in_transit = new_po['qty'] - new_po['d_qty']
			total = warehouse_stock["qty"] + in_transit
			
			stocks = frappe.db.sql("""select actual_qty,warehouse,stock_uom,stock_value from tabBin
			where item_code = '%s' """%(j["sku"]),as_dict=True)
			
			pos = frappe.db.sql("""select `tabPurchase Order Item`.item_code as item_code,`tabPurchase Order Item`.item_name as item_name,`tabPurchase Order`.supplier as supplier,sum(`tabPurchase Order Item`.qty) as qty,`tabPurchase Order Item`.rate as rate,`tabPurchase Order`.transaction_date as date,`tabPurchase Order`.name as po from `tabPurchase Order`
			left join `tabPurchase Order Item` on `tabPurchase Order`.name = `tabPurchase Order Item`.parent
			where `tabPurchase Order Item`.item_code = '%s' and `tabPurchase Order`.docstatus != 2 order by rate asc limit 1""" % (j["sku"]), as_dict=True)
			
			new_so = frappe.db.sql("""select sum(`tabSales Order Item`.qty) as qty,sum(`tabSales Order Item`.delivered_qty) as d_qty from `tabSales Order`
			left join `tabSales Order Item` on `tabSales Order`.name = `tabSales Order Item`.parent
			where `tabSales Order Item`.item_code = '%s' and `tabSales Order`.docstatus = 1  """ % (j["sku"]), as_dict=True)[0]
			
			if not new_so['qty']:
				new_so['qty'] = 0
			if not new_so['d_qty']:
				new_so['d_qty'] = 0
			del_total = new_so['qty'] - new_so['d_qty']
				
			i = 0
		   
			# for po in pos:
			#     if pos:
				
			data +='<tr>'
			data += '<td style="text-align:center;font-weight:bold;" colspan=1>%s</td>'%(j["sku"])
			
			data += '<td style="text-align:center;font-weight:bold;" colspan=1>%s</td>'%(frappe.db.get_value("Item Model",j["model"],['model']))
			# for stock in stocks:
				# if stock.actual_qty > 0:
			if company == "TSL COMPANY - Kuwait":
				kw = frappe.db.sql("""select actual_qty,warehouse,stock_uom,stock_value from tabBin
				where item_code = '%s' and warehouse = "%s" """%(j["sku"],"Kuwait - TSL"),as_dict=True)
				
				if kw:
					data += '<td style="text-align:center;font-weight:bold;" colspan=1>%s</td>'%(kw[0]["actual_qty"])
				else:
					data += '<td style="text-align:center;font-weight:bold;" colspan=1>%s</td>'%(0)

			dw = frappe.db.sql("""select actual_qty,warehouse,stock_uom,stock_value from tabBin
			where item_code = '%s' and warehouse = "%s" """%(j["sku"],"Dubai - TSL-UAE"),as_dict=True)
			
			if company == "TSL COMPANY - UAE":
				if dw:
					data += '<td style="text-align:center;font-weight:bold;" colspan=1>%s</td>'%(dw[0]["actual_qty"])
				else:
					data += '<td style="text-align:center;font-weight:bold;" colspan=1>%s</td>'%(0)

			if company == "TSL COMPANY - KSA":
				dm = frappe.db.sql("""select actual_qty,warehouse,stock_uom,stock_value from tabBin
				where item_code = '%s' and warehouse = "%s" """%(j["sku"],"Dammam - TSL - KSA"),as_dict=True)

				jed = frappe.db.sql("""select actual_qty,warehouse,stock_uom,stock_value from tabBin
				where item_code = '%s' and warehouse = "%s" """%(j["sku"],"Jeddah - TSL - KSA"),as_dict=True)

				ri = frappe.db.sql("""select actual_qty,warehouse,stock_uom,stock_value from tabBin
				where item_code = '%s' and warehouse = "%s" """%(j["sku"],"Riyadh - TSL - KSA"),as_dict=True)

				if dm:
					data += '<td style="text-align:center;font-weight:bold;" colspan=1>%s</td>'%(dm[0]["actual_qty"])
				else:
					data += '<td style="text-align:center;font-weight:bold;" colspan=1>%s</td>'%(0.0)
				
				if jed:
					data += '<td style="text-align:center;font-weight:bold;" colspan=1>%s</td>'%(jed[0]["actual_qty"])
				else:
					data += '<td style="text-align:center;font-weight:bold;" colspan=1>%s</td>'%(0.0)

				if ri:
					data += '<td style="text-align:center;font-weight:bold;" colspan=1>%s</td>'%(ri[0]["actual_qty"])
				else:
					data += '<td style="text-align:center;font-weight:bold;" colspan=1>%s</td>'%(0.0)

			data += '</tr>'
			i += 1
		data += '</table>'
					
		return data

@frappe.whitelist() # Stock against the Parts requested in initial Evaluation
def stock_ksa(item_details,company):
	item_details = json.loads(item_details)
	data = ''
	if item_details:
		data = ''
		data += '<h4><center><b>STOCK DETAILS</b></center></h4>'
		data += '<h6><B>Note:</h6>'
		data += '<table class="table table-bordered" style = font-size:12px width=100% ><tr><td>REPAIR KUWAIT - <b>TSL</b></td><td>REPAIR DUBAI - <b>TSL</b></td><td>RIYADH MAIN -<b>TSL SA</b></td><td>JEDDAH -<b>TSL SA</b></td><td>DAMMAM- <b>TSL SA</b></tr>'
		data += '</table>'

		data += '<table class="table table-bordered">'
		data += '<tr>'
		data += '<td style="width:07%;padding:1px;font-size:14px;font-size:12px;background-color:#3333ff;color:white;"><center><b>SKU</b><center></td>'
		# data += '<td style="width:07%;padding:1px;font-size:14px;font-size:12px;background-color:#3333ff;color:white;"><center><b>PART NUMBER</b><center></td>'
		if company == "TSL COMPANY - Kuwait":
			data += '<td style="width:07%;padding:1px;font-size:14px;font-size:12px;background-color:#3333ff;color:white;"><center><b>AVAILABLE STOCK - ( KW )</b><center></td>'
		if company == "TSL COMPANY - UAE":
			data += '<td style="width:07%;padding:1px;font-size:14px;font-size:12px;background-color:#3333ff;color:white;"><center><b>AVAILABLE STOCK-( DUBAI )</b><center></td>'
		if company == "TSL COMPANY - KSA":
			data += '<td style="width:07%;padding:1px;font-size:14px;font-size:12px;background-color:#3333ff;color:white;"><center><b>STOCK - ( RIYADH-SA )</b><center></td>'
			data += '<td style="width:07%;padding:1px;font-size:14px;font-size:12px;background-color:#3333ff;color:white;"><center><b>STOCK - ( JEDDAH-SA )</b><center></td>'
			data += '<td style="width:07%;padding:1px;font-size:14px;font-size:12px;background-color:#3333ff;color:white;"><center><b>STOCK - ( DAMMAM-SA )</b><center></td>'
		data += '</tr>'
		
		for j in item_details:
			country = frappe.get_value("Company",{"name":company},["country"])
			warehouse_stock = frappe.db.sql("""
			select sum(b.actual_qty) as qty from `tabBin` b join `tabWarehouse` wh on wh.name = b.warehouse join `tabCompany` c on c.name = wh.company where c.country = '%s' and b.item_code = '%s'
			""" % (country,j["sku"]),as_dict=True)[0]
			if not warehouse_stock["qty"]:
				warehouse_stock["qty"] = 0		
			new_po = frappe.db.sql("""select sum(`tabPurchase Order Item`.qty) as qty,sum(`tabPurchase Order Item`.received_qty) as d_qty from `tabPurchase Order` 
			left join `tabPurchase Order Item` on `tabPurchase Order`.name = `tabPurchase Order Item`.parent
			where `tabPurchase Order Item`.item_code = '%s' and `tabPurchase Order`.docstatus = 1 """ % (j["sku"]), as_dict=True)[0]
			if not new_po['qty']:
				new_po['qty'] = 0
			if not new_po['d_qty']:
				new_po['d_qty'] = 0
			in_transit = new_po['qty'] - new_po['d_qty']
			total = warehouse_stock["qty"] + in_transit
			
			stocks = frappe.db.sql("""select actual_qty,warehouse,stock_uom,stock_value from tabBin
			where item_code = '%s' """%(j["sku"]),as_dict=True)
			
			pos = frappe.db.sql("""select `tabPurchase Order Item`.item_code as item_code,`tabPurchase Order Item`.item_name as item_name,`tabPurchase Order`.supplier as supplier,sum(`tabPurchase Order Item`.qty) as qty,`tabPurchase Order Item`.rate as rate,`tabPurchase Order`.transaction_date as date,`tabPurchase Order`.name as po from `tabPurchase Order`
			left join `tabPurchase Order Item` on `tabPurchase Order`.name = `tabPurchase Order Item`.parent
			where `tabPurchase Order Item`.item_code = '%s' and `tabPurchase Order`.docstatus != 2 order by rate asc limit 1""" % (j["sku"]), as_dict=True)
			
			new_so = frappe.db.sql("""select sum(`tabSales Order Item`.qty) as qty,sum(`tabSales Order Item`.delivered_qty) as d_qty from `tabSales Order`
			left join `tabSales Order Item` on `tabSales Order`.name = `tabSales Order Item`.parent
			where `tabSales Order Item`.item_code = '%s' and `tabSales Order`.docstatus = 1  """ % (j["sku"]), as_dict=True)[0]
			
			if not new_so['qty']:
				new_so['qty'] = 0
			if not new_so['d_qty']:
				new_so['d_qty'] = 0
			del_total = new_so['qty'] - new_so['d_qty']
				
			i = 0
		   
			# for po in pos:
			#     if pos:
				
			data +='<tr>'
			data += '<td style="text-align:center;font-weight:bold;" colspan=1>%s</td>'%(j["sku"])
			
			# data += '<td style="text-align:center;font-weight:bold;" colspan=1>%s</td>'%(frappe.db.get_value("Item Model",j["model"],['model']))
			# for stock in stocks:
				# if stock.actual_qty > 0:
			if company == "TSL COMPANY - Kuwait":
				kw = frappe.db.sql("""select actual_qty,warehouse,stock_uom,stock_value from tabBin
				where item_code = '%s' and warehouse = "%s" """%(j["sku"],"Kuwait - TSL"),as_dict=True)
				
				if kw:
					data += '<td style="text-align:center;font-weight:bold;" colspan=1>%s</td>'%(kw[0]["actual_qty"])
				else:
					data += '<td style="text-align:center;font-weight:bold;" colspan=1>%s</td>'%(0)

			dw = frappe.db.sql("""select actual_qty,warehouse,stock_uom,stock_value from tabBin
			where item_code = '%s' and warehouse = "%s" """%(j["sku"],"Dubai - TSL-UAE"),as_dict=True)
			
			if company == "TSL COMPANY - UAE":
				if dw:
					data += '<td style="text-align:center;font-weight:bold;" colspan=1>%s</td>'%(dw[0]["actual_qty"])
				else:
					data += '<td style="text-align:center;font-weight:bold;" colspan=1>%s</td>'%(0)

			if company == "TSL COMPANY - KSA":
				dm = frappe.db.sql("""select actual_qty,warehouse,stock_uom,stock_value from tabBin
				where item_code = '%s' and warehouse = "%s" """%(j["sku"],"Dammam - TSL - KSA"),as_dict=True)

				jed = frappe.db.sql("""select actual_qty,warehouse,stock_uom,stock_value from tabBin
				where item_code = '%s' and warehouse = "%s" """%(j["sku"],"Jeddah - TSL - KSA"),as_dict=True)

				ri = frappe.db.sql("""select actual_qty,warehouse,stock_uom,stock_value from tabBin
				where item_code = '%s' and warehouse = "%s" """%(j["sku"],"Riyadh - TSL - KSA"),as_dict=True)

				if dm:
					data += '<td style="text-align:center;font-weight:bold;" colspan=1>%s</td>'%(dm[0]["actual_qty"])
				else:
					data += '<td style="text-align:center;font-weight:bold;" colspan=1>%s</td>'%(0.0)
				
				if jed:
					data += '<td style="text-align:center;font-weight:bold;" colspan=1>%s</td>'%(jed[0]["actual_qty"])
				else:
					data += '<td style="text-align:center;font-weight:bold;" colspan=1>%s</td>'%(0.0)

				if ri:
					data += '<td style="text-align:center;font-weight:bold;" colspan=1>%s</td>'%(ri[0]["actual_qty"])
				else:
					data += '<td style="text-align:center;font-weight:bold;" colspan=1>%s</td>'%(0.0)

			data += '</tr>'
			i += 1
		data += '</table>'
					
		return data



@frappe.whitelist() # part number details against the Parts requested in initial Evaluation
def get_model(item_details,company):
	item_details = json.loads(item_details)
	data = ''
	data += '<h4><center><b>STOCK DETAILS</b></center></h4>'

	data = ''
	data += '<h4><center><b>PART DETAILS</b></center></h4>'
		
	data += '<table class="table table-bordered">'
	data += '<tr>'
	data += '<td style="width:07%;padding:1px;font-size:14px;font-size:12px;background-color:#3333ff;color:white;"><center><b>SKU</b><center></td>'
	data += '<td style="width:07%;padding:1px;font-size:14px;font-size:12px;background-color:#3333ff;color:white;"><center><b>PART NUMBER</b><center></td>'
	data += '<td style="width:07%;padding:1px;font-size:14px;font-size:12px;background-color:#3333ff;color:white;"><center><b>CATEGORY</b><center></td>'
	data += '<td style="width:07%;padding:1px;font-size:14px;font-size:12px;background-color:#3333ff;color:white;"><center><b>SUB CATEGORY</b><center></td>'
	data += '<td style="width:07%;padding:1px;font-size:14px;font-size:12px;background-color:#3333ff;color:white;"><center><b>PACKAGE</b><center></td>'
	# data += '<td style="width:07%;padding:1px;font-size:14px;font-size:12px;background-color:#3333ff;color:white;"><center><b>VALUATION RATE</b><center></td>'
	   
	
	for j in item_details:
		md = frappe.get_value("Item Model",j["model"],'model')
		
		m = frappe.get_all("Item",{"model_num":md},["*"])
		for i in m:
			frappe.errprint(i.name)
			data += '<tr>'
			data += '<td style="padding:1px;font-size:14px;font-size:12px;"><center><b>%s</b><center></td>' %(i.name)
			data += '<td style="padding:1px;font-size:14px;font-size:12px;"><center><b>%s</b><center></td>' %(i.model_num)
			data += '<td style="padding:1px;font-size:14px;font-size:12px;"><center><b>%s</b><center></td>' %(i.category_)
			sc = frappe.get_value("Sub Category",i.sub_category,'sub_category')
			data += '<td style="padding:1px;font-size:14px;font-size:12px;"><center><b>%s</b><center></td>' %(sc)
			data += '<td style="padding:1px;font-size:14px;font-size:12px;"><center><b>%s</b><center></td>'%(i.package or '')
			# data += '<td style="padding:1px;font-size:14px;font-size:12px;"><center><b>%s</b><center></td>'%(i.valuation_rate)

			data += '</tr>'
		data += '<tr>'
		# data += '<tr>'
		# data += '<td style="padding:1px;font-size:14px;font-size:12px;"><center><b>-</b><center></td>' 
		# data += '<td style="padding:1px;font-size:14px;font-size:12px;"><center><b>-</b><center></td>' 
		# data += '<td style="padding:1px;font-size:14px;font-size:12px;"><center><b>-</b><center></td>' 
		# data += '<td style="padding:1px;font-size:14px;font-size:12px;"><center><b>-</b><center></td>'
		# data += '<td style="padding:1px;font-size:14px;font-size:12px;"><center><b>-</b><center></td>'
		# data += '</tr>'
	data += '</table>'
					
	return data

# def create_hooks():
#     job = frappe.db.exists('Scheduled Job Type', 'stock_reminder')
#     if not job:
#         sjt = frappe.new_doc("Scheduled Job Type")  
#         sjt.({
#             "method" : 'tsl.custom_py.utils.enque_qty',
#             "frequency" : 'Cron',
#             "cron_format" : '30 16 * * *'
#         })
#         sjt.save(ignore_permissions=True)

# def enque_qty():
# 	frappe.enqueue(
# 					item_qty, queue="long", enqueue_after_commit=True
# 				)


@frappe.whitelist() # Low Balance Email Trigger
def item_qty():
	item = frappe.db.sql("""select name,model,category,sub_category,qty from `tabItem` where qty < 5 """,as_dict=1)
	ir = 0
	data= ""
	data += '<table class="table table-bordered">'
	data += '<tr>'
	data += '<td style="width:07%;padding:1px;font-size:14px;font-size:12px;background-color:#3333ff;color:white;"><center><b>SKU</b><center></td>'
	data += '<td style="width:07%;padding:1px;font-size:14px;font-size:12px;background-color:#3333ff;color:white;"><center><b>PART NUMBER</b><center></td>'
	data += '<td style="width:07%;padding:1px;font-size:14px;font-size:12px;background-color:#3333ff;color:white;"><center><b>CATEGORY</b><center></td>'
	data += '<td style="width:07%;padding:1px;font-size:14px;font-size:12px;background-color:#3333ff;color:white;"><center><b>SUB - CATEGORY</b><center></td>'
	data += '<td style="width:07%;padding:1px;font-size:14px;font-size:12px;background-color:#3333ff;color:white;"><center><b>QTY</b><center></td>'
	data += '</tr>'


	for i in item:
		
		data +='<tr>'
		data += '<td style="text-align:center" colspan=1>%s</td>'%(i.name)
		data += '<td style="text-align:center" colspan=1>%s</td>'%(i.model)
		data += '<td style="text-align:center" colspan=1>%s</td>'%(i.category)
		data += '<td style="text-align:center" colspan=1>%s</td>'%(i.sub_category)
		data += '<td style="text-align:center;" colspan=1>%s</td>'%(i.qty)
		data += '</tr>'
		
	data += '</table>'
		# print(message)
	print(data)

	frappe.sendmail(
	recipients='yousuf@tsl-me.com',
	sender="info@tsl-me.com",
	subject="Low Stock",
	message=data
	
			
	)

@frappe.whitelist()
def amount(amount,currency,company):
	cc = frappe.get_value("Company",{"name":company},["default_currency"])
	# url = "https://api.exchangerate-api.com/v4/latest/%s"%(currency)
	# payload = {}
	# headers = {}
	# response = requests.request("GET", url, headers=headers, data=payload)
	# data = response.json()
	# frappe.errprint(data['rates'])
	# rate_kw = data['rates']['KWD']
	
	# conv_rate = float(amount) / rate_kw
	exr = get_exchange_rate(cc,currency)
	conv_rate = float(amount) * exr
	return conv_rate
	

@frappe.whitelist()
def get_work_orders(data):
	data = json.loads(data)
	wods = []
	sods = []
	for i in data:
		wo = frappe.db.sql(""" select DISTINCT `tabSales Invoice Item`.work_order_data as wo,`tabSales Invoice Item`.wod_no as w,
		`tabSales Invoice Item`.supply_order_data as so from `tabSales Invoice` 
		left join `tabSales Invoice Item` on `tabSales Invoice`.name = `tabSales Invoice Item`.parent 
		where `tabSales Invoice`.name = '%s' """ %(i["reference_name"]),as_dict =1)
		if wo:
			for j in wo:	
				wods.append(j)
	return wods

@frappe.whitelist()
def get_wod_purchase(data):
	data = json.loads(data)
	frappe.errprint(data)
	wods = []
	for i in data:
		pi = frappe.db.sql(""" select DISTINCT `tabPurchase Invoice`.name as p,`tabPurchase Invoice Item`.work_order_data as wo,`tabPurchase Invoice Item`.supply_order_data as so from `tabPurchase Invoice` 
		left join `tabPurchase Invoice Item` on `tabPurchase Invoice`.name = `tabPurchase Invoice Item`.parent 
		where `tabPurchase Invoice`.name = '%s' """ %(i["reference_name"]),as_dict =1)

		if pi:
			for j in pi:	
				wods.append(j)
	return wods

				
				


@frappe.whitelist()
def get_work_orders_purchase(data):
	data = json.loads(data)
	wods = []
	combined_data = []
	# sods = []
	for i in data:
		if i["reference_doctype"] == "Purchase Invoice":
			pi = frappe.db.sql(""" select DISTINCT `tabPurchase Invoice`.name as p,`tabPurchase Invoice Item`.work_order_data as wo,`tabPurchase Invoice Item`.supply_order_data as so from `tabPurchase Invoice` 
			left join `tabPurchase Invoice Item` on `tabPurchase Invoice`.name = `tabPurchase Invoice Item`.parent 
			where `tabPurchase Invoice`.name = '%s' """ %(i["reference_name"]),as_dict =1)
			for j in pi:
				frappe.errprint(i["reference_name"])
				combined_dict = {
				"doc_name":i["doc_name"],
				"reference_name": i["reference_name"],
				"work_order_data": j["wo"],
				"supply_order_data": j["so"]
			}
				combined_data.append(combined_dict)
	
	aggregated_data = {}

	for entry in combined_data:
		doc_name = entry['doc_name']
		ref_name = entry['reference_name']
		wo = entry['work_order_data']
		so = entry['supply_order_data']
		
		if ref_name not in aggregated_data:
			aggregated_data[ref_name] = {
				'work_order_data': set(),  # Use a set to avoid duplicates
				'supply_order_data': set()
			}
		
		if wo:
			aggregated_data[ref_name]['work_order_data'].add(wo)
		if so:
			aggregated_data[ref_name]['supply_order_data'].add(so)

	# Convert sets to lists for final output
	final_data = []
	for ref_name, values in aggregated_data.items():
		final_data.append({
		'reference_name': ref_name,
		'doc_name': doc_name,
		'work_order_data': ','.join(values['work_order_data']),
		'supply_order_data': ','.join(values['supply_order_data'])
	})

				
	frappe.errprint(final_data)
	
	return final_data


@frappe.whitelist()
def set_payment(data,pe,date):
	data = json.loads(data)
	for i in data:
		frappe.db.set_value("Work Order Data",i["work_order_data"],"payment_date",date)
		frappe.db.set_value("Work Order Data",i["work_order_data"],"payment_entry_reference",pe)
		wd = frappe.get_doc("Work Order Data",i["work_order_data"])
		wd.status = "P-Paid"
		# ldate = wd.status_duration_details[-1].date
		now = datetime.now()
		time_date = str(ldate).split(".")[0]
		format_data = "%Y-%m-%d %H:%M:%S"
		date = datetime.strptime(time_date, format_data)
		duration = now - date
		duration_in_s = duration.total_seconds()
		minutes = divmod(duration_in_s, 60)[0]/60
		data = str(minutes).split(".")[0]+"hrs "+str(minutes).split(".")[1][:2]+"min"
		frappe.db.set_value("Status Duration Details",wd.status_duration_details[-1].name,"duration",data)
		wd.append("status_duration_details",{
			"status":wd.status,
			"date":now,
		})
		doc = frappe.get_doc("Work Order Data",wd.name)
		doc.append("status_duration_details",{
			"status":wd.status,
			"date":now,
		})
		wd.save(ignore_permissions = 1)
		
		
		
@frappe.whitelist()
def get_receivable(customer,from_date,to_date,company):	

	data = ''
	data += '<table class="table table-bordered">'
	data += '<tr>'
	data += '<td align = center style="border:2px solid #dddddd;;font-size: 11px;width:15%"><b>Date</b></td>'
	data += '<td align = center style="border:2px solid #dddddd;;font-size: 11px;width:25%"><b>Invoice No</b></td>'
	data += '<td align = center style="border:2px solid #dddddd;;font-size: 11px;width:25%;"><b>Ref(WOD)</b></td>'
	data += '<td align = center style="border:2px solid #dddddd;;font-size: 11px;width:25%;"><b>Ref(PO)</b></td>'
	data += '<td align = center style="border:2px solid #dddddd;;font-size: 11px;width:15%;"><b>Invoiced</b></td>'
	data += '<td align = center style="border:2px solid #dddddd;;font-size: 11px;width:10%;"><b>Paid</b></td>'
	data += '<td align = center style="border:2px solid #dddddd;;font-size: 11px;width:15%;"><b>Outstanding</b></td>'
	data += '</tr>'
	# if company == "TSL COMPANY _KSA":
	# si = frappe.get_all("Sales Invoice",{"customer": customer,"posting_date": ["between", (from_date, to_date)],"status": ["in", ["Overdue", "Unpaid"]]})
	# else:
	si = frappe.get_all("Sales Invoice",{"company":company,"status": ["in", ["Overdue", "Unpaid"]],"customer":customer,"posting_date": ["between", (from_date,to_date)]},["*"],order_by="posting_date asc"  )
	os = 0
	# for i in si:
	for index, i in enumerate(si): 
		# frappe.errprint(index)
		os = os + i.outstanding_amount
		wo = frappe.db.sql(""" select DISTINCT `tabSales Invoice`.po_no as po_no,`tabSales Invoice Item`.work_order_data as wo,`tabSales Invoice Item`.previous_wod_no as pwo,`tabSales Invoice Item`.wod_no as w from `tabSales Invoice` 
		left join `tabSales Invoice Item` on `tabSales Invoice`.name = `tabSales Invoice Item`.parent 
		where `tabSales Invoice`.name = '%s' """ %(i.name),as_dict =1)
		wods = []

		for j in wo:
			if j["wo"]:
				original_string = str(j["wo"])
				substring = original_string[9:] 
				wods.append(substring)
				
			elif j["w"]:
				original_string = str(j["w"])
				substring = original_string[9:] 
				wods.append(substring)
			
			elif j["pwo"]:
				wods.append(j["pwo"])

			# else:
			# 	wods.append('')

		po_no = frappe.get_value("Sales Invoice",{"name":i.name},["po_no"])
		# if [po_no]:
		# 	wods.append(po_no)
		
		# Input date string
		input_date_string = str(i.due_date)

		# Convert string to datetime object
		input_date = datetime.strptime(input_date_string, "%Y-%m-%d")

		# Format the date in the desired format
		formatted_due_date = input_date.strftime("%d-%m-%Y")
		wod = str(wods).strip('[]')
		wod = wod.replace("'", '')
		if wod == None:
			wod = ''
		if index % 2 == 0:
			data += '<tr>'
			data += '<td style="font-size:10px;border:2px solid #dddddd;;">%s</td>'%(formatted_due_date)
			if company == "TSL COMPANY - Kuwait":
				data += '<td align = center style="border:2px solid #dddddd;;font-size: 10px;font-weight: bold;"><a href="https://erp.tsl-me.com/api/method/frappe.utils.print_format.download_pdf?doctype=Sales Invoice&name=%s&format=Sales Invoice&no_letterhead=0&letterhead=TSL New&settings={}&_lang=en">%s</a></td>'%(i.name,i.name)
			if company == "TSL COMPANY - UAE":
				data += '<td align = center style="border:2px solid #dddddd;;font-size: 10px;font-weight: bold;"><a href="https://erp.tsl-me.com/api/method/frappe.utils.print_format.download_pdf?doctype=Sales Invoice&name=%s&format=Sales Invoice - UAE&no_letterhead=0&letterhead=TSL New&settings={}&_lang=en">%s</a></td>'%(i.name,i.name)
			if company == "TSL COMPANY - KSA":
				data += '<td align = center style="border:2px solid #dddddd;;font-size: 10px;font-weight: bold;"><a href="https://erp.tsl-me.com/api/method/frappe.utils.print_format.download_pdf?doctype=Sales Invoice&name=%s&format=Sales Invoice KSA (R3)&no_letterhead=0&letterhead=TSL New&settings={}&_lang=en">%s</a></td>'%(i.name,i.name)
			
			data += '<td align = center style="font-size: 10px;border:2px solid #dddddd;;">%s</td>'%(wod)
			data += '<td align = center style="font-size: 10px;border:2px solid #dddddd;;">%s</td>'%(po_no)
			gt = "{:,.3f}".format(i.grand_total)
			p = i.grand_total-i.outstanding_amount
			paid = "{:,.3f}".format(p)
			outs= "{:,.3f}".format(i.outstanding_amount)
			data += '<td align = center style="font-size: 10px;border:2px solid #dddddd;;">%s</td>'%(gt)
			data += '<td align = center style="font-size: 10px;border:2px solid #dddddd;;">%s</td>'%("-")
			data += '<td align = center style="font-size: 10px;border:2px solid #dddddd;;">%s</td>'%(outs)
			data += '</tr>'

		else:
			data += '<tr>'
			data += '<td style="font-size: 10px;background-color:#CCCCCC;border:2px solid #dddddd;;">%s</td>'%(formatted_due_date)
			data += '<td align = center style="border:2px solid #dddddd;;background-color:#CCCCCC;font-size: 10px;font-weight: bold;"><a href="https://erp.tsl-me.com/api/method/frappe.utils.print_format.download_pdf?doctype=Sales Invoice&name=%s&format=Sales Invoice&no_letterhead=0&letterhead=TSL New&settings={}&_lang=en">%s</a></td>'%(i.name,i.name)
			data += '<td align = center style="background-color:#CCCCCC;font-size:10px;border:2px solid #dddddd;;">%s</td>'%(wod)
			data += '<td align = center style="background-color:#CCCCCC;font-size: 10px;border:2px solid #dddddd;;">%s</td>'%(po_no)
			gt = "{:,.3f}".format(i.grand_total)
			p = i.grand_total-i.outstanding_amount
			paid = "{:,.3f}".format(p)
			outs= "{:,.3f}".format(i.outstanding_amount)
			data += '<td align = center style="background-color:#CCCCCC;font-size: 10px;border:2px solid #dddddd;;">%s</td>'%(gt)
			data += '<td align = center style="background-color:#CCCCCC;font-size: 10px;border:2px solid #dddddd;;">%s</td>'%("-")
			data += '<td align = center style="background-color:#CCCCCC;font-size: 10px;border:2px solid #dddddd;;">%s</td>'%(outs)
			data += '</tr>'


		wods = []
	data += '<tr>'
	frmtd_num = "{:,.3f}".format(os)
	cur = frappe.get_value("Company",company,"default_currency")
	data += '<td align = right colspan = 6 style="border:2px solid #dddddd;;"><b>Balance Due(%s)</b></td>'%(cur)
	data += '<td align = center colspan = 2 style="border:2px solid #dddddd;;"><b>%s</b></td>'%(frmtd_num)
	data += '</tr>'
	data += '</table>'

	data += '<p><p>'


	return data

@frappe.whitelist()
def get_wod(from_date,to_date,company):
	data = ''
	data += '<div class="table-container">'
	data += '<table class="table table-bordered" width: 100%;>'
	data += '<tr>'
	data += '<td colspan = 3 style="border-color:#000000;"><img src = "/files/TSL Logo.png" align="left" width ="150"></td>'
	data += '<td colspan = 3 style="border-color:#000000;"><h2><center><b>TSL Company</b></center></h2></td>'
	data += '<td colspan = 3 style="border-color:#000000;"><center><img src = "/files/kuwait flag.jpg" width ="100"></center></td>'
	
	data += '</tr>'

	data += '<tr>'
	data += '<td colspan = 9 style="border-color:#000000;"><b>RSC & Quoted Summary</b></td>'
		
	data += '</tr>'

	data += '<tr>'
	data += '<td colspan = 1 style="border-color:#000000;"><center><b></b></center></td>'
	data += '<td colspan = 2 style="border-color:#000000;"><center><b>Total Amount(RSI)</b></center></td>'
	data += '<td colspan = 2 style="border-color:#000000;"><center><b>Total Amount(RSC)</b></center></td>'
	data += '<td colspan = 2 style="border-color:#000000;"><center><b>Total Amount(RS)</b></center></td>'
	data += '<td colspan = 2 style="border-color:#000000;"><center><b>Total Amount(Quoted)</b></center></td>'
	
	data += '</tr>'
	wo = frappe.db.sql(""" select DISTINCT sales_rep from `tabWork Order Data` where company = '%s' """ %(company) ,as_dict =1)
	data += '<tr>'
	data += '<td style="border-color:#000000;width:10%"><center><b>Salesman</b></center></td>'
	data += '<td style="border-color:#000000;width:10%;"><center><b>WO(in KD)</b></center></td>'
	data += '<td style="border-color:#000000;width:10%;"><center><b>SO(in KD)</b></center></td>'
	data += '<td style="border-color:#000000;width:10%;"><center><b>WO(in KD)</b></center></td>'
	data += '<td style="border-color:#000000;width:10%;"><center><b>SO(in KD)</b></center></td>'
	data += '<td style="border-color:#000000;width:10%;"><center><b>WO(in KD)</b></center></td>'
	data += '<td style="border-color:#000000;width:10%;"><center><b>SO(in KD)</b></center></td>'
	data += '<td style="border-color:#000000;width:10%;"><center><b>WO(in KD)</b></center></td>'  
	data += '<td style="border-color:#000000;width:10%;"><center><b>SO(in KD)</b></center></td>'	
	data += '</tr>'
	current_date_time = datetime.now()

	# Extract and print the current date
	current_date = to_date
	# Week_start = add_days(current_date,-6)
	Week_start = from_date

	total_rsi = 0
	total_rsc = 0
	total_rs = 0
	total_quotation = 0

	s_total_rsi = 0
	s_total_rsc = 0
	s_total_rs = 0
	s_total_quotation = 0
	
	for i in wo:
		if not i["sales_rep"] == None:
			if not i["sales_rep"] == 'Sales Team' and not i["sales_rep"] == 'Walkin' and not i["sales_rep"] == 'OMAR' and not i["sales_rep"] == '' and not i["sales_rep"] == 'Mazz':
				
				# wds = frappe.db.sql(""" select DISTINCT `tabWork Order Data` .name  from `tabWork Order Data` 
				# left join `tabStatus Duration Details` on `tabWork Order Data`.name = `tabStatus Duration Details`.parent
				# where  `tabStatus Duration Details`.status = "RSI-Repaired and Shipped Invoiced"
				# and `tabWork Order Data`.sales_rep = "%s" and `tabWork Order Data`.posting_date between "2015-01-01" and "2024-10-06" """ %(i["sales_rep"]) ,as_dict=1)

				wds = frappe.get_all("Work Order Data",{"sales_rep":i["sales_rep"],"status":"RSI-Repaired and Shipped Invoiced","posting_date": ["between", (Week_start,current_date)]},["*"])
				# RSI
				
				squot = 0
				for j in wds:
					rsi_q = frappe.db.sql(''' select DISTINCT `tabQuotation Item`.wod_no,`tabQuotation`.after_discount_cost, `tabQuotation`.Workflow_state,`tabQuotation Item`.margin_amount from `tabQuotation` 
					left join `tabQuotation Item` on  `tabQuotation`.name = `tabQuotation Item`.parent
					where `tabQuotation`.Workflow_state in ("Approved By Customer")  and `tabQuotation Item`.wod_no = %s  and `tabQuotation`.is_multiple_quotation = 0 ''',j.name,as_dict=1)
					if rsi_q:
						for k in rsi_q:
							squot = squot + k["after_discount_cost"]
					else:
						rsi_q_2 = frappe.db.sql(''' select DISTINCT `tabQuotation Item`.wod_no,`tabQuotation`.after_discount_cost, `tabQuotation`.Workflow_state,`tabQuotation Item`.margin_amount from `tabQuotation` 
						left join `tabQuotation Item` on  `tabQuotation`.name = `tabQuotation Item`.parent
						where `tabQuotation`.Workflow_state in ("Quoted to Customer")  and `tabQuotation Item`.wod_no = %s  and `tabQuotation`.is_multiple_quotation = 0 ''',j.name,as_dict=1)
						if rsi_q_2:
							for k in rsi_q_2:
								squot = squot + k["after_discount_cost"]
						else:
							rsi_q_3 = frappe.db.sql(''' select DISTINCT `tabQuotation Item`.wod_no,`tabQuotation`.after_discount_cost, `tabQuotation`.Workflow_state,`tabQuotation Item`.margin_amount from `tabQuotation` 
							left join `tabQuotation Item` on  `tabQuotation`.name = `tabQuotation Item`.parent
							where `tabQuotation`.Workflow_state in ("Rejected by Customer")  and `tabQuotation Item`.wod_no = %s  and `tabQuotation`.is_multiple_quotation = 0 ''',j.name,as_dict=1)
							if rsi_q_3:
								for k in rsi_q_3:
									squot = squot + k["after_discount_cost"]
							else:
								rsi_q_4 = frappe.db.sql(''' select DISTINCT `tabQuotation Item`.wod_no,`tabQuotation`.after_discount_cost, `tabQuotation`.Workflow_state,`tabQuotation Item`.margin_amount from `tabQuotation` 
								left join `tabQuotation Item` on  `tabQuotation`.name = `tabQuotation Item`.parent
								where `tabQuotation`.Workflow_state in ("Rejected")  and `tabQuotation Item`.wod_no = %s  and `tabQuotation`.is_multiple_quotation = 0 ''',j.name,as_dict=1)
								if rsi_q_4:
									for k in rsi_q_4:
										squot = squot + k["after_discount_cost"]
								else:
									rsi_q_5 = frappe.db.sql(''' select DISTINCT `tabQuotation Item`.wod_no,`tabQuotation`.after_discount_cost, `tabQuotation`.Workflow_state,`tabQuotation Item`.margin_amount from `tabQuotation` 
									left join `tabQuotation Item` on  `tabQuotation`.name = `tabQuotation Item`.parent
									where `tabQuotation`.Workflow_state in ("Approved By Management")  and `tabQuotation Item`.wod_no = %s  and `tabQuotation`.is_multiple_quotation = 0 and `tabQuotation`.quotation_type = "Revised Quotation - Repair" ''',j.name,as_dict=1)
									if rsi_q_5:
										for k in rsi_q_5:
											squot = squot + k["after_discount_cost"]



				mquot = 0
				for j in wds:
					rsi_mq = frappe.db.sql(''' select DISTINCT `tabQuotation Item`.wod_no,`tabQuotation`.after_discount_cost, `tabQuotation`.Workflow_state,`tabQuotation Item`.margin_amount from `tabQuotation` 
					left join `tabQuotation Item` on  `tabQuotation`.name = `tabQuotation Item`.parent
					where  `tabQuotation`.Workflow_state in ("Approved By Customer") and `tabQuotation Item`.wod_no = %s  and `tabQuotation`.is_multiple_quotation = 1 ''',j.name,as_dict=1)
					if rsi_mq:
						for k in rsi_mq:
							mquot = mquot + k["margin_amount"]
					else:
						rsi_mq_2 = frappe.db.sql(''' select DISTINCT `tabQuotation Item`.wod_no,`tabQuotation`.after_discount_cost, `tabQuotation`.Workflow_state,`tabQuotation Item`.margin_amount from `tabQuotation` 
						left join `tabQuotation Item` on  `tabQuotation`.name = `tabQuotation Item`.parent
						where  `tabQuotation`.Workflow_state in ("Quoted to Customer") and `tabQuotation Item`.wod_no = %s  and `tabQuotation`.is_multiple_quotation = 1 ''',j.name,as_dict=1)
						if rsi_mq_2:
							for k in rsi_mq_2:
								mquot = mquot + k["margin_amount"]
						else:
							rsi_mq_3 = frappe.db.sql(''' select DISTINCT `tabQuotation Item`.wod_no,`tabQuotation`.after_discount_cost, `tabQuotation`.Workflow_state,`tabQuotation Item`.margin_amount from `tabQuotation` 
							left join `tabQuotation Item` on  `tabQuotation`.name = `tabQuotation Item`.parent
							where  `tabQuotation`.Workflow_state in ("Rejected by Customer") and `tabQuotation Item`.wod_no = %s  and `tabQuotation`.is_multiple_quotation = 1 ''',j.name,as_dict=1)
							if rsi_mq_3:
								for k in rsi_mq_3:
									mquot = mquot + k["margin_amount"]
							else:
								rsi_mq_4 = frappe.db.sql(''' select DISTINCT `tabQuotation Item`.wod_no,`tabQuotation`.after_discount_cost, `tabQuotation`.Workflow_state,`tabQuotation Item`.margin_amount from `tabQuotation` 
								left join `tabQuotation Item` on  `tabQuotation`.name = `tabQuotation Item`.parent
								where  `tabQuotation`.Workflow_state in ("Rejected by Customer") and `tabQuotation Item`.wod_no = %s  and `tabQuotation`.is_multiple_quotation = 1 ''',j.name,as_dict=1)
								if rsi_mq_4:
									for k in rsi_mq_4:
										mquot = mquot + k["margin_amount"]
								else:
									rsi_mq_5 = frappe.db.sql(''' select DISTINCT `tabQuotation Item`.wod_no,`tabQuotation`.after_discount_cost, `tabQuotation`.Workflow_state,`tabQuotation Item`.margin_amount from `tabQuotation` 
									left join `tabQuotation Item` on  `tabQuotation`.name = `tabQuotation Item`.parent
									where  `tabQuotation`.Workflow_state in ("Approved By Management") and `tabQuotation Item`.wod_no = %s  and `tabQuotation`.is_multiple_quotation = 1 and `tabQuotation`.quotation_type = "Revised Quotation - Repair" ''',j.name,as_dict=1)
									if rsi_mq_5:
										for k in rsi_mq_5:
											mquot = mquot + k["margin_amount"]

				# old_wd_rsi = frappe.db.sql(""" select sum(`tabWork Order Data`.old_wo_q_amount) as rsi from `tabWork Order Data` where `tabWork Order Data`.sales_rep = "%s" and `tabWork Order Data`.posting_date between "%s" and "%s" and `tabWork Order Data`.status = "RSI-Repaired and Shipped Invoiced" """ %(i["sales_rep"],Week_start,current_date) ,as_dict=1)
				# old_wd_rsi = frappe.get_all("Work Order Data",{"sales_rep":i["sales_rep"],"status":"RSI-Repaired and Shipped Invoiced","posting_date": ["between", (Week_start,current_date)],"old_wo_q_amount": [">", 0]},["old_wo_q_amount"])
				 
				
				rsi_total = squot + mquot 


				#Invoiced
				sup_in = frappe.get_all("Supply Order Data",{"sales_rep":i["sales_rep"],"status":"Invoiced","posting_date": ["between", (Week_start,current_date)]})
			   
				sqm_i = 0
				for j in sup_in:
					sup_i_q = frappe.db.sql(''' select `tabQuotation`.after_discount_cost, `tabQuotation`.Workflow_state,`tabQuotation Item`.margin_amount from `tabQuotation` 
					left join `tabQuotation Item` on  `tabQuotation`.name = `tabQuotation Item`.parent
					where `tabQuotation`.Workflow_state in ("Approved By Customer","Quoted to Customer","Rejected by Customer")  and `tabQuotation Item`.supply_order_data = %s  and `tabQuotation`.is_multiple_quotation = 0 ''',j.name,as_dict=1)
					
					for k in sup_i_q:
						sqm_i = sqm_i + k["after_discount_cost"]
			

				sqm_i_2 = 0
				for j in sup_in:
					sup_i_q_2 = frappe.db.sql(''' select `tabQuotation`.after_discount_cost, `tabQuotation`.Workflow_state,`tabQuotation Item`.margin_amount from `tabQuotation` 
					left join `tabQuotation Item` on  `tabQuotation`.name = `tabQuotation Item`.parent
					where `tabQuotation`.Workflow_state in ("Approved By Customer","Quoted to Customer","Rejected by Customer")  and `tabQuotation Item`.supply_order_data = %s  and `tabQuotation`.is_multiple_quotation = 1 ''',j.name,as_dict=1)
					
					for k in sup_i_q_2:
						sqm_i_2 = sqm_i_2 + k["margin_amount"]
			
				sqm_total = sqm_i + sqm_i_2

				 #Delivered
				sup_del = frappe.get_all("Supply Order Data",{"sales_rep":i["sales_rep"],"status":"Delivered","posting_date": ["between", (Week_start,current_date)]})
			   
				sqm_d = 0
				for j in sup_del:
					sup_d_q = frappe.db.sql(''' select `tabQuotation`.after_discount_cost, `tabQuotation`.Workflow_state,`tabQuotation Item`.margin_amount from `tabQuotation` 
					left join `tabQuotation Item` on  `tabQuotation`.name = `tabQuotation Item`.parent
					where `tabQuotation`.Workflow_state in ("Approved By Customer","Quoted to Customer","Rejected by Customer")  and `tabQuotation Item`.supply_order_data = %s  and `tabQuotation`.is_multiple_quotation = 0 ''',j.name,as_dict=1)
					
					for k in sup_d_q:
						sqm_d = sqm_d + k["after_discount_cost"]
			

				sqm_d_2 = 0
				for j in sup_del:
					sup_d_q_2 = frappe.db.sql(''' select `tabQuotation`.after_discount_cost, `tabQuotation`.Workflow_state,`tabQuotation Item`.margin_amount from `tabQuotation` 
					left join `tabQuotation Item` on  `tabQuotation`.name = `tabQuotation Item`.parent
					where `tabQuotation`.Workflow_state in ("Approved By Customer","Quoted to Customer","Rejected by Customer")  and `tabQuotation Item`.supply_order_data = %s  and `tabQuotation`.is_multiple_quotation = 1 ''',j.name,as_dict=1)
					
					for k in sup_d_q_2:
						sqm_d_2 = sqm_d_2 + k["margin_amount"]
			
				sqm_del_total = sqm_d + sqm_d_2

				#Quoted
				sup_quo = frappe.get_all("Supply Order Data",{"sales_rep":i["sales_rep"],"status":"Quoted","posting_date": ["between", (Week_start,current_date)]})
			   
				sqm_quo = 0
				for j in sup_quo:
					sup_quo_q = frappe.db.sql(''' select `tabQuotation`.after_discount_cost, `tabQuotation`.Workflow_state,`tabQuotation Item`.margin_amount from `tabQuotation` 
					left join `tabQuotation Item` on  `tabQuotation`.name = `tabQuotation Item`.parent
					where `tabQuotation`.Workflow_state in ("Approved By Customer","Quoted to Customer","Rejected by Customer")  and `tabQuotation Item`.supply_order_data = %s  and `tabQuotation`.is_multiple_quotation = 0 ''',j.name,as_dict=1)
					
					for k in sup_quo_q:
						sqm_quo = sqm_quo + k["after_discount_cost"]
			

				sqm_quo_2 = 0
				for j in sup_del:
					sup_quo_q_2 = frappe.db.sql(''' select `tabQuotation`.after_discount_cost, `tabQuotation`.Workflow_state,`tabQuotation Item`.margin_amount from `tabQuotation` 
					left join `tabQuotation Item` on  `tabQuotation`.name = `tabQuotation Item`.parent
					where `tabQuotation`.Workflow_state in ("Approved By Customer","Quoted to Customer","Rejected by Customer")  and `tabQuotation Item`.supply_order_data = %s  and `tabQuotation`.is_multiple_quotation = 1 ''',j.name,as_dict=1)
					
					for k in sup_quo_q_2:
						sqm_quo_2 = sqm_quo_2 + k["margin_amount"]
			
				sqm_quo_total =sqm_quo + sqm_quo_2


				#Received
				sup_rec = frappe.get_all("Supply Order Data",{"sales_rep":i["sales_rep"],"status":"Received","posting_date": ["between", (Week_start,current_date)]})
			   
				sqm_rec = 0
				for j in sup_rec:
					sup_rec_q = frappe.db.sql(''' select `tabQuotation`.after_discount_cost, `tabQuotation`.Workflow_state,`tabQuotation Item`.margin_amount from `tabQuotation` 
					left join `tabQuotation Item` on  `tabQuotation`.name = `tabQuotation Item`.parent
					where `tabQuotation`.Workflow_state in ("Approved By Customer","Quoted to Customer","Rejected by Customer")  and `tabQuotation Item`.supply_order_data = %s  and `tabQuotation`.is_multiple_quotation = 0 ''',j.name,as_dict=1)
					
					for k in sup_rec_q:
						sqm_rec = sqm_rec + k["after_discount_cost"]
			

				sqm_rec_2 = 0
				for j in sup_rec:
					sup_rec_q_2 = frappe.db.sql(''' select `tabQuotation`.after_discount_cost, `tabQuotation`.Workflow_state,`tabQuotation Item`.margin_amount from `tabQuotation` 
					left join `tabQuotation Item` on  `tabQuotation`.name = `tabQuotation Item`.parent
					where `tabQuotation`.Workflow_state in ("Approved By Customer","Quoted to Customer","Rejected by Customer")  and `tabQuotation Item`.supply_order_data = %s  and `tabQuotation`.is_multiple_quotation = 1 ''',j.name,as_dict=1)
					
					for k in sup_rec_q_2:
						sqm_rec_2 = sqm_rec_2 + k["margin_amount"]
			
				sqm_rec_total =sqm_rec + sqm_rec_2



				# RSC	
				
				wds_rsc = frappe.get_all("Work Order Data",{"sales_rep":i["sales_rep"],"status":"RSC-Repaired and Shipped Client","posting_date": ["between", (Week_start,current_date)]},["*"])
				
				rsc_squot = 0
				for j in wds_rsc:
					rsc_q = frappe.db.sql(''' select DISTINCT `tabQuotation Item`.wod_no,`tabQuotation`.after_discount_cost, `tabQuotation`.Workflow_state,`tabQuotation Item`.margin_amount from `tabQuotation` 
					left join `tabQuotation Item` on  `tabQuotation`.name = `tabQuotation Item`.parent
					where `tabQuotation`.Workflow_state in ("Approved By Customer")  and `tabQuotation Item`.wod_no = %s  and `tabQuotation`.is_multiple_quotation = 0 ''',j.name,as_dict=1)
					if rsc_q:
						for k in rsc_q:
							rsc_squot = rsc_squot + k["after_discount_cost"]
					else:
						rsc_q_2 = frappe.db.sql(''' select DISTINCT `tabQuotation Item`.wod_no,`tabQuotation`.after_discount_cost, `tabQuotation`.Workflow_state,`tabQuotation Item`.margin_amount from `tabQuotation` 
						left join `tabQuotation Item` on  `tabQuotation`.name = `tabQuotation Item`.parent
						where `tabQuotation`.Workflow_state in ("Quoted to Customer")  and `tabQuotation Item`.wod_no = %s  and `tabQuotation`.is_multiple_quotation = 0 ''',j.name,as_dict=1)
						if rsc_q_2:
							for k in rsc_q_2:
								rsc_squot = rsc_squot + k["after_discount_cost"]
						# else:
						#     rsc_q_3 = frappe.db.sql(''' select DISTINCT `tabQuotation Item`.wod_no,`tabQuotation`.after_discount_cost, `tabQuotation`.Workflow_state,`tabQuotation Item`.margin_amount from `tabQuotation` 
						#     left join `tabQuotation Item` on  `tabQuotation`.name = `tabQuotation Item`.parent
						#     where `tabQuotation`.Workflow_state in ("Rejected by Customer")  and `tabQuotation Item`.wod_no = %s  and `tabQuotation`.is_multiple_quotation = 0 ''',j.name,as_dict=1)
						#     if rsc_q_3:
						#         for k in rsc_q_3:
						#             rsc_squot = rsc_squot + k["after_discount_cost"]
							# else:
							#     rsc_q_4 = frappe.db.sql(''' select DISTINCT `tabQuotation Item`.wod_no,`tabQuotation`.after_discount_cost, `tabQuotation`.Workflow_state,`tabQuotation Item`.margin_amount from `tabQuotation` 
							#     left join `tabQuotation Item` on  `tabQuotation`.name = `tabQuotation Item`.parent
							#     where `tabQuotation`.Workflow_state in ("Rejected")  and `tabQuotation Item`.wod_no = %s  and `tabQuotation`.is_multiple_quotation = 0 ''',j.name,as_dict=1)
							#     if rsc_q_4:
							#         for k in rsc_q_4:
							#             rsc_squot = rsc_squot + k["after_discount_cost"]
						
							#     else:
							#         rsc_q_5 = frappe.db.sql(''' select DISTINCT `tabQuotation Item`.wod_no,`tabQuotation`.after_discount_cost, `tabQuotation`.Workflow_state,`tabQuotation Item`.margin_amount from `tabQuotation` 
							#         left join `tabQuotation Item` on  `tabQuotation`.name = `tabQuotation Item`.parent
							#         where `tabQuotation`.Workflow_state in ("Approved By Management")  and `tabQuotation Item`.wod_no = %s  and `tabQuotation`.is_multiple_quotation = 0 ''',j.name,as_dict=1)
							#         if rsc_q_5:
							#             for k in rsc_q_5:
							#                 rsc_squot = rsc_squot + k["after_discount_cost"]
						


				rsc_mquot = 0
				for j in wds_rsc:
					rsc_mq = frappe.db.sql(''' select DISTINCT `tabQuotation Item`.wod_no,`tabQuotation`.after_discount_cost, `tabQuotation`.Workflow_state,`tabQuotation Item`.margin_amount from `tabQuotation` 
					left join `tabQuotation Item` on  `tabQuotation`.name = `tabQuotation Item`.parent
					where `tabQuotation`.Workflow_state in ("Approved By Customer")  and `tabQuotation Item`.wod_no = %s  and `tabQuotation`.is_multiple_quotation = 1 ''',j.name,as_dict=1)
					if rsc_mq:
						for k in rsc_mq:
							rsc_mquot = rsc_mquot + k["margin_amount"]
					else:
						rsc_mq_2 = frappe.db.sql(''' select DISTINCT `tabQuotation Item`.wod_no,`tabQuotation`.after_discount_cost, `tabQuotation`.Workflow_state,`tabQuotation Item`.margin_amount from `tabQuotation` 
						left join `tabQuotation Item` on  `tabQuotation`.name = `tabQuotation Item`.parent
						where `tabQuotation`.Workflow_state in ("Quoted to Customer")  and `tabQuotation Item`.wod_no = %s  and `tabQuotation`.is_multiple_quotation = 1 ''',j.name,as_dict=1)
						if rsc_mq_2:
							for k in rsc_mq_2:
								rsc_mquot = rsc_mquot + k["margin_amount"]
						# else:
						#     rsc_mq_3 = frappe.db.sql(''' select DISTINCT `tabQuotation Item`.wod_no,`tabQuotation`.after_discount_cost, `tabQuotation`.Workflow_state,`tabQuotation Item`.margin_amount from `tabQuotation` 
						#     left join `tabQuotation Item` on  `tabQuotation`.name = `tabQuotation Item`.parent
						#     where `tabQuotation`.Workflow_state in ("Rejected by Customer")  and `tabQuotation Item`.wod_no = %s  and `tabQuotation`.is_multiple_quotation = 1 ''',j.name,as_dict=1)
						#     if rsc_mq_3:
						#         for k in rsc_mq_3:
						#             rsc_mquot = rsc_mquot + k["margin_amount"]
							# else:
							#     rsc_mq_4 = frappe.db.sql(''' select DISTINCT `tabQuotation Item`.wod_no,`tabQuotation`.after_discount_cost, `tabQuotation`.Workflow_state,`tabQuotation Item`.margin_amount from `tabQuotation` 
							#     left join `tabQuotation Item` on  `tabQuotation`.name = `tabQuotation Item`.parent
							#     where `tabQuotation`.Workflow_state in ("Rejected")  and `tabQuotation Item`.wod_no = %s  and `tabQuotation`.is_multiple_quotation = 1 ''',j.name,as_dict=1)
							#     if rsc_mq_4:
							#         for k in rsc_mq_4:
							#             rsc_mquot = rsc_mquot + k["margin_amount"]
							#     else:
							#         rsc_mq_5 = frappe.db.sql(''' select DISTINCT `tabQuotation Item`.wod_no,`tabQuotation`.after_discount_cost, `tabQuotation`.Workflow_state,`tabQuotation Item`.margin_amount from `tabQuotation` 
							#         left join `tabQuotation Item` on  `tabQuotation`.name = `tabQuotation Item`.parent
							#         where `tabQuotation`.Workflow_state in ("Approved By Management")  and `tabQuotation Item`.wod_no = %s  and `tabQuotation`.is_multiple_quotation = 1 ''',j.name,as_dict=1)
							#         if rsc_mq_5:
							#             for k in rsc_mq_5:
							#                 rsc_mquot = rsc_mquot + k["margin_amount"]



				#RS				
				wds_rs = frappe.get_all("Work Order Data",{"sales_rep":i["sales_rep"],"status":"RS-Repaired and Shipped","posting_date": ["between", (Week_start,current_date)]},["*"])
				
				rs_squot = 0
				for j in wds_rs:
					rs_q = frappe.db.sql(''' select DISTINCT `tabQuotation`.after_discount_cost, `tabQuotation`.Workflow_state,`tabQuotation Item`.margin_amount from `tabQuotation` 
					left join `tabQuotation Item` on  `tabQuotation`.name = `tabQuotation Item`.parent
					where `tabQuotation`.Workflow_state in ("Approved By Customer")  and `tabQuotation Item`.wod_no = %s  and `tabQuotation`.is_multiple_quotation = 0 ''',j.name,as_dict=1)
					if rs_q:
						for k in rs_q:
							rs_squot = rs_squot + k["after_discount_cost"]
					else:
						rs_q_2 = frappe.db.sql(''' select DISTINCT `tabQuotation`.after_discount_cost, `tabQuotation`.Workflow_state,`tabQuotation Item`.margin_amount from `tabQuotation` 
						left join `tabQuotation Item` on  `tabQuotation`.name = `tabQuotation Item`.parent
						where `tabQuotation`.Workflow_state in ("Quoted to Customer")  and `tabQuotation Item`.wod_no = %s  and `tabQuotation`.is_multiple_quotation = 0 ''',j.name,as_dict=1)
						if rs_q_2:
							for k in rs_q_2:
								rs_squot = rs_squot + k["after_discount_cost"]
						# else:
						#     rs_q_3 = frappe.db.sql(''' select DISTINCT `tabQuotation`.after_discount_cost, `tabQuotation`.Workflow_state,`tabQuotation Item`.margin_amount from `tabQuotation` 
						#     left join `tabQuotation Item` on  `tabQuotation`.name = `tabQuotation Item`.parent
						#     where `tabQuotation`.Workflow_state in ("Rejected")  and `tabQuotation Item`.wod_no = %s  and `tabQuotation`.is_multiple_quotation = 0 ''',j.name,as_dict=1)
						#     if rs_q_3:
						#         for k in rs_q_3:
						#             rs_squot = rs_squot + k["after_discount_cost"]


				rs_mquot = 0
				for j in wds_rs:
					rs_mq = frappe.db.sql(''' select DISTINCT `tabQuotation`.after_discount_cost, `tabQuotation`.Workflow_state,`tabQuotation Item`.margin_amount from `tabQuotation` 
					left join `tabQuotation Item` on  `tabQuotation`.name = `tabQuotation Item`.parent
					where `tabQuotation`.Workflow_state in ("Approved By Customer")  and `tabQuotation Item`.wod_no = %s  and `tabQuotation`.is_multiple_quotation = 1 ''',j.name,as_dict=1)
					if rs_mq:
						for k in rs_mq:
							rs_mquot = rs_mquot + k["margin_amount"]
					else:
						rs_mq_2 = frappe.db.sql(''' select DISTINCT `tabQuotation`.after_discount_cost, `tabQuotation`.Workflow_state,`tabQuotation Item`.margin_amount from `tabQuotation` 
						left join `tabQuotation Item` on  `tabQuotation`.name = `tabQuotation Item`.parent
						where `tabQuotation`.Workflow_state in ("Quoted to Customer")  and `tabQuotation Item`.wod_no = %s  and `tabQuotation`.is_multiple_quotation = 1 ''',j.name,as_dict=1)
						if rs_mq_2:
							for k in rs_mq_2:
								rs_mquot = rs_mquot + k["margin_amount"]
						
						# else:
						#     rs_mq_3 = frappe.db.sql(''' select DISTINCT `tabQuotation`.after_discount_cost, `tabQuotation`.Workflow_state,`tabQuotation Item`.margin_amount from `tabQuotation` 
						#     left join `tabQuotation Item` on  `tabQuotation`.name = `tabQuotation Item`.parent
						#     where `tabQuotation`.Workflow_state in ("Rejected")  and `tabQuotation Item`.wod_no = %s  and `tabQuotation`.is_multiple_quotation = 1 ''',j.name,as_dict=1)
						#     if rs_mq_3:
						#         for k in rs_mq_3:
						#             rs_mquot = rs_mquot + k["margin_amount"]



				
				#rsi
				rsi_old = 0
				for i in wds:
					rsi_old = rsi_old + i.old_wo_q_amount
			  
				rsi_total = squot + mquot + rsi_old

				#rsc
				rsc_old = 0
				for i in wds_rsc:
					rsc_old = rsc_old + i.old_wo_q_amount
			  
				rsc_total = rsc_squot + rsc_mquot + rsc_old

				#rs
				rs_old = 0
				for i in wds_rs:
					rs_old = rs_old + i.old_wo_q_amount
			  
				rs_total = rs_squot + rs_mquot + rs_old

				total_rsi = total_rsi + rsi_total
				total_rsc = total_rsc + rsc_total
				total_rs = total_rs + rs_total 

				s_total_rsi = s_total_rsi + sqm_total
				s_total_rsc = s_total_rsc + sqm_del_total
				s_total_rs = s_total_rs + sqm_rec_total
				s_total_quotation =  s_total_quotation + sqm_quo_total

				total_q = frappe.db.sql(''' select sum(after_discount_cost) as t from `tabQuotation` 
				where Workflow_state in ("Approved By Customer","Quoted to Customer","Rejected by Customer")  and sales_rep = %s ''',i["sales_rep"],as_dict=1)
				print(total_q[0]["t"])
				if total_q[0]["t"] == None:
					total_q[0]["t"] = 0

				total_quotation = total_quotation + total_q[0]["t"] 

				
				sales = frappe.get_value("Sales Person",i["sales_rep"],["user"])
				if sales:
					gt = 0
					gr = frappe.get_all("Sales Invoice",{"department":"Repair - TSL","status": ["in", ["Unpaid","Overdue"]]},["*"])
					frappe.errprint(gr)
					frappe.errprint("jhiku")
					if gr.grand_total:
						frappe.errprint("jhiku")

						for i in gr.grand_total:
							frappe.errprint("jhiku")

						# gt = gt + i.grand_total

				data += '<tr>'
				data += '<td style="border-color:#000000;"><center><b>%s</b></center></td>' %(i["sales_rep"])
				data += '<td style="border-color:#000000;"><center><b>%s</b></center></td>' %("{:,.2f}".format(rsi_total))
				data += '<td style="border-color:#000000;"><center><b>%s</b></center></td>' %("{:,.2f}".format(sqm_total))
				data += '<td style="border-color:#000000;"><center><b>%s</b></center></td>' %("{:,.2f}".format(rsc_total))
				data += '<td style="border-color:#000000;"><center><b>%s</b></center></td>' %("{:,.2f}".format(sqm_del_total))
				data += '<td style="border-color:#000000;"><center><b>%s</b></center></td>' %("{:,.2f}".format(rs_total))
				data += '<td style="border-color:#000000;"><center><b>%s</b></center></td>' %("{:,.2f}".format(sqm_rec_total)) 
				data += '<td style="border-color:#000000;"><center><b>%s</b></center></td>' %("{:,.2f}".format(total_q[0]["t"]))
				data += '<td style="border-color:#000000;"><center><b>%s</b></center></td>'	 %("{:,.2f}".format(sqm_quo_total))
				data += '</tr>'
	data += '<tr>'
	data += '<td style="border-color:#000000;"><center><b>Total</b></center></td>'
	data += '<td style="border-color:#000000;"><center><b>%s</b></center></td>' %("{:,.2f}".format(total_rsi)) 
	data += '<td style="border-color:#000000;"><center><b>%s</b></center></td>'  %("{:,.2f}".format(s_total_rsi))
	data += '<td style="border-color:#000000;"><center><b>%s</b></center></td>'  %("{:,.2f}".format(total_rsc))
	data += '<td style="border-color:#000000;"><center><b>%s</b></center></td>'  %("{:,.2f}".format(s_total_rsc))
	data += '<td style="border-color:#000000;"><center><b>%s</b></center></td>'  %("{:,.2f}".format(total_rs))
	data += '<td style="border-color:#000000;"><center><b>%s</b></center></td>' %("{:,.2f}".format(s_total_rs))
	data += '<td style="border-color:#000000;"><center><b>%s</b></center></td>' %("{:,.2f}".format(total_quotation)) 
	data += '<td style="border-color:#000000;"><center><b>%s</b></center></td>' %("{:,.2f}".format(s_total_quotation))
	data += '</tr>'

	data += '</table>'
	data += '</div>'
	
	return data


@frappe.whitelist()
def get_wrk_ord(from_date,to_date,customer,work_order_data,sales_person,report):
	if report == "RSI-Repaired and Shipped Invoiced":
		data= ""
		data += '<table class="table table-bordered">'

		data += '<tr>'
		data += '<td colspan = 3 style="border-color:#000000;"><img src = "/files/TSL Logo.png" align="left" width ="150"></td>'
		data += '<td colspan = 4 style="border-color:#000000;"><h1><center><b>Sales Summary Report</b></center></h1></td>'
		
		data += '</tr>'


		data += '<tr>'
		data += '<td style="border-color:#000000;width:5%;padding:1px;font-size:12px;background-color:#4169E1;color:white;text-align:center;font-weight:bold;">S.No</td>'
		data += '<td style="border-color:#000000;width:10%;padding:1px;font-size:12px;background-color:#4169E1;color:white;text-align:center;font-weight:bold;">Date</td>'
		data += '<td style="border-color:#000000;width:10%;padding:1px;font-size:12px;background-color:#4169E1;color:white;text-align:center;font-weight:bold;">Work Order</td>'
		data += '<td style="border-color:#000000;width:10%;padding:1px;font-size:12px;background-color:#4169E1;color:white;text-align:center;font-weight:bold;">Sales Person</td>'
		data += '<td style="border-color:#000000;width:15%;padding:1px;font-size:12px;background-color:#4169E1;color:white;text-align:center;font-weight:bold;">Sales Invoice</td>'
		data += '<td style="border-color:#000000;width:35%;padding:1px;font-size:12px;background-color:#4169E1;color:white;text-align:center;font-weight:bold;">Customer</td>'
		data += '<td style="border-color:#000000;width:15%;padding:1px;font-size:12px;background-color:#4169E1;color:white;text-align:center;font-weight:bold;">Amount</td>'
		data += '</tr>'
		
		if customer:
			si = frappe.db.sql(""" select DISTINCT `tabSales Invoice`.due_date, `tabSales Invoice Item`.work_order_data as wo,`tabSales Invoice Item`.wod_no as w from `tabSales Invoice` 
			left join `tabSales Invoice Item` on `tabSales Invoice`.name = `tabSales Invoice Item`.parent 
			where `tabSales Invoice`.status = "Overdue" and `tabSales Invoice`.due_date between '%s' and '%s' and `tabSales Invoice`.customer = '%s' """ %(from_date,to_date,customer),as_dict =1)
		
		elif sales_person:
			si = frappe.db.sql(""" select DISTINCT  `tabSales Invoice`.due_date,`tabSales Team`.sales_person,`tabSales Invoice Item`.work_order_data as wo,`tabSales Invoice Item`.wod_no as w from `tabSales Invoice` 
			left join `tabSales Invoice Item` on `tabSales Invoice`.name = `tabSales Invoice Item`.parent 
			right join `tabSales Team` on `tabSales Invoice`.name = `tabSales Team`.parent
			where `tabSales Invoice`.status = "Overdue" and `tabSales Team`.sales_person = '%s' and `tabSales Invoice`.due_date between '%s' and '%s' ORDER BY `tabSales Invoice`.due_date ASC """ %(sales_person,from_date,to_date),as_dict =1)

		elif work_order_data:
			si = frappe.db.sql(""" select DISTINCT `tabSales Invoice`.due_date,`tabSales Invoice Item`.work_order_data as wo,`tabSales Invoice Item`.wod_no as w from `tabSales Invoice` 
			left join `tabSales Invoice Item` on `tabSales Invoice`.name = `tabSales Invoice Item`.parent 
			where `tabSales Invoice`.status = "Overdue" and `tabSales Invoice`.due_date between '%s' and '%s' and `tabSales Invoice Item`.work_order_data = '%s' """ %(from_date,to_date,work_order_data),as_dict =1)
			if not si:
				si = frappe.db.sql(""" select DISTINCT `tabSales Invoice`.due_date,`tabSales Invoice Item`.work_order_data as wo,`tabSales Invoice Item`.wod_no as w from `tabSales Invoice` 
				left join `tabSales Invoice Item` on `tabSales Invoice`.name = `tabSales Invoice Item`.parent 
				where `tabSales Invoice`.due_date between '%s' and '%s' and `tabSales Invoice Item`.wod_no = '%s' """ %(from_date,to_date,work_order_data),as_dict =1)


		else:
			si = frappe.db.sql(""" select DISTINCT `tabSales Invoice`.due_date, `tabSales Invoice Item`.work_order_data as wo,`tabSales Invoice Item`.wod_no as w from `tabSales Invoice` 
			left join `tabSales Invoice Item` on `tabSales Invoice`.name = `tabSales Invoice Item`.parent 
			where `tabSales Invoice`.status = "Overdue" and `tabSales Invoice`.due_date between '%s' and '%s' """ %(from_date,to_date),as_dict =1)


			
		wod_list = []
		for j in si:
			if j["wo"]:
				wod_list.append(j["wo"])
			else:
				wod_list.append(j["w"])
				
		sn = 0
		total_amt = 0
		for i in wod_list:
			st = frappe.get_value("Work Order Data",{"name":i},["Status"])
			
			if i:
				sid_1 = frappe.db.sql(""" select `tabSales Invoice`.due_date from `tabSales Invoice` 
				left join `tabSales Invoice Item` on `tabSales Invoice`.name = `tabSales Invoice Item`.parent 
				where `tabSales Invoice Item`.work_order_data = '%s' """ %(i),as_dict =1)

				sid_2 = frappe.db.sql(""" select `tabSales Invoice`.due_date from `tabSales Invoice` 
				left join `tabSales Invoice Item` on `tabSales Invoice`.name = `tabSales Invoice Item`.parent 
				where `tabSales Invoice Item`.wod_no = '%s' """ %(i),as_dict =1)
			

				sales_rep_1 = frappe.db.sql(""" select `tabSales Team`.sales_person from `tabSales Invoice` 
				left join `tabSales Invoice Item` on `tabSales Invoice`.name = `tabSales Invoice Item`.parent 
				right join `tabSales Team` on `tabSales Invoice`.name = `tabSales Team`.parent
				where `tabSales Invoice Item`.work_order_data = '%s' """ %(i),as_dict =1)
					
				sales_rep_2 = frappe.db.sql(""" select `tabSales Team`.sales_person from `tabSales Invoice` 
				left join `tabSales Invoice Item` on `tabSales Invoice`.name = `tabSales Invoice Item`.parent 
				right join `tabSales Team` on `tabSales Invoice`.name = `tabSales Team`.parent
				where `tabSales Invoice Item`.wod_no = '%s' """ %(i),as_dict =1)
					
				cust = frappe.get_value("Work Order Data",{"name":i},["customer"])
			
				if sid_1:
					input_date_string = str(sid_1[0]["due_date"])
				else:
					input_date_string = str(sid_2[0]["due_date"])

				if input_date_string:
					input_date = datetime.strptime(input_date_string, "%Y-%m-%d")

				
				formatted_date = input_date.strftime("%d-%m-%Y")
				
				
				wo = frappe.db.sql(""" select DISTINCT `tabSales Invoice`.name, `tabSales Invoice Item`.amount as amount from `tabSales Invoice` 
				left join `tabSales Invoice Item` on `tabSales Invoice`.name = `tabSales Invoice Item`.parent 
				where `tabSales Invoice Item`.work_order_data = '%s' """ %(i),as_dict =1)
				amt = 0
				n = []
				if wo:
					amt = wo[0]["amount"]
					n.append(wo[0]["name"])
					n = str(n).strip('[]')
					n = n.replace("'", '')
					
				
				else:
					wos = frappe.db.sql(""" select DISTINCT `tabSales Invoice`.name,`tabSales Invoice Item`.amount as amount from `tabSales Invoice` 
					left join `tabSales Invoice Item` on `tabSales Invoice`.name = `tabSales Invoice Item`.parent 
					where `tabSales Invoice Item`.wod_no = '%s' """ %(i),as_dict =1)
					amt = wos[0]["amount"]
					n.append(wos[0]["name"])
					n = str(n).strip('[]')
					n = n.replace("'", '')
				
				
				total_amt = total_amt + amt
				sn = sn + 1
				data += '<tr>'
				data += '<td style="border-color:#000000;padding:1px;font-size:14px;font-size:12px;"><center>%s<center></td>'%(sn)
				data += '<td style="border-color:#000000;padding:1px;font-size:14px;font-size:12px;"><center>%s<center></td>'%(formatted_date)
					
				
				data += '<td style="border-color:#000000;padding:1px;font-size:14px;font-size:12px;"><center><a href="https://erp.tsl-me.com/app/work-order-data/%s"target="_blank">%s</a><center></td>'%(i,i)
				if sales_rep_1:
					data += '<td style="border-color:#000000;padding:1px;font-size:14px;font-size:12px;"><center>%s<center></td>'%(sales_rep_1[0]["sales_person"] or "-" )
				elif sales_rep_2:
					data += '<td style="border-color:#000000;padding:1px;font-size:14px;font-size:12px;"><center>%s<center></td>'%(sales_rep_2[0]["sales_person"] or "-" )
				else:
					data += '<td style="border-color:#000000;padding:1px;font-size:14px;font-size:12px;"><center>%s<center></td>'%("-" )

				data += '<td style="border-color:#000000;padding:1px;font-size:14px;font-size:12px;"><center><a href="https://erp.tsl-me.com/app/sales-invoice/%s"target="_blank">%s</a><center></td>'%(n,n)
				data += '<td style="border-color:#000000;padding:1px;font-size:14px;font-size:12px;"><center>%s<center></td>'%(cust)
				data += '<td style="border-color:#000000;padding:1px;font-size:14px;font-size:12px;"><center>%s<center></td>' %(format(amt, ".2f"))
			
		data += '<tr>'
		data += '<td style="border-color:#000000;padding:1px;font-size:14px;font-size:12px;background-color:#4169E1;color:white;text-align:center;font-weight:bold;"></td>'
		data += '<td style="border-color:#000000;padding:1px;font-size:14px;font-size:12px;background-color:#4169E1;color:white;text-align:center;font-weight:bold;"></td>'
		data += '<td style="border-color:#000000;padding:1px;font-size:14px;font-size:12px;background-color:#4169E1;color:white;text-align:center;font-weight:bold;"></td>'
		data += '<td style="border-color:#000000;padding:1px;font-size:14px;font-size:12px;background-color:#4169E1;color:white;text-align:center;font-weight:bold;"></td>'
		data += '<td style="border-color:#000000;padding:1px;font-size:14px;font-size:12px;background-color:#4169E1;color:white;text-align:center;font-weight:bold;"></td>'
		data += '<td style="border-color:#000000;padding:1px;font-size:14px;font-size:12px;background-color:#4169E1;color:white;text-align:center;font-weight:bold;">Total</td>'
		data += '<td style="border-color:#000000;padding:1px;font-size:14px;font-size:12px;background-color:#4169E1;color:white;text-align:center;font-weight:bold;">%s</td>'%(format(total_amt, ".2f"))
		data += '</tr>'
		data += '</table>'
	
	if report == "RSC-Repaired and Shipped Client":
		data= ""
		data += '<table class="table table-bordered">'

		data += '<tr>'
		data += '<td colspan = 3 style="border-color:#000000;"><img src = "/files/TSL Logo.png" align="left" width ="150"></td>'
		data += '<td colspan = 5 style="border-color:#000000;"><h1><center><b>Sales Summary Report</b></center></h1></td>'
		
		data += '</tr>'

		data += '</tr>'
		data += '<tr>'
		data += '<td style="border-color:#000000;width:5%;padding:1px;font-size:12px;background-color:#4169E1;color:white;text-align:center;font-weight:bold;">S.No</td>'
		data += '<td style="border-color:#000000;width:10%;padding:1px;font-size:12px;background-color:#4169E1;color:white;text-align:center;font-weight:bold;">Date</td>'
		data += '<td style="border-color:#000000;width:15%;padding:1px;font-size:12px;background-color:#4169E1;color:white;text-align:center;font-weight:bold;">Work Order</td>'
		data += '<td style="border-color:#000000;width:10%;padding:1px;font-size:12px;background-color:#4169E1;color:white;text-align:center;font-weight:bold;">Sales Person</td>'
		data += '<td style="border-color:#000000;width:30%;padding:1px;font-size:12px;background-color:#4169E1;color:white;text-align:center;font-weight:bold;">Customer</td>'
		data += '<td style="border-color:#000000;width:10%;padding:1px;font-size:12px;background-color:#4169E1;color:white;text-align:center;font-weight:bold;">Delivery Note</td>'
		data += '<td style="border-color:#000000;width:10%;padding:1px;font-size:12px;background-color:#4169E1;color:white;text-align:center;font-weight:bold;">Delivery Date</td>'
		data += '<td style="border-color:#000000;width:10%;padding:1px;font-size:12px;background-color:#4169E1;color:white;text-align:center;font-weight:bold;">Amount</td>'
		data += '</tr>'
		
		if customer:
			wd = frappe.get_all("Work Order Data",{"posting_date": ["between", (from_date,to_date)],"status":"RSC-Repaired and Shipped Client","customer":customer},["*"])

		elif sales_person:
			wd = frappe.get_all("Work Order Data",{"posting_date": ["between", (from_date,to_date)],"status":"RSC-Repaired and Shipped Client","sales_rep":sales_person},["*"])

		elif work_order_data:
			wd = frappe.get_all("Work Order Data",{"posting_date": ["between", (from_date,to_date)],"status":"RSC-Repaired and Shipped Client","name":work_order_data},["*"])

		else:
			wd = frappe.get_all("Work Order Data",{"posting_date": ["between", (from_date,to_date)],"status":"RSC-Repaired and Shipped Client"},["*"])
		
		sn = 0
		total_rsc = 0
		for i in wd:
			amount = 0
			amt = frappe.db.sql(""" select `tabQuotation`.after_discount_cost as cost  from `tabQuotation` 
			left join `tabQuotation Item` on `tabQuotation`.name = `tabQuotation Item`.parent 
			where `tabQuotation Item`.wod_no = '%s' and `tabQuotation`.workflow_state = 'Approved By Customer' """ %(i.name),as_dict =1)
			if amt:
				amount = amt[0]["cost"]
			else:
				am = frappe.db.sql(""" select `tabQuotation`.after_discount_cost as cos  from `tabQuotation` 
				left join `tabQuotation Item` on `tabQuotation`.name = `tabQuotation Item`.parent 
				where `tabQuotation Item`.wod_no = '%s' and `tabQuotation`.workflow_state = 'Approved By Management' """ %(i.name),as_dict =1)
				if am:
					amount = am[0]["cos"]
					
			sn = sn+1

			input_date_string_1 = str(i.posting_date)
			input_date_string_2 = str(i.dn_date)

			input_date_1 = datetime.strptime(input_date_string_1, "%Y-%m-%d")
			if i.dn_date:
				input_date_2 = datetime.strptime(input_date_string_2, "%Y-%m-%d")


			formatted_date_1 = input_date_1.strftime("%d-%m-%Y")
			if input_date_2:
				formatted_date_2 = input_date_2.strftime("%d-%m-%Y")
			
				
			data += '<tr>'
			data += '<td style="border-color:#000000;padding:1px;font-size:14px;font-size:12px;"><center>%s<center></td>'%(sn)
				
			data += '<td style="border-color:#000000;padding:1px;font-size:14px;font-size:12px;"><center>%s<center></td>'%(formatted_date_1)
			
			data += '<td style="border-color:#000000;padding:1px;font-size:14px;font-size:12px;"><center><a href="https://erp.tsl-me.com/app/work-order-data/%s"target="_blank">%s</a><center></td>'%(i.name,i.name)
				
			data += '<td style="border-color:#000000;padding:1px;font-size:14px;font-size:12px;"><center>%s<center></td>'%(i.sales_rep)

			data += '<td style="border-color:#000000;padding:1px;font-size:14px;font-size:12px;"><center>%s<center></td>'%(i.customer)
				
			data += '<td style="border-color:#000000;padding:1px;font-size:14px;font-size:12px;"><center><a href="https://erp.tsl-me.com/app/delivery-note/%s"target="_blank">%s</a><center></td>'%(i.dn_no,i.dn_no)					

			data += '<td style="border-color:#000000;padding:1px;font-size:14px;font-size:12px;"><center>%s<center></td>'%(formatted_date_2)

			data += '<td style="border-color:#000000;padding:1px;font-size:14px;font-size:12px;"><center>%s<center></td>'%(format(amount, ".2f"))
			total_rsc = total_rsc + amount
			
		data += '<tr>'
		data += '<td style="border-color:#000000;padding:1px;font-size:14px;font-size:12px;background-color:#3333ff;color:white;"></td>'
		data += '<td style="border-color:#000000;padding:1px;font-size:14px;font-size:12px;background-color:#3333ff;color:white;"></td>'
		data += '<td style="border-color:#000000;padding:1px;font-size:14px;font-size:12px;background-color:#3333ff;color:white;"></td>'
		data += '<td style="border-color:#000000;padding:1px;font-size:14px;font-size:12px;background-color:#3333ff;color:white;"></td>'
		data += '<td style="border-color:#000000;padding:1px;font-size:14px;font-size:12px;background-color:#3333ff;color:white;"></td>'
		data += '<td style="border-color:#000000;padding:1px;font-size:14px;font-size:12px;background-color:#3333ff;color:white;"></td>'
		data += '<td align = center style="border-color:#000000;padding:1px;font-size:14px;font-size:12px;background-color:#3333ff;color:white;font-weight:bold;">Total</td>'
		data += '<td align = center style="border-color:#000000;padding:1px;font-size:14px;font-size:12px;background-color:#3333ff;color:white;font-weight:bold;">%s</td>'%(format(total_rsc, ".2f"))
		data += '</tr>'

	if report == "Q-Quoted":
		data= ""
		data += '<table class="table table-bordered">'
		
		data += '<tr>'
		data += '<td colspan = 3 style="border-color:#000000;"><img src = "/files/TSL Logo.png" align="left" width ="150"></td>'
		data += '<td colspan = 4 style="border-color:#000000;"><h1><center><b>Sales Summary Report</b></center></h1></td>'
		
		data += '</tr>'
		
		data += '<tr>'
		data += '<td style="border-color:#000000;width:5%;padding:1px;font-size:14px;font-size:12px;background-color:#4169E1;color:white;text-align:center;font-weight:bold;">S.No</td>'
		data += '<td style="border-color:#000000;width:10%;padding:1px;font-size:14px;font-size:12px;background-color:#4169E1;color:white;text-align:center;font-weight:bold;">Date</td>'
		data += '<td style="border-color:#000000;width:15%;padding:1px;font-size:14px;font-size:12px;background-color:#4169E1;color:white;text-align:center;font-weight:bold;">Work Order</td>'
		data += '<td style="border-color:#000000;width:10%;padding:1px;font-size:14px;font-size:12px;background-color:#4169E1;color:white;text-align:center;font-weight:bold;">Sales Person</td>'
		data += '<td style="border-color:#000000;width:15%;padding:1px;font-size:14px;font-size:12px;background-color:#4169E1;color:white;text-align:center;font-weight:bold;">Quotation</td>'
		data += '<td style="border-color:#000000;width:35%;padding:1px;font-size:14px;font-size:12px;background-color:#4169E1;color:white;text-align:center;font-weight:bold;">Customer</td>'
		data += '<td style="border-color:#000000;width:10%;padding:1px;font-size:14px;font-size:12px;background-color:#4169E1;color:white;text-align:center;font-weight:bold;">Amount</td>'
		data += '</tr>'
		
		if customer:
			wd = frappe.get_all("Work Order Data",{"posting_date": ["between", (from_date,to_date)],"status":"Q-Quoted","customer":customer},["*"])

		elif sales_person:
			wd = frappe.get_all("Work Order Data",{"posting_date": ["between", (from_date,to_date)],"status":"Q-Quoted","sales_rep":sales_person},["*"])

		elif work_order_data:
			wd = frappe.get_all("Work Order Data",{"posting_date": ["between", (from_date,to_date)],"status":"Q-Quoted","name":work_order_data},["*"])

		else:
			wd = frappe.get_all("Work Order Data",{"posting_date": ["between", (from_date,to_date)],"status":"Q-Quoted"},["*"])
		
		sn = 0
		total_amt = 0
		for i in wd:
			sn = sn+1
			ada  = si = frappe.db.sql(""" select DISTINCT `tabQuotation`.is_multiple_quotation as multi,`tabQuotation`.name as qname,`tabQuotation Item`.margin_amount as ma,`tabQuotation`.after_discount_cost as adc from `tabQuotation` 
			left join `tabQuotation Item` on `tabQuotation`.name = `tabQuotation Item`.parent 
			where `tabQuotation Item`.wod_no = '%s' """ %(i.name),as_dict =1)
			amt = 0
			if ada:
				if ada[0]["multi"] == 1:
					amt = ada[0]["ma"]
				else:
					amt = ada[0]["adc"]

			total_amt = total_amt + amt

			
			input_date_string = str(i.posting_date)

			input_date = datetime.strptime(input_date_string, "%Y-%m-%d")

			formatted_date = input_date.strftime("%d-%m-%Y")
				

			data += '<tr>'
			data += '<td style="border-color:#000000;padding:1px;font-size:14px;font-size:12px;"><center>%s<center></td>'%(sn)
				
			data += '<td style="border-color:#000000;padding:1px;font-size:14px;font-size:12px;"><center>%s<center></td>'%(formatted_date)
			
			data += '<td style="border-color:#000000;padding:1px;font-size:14px;font-size:12px;"><center><a href="https://erp.tsl-me.com/app/work-order-data/%s"target="_blank">%s</a><center></td>'%(i.name,i.name)
				
			data += '<td style="border-color:#000000;padding:1px;font-size:14px;font-size:12px;"><center>%s<center></td>'%(i.sales_rep)

			data += '<td style="border-color:#000000;padding:1px;font-size:14px;font-size:12px;"><center><a href="https://erp.tsl-me.com/app/quotation/%s"target="_blank">%s</a><center></td>'%(ada[0]["qname"],ada[0]["qname"])					

			data += '<td style="border-color:#000000;padding:1px;font-size:14px;font-size:12px;"><center>%s<center></td>'%(i.customer)
				
		
			data += '<td style="border-color:#000000;padding:1px;font-size:14px;font-size:12px;"><center>%s<center></td>'%(amt)
	
		data += '<tr>'
		data += '<td style="border-color:#000000;padding:1px;font-size:14px;font-size:12px;background-color:#4169E1;color:white;text-align:center;font-weight:bold;"></td>'
		data += '<td style="border-color:#000000;padding:1px;font-size:14px;font-size:12px;background-color:#4169E1;color:white;text-align:center;font-weight:bold;"></td>'
		data += '<td style="border-color:#000000;padding:1px;font-size:14px;font-size:12px;background-color:#4169E1;color:white;text-align:center;font-weight:bold;"></td>'
		data += '<td style="border-color:#000000;padding:1px;font-size:14px;font-size:12px;background-color:#4169E1;color:white;text-align:center;font-weight:bold;"></td>'
		data += '<td style="border-color:#000000;padding:1px;font-size:14px;font-size:12px;background-color:#4169E1;color:white;text-align:center;font-weight:bold;"></td>'
		data += '<td style="border-color:#000000;padding:1px;font-size:14px;font-size:12px;background-color:#4169E1;color:white;text-align:center;font-weight:bold;">Total</td>'
		data += '<td style="border-color:#000000;padding:1px;font-size:14px;font-size:12px;background-color:#4169E1;color:white;text-align:center;font-weight:bold;">%s</td>' %(total_amt)
		data += '</tr>'

	
	return data


@frappe.whitelist()
def change_shipping_cost(currency,sq):
	items = json.loads(sq)
	d = []
	data = []
	t_amt = 0

	for j in items:
		d.append(j["supply_order_data"])
	
	uv = list(set(d))

	shp_list = 0
	for k in uv:
		sq   = frappe.get_all("Supplier Quotation",{"supply_order_data":k,"Workflow_state":"Approved By Management"},["*"])
		for s in sq:
			cur = frappe.get_value("Supplier",{"name":s.supplier},["default_currency"])
			exr = get_exchange_rate(cur,currency)
			s_cost = s.shipping_cost * exr
			shp_list = shp_list + s_cost
	
	total_cost = 0
	for j in items:

		sup = frappe.db.sql(""" select `tabSupplier Quotation`.supplier,`tabSupplier Quotation`.name,
		`tabSupplier Quotation Item`.base_rate,`tabSupplier Quotation Item`.supply_order_data,
		`tabSupplier Quotation`.shipping_cost from `tabSupplier Quotation` join 
		`tabSupplier Quotation Item` on `tabSupplier Quotation Item`.parent = `tabSupplier Quotation`.name where
		`tabSupplier Quotation Item`.supply_order_data = '%s' and `tabSupplier Quotation Item`.item_code = '%s' and `tabSupplier Quotation`.workflow_state = "Approved By Management" """ %(j["supply_order_data"],j["item_code"]),as_dict =1)

		if sup:
			total_cost = total_cost + (sup[0]["base_rate"]*j["qty"])
			cur = frappe.get_value("Supplier",{"name":sup[0]["supplier"]},["default_currency"])
			exr = get_exchange_rate(cur,currency)
			cost = sup[0]["shipping_cost"] * exr
			row = [sup[0]["supply_order_data"],sup[0]["name"],sup[0]["supplier"],sup[0]["base_rate"]*j["qty"],cost]
			data.append(row)

	t_amt = shp_list + total_cost       
	return data,t_amt
		

# @frappe.whitelist()
# def dlt_wo():
	# frappe.db.sql("""
	# UPDATE `tabSales Taxes and Charges` AS sii
	# JOIN `tabQuotation` AS si ON sii.parent = si.name
	# SET sii.branch = si.branch_name
	# WHERE si.branch_name = "Riyadh - TSL- KSA"
	# """)

	# wo = frappe.db.sql(""" UPDATE `tabEvaluation Report` SET branch = "Kuwait - TSL" WHERE company = "TSL COMPANY - Kuwait"; """)
	# wo = frappe.db.sql(""" delete from `tabSupply Order Data` where company = "TSL COMPANY - KSA" """)

@frappe.whitelist()
def update_wo():
	wo = frappe.db.sql(""" delete from `tabEvaluation Report` WHERE name = "EVAL-R-25-00038" """)



@frappe.whitelist()
def set_alert_urgent():
	w = frappe.get_all("Work Order Data",["*"])
	tody = datetime.now().date()
	data = ""
	data += '<p>Hi</p>'
	data += '<p>Kindly find the below Work orders are pending for 2 days.Please take necessary action to click Work order</p>'
	count = 0
	for i in w:
		if not i.old_wo_no and i.priority_status == "Urgent":
			qu = frappe.db.sql(""" select `tabQuotation`.name  from `tabQuotation` left join 
			`tabQuotation Item` on `tabQuotation`.name = `tabQuotation Item`.parent
			where `tabQuotation`.workflow_state  != "Rejected by Customer" and `tabQuotation Item`.wod_no = '%s' """ %(i.name) ,as_dict=1)
			if not qu:
				d = add_days(i.posting_date,2)
				if d <= tody:
					if not i.status == "RNA-Return Not Approved" and not i.status == "RNF-Return No Fault":
						count = count + 1
						data += '<p><a href="https://erp.tsl-me.com/app/work-order-data/%s">%s-%s</a></p>' % (i.name,i.name,i.status)
					# data += '<table class="table table-bordered">'
					# data += '<tr>'
					# data += '<td colspan = "2" style="border-color:#000000;padding:1px;font-size:14px;font-size:12px;b"><center><b>Work Orders</b><center></td>'
					# data += '</tr>'

					# data += '<tr>'
					# data += '<td colspan = "2" style="border-color:#000000;padding:1px;font-size:14px;font-size:12px;b"><center><b>%s</b><center></td>' %(i.name)
					# data += '</tr>'
					# data += '</table>'

	
	# message = """{}""".format(data)
	
	if count > 0:
		frappe.sendmail(recipients=["omar@tsl-me.com"],
						sender="Notification from TSL <info@tsl-me.com>",
						subject="Work Orders - Pending",
						message=data)

@frappe.whitelist()
def set_alert_normal():
	w = frappe.get_all("Work Order Data",["*"])
	tody = datetime.now().date()
	data = ""
	data += '<p>Hi</p>'
	data += '<p>Kindly find the below Work orders are pending for 6 days.Please take necessary action to click Work order</p>'
	count = 0
	for i in w:	
		if not i.old_wo_no and i.priority_status == "Normal":
			qu = frappe.db.sql(""" select `tabQuotation`.name  from `tabQuotation` left join 
			`tabQuotation Item` on `tabQuotation`.name = `tabQuotation Item`.parent
			where `tabQuotation`.workflow_state  != "Rejected by Customer" and `tabQuotation Item`.wod_no = '%s' """ %(i.name) ,as_dict=1)
			if not qu:
				d = add_days(i.posting_date,6)
				if d <= tody:
					if not i.status == "RNA-Return Not Approved" and not i.status == "RNF-Return No Fault":
						count = count + 1
						data += '<p><a href="https://erp.tsl-me.com/app/work-order-data/%s">%s</a></p>' % (i.name,i.name)

					# data += '<table class="table table-bordered">'
					# data += '<tr>'
					# data += '<td colspan = "2" style="border-color:#000000;padding:1px;font-size:14px;font-size:12px;b"><center><b>Work Orders</b><center></td>'
					# data += '</tr>'

					# data += '<tr>'
					# data += '<td colspan = "2" style="border-color:#000000;padding:1px;font-size:14px;font-size:12px;b"><center><b>%s</b><center></td>' %(i.name)
					# data += '</tr>'
					# data += '</table>'

	
	# message = """{}""".format(data)
	# print(message)
	if count > 0:
		frappe.sendmail(recipients=["omar@tsl-me.com"],
						sender="Notification from TSL <info@tsl-me.com>",
						subject="Work Orders - Pending",
						message=data)


@frappe.whitelist()
def send_mail_to_customer(q,type,email,company,branch):
	if type == "Customer Quotation - Repair" and branch =="Kuwait - TSL":
		if email:
			frappe.sendmail(recipients=[email],
			sender="Notification from TSL <info@tsl-me.com>",
			subject="Quotation from TSL",
			message=""" 
		   
			<p>Dear Sir,</p><br><br>
				I hope this message ﬁnds you well.<br><br>
				We are pleased to provide the below Quotation for the reference (Wo #No.)<br><br>
				Awaitng your approval for the same.<br><br>
				Thank you for choosing TSL COMPANY.<br><br>
				Best regards,<br>

				<div><style>.sh-src a{text-decoration:none!important;}</style></div> <br> <table cellpadding="0" cellspacing="0" border="0" class="sh-src" style="margin: 0px; border-collapse: collapse;"><tr><td style="padding: 0px 1px 0px 0px;"><table cellpadding="0" cellspacing="0" border="0" style="margin: 0px; border-collapse: collapse;"><tr><td align="center" style="padding: 0px 18px 0px 0px; vertical-align: top;"><table cellpadding="0" cellspacing="0" border="0" style="margin: 0px; border-collapse: collapse;"><tr><td style="padding: 0px 1px 13px 0px;"><p style="margin: 1px;"><img src="https://signaturehound.com/api/v1/file/5r1rjllxn18j0p" alt="" title="Profile Picture" width="100" height="100" class="" style="display: block; border: 0px; max-width: 100px;"></p></td></tr></table> <table cellpadding="0" cellspacing="0" border="0" style="margin: 0px; border-collapse: collapse;"><tr><td style="padding: 0px 1px 0px 0px;"><p style="margin: 1px;"><a href="https://tsl-me.com/" target="_blank"><img src="https://signaturehound.com/api/v1/file/3lwizllxn10tyt" alt="" title="Logo" width="150" height="50" style="display: block; border: 0px; max-width: 150px;"></a></p></td></tr></table></td> <td width="5" style="padding: 1px 0px 0px;"></td> <td style="padding: 0px 1px 0px 0px; vertical-align: top;"><table cellpadding="0" cellspacing="0" border="0" style="margin: 0px; border-collapse: collapse;"><tr><td style="padding: 0px 1px 10px 0px; border-bottom: 2px solid rgb(0,92,163); font-family: Arial, sans-serif; font-size: 13px; line-height: 15px; white-space: nowrap;"><p style="font-family: Arial, sans-serif; font-size: 13px; line-height: 15px; font-weight: 700; color: rgb(0,92,163); white-space: nowrap; margin: 1px;">Mohammed K Daddy</p> <p style="font-family: Arial, sans-serif; font-size: 13px; line-height: 15px; white-space: nowrap; color: rgb(136,136,136); margin: 1px;">Customer Support Officer</p> <!----> <p style="font-family: Arial, sans-serif; font-size: 13px; line-height: 15px; white-space: nowrap; color: rgb(136,136,136); margin: 1px;">TSL Group | Kuwait</p></td></tr> <tr><td style="padding: 10px 1px 10px 0px; border-bottom: 2px solid rgb(0,92,163);"><table cellpadding="0" cellspacing="0" border="0" style="margin: 0px; border-collapse: collapse;"><tr><td valign="middle" style="padding: 1px 5px 1px 0px; vertical-align: middle;"><p style="margin: 1px;"><img src="https://signaturehound.com/api/v1/png/email/round-outlined/0088cc.png" alt="" width="22" height="22" style="display: block; border: 0px; width: 22px; height: 22px;"></p></td> <td style="font-family: Arial, sans-serif; font-size: 13px; line-height: 15px; white-space: nowrap; color: rgb(136,136,135) !important; padding: 1px 0px; vertical-align: middle;"><p style="margin: 1px;"><a href="mailto:info@tsl-me.com" target="_blank" style="font-family: Arial, sans-serif; font-size: 13px; line-height: 15px; white-space: nowrap; color: rgb(136,136,136); text-decoration: none !important;"><span style="font-family: Arial, sans-serif; font-size: 13px; line-height: 15px; white-space: nowrap; color: rgb(136,136,136); text-decoration: none !important;">info@tsl-me.com</span></a></p></td></tr> <tr><td valign="middle" style="padding: 1px 5px 1px 0px; vertical-align: middle;"><p style="margin: 1px;"><img src="https://signaturehound.com/api/v1/png/direct/round-outlined/0088cc.png" alt="" width="22" height="22" style="display: block; border: 0px; width: 22px; height: 22px;"></p></td> <td style="font-family: Arial, sans-serif; font-size: 13px; line-height: 15px; white-space: nowrap; color: rgb(136,136,135) !important; padding: 1px 0px; vertical-align: middle;"><p style="margin: 1px;"><a href="tel:+96524741313" target="_blank" style="font-family: Arial, sans-serif; font-size: 13px; line-height: 15px; white-space: nowrap; color: rgb(136,136,136); text-decoration: none !important;"><span style="font-family: Arial, sans-serif; font-size: 13px; line-height: 15px; white-space: nowrap; color: rgb(136,136,136); text-decoration: none !important;">+965 24741313</span></a></p></td></tr><tr><td valign="middle" style="padding: 1px 5px 1px 0px; vertical-align: middle;"><p style="margin: 1px;"><img src="https://signaturehound.com/api/v1/png/fax/round-outlined/0088cc.png" alt="" width="22" height="22" style="display: block; border: 0px; width: 22px; height: 22px;"></p></td> <td style="font-family: Arial, sans-serif; font-size: 13px; line-height: 15px; white-space: nowrap; color: rgb(136,136,135) !important; padding: 1px 0px; vertical-align: middle;"><p style="margin: 1px;"><a href="tel:+96524741311" target="_blank" style="font-family: Arial, sans-serif; font-size: 13px; line-height: 15px; white-space: nowrap; color: rgb(136,136,136); text-decoration: none !important;"><span style="font-family: Arial, sans-serif; font-size: 13px; line-height: 15px; white-space: nowrap; color: rgb(136,136,136); text-decoration: none !important;">+965 24741311</span></a></p></td></tr> <tr><td valign="top" style="padding: 1px 5px 1px 0px; vertical-align: top;"><p style="margin: 1px;">
				<img src="https://signaturehound.com/api/v1/png/map/round-outlined/0088cc.png" alt="" width="22" height="22" style="display: block; border: 0px; width: 22px; height: 22px;"></p></td> <td style="font-family: Arial, sans-serif; font-size: 13px; line-height: 15px; white-space: nowrap; color: rgb(136,136,135) !important; padding: 1px 0px; vertical-align: middle;"><p style="margin: 1px;"><a href="https://www.google.com/maps/place/TSL+For+Industrial+Electronics+Repairing+%26+Supply+-+Kuwait/@29.3082683,47.9352691,18z/data=!4m6!3m5!1s0x3fcf9a9068440231:0x7321607de759fdc1!8m2!3d29.3082309!4d47.9365244!16s%2Fg%2F11c0rlsd_t?entry=ttu" target="_blank" style="font-family: Arial, sans-serif; font-size: 13px; line-height: 15px; white-space: nowrap; color: rgb(136,136,136); text-decoration: none !important;"><span style="font-family: Arial, sans-serif; font-size: 13px; line-height: 15px; white-space: nowrap; color: rgb(136,136,136); text-decoration: none !important;">Bldg: 1473, Unit: 13, Street: 24, Block: 1,<br>Al Rai Industrial Area, Kuwait</span></a></p></td></tr> <tr><td valign="middle" style="padding: 1px 5px 1px 0px; vertical-align: middle;"><p style="margin: 1px;"><img src="https://signaturehound.com/api/v1/png/website/round-outlined/0088cc.png" alt="" width="22" height="22" style="display: block; border: 0px; width: 22px; height: 22px;"></p></td> <td style="font-family: Arial, sans-serif; font-size: 13px; line-height: 15px; white-space: nowrap; color: rgb(0,92,162) !important; font-weight: 700; padding: 1px 0px; vertical-align: middle;"><p style="margin: 1px;"><a href="https://tsl-me.com/" target="_blank" style="font-family: Arial, sans-serif; font-size: 13px; line-height: 15px; white-space: nowrap; color: rgb(0,92,163); font-weight: 700; text-decoration: none !important;"><span style="font-family: Arial, sans-serif; font-size: 13px; line-height: 15px; white-space: nowrap; color: rgb(0,92,163); font-weight: 700; text-decoration: none !important;">tsl-me.com</span></a></p></td></tr></table></td></tr> <tr><td style="padding: 0px 1px 0px 0px;"><table cellpadding="0" cellspacing="0" border="0" style="margin: 0px; border-collapse: collapse;"><tr><td width="27" style="font-size: 0px; line-height: 0px; padding: 13px 1px 0px 0px;"><p style="margin: 1px;"><a href="https://www.linkedin.com/company/tsl-me/mycompany/" target="_blank"><img src="https://signaturehound.com/api/v1/png/linkedin/round/0088cc.png" alt="" width="27" height="27" style="display: block; border: 0px; width: 27px; height: 27px;"></a></p></td> <td width="3" style="padding: 0px 0px 1px;"></td><td width="27" style="font-size: 0px; line-height: 0px; padding: 13px 1px 0px 0px;"><p style="margin: 1px;">
				<a href="https://x.com/tsl_mecompany?s=11&amp;t=Zxza0-9Q_18nsDCddfTQPw" target="_blank"><img src="https://signaturehound.com/api/v1/png/x/round/0088cc.png" alt="" width="27" height="27" style="display: block; border: 0px; width: 27px; height: 27px;"></a></p></td> <td width="3" style="padding: 0px 0px 1px;"></td><td width="27" style="font-size: 0px; line-height: 0px; padding: 13px 1px 0px 0px;"><p style="margin: 1px;"><a href="https://www.instagram.com/tslcom/?igshid=MzRlODBiNWFlZA%3D%3D" target="_blank"><img src="https://signaturehound.com/api/v1/png/instagram/round/0088cc.png" alt="" width="27" height="27" style="display: block; border: 0px; width: 27px; height: 27px;"></a></p></td> <td width="3" style="padding: 0px 0px 1px;"></td><td width="27" style="font-size: 0px; line-height: 0px; padding: 13px 1px 0px 0px;"><p style="margin: 1px;"><a href="https://www.facebook.com/people/TSL-Industrial-Electronics-Services/61550277093129/" target="_blank"><img src="https://signaturehound.com/api/v1/png/facebook/round/0088cc.png" alt="" width="27" height="27" style="display: block; border: 0px; width: 27px; height: 27px;"></a></p></td> <td width="3" style="padding: 0px 0px 1px;"></td><td width="27" style="font-size: 0px; line-height: 0px; padding: 13px 1px 0px 0px;"><p style="margin: 1px;"><a href="https://www.youtube.com/@TSLELECTRONICSSERVICES" target="_blank"><img src="https://signaturehound.com/api/v1/png/youtube/round/0088cc.png" alt="" width="27" height="27" style="display: block; border: 0px; width: 27px; height: 27px;"></a></p></td> <td width="3" style="padding: 0px 0px 1px;"></td></tr></table></td></tr></table></td></tr></table></td></tr> <!----> <tr><td style="padding: 0px 1px 0px 0px;"><table cellpadding="0" cellspacing="0" border="0" style="max-width: 600px; margin: 0px; border-collapse: collapse;"><tr><td style="padding: 15px 1px 0px 0px; font-family: Arial, sans-serif; font-size: 10px; line-height: 12px; color: rgb(136,136,136);"><p style="font-family: Arial, sans-serif; font-size: 10px; line-height: 12px; color: rgb(136,136,136); margin: 1px;">The content of this email is confidential and intended for the recipient specified in message only. It is strictly forbidden to share any part of this message with any third party, without a written consent of the sender. If you received this message by mistake, please reply to this message and follow with its deletion, so that we can ensure such a mistake does not occur in the future.</p></td></tr></table></td></tr> <tr><td style="padding: 0px 1px 0px 0px;"><table cellpadding="0" cellspacing="0" border="0" style="margin: 0px; border-collapse: collapse;"><tr><td valign="middle" style="padding: 15px 4px 1px 0px; vertical-align: middle;"><p style="margin: 1px;"></p></td> <td style="font-family: Arial, sans-serif; color: rgb(136,136,135) !important; font-weight: 700; font-size: 10px; line-height: 15px; padding: 15px 0px 1px; vertical-align: middle;"><p style="margin: 1px;"</p></td></tr></table></td></tr> <!----></table>
				""",
			
			attachments=get_attachments(q,"Quotation")
			)
			frappe.msgprint("Mail Successfully Sent to Customer")

	if type == "Customer Quotation - Repair" and branch == "Jeddah - TSL-SA":
		
		if email:
			msg1 = """<p>Dear Sir,</p><br>
				I hope this message ﬁnds you well.<br><br>
				We are pleased to provide the below Quotation for the reference (%s)<br><br>
				Awaitng your approval for the same.<br><br>
				Thank you for choosing TSL COMPANY.<br><br>
				Best regards,<br>"""%(q)
			msg2 = """<div><style>.sh-src a{text-decoration:none!important;}</style></div> <br> <table cellpadding="0" cellspacing="0" border="0" class="sh-src" style="margin: 0px; border-collapse: collapse;"><tr><td style="padding: 0px 1px 0px 0px;"><table cellpadding="0" cellspacing="0" border="0" style="margin: 0px; border-collapse: collapse;"><tr><td align="center" style="padding: 0px 18px 0px 0px; vertical-align: top;">
				<table cellpadding="0" cellspacing="0" border="0" style="margin: 0px; border-collapse: collapse;"><tr><td style="padding: 0px 1px 13px 0px;"><p style="margin: 1px;"><img src="https://signaturehound.com/api/v1/file/5r1rjllxn0zdme" alt="" title="Profile Picture" width="100" height="100" class="" style="display: block; border: 0px; max-width: 100px;"></p></td></tr></table> <table cellpadding="0" cellspacing="0" border="0" style="margin: 0px; border-collapse: collapse;"><tr><td style="padding: 0px 1px 0px 0px;"><p style="margin: 1px;"><a href="https://tsl-me.com/" target="_blank"><img src="https://signaturehound.com/api/v1/file/137twgllxltmdmv" alt="" title="Logo" width="150" height="50" style="display: block; border: 0px; max-width: 150px;"></a></p></td></tr></table></td> <td width="5" style="padding: 1px 0px 0px;"></td> <td style="padding: 0px 1px 0px 0px; vertical-align: top;"><table cellpadding="0" cellspacing="0" border="0" style="margin: 0px; border-collapse: collapse;"><tr><td style="padding: 0px 1px 10px 0px; border-bottom: 2px solid rgb(0,92,163); font-family: Arial, sans-serif; font-size: 13px; line-height: 15px; white-space: nowrap;"><p style="font-family: Arial, sans-serif; font-size: 13px; line-height: 15px; font-weight: 700; color: rgb(0,92,163); white-space: nowrap; margin: 1px;">Ajai
				</p> <p style="font-family: Arial, sans-serif; font-size: 13px; line-height: 15px; white-space: nowrap; color: rgb(136,136,136); margin: 1px;">Admin &amp; Customer Support</p> <!----> <p style="font-family: Arial, sans-serif; font-size: 13px; line-height: 15px; white-space: nowrap; color: rgb(136,136,136); margin: 1px;">
					TSL Group | KSA - Jeddah</p> <!----></td></tr> <tr><td style="padding: 10px 1px 10px 0px; border-bottom: 2px solid rgb(0,92,163);"><table cellpadding="0" cellspacing="0" border="0" style="margin: 0px; border-collapse: collapse;"><tr><td valign="middle" style="padding: 1px 5px 1px 0px; vertical-align: middle;"><p style="margin: 1px;"><img src="https://signaturehound.com/api/v1/png/email/round-outlined/0088cc.png" alt="" width="22" height="22" style="display: block; border: 0px; margin: 0px; width: 22px; height: 22px;"></p></td> <td style="font-family: Arial, sans-serif; font-size: 13px; line-height: 15px; white-space: nowrap; color: rgb(136,136,135) !important; padding: 1px 0px; vertical-align: middle;"><p style="margin: 1px;"><a href="mailto:info-jed@tsl-me.com" target="_blank" style="font-family: Arial, sans-serif; font-size: 13px; line-height: 15px; white-space: nowrap; color: rgb(136,136,136); text-decoration: none !important;"><span style="font-family: Arial, sans-serif; font-size: 13px; line-height: 15px; white-space: nowrap; color: rgb(136,136,136); text-decoration: none !important;">info-jed@tsl-me.com</span></a></p></td></tr> <tr><td valign="middle" style="padding: 1px 5px 1px 0px; vertical-align: middle;"><p style="margin: 1px;"><img src="https://signaturehound.com/api/v1/png/mobile/round-outlined/0088cc.png" alt="" width="22" height="22" style="display: block; border: 0px; margin: 0px; width: 22px; height: 22px;"></p></td> <td style="font-family: Arial, sans-serif; font-size: 13px; line-height: 15px; white-space: nowrap; color: rgb(136,136,135) !important; padding: 1px 0px; vertical-align: middle;"><p style="margin: 1px;"><a href="tel:+966558803522" target="_blank" style="font-family: Arial, sans-serif; font-size: 13px; line-height: 15px; white-space: nowrap; color: rgb(136,136,136); text-decoration: none !important;"><span style="font-family: Arial, sans-serif; font-size: 13px; line-height: 15px; white-space: nowrap; color: rgb(136,136,136); text-decoration: none !important;">+966 55 880 3522</span></a></p></td></tr> <tr><td valign="middle" style="padding: 1px 5px 1px 0px; vertical-align: middle;"><p style="margin: 1px;"><img src="https://signaturehound.com/api/v1/png/map/round-outlined/0088cc.png" alt="" width="22" height="22" style="display: block; border: 0px; margin: 0px; width: 22px; height: 22px;"></p></td> <td style="font-family: Arial, sans-serif; font-size: 13px; line-height: 15px; white-space: nowrap; color: rgb(136,136,135) !important; padding: 1px 0px; vertical-align: middle;"><p style="margin: 1px;"><a href="https://www.google.com/maps/dir/21.3180818,39.227158/TSL+Industrial+Electronics+for+Repairing+%26+Supply+-+Jeddah,+80th+street%D8%8C+Al-Qarinia+District%D8%8C+Jeddah+22535,+Saudi+Arabia%E2%80%AD/@21.3172514,39.1901511,13z/data=!3m1!4b1!4m9!4m8!1m1!4e1!1m5" target="_blank" style="font-family: Arial, sans-serif; font-size: 13px; line-height: 15px; white-space: nowrap; color: rgb(136,136,136); text-decoration: none !important;"><span style="font-family: Arial, sans-serif; font-size: 13px; line-height: 15px; white-space: nowrap; color: rgb(136,136,136); text-decoration: none !important;">80th Street, Al-Qrainia, Jeddah, Saudi Arabia.</span></a></p></td></tr> <tr><td valign="middle" style="padding: 1px 5px 1px 0px; vertical-align: middle;"><p style="margin: 1px;"><img src="https://signaturehound.com/api/v1/png/website/round-outlined/0088cc.png" alt="" width="22" height="22" style="display: block; border: 0px; margin: 0px; width: 22px; height: 22px;"></p></td> <td style="font-family: Arial, sans-serif; font-size: 13px; line-height: 15px; white-space: nowrap; color: rgb(0,92,162) !important; font-weight: 700; padding: 1px 0px; vertical-align: middle;"><p style="margin: 1px;"><a href="https://tsl-me.com/" target="_blank" style="font-family: Arial, sans-serif; font-size: 13px; line-height: 15px; white-space: nowrap; color: rgb(0,92,163); font-weight: 700; text-decoration: none !important;"><span style="font-family: Arial, sans-serif; font-size: 13px; line-height: 15px; white-space: nowrap; color: rgb(0,92,163); font-weight: 700; text-decoration: none !important;">tsl-me.com</span></a></p></td></tr></table></td></tr> <tr><td style="padding: 0px 1px 0px 0px;"><table cellpadding="0" cellspacing="0" border="0" style="margin: 0px; border-collapse: collapse;"><tr><td width="27" style="font-size: 0px; line-height: 0px; padding: 13px 1px 0px 0px;"><p style="margin: 1px;"><a href="https://www.linkedin.com/company/tsl-me/mycompany/" target="_blank"><img src="https://signaturehound.com/api/v1/png/linkedin/round/0088cc.png" alt="" width="27" height="27" style="display: block; border: 0px; margin: 0px; width: 27px; height: 27px;"></a></p></td> <td width="3" style="padding: 0px 0px 1px;"></td><td width="27" style="font-size: 0px; line-height: 0px; padding: 13px 1px 0px 0px;"><p style="margin: 1px;"><a href="https://x.com/tsl_mecompany?s=11&amp;t=Zxza0-9Q_18nsDCddfTQPw" target="_blank"><img src="https://signaturehound.com/api/v1/png/x/round/0088cc.png" alt="" width="27" height="27" style="display: block; border: 0px; margin: 0px; width: 27px; height: 27px;"></a></p></td> <td width="3" style="padding: 0px 0px 1px;"></td><td width="27" style="font-size: 0px; line-height: 0px; padding: 13px 1px 0px 0px;"><p style="margin: 1px;"><a href="https://www.instagram.com/tslcom/?igshid=MzRlODBiNWFlZA%3D%3D" target="_blank"><img src="https://signaturehound.com/api/v1/png/instagram/round/0088cc.png" alt="" width="27" height="27" style="display: block; border: 0px; margin: 0px; width: 27px; height: 27px;"></a></p></td> <td width="3" style="padding: 0px 0px 1px;"></td><td width="27" style="font-size: 0px; line-height: 0px; padding: 13px 1px 0px 0px;"><p style="margin: 1px;"><a href="https://www.facebook.com/people/TSL-Industrial-Electronics-Services/61550277093129/" target="_blank"><img src="https://signaturehound.com/api/v1/png/facebook/round/0088cc.png" alt="" width="27" height="27" style="display: block; border: 0px; margin: 0px; width: 27px; height: 27px;"></a></p></td> <td width="3" style="padding: 0px 0px 1px;"></td><td width="27" style="font-size: 0px; line-height: 0px; padding: 13px 1px 0px 0px;"><p style="margin: 1px;"><a href="https://www.youtube.com/@TSLELECTRONICSSERVICES" target="_blank"><img src="https://signaturehound.com/api/v1/png/youtube/round/0088cc.png" alt="" width="27" height="27" style="display: block; border: 0px; margin: 0px; width: 27px; height: 27px;"></a></p></td> <td width="3" style="padding: 0px 0px 1px;"></td>
					</tr></table></td></tr></table></td></tr></table></td></tr> <!----> <tr><td style="padding: 0px 1px 0px 0px;"><table cellpadding="0" cellspacing="0" border="0" style="max-width: 600px; margin: 0px; border-collapse: collapse;"><tr><td style="padding: 15px 1px 0px 0px; font-family: Arial, sans-serif; font-size: 10px; line-height: 12px; color: rgb(136,136,136);"><p style="font-family: Arial, sans-serif; font-size: 10px; line-height: 12px; color: rgb(136,136,136); margin: 1px;">The content of this email is confidential and intended for the recipient specified in message only. It is strictly forbidden to share any part of this message with any third party, without a written consent of the sender. If you received this message by mistake, please reply to this message and follow with its deletion, so that we can ensure such a mistake does not occur in the future.</p></td></tr></table></td></tr> """
					
			frappe.sendmail(recipients=[email],
			sender="info-jed@tsl-me.com",
			subject="Quotation from TSL",
			message= msg1+msg2,
			attachments=get_jed_att(q,"Quotation")
			)
			frappe.msgprint("Mail Successfully Sent to Customer")
	# Supply Mail
	if type == "Customer Quotation - Supply":
		if email:
			frappe.sendmail(recipients=[email],
			sender="Notification from TSL <info@tsl-me.com>",
			subject="Quotation from TSL",
			message=""" 
		   
			<p>Dear Sir</p><br><br>,
				I hope this message ﬁnds you well.<br><br>
				We are pleased to provide the below Quotation for the reference (Wo #No.)<br><br>
				Awaitng your approval for the same.<br><br>
				Thank you for choosing TSL COMPANY.<br><br>
				Best regards,<br>

				<div><style>.sh-src a{text-decoration:none!important;}</style></div> <br> <table cellpadding="0" cellspacing="0" border="0" class="sh-src" style="margin: 0px; border-collapse: collapse;"><tr><td style="padding: 0px 1px 0px 0px;"><table cellpadding="0" cellspacing="0" border="0" style="margin: 0px; border-collapse: collapse;"><tr><td align="center" style="padding: 0px 18px 0px 0px; vertical-align: top;"><table cellpadding="0" cellspacing="0" border="0" style="margin: 0px; border-collapse: collapse;"><tr><td style="padding: 0px 1px 13px 0px;"><p style="margin: 1px;"><img src="https://signaturehound.com/api/v1/file/5r1rjllxn18j0p" alt="" title="Profile Picture" width="100" height="100" class="" style="display: block; border: 0px; max-width: 100px;"></p></td></tr></table> <table cellpadding="0" cellspacing="0" border="0" style="margin: 0px; border-collapse: collapse;"><tr><td style="padding: 0px 1px 0px 0px;"><p style="margin: 1px;"><a href="https://tsl-me.com/" target="_blank"><img src="https://signaturehound.com/api/v1/file/3lwizllxn10tyt" alt="" title="Logo" width="150" height="50" style="display: block; border: 0px; max-width: 150px;"></a></p></td></tr></table></td> <td width="5" style="padding: 1px 0px 0px;"></td> <td style="padding: 0px 1px 0px 0px; vertical-align: top;"><table cellpadding="0" cellspacing="0" border="0" style="margin: 0px; border-collapse: collapse;"><tr><td style="padding: 0px 1px 10px 0px; border-bottom: 2px solid rgb(0,92,163); font-family: Arial, sans-serif; font-size: 13px; line-height: 15px; white-space: nowrap;"><p style="font-family: Arial, sans-serif; font-size: 13px; line-height: 15px; font-weight: 700; color: rgb(0,92,163); white-space: nowrap; margin: 1px;">Mohammed K Daddy</p> <p style="font-family: Arial, sans-serif; font-size: 13px; line-height: 15px; white-space: nowrap; color: rgb(136,136,136); margin: 1px;">Customer Support Officer</p> <!----> <p style="font-family: Arial, sans-serif; font-size: 13px; line-height: 15px; white-space: nowrap; color: rgb(136,136,136); margin: 1px;">TSL Group | Kuwait</p></td></tr> <tr><td style="padding: 10px 1px 10px 0px; border-bottom: 2px solid rgb(0,92,163);"><table cellpadding="0" cellspacing="0" border="0" style="margin: 0px; border-collapse: collapse;"><tr><td valign="middle" style="padding: 1px 5px 1px 0px; vertical-align: middle;"><p style="margin: 1px;"><img src="https://signaturehound.com/api/v1/png/email/round-outlined/0088cc.png" alt="" width="22" height="22" style="display: block; border: 0px; width: 22px; height: 22px;"></p></td> <td style="font-family: Arial, sans-serif; font-size: 13px; line-height: 15px; white-space: nowrap; color: rgb(136,136,135) !important; padding: 1px 0px; vertical-align: middle;"><p style="margin: 1px;"><a href="mailto:info@tsl-me.com" target="_blank" style="font-family: Arial, sans-serif; font-size: 13px; line-height: 15px; white-space: nowrap; color: rgb(136,136,136); text-decoration: none !important;"><span style="font-family: Arial, sans-serif; font-size: 13px; line-height: 15px; white-space: nowrap; color: rgb(136,136,136); text-decoration: none !important;">info@tsl-me.com</span></a></p></td></tr> <tr><td valign="middle" style="padding: 1px 5px 1px 0px; vertical-align: middle;"><p style="margin: 1px;"><img src="https://signaturehound.com/api/v1/png/direct/round-outlined/0088cc.png" alt="" width="22" height="22" style="display: block; border: 0px; width: 22px; height: 22px;"></p></td> <td style="font-family: Arial, sans-serif; font-size: 13px; line-height: 15px; white-space: nowrap; color: rgb(136,136,135) !important; padding: 1px 0px; vertical-align: middle;"><p style="margin: 1px;"><a href="tel:+96524741313" target="_blank" style="font-family: Arial, sans-serif; font-size: 13px; line-height: 15px; white-space: nowrap; color: rgb(136,136,136); text-decoration: none !important;"><span style="font-family: Arial, sans-serif; font-size: 13px; line-height: 15px; white-space: nowrap; color: rgb(136,136,136); text-decoration: none !important;">+965 24741313</span></a></p></td></tr><tr><td valign="middle" style="padding: 1px 5px 1px 0px; vertical-align: middle;"><p style="margin: 1px;"><img src="https://signaturehound.com/api/v1/png/fax/round-outlined/0088cc.png" alt="" width="22" height="22" style="display: block; border: 0px; width: 22px; height: 22px;"></p></td> <td style="font-family: Arial, sans-serif; font-size: 13px; line-height: 15px; white-space: nowrap; color: rgb(136,136,135) !important; padding: 1px 0px; vertical-align: middle;"><p style="margin: 1px;"><a href="tel:+96524741311" target="_blank" style="font-family: Arial, sans-serif; font-size: 13px; line-height: 15px; white-space: nowrap; color: rgb(136,136,136); text-decoration: none !important;"><span style="font-family: Arial, sans-serif; font-size: 13px; line-height: 15px; white-space: nowrap; color: rgb(136,136,136); text-decoration: none !important;">+965 24741311</span></a></p></td></tr> <tr><td valign="top" style="padding: 1px 5px 1px 0px; vertical-align: top;"><p style="margin: 1px;">
				<img src="https://signaturehound.com/api/v1/png/map/round-outlined/0088cc.png" alt="" width="22" height="22" style="display: block; border: 0px; width: 22px; height: 22px;"></p></td> <td style="font-family: Arial, sans-serif; font-size: 13px; line-height: 15px; white-space: nowrap; color: rgb(136,136,135) !important; padding: 1px 0px; vertical-align: middle;"><p style="margin: 1px;"><a href="https://www.google.com/maps/place/TSL+For+Industrial+Electronics+Repairing+%26+Supply+-+Kuwait/@29.3082683,47.9352691,18z/data=!4m6!3m5!1s0x3fcf9a9068440231:0x7321607de759fdc1!8m2!3d29.3082309!4d47.9365244!16s%2Fg%2F11c0rlsd_t?entry=ttu" target="_blank" style="font-family: Arial, sans-serif; font-size: 13px; line-height: 15px; white-space: nowrap; color: rgb(136,136,136); text-decoration: none !important;"><span style="font-family: Arial, sans-serif; font-size: 13px; line-height: 15px; white-space: nowrap; color: rgb(136,136,136); text-decoration: none !important;">Bldg: 1473, Unit: 13, Street: 24, Block: 1,<br>Al Rai Industrial Area, Kuwait</span></a></p></td></tr> <tr><td valign="middle" style="padding: 1px 5px 1px 0px; vertical-align: middle;"><p style="margin: 1px;"><img src="https://signaturehound.com/api/v1/png/website/round-outlined/0088cc.png" alt="" width="22" height="22" style="display: block; border: 0px; width: 22px; height: 22px;"></p></td> <td style="font-family: Arial, sans-serif; font-size: 13px; line-height: 15px; white-space: nowrap; color: rgb(0,92,162) !important; font-weight: 700; padding: 1px 0px; vertical-align: middle;"><p style="margin: 1px;"><a href="https://tsl-me.com/" target="_blank" style="font-family: Arial, sans-serif; font-size: 13px; line-height: 15px; white-space: nowrap; color: rgb(0,92,163); font-weight: 700; text-decoration: none !important;"><span style="font-family: Arial, sans-serif; font-size: 13px; line-height: 15px; white-space: nowrap; color: rgb(0,92,163); font-weight: 700; text-decoration: none !important;">tsl-me.com</span></a></p></td></tr></table></td></tr> <tr><td style="padding: 0px 1px 0px 0px;"><table cellpadding="0" cellspacing="0" border="0" style="margin: 0px; border-collapse: collapse;"><tr><td width="27" style="font-size: 0px; line-height: 0px; padding: 13px 1px 0px 0px;"><p style="margin: 1px;"><a href="https://www.linkedin.com/company/tsl-me/mycompany/" target="_blank"><img src="https://signaturehound.com/api/v1/png/linkedin/round/0088cc.png" alt="" width="27" height="27" style="display: block; border: 0px; width: 27px; height: 27px;"></a></p></td> <td width="3" style="padding: 0px 0px 1px;"></td><td width="27" style="font-size: 0px; line-height: 0px; padding: 13px 1px 0px 0px;"><p style="margin: 1px;">
				<a href="https://x.com/tsl_mecompany?s=11&amp;t=Zxza0-9Q_18nsDCddfTQPw" target="_blank"><img src="https://signaturehound.com/api/v1/png/x/round/0088cc.png" alt="" width="27" height="27" style="display: block; border: 0px; width: 27px; height: 27px;"></a></p></td> <td width="3" style="padding: 0px 0px 1px;"></td><td width="27" style="font-size: 0px; line-height: 0px; padding: 13px 1px 0px 0px;"><p style="margin: 1px;"><a href="https://www.instagram.com/tslcom/?igshid=MzRlODBiNWFlZA%3D%3D" target="_blank"><img src="https://signaturehound.com/api/v1/png/instagram/round/0088cc.png" alt="" width="27" height="27" style="display: block; border: 0px; width: 27px; height: 27px;"></a></p></td> <td width="3" style="padding: 0px 0px 1px;"></td><td width="27" style="font-size: 0px; line-height: 0px; padding: 13px 1px 0px 0px;"><p style="margin: 1px;"><a href="https://www.facebook.com/people/TSL-Industrial-Electronics-Services/61550277093129/" target="_blank"><img src="https://signaturehound.com/api/v1/png/facebook/round/0088cc.png" alt="" width="27" height="27" style="display: block; border: 0px; width: 27px; height: 27px;"></a></p></td> <td width="3" style="padding: 0px 0px 1px;"></td><td width="27" style="font-size: 0px; line-height: 0px; padding: 13px 1px 0px 0px;"><p style="margin: 1px;"><a href="https://www.youtube.com/@TSLELECTRONICSSERVICES" target="_blank"><img src="https://signaturehound.com/api/v1/png/youtube/round/0088cc.png" alt="" width="27" height="27" style="display: block; border: 0px; width: 27px; height: 27px;"></a></p></td> <td width="3" style="padding: 0px 0px 1px;"></td></tr></table></td></tr></table></td></tr></table></td></tr> <!----> <tr><td style="padding: 0px 1px 0px 0px;"><table cellpadding="0" cellspacing="0" border="0" style="max-width: 600px; margin: 0px; border-collapse: collapse;"><tr><td style="padding: 15px 1px 0px 0px; font-family: Arial, sans-serif; font-size: 10px; line-height: 12px; color: rgb(136,136,136);"><p style="font-family: Arial, sans-serif; font-size: 10px; line-height: 12px; color: rgb(136,136,136); margin: 1px;">The content of this email is confidential and intended for the recipient specified in message only. It is strictly forbidden to share any part of this message with any third party, without a written consent of the sender. If you received this message by mistake, please reply to this message and follow with its deletion, so that we can ensure such a mistake does not occur in the future.</p></td></tr></table></td></tr> <tr><td style="padding: 0px 1px 0px 0px;"><table cellpadding="0" cellspacing="0" border="0" style="margin: 0px; border-collapse: collapse;"><tr><td valign="middle" style="padding: 15px 4px 1px 0px; vertical-align: middle;"><p style="margin: 1px;"></p></td> <td style="font-family: Arial, sans-serif; color: rgb(136,136,135) !important; font-weight: 700; font-size: 10px; line-height: 15px; padding: 15px 0px 1px; vertical-align: middle;"><p style="margin: 1px;"</p></td></tr></table></td></tr> <!----></table>
			""",
			attachments=get_jed_att(q,"Quotation")
			)
			frappe.msgprint("Mail Successfully Sent to Customer")
		


def get_attachments(name,doctype):
	attachments = frappe.attach_print(doctype, name,file_name=doctype, print_format="Quotation TSL")
	return [attachments]
def get_jed_att(name,doctype):
	attachments = frappe.attach_print(doctype, name,file_name=doctype, print_format="Quotation TSL - KSA - Long")
	return [attachments]

def get_attach(name,doctype):
	attachments = frappe.attach_print(doctype, name,file_name=doctype, print_format="Supply Quotation")
	return [attachments]

@frappe.whitelist()
def customer_notification():
	si = frappe.get_all("Sales Invoice",["*"])
	tody = datetime.now().date()
	for i in si:
		d = add_days(i.posting_date,6)
		if tody == d and i.send_later:
			email = frappe.get_value("Customer",{"name":i.customer},["email_id"])
			if email:
				frappe.sendmail(recipients=["karthiksrinivasan1996.ks@gmail.com"],
				sender="Notification from TSL <info@tsl-me.com>",
				subject="Quotatio from TSL - ",
				message=""" 
				<p>Dear Mr / ms</p><br>
				Please find the attached Invoice and delivery copy as requested.
				Kindly issue the payment as soon as possible.</p>
				""",
				attachments=get_attachments(i.name,"Sales Invoice")
				)



@frappe.whitelist()
def cron_job_allocation():
	job = frappe.db.exists('Scheduled Job Type', 'urgent_work_order')
	if not job:
		sjt = frappe.new_doc("Scheduled Job Type")  
		sjt.update({
			"method" : 'tsl.custom_py.utils.set_alert_normal',
			"frequency" : 'Daily',
			"cron_format" : '0 8 * * *'
		})
		sjt.save(ignore_permissions=True)



@frappe.whitelist()
def schedule_allocate_leave_on_anniversary():
	job = frappe.db.exists('Scheduled Job Type', 'utils.allocate_leave_on_anniversary')
	if not job:
		sjt = frappe.new_doc("Scheduled Job Type")  
		sjt.update({
			"method" : 'tsl.custom_py.utils.allocate_leave_on_anniversary',
			"frequency" : 'Daily',
		})
		sjt.save(ignore_permissions=True)

@frappe.whitelist()
def crt_item(import_file):
	from datetime import datetime
	filepath = get_file(import_file)
	data = read_csv_content(filepath[1])
	c = 0
	for i in data:
		s = frappe.db.exists("Item",{"name":i[1],"published_in_website":0})
		if s:
			frappe.db.set_value("Item",i[1],"published_in_website",1)
			c = c +1
			print(c)
	
# @frappe.whitelist()
# def update_ner(wod):
#     frappe.errprint(wod)
	# frappe.db.set_value("Work Order Data",wod,"status_cap","")
	# ev = frappe.get_doc("Evaluation Report",{"work_order_data":wod})
	# ev.ner_field = ""
	# ev.save(ignore_permissions =1)

@frappe.whitelist()
def update_item(model,item):
	i = frappe.db.exists("Item",item)
	if i:
		it = frappe.get_doc("Item",item)
		it.model = model
		it.save(ignore_permissions = 1)
		frappe.msgprint("Model Number Updated")

@frappe.whitelist()
def salary_register(month_name,company,cyrix_employee,civil_id_no,docstatus):
	from datetime import timedelta
	current_year = datetime.now().year
	month_number = datetime.strptime(month_name, '%B').month
	
	first_date = datetime(current_year, month_number, 1)
	if month_number == 12:
		last_date = datetime(current_year, month_number, 31)
	else:
		next_month = datetime(current_year, month_number + 1, 1)
		last_date = next_month - timedelta(days=1)
	if company == "TSL COMPANY - Kuwait":
		filters = {'from_date': first_date.date(), 'to_date': last_date.date(), 'currency': 'KWD', 'company': company, 'docstatus': docstatus}
	if company == "TSL COMPANY - KSA":
		filters = {'from_date': first_date.date(), 'to_date': last_date.date(), 'currency': 'SAR', 'company': company, 'docstatus': docstatus}
	if company == "TSL COMPANY - UAE":
		filters = {'from_date': first_date.date(), 'to_date': last_date.date(), 'currency': 'AED', 'company': company, 'docstatus': docstatus}
	from tsl.tsl.report.salary_register_report.salary_register_report import execute
	result = execute(filters)
	headers = result[0]
	records = result[1]
	skip_columns = ['total_loan_repayment','company','salary_slip_id', 'branch', 'date_of_joining', 'start_date', 'end_date','currency','data_of_joining','department','leave_without_pay']
	int_columns = ['leave_without_pay', 'payment_days',"civil_id_no"]
	currency_columns = [header['fieldname'] for header in headers if header.get('fieldtype') == 'Currency']
	filtered_headers = []
	for header in headers:
		if header['fieldname'] in skip_columns:
			continue
		if header['fieldname'] == 'loan':
			header['label'] = 'Loan Deduction'
		filtered_headers.append(header)


	# Add Civil ID header after employee_name
	for index, header in enumerate(filtered_headers):
		if header['fieldname'] == 'employee_name' and company == "TSL COMPANY - KSA":
			# Create a new header for Civil ID
			civil_id_header = {
				'fieldname': 'civil_id_no',
				'label': 'Civil ID',
				'fieldtype': 'Data'  # Assuming it's a string
			}
			filtered_headers.insert(index + 1, civil_id_header)  # Insert after employee_name
			
		if header['fieldname'] == 'designation':
			ctc_header = {
				'fieldname': 'ctc',
				'label': 'Total Salary',
				'fieldtype': 'Currency'
			}
			filtered_headers.insert(index + 1, ctc_header)  # Insert after designation
			break

	totals = {fieldname: 0.0 for fieldname in currency_columns}

	# Generate HTML table headers
	data = '<table border="1" width="100%"><tr><td style="text-align:right;font-size:6px;font-weight:bold">Sl.No.</td>'
	data += ''.join(f'<td style="font-size:7px;font-weight:bold">{header["label"]}</td>' for header in filtered_headers)
	data += '</tr>'
	records = sorted(records, key=lambda x: int(x.get('employee', 0)))
	# Generate HTML table rows
	sl_no = 1
	for record in records:
		cyrix = frappe.db.get_value("Employee",record.get('employee'),'cyrix_employee_')
		if cyrix == cyrix_employee:
			data += '<tr>'
			data += f'<td style="font-size:6px">{sl_no}</td>'

			for header in filtered_headers:
				fieldname = header['fieldname']
				if fieldname == "civil_id_no" and company == "TSL COMPANY - KSA":
					fieldtype = "Data"
					cell_value = frappe.db.get_value("Employee",record.get("employee"),'civil_id_no')  # Default to 0 if None
				
				elif fieldname == "ctc":
					fieldtype = "Currency"
					cell_value = frappe.db.get_value("Employee",record.get("employee"),'ctc')  # Default to 0 if None
				
				else:
					fieldtype = header.get('fieldtype', '')
					cell_value = record.get(fieldname, 0)  # Default to 0 if None
				
				# Update totals for currency columns
				if fieldname in currency_columns:
					totals[fieldname] += float(cell_value) if isinstance(cell_value, (float, int)) else 0
				
				# Format cell value
				if fieldtype == 'Currency':
					cell_value = "{:,.2f}".format(cell_value) if isinstance(cell_value, (float, int)) else "0.000"
				elif isinstance(cell_value, float):
					# cell_value = f"{cell_value:.2f}"
					cell_value = "{:,.2f}".format(cell_value) 
				data += f'<td style="font-size:6px">{cell_value or ""}</td>'
			data += '</tr>'
			sl_no += 1

	# Generate totals row for currency columns
	data += '<tr>'

	# Determine the number of non-currency columns before the first currency column
	first_currency_index = next((i for i, header in enumerate(filtered_headers) if header['fieldname'] in currency_columns), len(filtered_headers))
	non_currency_colspan = first_currency_index + 1

	# Add merged cell for 'Total' title
	if non_currency_colspan > 0:
		data += f'<td style="font-size:6px;font-weight:bold;text-align:center" colspan="{non_currency_colspan}">Total</td>'

	# Add empty cells for any columns between the merged 'Total' cell and the first currency column
	for i in range(non_currency_colspan, first_currency_index):
		data += '<td></td>'

	# Add total values for currency columns and empty cells for non-currency columns after 'Total'
	for i, header in enumerate(filtered_headers[first_currency_index:], start=first_currency_index):
		fieldname = header['fieldname']
		if fieldname in currency_columns:
			total_value = "{:,.2f}".format(totals[fieldname])
		else:
			total_value = ''
		data += f'<td style="font-size:6px;font-weight:bold">{total_value}</td>'
	data += '</tr>'

	data += '</table>'

	return data


@frappe.whitelist()
def update_wd():
	s = frappe.db.sql(""" UPDATE `tabWork Order Data` SET tech_id = 5 WHERE technician = "sami@tsl-me.com" """)
	

@frappe.whitelist()
def get_incentive(from_date,to_date,company,branch):	
	d = datetime.now().date()
	ogdate = datetime.strptime(str(d),"%Y-%m-%d")
	fr =datetime.strptime(str(from_date),"%Y-%m-%d")
	td = datetime.strptime(str(to_date),"%Y-%m-%d")
	# Format the date as a string in the desired format
	formatted_date = ogdate.strftime("%d-%m-%Y")
	formatted_date_1 = fr.strftime("%d-%m-%Y")
	formatted_date_2 = td.strftime("%d-%m-%Y")
	data= ""
	data += '<table class="table table-bordered">'
	data += '<tr>'
	data += '<td style="width:30%;border-color:#000000;"><img src = "/files/TSL Logo.png" align="left" width ="200"></td>'
	if company == "TSL COMPANY - Kuwait":
		data += '<td style="width:30%;border-color:#000000;color:#055c9d;"><center><b style="color:#055c9d;font-size:19px;font-weight:bold;">TSL Company</b></center><br><center style="font-weight:bold;font-size:15px;color:#055c9d;">Branch - Kuwait</center></td>' 
	
	if company == "TSL COMPANY - UAE":
		data += '<td style="width:30%;border-color:#000000;color:#055c9d;"><center><b style="color:#055c9d;font-size:19px;font-weight:bold;">TSL Company</b></center><br><center style="font-weight:bold;font-size:15px;color:#055c9d;">Branch - UAE</center></td>' 

	if company == "TSL COMPANY - KSA":
		if branch == "Jeddah - TSL-SA":
			data += '<td style="width:30%;border-color:#000000;color:#055c9d;"><center><b style="color:#055c9d;font-size:19px;font-weight:bold;">TSL Company</b></center><br><center style="font-weight:bold;font-size:15px;color:#055c9d;">Branch - Jeddah</center></td>' 

		if branch == "Dammam - TSL-SA":
			data += '<td style="width:30%;border-color:#000000;color:#055c9d;"><center><b style="color:#055c9d;font-size:19px;font-weight:bold;">TSL Company</b></center><br><center style="font-weight:bold;font-size:15px;color:#055c9d;">Branch - Dammam</center></td>' 
	
		if branch == "Riyadh - TSL- KSA":
			data += '<td style="width:30%;border-color:#000000;color:#055c9d;"><center><b style="color:#055c9d;font-size:19px;font-weight:bold;">TSL Company</b></center><br><center style="font-weight:bold;font-size:15px;color:#055c9d;">Branch - Riyadh</center></td>' 
	
	if company == "TSL COMPANY - Kuwait":
		data += '<td style="width:30%;border-color:#000000;"><center><img src = "/files/kuwait flag.jpg" width ="120"></center></td>'

	if company == "TSL COMPANY - UAE":
		data += '<td style="width:30%;border-color:#000000;"><center><img src = "files/Flag_of_the_United_Arab_Emirates.svg.jpg" width ="120"></center></td>'
	
	if company == "TSL COMPANY - KSA":
		data += '<td style="width:30%;border-color:#000000;"><center></center></td>'
	
	data += '</tr>'
	data += '<tr>'
	data += '<td colspan = 2 align = left style="border-color:#000000;padding:1px;font-size:14px;font-size:12px;background-color:#0e86d4;color:white;"><b style="color:white;">From : %s    To : %s</b></td>' %(formatted_date_1,formatted_date_2)
	data += '<td colspan = 2 align = right style="border-left:hidden;border-color:#000000;padding:1px;font-size:14px;font-size:12px;background-color:#0e86d4;color:white;"><b style="color:white;">Generation Date: %s</b></td>' %(formatted_date)
	data += '</tr>'
	data += '</table>'

	data += '<table class="table table-bordered">'

	

	data += '<tr>'
	data += '<td style="border-color:#000000;width:10%;padding:1px;font-size:14px;font-size:11px;background-color:#0e86d4;color:white;font-weight:bold;text-align:center;">TECHNICIAN</td>'

	data += '<td style="border-color:#000000;width:7%;padding:1px;font-size:11px;background-color:#0e86d4;color:white;font-weight:bold;text-align:center;">Total RS WO Count</td>'
	data += '<td style="border-color:#000000;width:7%;padding:1px;font-size:11px;background-color:#0e86d4;color:white;font-weight:bold;text-align:center;">Total RS Amount</td>'
	data += '<td style="border-color:#000000;width:7%;padding:1px;font-size:11px;background-color:#0e86d4;color:white;font-weight:bold;text-align:center;">Total Material Cost</td>'
	data += '<td style="border-color:#000000;width:7%;padding:1px;font-size:11px;background-color:#0e86d4;color:white;font-weight:bold;text-align:center;">NET Amount</td>'

	data += '<td style="border-color:#000000;width:7%;padding:1px;font-size:11px;background-color:#0e86d4;color:white;font-weight:bold;text-align:center;">NER COUNT</td>'
	data += '<td style="border-color:#000000;width:7%;padding:1px;font-size:11px;background-color:#0e86d4;color:white;font-weight:bold;text-align:center;">NER % Against RS</td>'
	
	data += '<td style="border-color:#000000;width:7%;padding:1px;font-size:11px;background-color:#0e86d4;color:white;font-weight:bold;text-align:center;">NER DEDUCTION AMT</td>'
	data += '<td style="border-color:#000000;width:7%;padding:1px;font-size:11px;background-color:#0e86d4;color:white;font-weight:bold;text-align:center;">NER DEDUCTION COMMISSION</td>'
	data += '<td style="border-color:#000000;width:7%;padding:1px;font-size:11px;background-color:#0e86d4;color:white;font-weight:bold;text-align:center;">AMT AFTER DEDUCTION</td>'
	data += '<td style="border-color:#000000;width:7%;padding:1px;font-size:11px;background-color:#0e86d4;color:white;font-weight:bold;text-align:center;">COMMISSION</td>'
	data += '<td style="border-color:#000000;width:7%;padding:1px;font-size:11px;background-color:#0e86d4;color:white;font-weight:bold;text-align:center;">AFTER DEDUCTION COMMISSION</td>'
	# data += '<td style="border-color:#000000;width:7%;padding:1px;font-size:14px;font-size:12px;background-color:#0e86d4;color:white;"><center><b>TOTAL</b><center></td>'
	# data += '<td style="border-color:#000000;width:7%;padding:1px;font-size:14px;font-size:12px;background-color:#0e86d4;color:white;"><center><b>SUB TOTAL</b><center></td>'
	data += '</tr>'	


	
	sp = frappe.get_all("Employee",{"designation": ["in", ["Technician",'Senior Technician',"Automation Technician"]],"company":company,"status":"Active"},["*"])
	# wd = frappe.get_all("Work Order Data",{"status":"RSI-Repaired and Shipped Invoiced","posting_date": ["between", (from_date,to_date)]},["*"])
	for i in sp:
		data += '<tr>'
		data += '<td style="border-color:#000000;padding:1px;font-size:14px;font-size:12px;"><center><b>%s</b><center></td>' %(i.employee_name)
		wd = frappe.get_all("Work Order Data",{"technician":i.user_id,"invoice_no": ["!=", ""],"status_cap": ["=",""]},["*"])

		rs = frappe.db.sql(""" select count(DISTINCT `tabWork Order Data` .name) as ct from `tabWork Order Data` 
			left join `tabStatus Duration Details` on `tabWork Order Data`.name = `tabStatus Duration Details`.parent
			where  `tabStatus Duration Details`.status = "RS-Repaired and Shipped"  and `tabWork Order Data`.status_cap IS NULL
			and `tabWork Order Data`.technician = "%s" and DATE(`tabStatus Duration Details`.date) between '%s' and '%s' """ %(i.user_id,from_date,to_date) ,as_dict=1)

		ner = frappe.db.sql(""" select count(DISTINCT `tabWork Order Data` .name) as ct from `tabWork Order Data` 
			left join `tabStatus Duration Details` on `tabWork Order Data`.name = `tabStatus Duration Details`.parent
			where  `tabStatus Duration Details`.status = "NER-Need Evaluation Return"
			and `tabWork Order Data`.technician = "%s" and DATE(`tabStatus Duration Details`.date) between '%s' and '%s' """ %(i.user_id,from_date,to_date) ,as_dict=1)

	
		data += '<td style="border-color:#000000;padding:1px;font-size:14px;font-size:12px;"><center><b>%s</b><center></td>' %(rs[0]["ct"])
		
		count_a = 0
		count_b = 0
		count_c = 0
		count_d = 0

		wd = frappe.db.sql(""" select DISTINCT `tabWork Order Data` .name as ct from `tabWork Order Data` 
		left join `tabStatus Duration Details` on `tabWork Order Data`.name = `tabStatus Duration Details`.parent
		where  `tabStatus Duration Details`.status = "RS-Repaired and Shipped" and `tabWork Order Data`.status_cap IS NULL
		and `tabWork Order Data`.technician = "%s" and DATE(`tabStatus Duration Details`.date) between '%s' and '%s' """ %(i.user_id,from_date,to_date) ,as_dict=1)
		
		wds = frappe.db.sql(""" select DISTINCT `tabWork Order Data` .name as ct from `tabWork Order Data` 
			left join `tabStatus Duration Details` on `tabWork Order Data`.name = `tabStatus Duration Details`.parent
			where  `tabStatus Duration Details`.status = "RS-Repaired and Shipped"
			and `tabWork Order Data`.technician = "%s" and DATE(`tabStatus Duration Details`.date) between '%s' and '%s' """ %(i.user_id,from_date,to_date) ,as_dict=1)
			
		
		q_m = 0
		s_total = 0
		inv_total = 0
		
		for j in wds:
			ev = frappe.db.sql(""" select  `tabPart Sheet Item`.total as t from `tabEvaluation Report` 
			left join `tabPart Sheet Item` on `tabEvaluation Report`.name = `tabPart Sheet Item`.parent
			where  `tabEvaluation Report`.work_order_data = '%s' """ %(j.ct) ,as_dict=1)
			if ev:
				for e in ev:
					if e["t"]:
						inv_total = inv_total + e["t"]
			
			# s_amt = frappe.get_all("Supplier Quotation",{"work_order_data":j.ct,"workflow_state":"Approved By Management"},["*"])
			s_amt= frappe.db.sql(''' select base_total as b_am,shipping_cost as ship from `tabSupplier Quotation` 
			where Workflow_state in ("Approved By Management") and
			work_order_data = '%s' ''' %(j.ct) ,as_dict=1)
			if s_amt:
				for s in s_amt:
					# s_total = s_total + s.base_total
					s_cur = frappe.get_value("Supplier",{"name":s.supplier},["default_currency"])
					exr = get_exchange_rate(s_cur,"KWD")
					if s.shipping_cost:
						s_total = s_total + (s.shipping_cost * exr)
				# s_total = s_total + s_amt[0]["b_am"]

			

			
		for j in wd:
			
			q_amt= frappe.db.sql(''' select `tabQuotation`.name as q_name,
			`tabQuotation`.default_discount_percentage as dis,
			`tabQuotation`.is_multiple_quotation as is_m,
			`tabQuotation`.after_discount_cost as adc,
			`tabQuotation`.Workflow_state,
			`tabQuotation Item`.unit_price as up,
			`tabQuotation`.grand_total as gt,
			`tabQuotation Item`.margin_amount as mar,
			`tabQuotation Item`.margin_amount as ma from `tabQuotation` 
			left join `tabQuotation Item` on  `tabQuotation`.name = `tabQuotation Item`.parent
			where `tabQuotation`.Workflow_state in ("Approved By Customer") and
			`tabQuotation`.quotation_type in ("Customer Quotation - Repair","Revised Quotation - Repair")
			and `tabQuotation Item`.wod_no = '%s' ''' %(j.ct) ,as_dict=1)

			if company == "TSL Company - Kuwait":
				if q_amt:
					
					for k in q_amt:
						if k.is_m == 1:
							
							per = (k.up * k.dis)/100
							amt = k.up - per
							
							q_m = q_m + amt


						else:
							amt = k.adc
						
							q_m = q_m + amt

			else:
				if q_amt:			
					q_m = q_m + q_amt[0]["gt"]

		data += '<td style="border-color:#000000;padding:1px;font-size:14px;font-size:12px;"><center><b>%s</b><center></td>' %(f"{round(q_m):,}")
		data += '<td style="border-color:#000000;padding:1px;font-size:14px;font-size:12px;"><center><b>%s</b><center></td>' %(f"{round(s_total + inv_total):,}")
		data += '<td style="border-color:#000000;padding:1px;font-size:14px;font-size:12px;"><center><b>%s</b><center></td>' %(f"{round(c):,}") 
		data += '<td style="border-color:#000000;padding:1px;font-size:14px;font-size:12px;"><center><b>%s</b><center></td>' %(ner[0]["ct"])

		if rs[0]["ct"] == 0:
			data += '<td style="border-color:#000000;padding:1px;font-size:14px;font-size:12px;"><center><b>%s%s</b><center></td>' %(rs[0]["ct"],"%")
		else:
			data += '<td style="border-color:#000000;padding:1px;font-size:14px;font-size:12px;"><center><b>%s%s</b><center></td>' %(round((ner[0]["ct"]/rs[0]["ct"])*100),"%")
		# data += '<td style="border-color:#000000;padding:1px;font-size:14px;font-size:12px;"><center><b>%s</b><center></td>' %(ner[0]["ct"])

		nar = 0
		if rs[0]["ct"] == 0:
			nar = 0
		else:
			nar = round((ner[0]["ct"]/rs[0]["ct"])*100)
			
		to_rs = round(q_m,2)
		ner_deduct = (nar * to_rs)/100
		net_amount =  round((q_m - (s_total + inv_total)),2)
		data += '<td style="border-color:#000000;padding:1px;font-size:14px;font-size:12px;"><center><b>%s</b><center></td>' %(f"{round(ner_deduct):,}")
		data += '<td style="border-color:#000000;padding:1px;font-size:14px;font-size:12px;"><center><b>%s</b><center></td>' %(f"{round(ner_deduct*2/100):,}")
		aad = round(net_amount - ner_deduct,2)
		data += '<td style="border-color:#000000;padding:1px;font-size:14px;font-size:12px;"><center><b>%s</b><center></td>' %(f"{round(aad):,}")
		data += '<td style="border-color:#000000;padding:1px;font-size:14px;font-size:12px;"><center><b>%s</b><center></td>' %(f"{round(aad*2/100):,}") 

		ner_d_com =  round(ner_deduct*2/100)
		aad_d_com = round(aad*2/100)		
		data += '<td style="border-color:#000000;padding:1px;font-size:14px;font-size:12px;"><center><b>%s</b><center></td>' %(f"{aad_d_com-ner_d_com:,}") 

		# frappe.errprint(s_total)
		# 	sales = frappe.db.sql(""" select `tabSales Invoice`.posting_date,`tabSales Invoice Item`.amount  from `tabSales Invoice`
		# 	left join `tabSales Invoice Item` on `tabSales Invoice Item`.parent = `tabSales Invoice`.name
		# 	where `tabSales Invoice Item`.wod_no = '%s' or `tabSales Invoice Item`.work_order_data = '%s' and `tabSales Invoice`.status IN ('Paid', 'Overdue','Unpaid') """ %(j.name,j.name),as_dict = 1)
		# 	if sales:
		# 		from_date = datetime.strptime(str(from_date), "%Y-%m-%d").date()
		# 		to_date = datetime.strptime(str(to_date), "%Y-%m-%d").date()
		# 		if sales[0]["posting_date"] >= from_date and sales[0]["posting_date"] <= to_date:
		# 			if sales[0]["amount"] > 0 and sales[0]["amount"] < 320:
		# 				count_a = count_a + 1
						
		# 			elif sales[0]["amount"] >= 320 and sales[0]["amount"] < 640:
		# 				count_b = count_b + 2
						
		# 			elif sales[0]["amount"] >=640 and sales[0]["amount"] < 1200:
		# 				count_c = count_c + 3
						
		# 			elif sales[0]["amount"] >= 1200:
		# 				count_d = count_d + 4
						

		# data += '<td style="border-color:#000000;padding:1px;font-size:14px;font-size:12px;"><center><b>%s</b><center></td>' %(count_a)
		# data += '<td style="border-color:#000000;padding:1px;font-size:14px;font-size:12px;"><center><b>%s</b><center></td>' %(count_b)
		# data += '<td style="border-color:#000000;padding:1px;font-size:14px;font-size:12px;"><center><b>%s</b><center></td>' %(count_c)
		# data += '<td style="border-color:#000000;padding:1px;font-size:14px;font-size:12px;"><center><b>%s</b><center></td>' %(count_d)
		# t_cnt = count_a + count_b + count_c + count_d
		# data += '<td style="border-color:#000000;padding:1px;font-size:14px;font-size:12px;"><center><b>%s</b><center></td>' %(t_cnt)
		
		
		
		# ner = frappe.get_all("Work Order Data",{"technician":i.user_id,"status_cap_date": ["between", (from_date,to_date)]},["*"])
		# ner_count_a = 0
		# ner_count_b = 0
		# ner_count_c = 0
		# ner_count_d = 0
		
		# for k in ner:
		# 	sales = frappe.db.sql(""" select `tabSales Invoice`.name,`tabSales Invoice Item`.amount  from `tabSales Invoice`
		# 	left join `tabSales Invoice Item` on `tabSales Invoice Item`.parent = `tabSales Invoice`.name
		# 	where `tabSales Invoice Item`.wod_no = '%s' or `tabSales Invoice Item`.work_order_data = '%s' and `tabSales Invoice`.status IN ('Paid', 'Overdue','Unpaid') """ %(j.name,j.name),as_dict = 1)
		# 	if sales:
		# 		if sales[0]["amount"] > 0 and sales[0]["amount"] < 320:
		# 			ner_count_a = ner_count_a + 1
		# 		if sales[0]["amount"] >= 320 and sales[0]["amount"] < 640:
		# 			ner_count_b = ner_count_b + 2
		# 		if sales[0]["amount"] >=640 and sales[0]["amount"] <1200:
		# 			ner_count_c = ner_count_c + 3
		# 		if sales[0]["amount"] >= 1200:
		# 			ner_count_d = ner_count_d + 4
		
		# data += '<td style="border-color:#000000;padding:1px;font-size:14px;font-size:12px;"><center><b>%s</b><center></td>' %(ner_count_a)
		# data += '<td style="border-color:#000000;padding:1px;font-size:14px;font-size:12px;"><center><b>%s</b><center></td>' %(ner_count_b)
		# data += '<td style="border-color:#000000;padding:1px;font-size:14px;font-size:12px;"><center><b>%s</b><center></td>' %(ner_count_c)
		# data += '<td style="border-color:#000000;padding:1px;font-size:14px;font-size:12px;"><center><b>%s</b><center></td>' %(ner_count_d)
		# ner_t_cnt = ner_count_a + ner_count_b + ner_count_c + ner_count_d
		# data += '<td style="border-color:#000000;padding:1px;font-size:14px;font-size:12px;"><center><b>%s</b><center></td>' %(ner_t_cnt)
		
		# data += '<td style="border-color:#000000;padding:1px;font-size:14px;font-size:12px;"><center><b>%s</b><center></td>' %(t_cnt - ner_t_cnt)
		
		data += '</tr>'	
	data += '</table>'

	return data



from datetime import datetime, timedelta, date

@frappe.whitelist()
def get_leave_application(leave_application):
	from hrms.hr.doctype.leave_application.leave_application import get_leave_details
	leave_app = frappe.db.sql(""" select * from `tabLeave Application Form` where name = '%s' """%(leave_application),as_dict=1)
	if leave_app:
		
		for lap in leave_app:
			remaining_leaves = 0
			val = get_leave_details(lap.employee,today())
			if val['leave_allocation'].get("Annual Leave", {}).get("remaining_leaves"):
				remaining_leaves = (val['leave_allocation'].get("Annual Leave", {}).get("remaining_leaves"))
			from_date = lap.leave_start_date or lap.from_date
			first_of_month = from_date.replace(day=1)
			if first_of_month != from_date:
				before_day = from_date - timedelta(days=1)
				to_date = lap.leave_end_date
				worked_days = validate_balance_leaves(lap.company,first_of_month,before_day,lap.employee,lap.leave_type)
				leave_days = date_diff(to_date,from_date) +1
				holiday_count = get_holidays_no(lap.employee, lap.leave_start_date,lap.leave_end_date , holiday_list=None, company = lap.company)
				return first_of_month,before_day,lap.leave_start_date,lap.leave_end_date,worked_days,lap.leave_days,lap.leave_balance,holiday_count
			else:
				to_date = lap.leave_end_date
				return '','',lap.leave_start_date,lap.leave_end_date,0,lap.leave_days,lap.leave_balance,0

def get_holidays_no(employee, from_date, to_date, holiday_list=None, company = None):
	"""get holidays between two dates for the given employee"""
	from erpnext.setup.doctype.employee.employee import get_holiday_list_for_employee
	if not holiday_list:
		holiday_list = get_holiday_list_for_employee(employee)
	holidays = frappe.db.sql(
		"""select count(distinct holiday_date) from `tabHoliday` h1, `tabHoliday List` h2
		where h1.parent = h2.name and h1.holiday_date between %s and %s
		and h2.name = %s and h1.weekly_off = 0 """,
		(from_date, to_date, holiday_list),
	)[0][0]

	return holidays


import openpyxl
from openpyxl import Workbook
from six import BytesIO, string_types
from openpyxl.styles import Font, Alignment, Border, Side
@frappe.whitelist()
def salary_register_excel():
	args = frappe.local.form_dict
	filename = "Salary Summary"
	test = build_xlsx_response(filename)

def make_xlsx(data, sheet_name=None, wb=None, column_widths=None):
	args = frappe.local.form_dict
	column_widths = column_widths or []
	if wb is None:
		wb = openpyxl.Workbook()
	ws = wb.create_sheet(sheet_name, 0)
	month_name = args.month_name
	company = args.company
	docstatus = args.doc_status
	cyrix_employee = args.cyrix_employee
	from datetime import timedelta
	current_year = datetime.now().year
	month_number = datetime.strptime(month_name, '%B').month    
	first_date = datetime(current_year, month_number, 1)
	if month_number == 12:
		last_date = datetime(current_year, month_number, 31)
	else:
		next_month = datetime(current_year, month_number + 1, 1)
		last_date = next_month - timedelta(days=1)
	if company == "TSL COMPANY - Kuwait":
		filters = {'from_date': first_date.date(), 'to_date': last_date.date(), 'currency': 'KWD', 'company': company, 'docstatus': docstatus}
	if company == "TSL COMPANY - KSA":
		filters = {'from_date': first_date.date(), 'to_date': last_date.date(), 'currency': 'SAR', 'company': company, 'docstatus': docstatus}
	if company == "TSL COMPANY - UAE":
		filters = {'from_date': first_date.date(), 'to_date': last_date.date(), 'currency': 'AED', 'company': company, 'docstatus': docstatus}
	from tsl.tsl.report.salary_register_report.salary_register_report import execute
	result = execute(filters)
	headers = result[0]
	records = result[1]
	skip_columns = ['total_loan_repayment','company','salary_slip_id', 'branch', 'date_of_joining', 'start_date', 'end_date','currency','data_of_joining','department','leave_without_pay']
	int_columns = ['leave_without_pay', 'payment_days']
	currency_columns = [header['fieldname'] for header in headers if header.get('fieldtype') == 'Currency']
	filtered_headers = []
	for header in headers:
		if header['fieldname'] in skip_columns:
			continue
		if header['fieldname'] == 'loan':
			header['label'] = 'Loan Deduction'
		filtered_headers.append(header)
	for index, header in enumerate(filtered_headers):
		if company == "TSL COMPANY - KSA" and header['fieldname'] == 'employee_name':
			# Create a new header for Civil ID
			civil_id_header = {
				'fieldname': 'civil_id_no',
				'label': 'Civil ID',
				'fieldtype': 'Data'  # Assuming it's a string
			}
			filtered_headers.insert(index + 1, civil_id_header)  # Insert after employee_name

		if header['fieldname'] == 'designation':
			ctc_header = {
				'fieldname': 'ctc',
				'label': 'Total Salary',
				'fieldtype': 'Currency'
			}
			filtered_headers.insert(index + 1, ctc_header)
			break
	totals = {fieldname: 0.0 for fieldname in currency_columns}
	ws.append(["SALARY SUMMARY"])
	ws.append([""])
	if int(cyrix_employee) == 1:
		ws.append([company + " (CYRIX)"])
	else:
		ws.append([company])

	title = ["Sl.No."]
	for header in filtered_headers:
		title.append(header["label"])
	ws.append(title)
	cl_no = len(filtered_headers)+1
	for header in ws.iter_rows(min_row=1 , max_row=4, min_col=1, max_col=cl_no):
		for cell in header:
			cell.font = Font(bold=True)
	records = sorted(records, key=lambda x: int(x.get('employee', 0)))
	sl_no = 1
	for record in records:
		columns = [sl_no]
		cyrix = frappe.db.get_value("Employee",record.get('employee'),'cyrix_employee_')
		if cyrix == int(cyrix_employee):
			for header in filtered_headers:
				fieldname = header['fieldname']
				if fieldname == "civil_id_no" and company == "TSL COMPANY - KSA":
					fieldtype = "Data"
					cell_value = frappe.db.get_value("Employee",record.get("employee"),'civil_id_no')  # Default to 0 if None

				elif fieldname == "ctc":
					fieldtype = "Currency"
					cell_value = frappe.db.get_value("Employee",record.get("employee"),'ctc')

				else:
					fieldtype = header.get('fieldtype', '')
					cell_value = record.get(fieldname, 0)  # Default to 0 if None
				if fieldname in currency_columns:
					totals[fieldname] += float(cell_value) if isinstance(cell_value, (float, int)) else 0
				if fieldtype == 'Currency':
					cell_value = "{:,.2f}".format(cell_value) if isinstance(cell_value, (float, int)) else "0.000"
				elif isinstance(cell_value, float):
					cell_value = "{:,.2f}".format(cell_value)
				columns.append(cell_value)
			ws.append(columns)
			sl_no += 1
	total_row = []
	first_currency_index = next((i for i, header in enumerate(filtered_headers) if header['fieldname'] in currency_columns), len(filtered_headers))
	non_currency_colspan = first_currency_index + 1
	total_row.append("Total")
	if non_currency_colspan > 0:
		for _ in range(non_currency_colspan-1):
			total_row.append("")
	for i in range(non_currency_colspan, first_currency_index):
		data += '<td></td>'
		total_row.append("")
	for i, header in enumerate(filtered_headers[first_currency_index:], start=first_currency_index):
		fieldname = header['fieldname']
		if fieldname in currency_columns:
			total_value = "{:,.2f}".format(totals[fieldname])
		else:
			total_value = ''
		total_row.append(total_value)
	ws.append(total_row)
	align_center = Alignment(horizontal='center',vertical='center')
	ws.merge_cells(start_row=sl_no+4, start_column=1, end_row=sl_no+4, end_column=non_currency_colspan )
	ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=cl_no )
	ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=cl_no )

	for header in ws.iter_rows(min_row=sl_no+4 , max_row=sl_no+4, min_col=1, max_col=non_currency_colspan):
		for cell in header:
			cell.font = Font(bold=True)
	border_thin = Border(
	left=Side(style='thin'),
	right=Side(style='thin'),
	top=Side(style='thin'),
	bottom=Side(style='thin'))
	header_range = ws['A1':ws.cell(row=sl_no+4, column=cl_no).coordinate]
	for row in header_range:
		for cell in row:
			cell.border = border_thin
			cell.alignment = align_center
	xlsx_file = BytesIO()
	wb.save(xlsx_file)
	return xlsx_file

def build_xlsx_response(filename):
	xlsx_file = make_xlsx(filename)
	frappe.response['filename'] = filename + '.xlsx'
	frappe.response['filecontent'] = xlsx_file.getvalue()
	frappe.response['type'] = 'binary'

def build_xlsx_response2(filename):
	xlsx_file = make_xlsx2(filename)
	frappe.response['filename'] = filename + '.xlsx'
	frappe.response['filecontent'] = xlsx_file.getvalue()
	frappe.response['type'] = 'binary'


@frappe.whitelist()
def rfq_excel():
	args = frappe.local.form_dict
	filename = args.name

	test = build_xlsx_response2(filename)

def make_xlsx2(data, sheet_name=None, wb=None, column_widths=None):
	args = frappe.local.form_dict
	column_widths = column_widths or []
	if wb is None:
		wb = openpyxl.Workbook()
	ws = wb.create_sheet(sheet_name, 0)
	
	headers = []
	records = []
	ws.append(["Request for Quotation"])
	ws.append([""])
	rfq = frappe.get_doc("Request for Quotation",args.name)
	ws.append(["Sr","Sku","Category","Sub Category","Model / Part Number","Description","Package","Qty","Supplier Price"])
	sno = 0
	for i in rfq.items:
		sno = sno + 1
		cat = frappe.get_value("Item",i.item_code,["category_"])
		sub_cat = frappe.get_value("Item",i.item_code,["sub_category_name"])
		model = frappe.get_value("Item",i.item_code,["model_num"])
		des = frappe.get_value("Item",i.item_code,["description"])
		pac = frappe.get_value("Item",i.item_code,["package"])
		ws.append([sno,i.item_code,cat,sub_cat,model,des,pac,i.qty])
	
	for header in ws.iter_rows(min_row=1 , max_row=3, min_col=1, max_col=9):
		for cell in header:
			cell.font = Font(bold=True)

	xlsx_file = BytesIO()
	wb.save(xlsx_file)
	return xlsx_file


@frappe.whitelist()
def get_pi(posting_date,name,party_name,amount_in,total_allocated_amount,currency_paid,cost_center,references,remarks ):
	data = ""
	data+= '<table class="table table-bordered" style="border:1px solid black;" >'
	data+= '<tr><td colspan = 6><center><b style = "color:blue;font-size:15px">TSL COMPANY</b></center></td></tr>'
	data+= '<tr><td colspan = 2><center><b style = "color:red";>PAYMENT TRANSFER APPROVAL FORM</b></center></td></tr>'
	data+= '<tr><td>Date</td><td>%s</td></tr>' %(posting_date)
	data+='<tr><td>REF NO</td><td>%s</td></tr>' %(name)
	data+='<tr>  <td>Supplier Name</td><td>%s</td></tr>' %(party_name)
	data+='<tr> <td>Amount</td><td>%s</td></tr>' %("{:,.2f}".format(amount_in or total_allocated_amount))
	data+='<tr><td>Currency</td><td>%s</td></tr>' %(currency_paid)
	data+='<tr><td>Department</td><td>%s</td></tr>' %(cost_center)
	data+='<tr><td >Remarks</td><td>%s</td></tr>' %(remarks)
	for i in references:
		if i.reference_doctype == "Purchase Invoice" or i.reference_doctype == "Sales Invoice":
			pat = frappe.get_value("Purchase Invoice",{"name":i.reference_name},["supplier_invoice_attach"])
			pi = frappe.db.sql(""" select DISTINCT `tabPurchase Invoice`.name as p,`tabPurchase Invoice Item`.work_order_data as wo,`tabPurchase Invoice Item`.supply_order_data as so from `tabPurchase Invoice` 
			left join `tabPurchase Invoice Item` on `tabPurchase Invoice`.name = `tabPurchase Invoice Item`.parent 
			where `tabPurchase Invoice`.name = '%s' """ %(i.reference_name),as_dict =1)
			data+='<tr><td>Attached With Supporting Document</td><td><b>%s</b>/ <a href="%s"><u><b style = "color:red">Supplier Invoice Link</b></u></a></td><br>'%(i.reference_name,pat)
			for j in pi:
				if j["wo"]:
					data+='<tr><td></td><td><a href="https://erp.tsl-me.com/app/work-order-data/%s">%s</a></td></tr>' %(j["wo"],j["wo"])
				if j["so"]:
					data+='<tr><td></td><td><a href="https://erp.tsl-me.com/app/supply-order-data/%s">%s</a></td></tr>' %(j["so"],j["so"])

		if i.reference_doctype == "Journal Entry":
			je_attach = frappe.get_value("Journal Entry",{"name":i.reference_name},["attach"])
			data+='<tr><td>Attached With Supporting Document</td><td><b>%s</b>/ <a href="%s"><u><b style = "color:red">Journal Entry Attachment</b></u></a></td><br>'%(i.reference_name,je_attach)

	data+= "</table>"
	return data

@frappe.whitelist()
def get_q(name):
	qu = frappe.db.sql(""" select `tabQuotation Item`.wod_no,`tabQuotation`.name as q  from `tabQuotation` left join 
	`tabQuotation Item` on `tabQuotation`.name = `tabQuotation Item`.parent where `tabQuotation`.name = '%s' """ %(name) ,as_dict=1)
	count = 0
	for i in qu:
		if i['wod_no']:
			ev = frappe.db.exists("Evaluation Report",{"work_order_data":i['wod_no']})
			if not ev:
				count = count + 1
	
	if count > 0:     
		data = ""
		data+= '<table class="table table-bordered" style="border:1px solid black;" >'
		data+= '<tr><td colspan = 8><center><b style = "color:blue;font-size:15px">TSL COMPANY</b></center></td></tr>'
		data+= '<tr><td colspan = 8><center><b style = "color:red";>PAYMENT TRANSFER APPROVAL FORM</b></center></td></tr>'
		data+= "</table>"
		return "Yes"




@frappe.whitelist()
def get_sales_table(name):
	ic = frappe.get_doc("Invoice Cancellation",name)

	if ic:
		for i in ic.cancellation_list:
			customer = frappe.get_value("Sales Invoice",i.invoice_no,"customer")
			data = ""
			ogdate = datetime.strptime(str(ic.date),"%Y-%m-%d")
			formatted_date = ogdate.strftime("%d-%m-%Y")
			data+= '<p><span style="font-size:14px;font-weight:bold">Customer : %s</span>&emsp;&emsp;&emsp;&emsp;&emsp;&emsp;&emsp;&emsp;&emsp;&emsp;&emsp;&emsp;&emsp;&emsp;&emsp;&emsp;&emsp;&emsp;&emsp;&emsp;<b style = "text-align:left">Date : %s</b></p><br><br>' %(customer,formatted_date)
			
	
	data+= '<tr>'
	data+= '<td style = "text-align:center;font-weight:bold;"><b>Sr</b></td>'
	data+= '<td style = "text-align:center;font-weight:bold;"><b>Invoice No</b></td>'
	if ic:
		for i in ic.cancellation_list:
			si = frappe.get_doc("Sales Invoice",i.invoice_no)
			if si.cost_center == "Repair - TSL":
				data+= '<td style = "text-align:center;font-weight:bold;"><b>Work Order</b></td>'
			if si.cost_center == "Supply - TSL":
				data+= '<td style = "text-align:center;font-weight:bold;"><b>Supply Order</b></td>'
	data+= '<td style = "text-align:center;font-weight:bold;"><b>Model</b></td>'
	data+= '<td style = "text-align:center;font-weight:bold;"><b>Mfg</b></td>'
	data+= '<td style = "text-align:center;font-weight:bold;"><b>Serial No</b></td>'
	data+= '<td style = "text-align:center;font-weight:bold;"><b>Amount</b></td>'
	data+='</tr>'

	if ic:
		for i in ic.cancellation_list:
			if i.invoice_no:
				si = frappe.get_doc("Sales Invoice",i.invoice_no)
				count = 0
				for idx, k in enumerate(si.items):
					count = count + 1
					data+= '<tr>'
					data+= '<td style = "text-align:center;">%s</td>' %(count)
					if idx == 0:
						data+= '<td style = "text-align:center;">%s</td>' %(i.invoice_no)
					else:
						data+= '<td style = "text-align:center;">%s</td>' %("")
					if idx == 0:
						if k.wod_no or k.work_order_data:
							data += '<td style="text-align:center;">%s</td>' % (k.wod_no or k.work_order_data)
						if k.supply_order_data:
							data += '<td style="text-align:center;">%s</td>' % (k.supply_order_data)
					else:
						if k.wod_no or k.work_order_data:
							data += '<td style="text-align:center;">%s</td>' % ("")
						if k.supply_order_data:
							data += '<td style="text-align:center;">%s</td>' % ("")

					md = frappe.get_value("Item Model",k.model,"model")
					data+= '<td style = "text-align:center;">%s</td>' %(md)
					data+= '<td style = "text-align:center;">%s</td>' %(k.manufacturer)
					data+= '<td style = "text-align:center;">%s</td>' %(k.serial_number or "-")
					data+= '<td style = "text-align:center;">%s</td>'  % (f"{k.amount or 0:,.2f}")
					data+='</tr>'



	return data

@frappe.whitelist()
def get_mc(name):
	qu = frappe.db.sql(""" select `tabTechnician Hours Spent`.work_order_data,`tabTechnician Hours Spent`.total_price from `tabQuotation` left join 
	`tabTechnician Hours Spent` on `tabQuotation`.name = `tabTechnician Hours Spent`.parent where `tabQuotation`.name = '%s' ORDER BY 
	`tabTechnician Hours Spent`.work_order_data """ %(name) ,as_dict=1)

	qi = frappe.db.sql(""" select `tabQuotation Item`.wod_no from `tabQuotation` left join 
	`tabQuotation Item` on `tabQuotation`.name = `tabQuotation Item`.parent where `tabQuotation`.name = '%s' """ %(name) ,as_dict=1)

	data = ""
	# data+= '<table class="table table-bordered" style= "border:1px solid grey;">'
	data+= '<tr>'
	data+= '<td style="text-align:center;background-color:#E0E0E0"><b>WO-NO</b></td>'
	data+= '<td style="text-align:center;background-color:#E0E0E0"><b>Labour Cost</b></td>'
	data+= '<td style="text-align:center;background-color:#E0E0E0"><b>Material Cost</b></td>'
	data+= '<td style="text-align:center;background-color:#E0E0E0"><b>Unit Price</b></td>'
	data+='</tr>'

	dat= []
	for i in qu:
		dat.append(i.work_order_data)
		
		mc = frappe.db.sql(""" select sum(`tabItem Price Details`.amount) as mcost from `tabQuotation` left join 
		`tabItem Price Details` on `tabQuotation`.name = `tabItem Price Details`.parent 
		where `tabItem Price Details`.work_order_data = '%s' and `tabItem Price Details`.parent = '%s' """ %(i.work_order_data,name) ,as_dict=1)
		frappe.errprint(mc)
		up = frappe.db.sql(""" select `tabQuotation Item`.unit_price as up from `tabQuotation` left join 
		`tabQuotation Item` on `tabQuotation`.name =  `tabQuotation Item`.parent
		where  `tabQuotation Item`.wod_no= '%s' and  `tabQuotation Item`.parent = '%s' """ %(i.work_order_data,name) ,as_dict=1)
		
		import re
		input_string = i.work_order_data
		numeric_part = re.search(r'\d+$', input_string).group()

		data+= '<tr>'
		data+= '<td style="text-align:center;">%s</td>' %(numeric_part)
		data+= '<td style="text-align:center;"><b style = "color:red">%s</b></td>' %(f"{i.total_price:.2f}")
		mc[0]['mcost'] = (mc[0]['mcost'] or 0)
		frappe.errprint(type(mc[0]['mcost']))
		data+= '<td style="text-align:center;"><b style = "color:red">%s</b></td>' %(f"{mc[0]['mcost']:.2f}")
		data+= '<td style="text-align:center;"><b style = "color:red">%s</b></td>' %(f"{up[0]['up']:.2f}")
		data+= '</tr>'
	
	for k in qi:
		if not k.wod_no in dat:
			up_1 = frappe.db.sql(""" select `tabQuotation Item`.unit_price as up from `tabQuotation` left join 
			`tabQuotation Item` on `tabQuotation`.name =  `tabQuotation Item`.parent
			where  `tabQuotation Item`.wod_no= '%s' and  `tabQuotation Item`.parent = '%s' """ %(k.wod_no,name) ,as_dict=1)

			import re
			input_string = k.wod_no
			numeric_part = re.search(r'\d+$', input_string).group()

			data+= '<tr>'
			data+= '<td style="text-align:center;">%s</td>' %(numeric_part)
			data+= '<td style="text-align:center;"><b style = "color:red">%s</b></td>' %("0.00")
			data+= '<td style="text-align:center;"><b style = "color:red">%s</b></td>'%("0.00")
			data+= '<td style="text-align:center;"><b style = "color:red">%s</b></td>' %(f"{up_1[0]['up']:.2f}")
			data+= '</tr>'
	# data+= "</table>"
	 
	
	return data

@frappe.whitelist()
def get_mc_2(name):
	qu = frappe.db.sql(""" select `tabTechnician Hours Spent`.work_order_data,`tabTechnician Hours Spent`.total_price from `tabQuotation` left join 
	`tabTechnician Hours Spent` on `tabQuotation`.name = `tabTechnician Hours Spent`.parent where `tabQuotation`.name = '%s' ORDER BY 
	`tabTechnician Hours Spent`.work_order_data """ %(name) ,as_dict=1)

	qi = frappe.db.sql(""" select `tabQuotation Item`.wod_no from `tabQuotation` left join 
	`tabQuotation Item` on `tabQuotation`.name = `tabQuotation Item`.parent where `tabQuotation`.name = '%s' """ %(name) ,as_dict=1)

	data = ""
	# data+= '<table class="table table-bordered" style= "border:1px solid grey;">'
	data+= '<tr>'
	data+= '<td style="text-align:center;background-color:#E0E0E0"><b>WO-NO</b></td>'
	data+= '<td style="text-align:center;background-color:#E0E0E0"><b>Labour Cost</b></td>'
	data+= '<td style="text-align:center;background-color:#E0E0E0"><b>Material Cost</b></td>'
	data+= '<td style="text-align:center;background-color:#E0E0E0"><b>Unit Price</b></td>'
	data+='</tr>'

	dat= []
	for i in qu:
		dat.append(i.work_order_data)
		
		up = frappe.db.sql(""" select `tabQuotation Item`.unit_price as up from `tabQuotation` left join 
		`tabQuotation Item` on `tabQuotation`.name =  `tabQuotation Item`.parent
		where  `tabQuotation Item`.wod_no= '%s' and  `tabQuotation Item`.parent = '%s' """ %(i.work_order_data,name) ,as_dict=1)
		
		import re
		input_string = i.work_order_data
		numeric_part = re.search(r'\d+$', input_string).group()

		data+= '<tr>'
		data+= '<td style="text-align:center;">%s</td>' %(numeric_part)
		data+= '<td style="text-align:center;"><b style = "color:red">%s</b></td>' %(f"{i.total_price:.2f}")
		data+= '<td style="text-align:center;"><b style = "color:red">%s</b></td>' %("0.00")
		data+= '<td style="text-align:center;"><b style = "color:red">%s</b></td>' %(f"{up[0]['up']:.2f}")
		data+= '</tr>'
	
	for k in qi:
		if not k.wod_no in dat:
			up_1 = frappe.db.sql(""" select `tabQuotation Item`.unit_price as up from `tabQuotation` left join 
			`tabQuotation Item` on `tabQuotation`.name =  `tabQuotation Item`.parent
			where  `tabQuotation Item`.wod_no= '%s' and  `tabQuotation Item`.parent = '%s' """ %(k.wod_no,name) ,as_dict=1)

			import re
			input_string = k.wod_no
			numeric_part = re.search(r'\d+$', input_string).group()

			data+= '<tr>'
			data+= '<td style="text-align:center;">%s</td>' %(numeric_part)
			data+= '<td style="text-align:center;"><b style = "color:red">%s</b></td>' %("0.00")
			data+= '<td style="text-align:center;"><b style = "color:red">%s</b></td>'%("0.00")
			data+= '<td style="text-align:center;"><b style = "color:red">%s</b></td>' %(f"{up_1[0]['up']:.2f}")
			data+= '</tr>'
	# data+= "</table>"
	 
	
	return data

@frappe.whitelist()
def crt_items(import_file):
	from datetime import datetime
	filepath = get_file(import_file)
	data = read_csv_content(filepath[1])
	for i in data:
		s = frappe.db.exists("Item",{"model":i[6],"category":i[4],"sub_category":i[5]})
		if s:
			# k = frappe.get_value("Item",{"model":i[6]})
			print(s)
	# ct = 0
	# s = frappe.new_doc("Stock Reconciliation")
	# s.company = "TSL COMPANY - UAE"
	# s.purpose = "Opening Stock"
	# s.expense_account = "Test Account - TSL-UAE"
	# for i in data:
	#     print(i[0])
	#     s.append("items", {
	#     'item_code':i[0],
	#     'warehouse':"Dubai - TSL - TSL-UAE",
	#     "qty" :i[2],
	#     "valuation_rate":i[3]
					
	#     })
	
	# s.save(ignore_permissions = 1)
	# print("yes")

	# for i in data[1:]:
	#     if i[6]:
	#         c = frappe.db.exists("Item",{"model":i[6]})
	#         if not c:
	#             print(i[6])
	#             ct = ct+1
	#             cat = frappe.new_doc("Item")
	#             cat.naming_series = "P.######"
	#             md = frappe.get_value("Item Model",{"model":i[6]},["name"])
	#             # print(md)
	#             cat.model = md
	#             cat.model_num = i[6]
	#             k = frappe.db.exists("Sub Category",{"sub_category":i[5],"category":i[4]})
	#             if k:
	#                 sc = frappe.get_value("Sub Category",k)
	#                 cat.sub_category = sc
	#             cat.category_ = i[4]
	#             cat.package = i[7]
	#             cat.valuation_rate = i[9]
	#             cat.item_group = "Components"
	#             cat.opening_qty = i[8]
	#             cat.save(ignore_permissions = 1)
	# print(ct)


@frappe.whitelist()
def get_documents(name):
	dn = frappe.get_value("Work Order Data",{"name":name},["dn_no"])
	si = frappe.get_value("Work Order Data",{"name":name},["invoice_no"])
	per = frappe.get_value("Work Order Data",{"name":name},["payment_entry_reference"])
	# frappe.errprint(dn)
	data = ""
	data+= '<table class="table table-bordered">'
	if dn or si or per: 
		data+= '<tr><td colspan = 6 align = center ><b style= "color:blue;font-size:14px">Linked Documents</b></td></tr>'
		if dn:  
			data+= '<tr><td><b style = "color:blue;font-size:14px">Delivery Note</b></td><td style = "color:red;font-size:14px">%s</td><td><a style = "color:red;font-size:14px;" href="https://erp.tsl-me.com/app/delivery-note/%s" target="_blank" >VIEW</a><span>&#128065;</span></td></tr>' %(dn,dn)
		if si:
			data+= '<tr><td ><b style = "color:blue;font-size:14px">Sales Invoice</b></td><td style = "color:red;font-size:14px">%s</td><td><a style = "color:red;font-size:14px;" href="https://erp.tsl-me.com/app/sales-invoice/%s" target="_blank" >VIEW</a><span>&#128065;</span></td></tr>' %(si,si)
		if per:
			data+= '<tr><td><b style = "color:blue;font-size:14px">Payment Entry</b></td><td style = "color:red;font-size:14px">%s</td><td><a style = "color:red;font-size:14px;" href="https://erp.tsl-me.com/app/payment-entry/%s" target="_blank" >VIEW</a><span>&#128065;</span></td></tr>' %(per,per)
		data+= "</table>"
	
	return data


# @frappe.whitelist()
# def get_pre_eval():
#     qu = frappe.db.sql(""" select `tabQuotation Item`.wod_no,`tabQuotation`.name as q  from `tabQuotation` left join 
#     `tabQuotation Item` on `tabQuotation`.name = `tabQuotation Item`.parent """  ,as_dict=1)
#     count = 0
#     for i in qu:
#         if i['wod_no']:
#             ev = frappe.db.exists("Evaluation Report",{"work_order_data":i['wod_no']})
#             if not ev:
#                 print(i["wod_no"])
			#     count = count + 1
	
@frappe.whitelist()
def get_active_customer():
	cus = frappe.get_all("Customer",["*"])
	c = []
	count = 0
	wod = frappe.get_all("Work Order Data",{"posting_date": ["between", ("2024-01-01","2024-09-17")]},["*"])
	for k in wod:
	
		c.append(k.customer)
	# print(c)

	for i in cus:
		if i.name not in c:
			# print(i.name)
			count = count + 1
			s = frappe.db.exists("Sales Team",{"parent":i.name})
			sales = "No Sales Person"
			if s:
				sl = frappe.get_doc("Sales Team",s)
				# print(sl.sales_person)
				sales = sl.sales_person
			# else:
			#     print("No Sales Person")

			print(i.name,",",sales,",",i.customer_primary_contact,",",i.mobile_no)
			# print("Primary Contact :",i.customer_primary_contact)
			# print("Mobile:" ,i.mobile_no)
			print("")
	print(count)

@frappe.whitelist()
def update_item_list():
	# count = 0
	wo = frappe.db.sql(""" UPDATE `tabItem` SET has_serial_no = 0 WHERE name = "002304" ; """)
	# old_wd_rsi = frappe.db.sql(""" select sum(`tabWork Order Data`.old_wo_q_amount) as rsi from `tabWork Order Data` where `tabWork Order Data`.sales_rep = "%s" and `tabWork Order Data`.posting_date between "%s" and "%s" and `tabWork Order Data`.status = "RSI-Repaired and Shipped Invoiced" """ %("Vazeem","2015-01-01","2024-10-06") ,as_dict=1)
	# old_wd_rsi = frappe.get_all("Work Order Data",{"sales_rep":"Vazeem","status":"RSI-Repaired and Shipped Invoiced","posting_date": ["between", ("2015-01-01","2024-10-06")],"old_wo_q_amount": [">", 0]},["old_wo_q_amount"])
	# rsi_old = 0
	# for i in old_wd_rsi:
	#     rsi_old = rsi_old + i.old_wo_q_amount
	# print(rsi_old)
	# wds = frappe.db.sql(""" select DISTINCT `tabWork Order Data` .name  from `tabWork Order Data` 
	# left join `tabStatus Duration Details` on `tabWork Order Data`.name = `tabStatus Duration Details`.parent
	# where  `tabStatus Duration Details`.status = "RSI-Repaired and Shipped Invoiced"
	# and `tabWork Order Data`.sales_rep = "%s" and `tabWork Order Data`.posting_date between "2015-01-01" and "2024-10-06" """ %("Vazeem") ,as_dict=1)
	
	# for i in wds:
	#     print(i.name)
	#     count = count +1
	# print(count)

@frappe.whitelist()
def check_wod():

	from_date = "2015-01-01"
	to_date = "2024-10-12"
	company = "TSL COMPANY - Kuwait"
	data = ''
	data += '<div class="table-container">'
	data += '<table class="table table-bordered" width: 100%;>'
	data += '<tr>'
	data += '<td colspan = 3 style="border-color:#000000;"><img src = "/files/TSL Logo.png" align="left" width ="150"></td>'
	data += '<td colspan = 3 style="border-color:#000000;"><h2><center><b>TSL Company</b></center></h2></td>'
	data += '<td colspan = 3 style="border-color:#000000;"><center><img src = "/files/kuwait flag.jpg" width ="100"></center></td>'
	
	data += '</tr>'

	data += '<tr>'
	data += '<td colspan = 9 style="border-color:#000000;"><b>RSC & Quoted Summary</b></td>'
		
	data += '</tr>'

	data += '<tr>'
	data += '<td colspan = 1 style="border-color:#000000;"><center><b></b></center></td>'
	data += '<td colspan = 2 style="border-color:#000000;"><center><b>Total Amount(RSI)</b></center></td>'
	data += '<td colspan = 2 style="border-color:#000000;"><center><b>Total Amount(RSC)</b></center></td>'
	data += '<td colspan = 2 style="border-color:#000000;"><center><b>Total Amount(RS)</b></center></td>'
	data += '<td colspan = 2 style="border-color:#000000;"><center><b>Total Amount(Quoted)</b></center></td>'
	
	data += '</tr>'
	wo = frappe.db.sql(""" select DISTINCT sales_rep from `tabWork Order Data` where company = '%s' """ %(company) ,as_dict =1)
	data += '<tr>'
	data += '<td style="border-color:#000000;width:10%"><center><b>Salesman</b></center></td>'
	data += '<td style="border-color:#000000;width:10%;"><center><b>WO(in KD)</b></center></td>'
	data += '<td style="border-color:#000000;width:10%;"><center><b>SO(in KD)</b></center></td>'
	data += '<td style="border-color:#000000;width:10%;"><center><b>WO(in KD)</b></center></td>'
	data += '<td style="border-color:#000000;width:10%;"><center><b>SO(in KD)</b></center></td>'
	data += '<td style="border-color:#000000;width:10%;"><center><b>WO(in KD)</b></center></td>'
	data += '<td style="border-color:#000000;width:10%;"><center><b>SO(in KD)</b></center></td>'
	data += '<td style="border-color:#000000;width:10%;"><center><b>WO(in KD)</b></center></td>'  
	data += '<td style="border-color:#000000;width:10%;"><center><b>SO(in KD)</b></center></td>'	
	data += '</tr>'
	current_date_time = datetime.now()

	# Extract and print the current date
	current_date = to_date
	# Week_start = add_days(current_date,-6)
	Week_start = from_date

	total_rsi = 0
	total_rsc = 0
	total_rs = 0
	total_quotation = 0

	s_total_rsi = 0
	s_total_rsc = 0
	s_total_rs = 0
	s_total_quotation = 0
	
	for i in wo:
		if not i["sales_rep"] == None:
			# if not i["sales_rep"] == 'Sales Team' and not i["sales_rep"] == 'Walkin' and not i["sales_rep"] == 'Omar' and not i["sales_rep"] == '':
			if i["sales_rep"] == 'Vazeem':
	 
				# wds = frappe.db.sql(""" select DISTINCT `tabWork Order Data` .name  from `tabWork Order Data` 
				# left join `tabStatus Duration Details` on `tabWork Order Data`.name = `tabStatus Duration Details`.parent
				# where  `tabStatus Duration Details`.status = "RSI-Repaired and Shipped Invoiced"
				# and `tabWork Order Data`.sales_rep = "%s" and `tabWork Order Data`.posting_date between "2015-01-01" and "2024-10-06" """ %(i["sales_rep"]) ,as_dict=1)

				wds = frappe.get_all("Work Order Data",{"sales_rep":i["sales_rep"],"status":"RSI-Repaired and Shipped Invoiced","posting_date": ["between", (Week_start,current_date)]},["*"])
				# RSI
				
				squot = 0
				for j in wds:
					rsi_q = frappe.db.sql(''' select DISTINCT `tabQuotation Item`.wod_no,`tabQuotation`.after_discount_cost, `tabQuotation`.Workflow_state,`tabQuotation Item`.margin_amount from `tabQuotation` 
					left join `tabQuotation Item` on  `tabQuotation`.name = `tabQuotation Item`.parent
					where `tabQuotation`.Workflow_state in ("Approved By Customer","Quoted to Customer")  and `tabQuotation Item`.wod_no = %s  and `tabQuotation`.is_multiple_quotation = 0 ''',j.name,as_dict=1)
					if rsi_q:
						for k in rsi_q:
							squot = squot + k["after_discount_cost"]
						   
					else:
						rsi_q_2 = frappe.db.sql(''' select DISTINCT `tabQuotation Item`.wod_no,`tabQuotation`.after_discount_cost, `tabQuotation`.Workflow_state,`tabQuotation Item`.margin_amount from `tabQuotation` 
						left join `tabQuotation Item` on  `tabQuotation`.name = `tabQuotation Item`.parent
						where `tabQuotation`.Workflow_state in ("Rejected by Customer")  and `tabQuotation Item`.wod_no = %s  and `tabQuotation`.is_multiple_quotation = 0 ''',j.name,as_dict=1)
						if rsi_q_2:
							for k in rsi_q_2:
								squot = squot + k["after_discount_cost"]
							 

						# else:
						#     rsi_q_3 = frappe.db.sql(''' select DISTINCT `tabQuotation Item`.wod_no,`tabQuotation`.after_discount_cost, `tabQuotation`.Workflow_state,`tabQuotation Item`.margin_amount from `tabQuotation` 
						#     left join `tabQuotation Item` on  `tabQuotation`.name = `tabQuotation Item`.parent
						#     where `tabQuotation`.Workflow_state in ("Rejected")  and `tabQuotation Item`.wod_no = %s  and `tabQuotation`.is_multiple_quotation = 0 ''',j.name,as_dict=1)
						#     if rsi_q_3:
						#         for k in rsi_q_3:
						#             squot = squot + k["after_discount_cost"]

				mquot = 0
				for j in wds:
					rsi_mq = frappe.db.sql(''' select DISTINCT `tabQuotation Item`.wod_no,`tabQuotation`.after_discount_cost, `tabQuotation`.Workflow_state,`tabQuotation Item`.margin_amount from `tabQuotation` 
					left join `tabQuotation Item` on  `tabQuotation`.name = `tabQuotation Item`.parent
					where  `tabQuotation`.Workflow_state in ("Approved By Customer","Quoted to Customer") and `tabQuotation Item`.wod_no = %s  and `tabQuotation`.is_multiple_quotation = 1 ''',j.name,as_dict=1)
					if rsi_mq:
						for k in rsi_mq:
							mquot = mquot + k["margin_amount"]
						   

					else:
						rsi_mq_2 = frappe.db.sql(''' select DISTINCT `tabQuotation Item`.wod_no,`tabQuotation`.after_discount_cost, `tabQuotation`.Workflow_state,`tabQuotation Item`.margin_amount from `tabQuotation` 
						left join `tabQuotation Item` on  `tabQuotation`.name = `tabQuotation Item`.parent
						where  `tabQuotation`.Workflow_state in ("Rejected by Customer") and `tabQuotation Item`.wod_no = %s  and `tabQuotation`.is_multiple_quotation = 1 ''',j.name,as_dict=1)
						if rsi_mq_2:
							for k in rsi_mq_2:
								mquot = mquot + k["margin_amount"]
							   


						
				
				# old_wd_rsi = frappe.db.sql(""" select sum(`tabWork Order Data`.old_wo_q_amount) as rsi from `tabWork Order Data` where `tabWork Order Data`.sales_rep = "%s" and `tabWork Order Data`.posting_date between "%s" and "%s" and `tabWork Order Data`.status = "RSI-Repaired and Shipped Invoiced" """ %(i["sales_rep"],Week_start,current_date) ,as_dict=1)
				# old_wd_rsi = frappe.get_all("Work Order Data",{"sales_rep":i["sales_rep"],"status":"RSI-Repaired and Shipped Invoiced","posting_date": ["between", (Week_start,current_date)],"old_wo_q_amount": [">", 0]},["old_wo_q_amount"])
				 
				
				rsi_total = squot + mquot 
			   
				#Invoiced
				sup_in = frappe.get_all("Supply Order Data",{"sales_rep":i["sales_rep"],"status":"Invoiced","posting_date": ["between", (Week_start,current_date)]})
			   
				sqm_i = 0
				for j in sup_in:
					sup_i_q = frappe.db.sql(''' select `tabQuotation`.after_discount_cost, `tabQuotation`.Workflow_state,`tabQuotation Item`.margin_amount from `tabQuotation` 
					left join `tabQuotation Item` on  `tabQuotation`.name = `tabQuotation Item`.parent
					where `tabQuotation`.Workflow_state in ("Approved By Customer","Quoted to Customer","Rejected by Customer")  and `tabQuotation Item`.supply_order_data = %s  and `tabQuotation`.is_multiple_quotation = 0 ''',j.name,as_dict=1)
					
					for k in sup_i_q:
						sqm_i = sqm_i + k["after_discount_cost"]
			

				sqm_i_2 = 0
				for j in sup_in:
					sup_i_q_2 = frappe.db.sql(''' select `tabQuotation`.after_discount_cost, `tabQuotation`.Workflow_state,`tabQuotation Item`.margin_amount from `tabQuotation` 
					left join `tabQuotation Item` on  `tabQuotation`.name = `tabQuotation Item`.parent
					where `tabQuotation`.Workflow_state in ("Approved By Customer","Quoted to Customer","Rejected by Customer")  and `tabQuotation Item`.supply_order_data = %s  and `tabQuotation`.is_multiple_quotation = 1 ''',j.name,as_dict=1)
					
					for k in sup_i_q_2:
						sqm_i_2 = sqm_i_2 + k["margin_amount"]
			
				sqm_total = sqm_i + sqm_i_2

				 #Delivered
				sup_del = frappe.get_all("Supply Order Data",{"sales_rep":i["sales_rep"],"status":"Delivered","posting_date": ["between", (Week_start,current_date)]})
			   
				sqm_d = 0
				for j in sup_del:
					sup_d_q = frappe.db.sql(''' select `tabQuotation`.after_discount_cost, `tabQuotation`.Workflow_state,`tabQuotation Item`.margin_amount from `tabQuotation` 
					left join `tabQuotation Item` on  `tabQuotation`.name = `tabQuotation Item`.parent
					where `tabQuotation`.Workflow_state in ("Approved By Customer","Quoted to Customer","Rejected by Customer")  and `tabQuotation Item`.supply_order_data = %s  and `tabQuotation`.is_multiple_quotation = 0 ''',j.name,as_dict=1)
					
					for k in sup_d_q:
						sqm_d = sqm_d + k["after_discount_cost"]
			

				sqm_d_2 = 0
				for j in sup_del:
					sup_d_q_2 = frappe.db.sql(''' select `tabQuotation`.after_discount_cost, `tabQuotation`.Workflow_state,`tabQuotation Item`.margin_amount from `tabQuotation` 
					left join `tabQuotation Item` on  `tabQuotation`.name = `tabQuotation Item`.parent
					where `tabQuotation`.Workflow_state in ("Approved By Customer","Quoted to Customer","Rejected by Customer")  and `tabQuotation Item`.supply_order_data = %s  and `tabQuotation`.is_multiple_quotation = 1 ''',j.name,as_dict=1)
					
					for k in sup_d_q_2:
						sqm_d_2 = sqm_d_2 + k["margin_amount"]
			
				sqm_del_total = sqm_d + sqm_d_2

				#Quoted
				sup_quo = frappe.get_all("Supply Order Data",{"sales_rep":i["sales_rep"],"status":"Quoted","posting_date": ["between", (Week_start,current_date)]})
			   
				sqm_quo = 0
				for j in sup_quo:
					sup_quo_q = frappe.db.sql(''' select DISTINCT `tabQuotation`.after_discount_cost, `tabQuotation`.Workflow_state,`tabQuotation Item`.margin_amount from `tabQuotation` 
					left join `tabQuotation Item` on  `tabQuotation`.name = `tabQuotation Item`.parent
					where `tabQuotation`.Workflow_state in ("Approved By Customer","Quoted to Customer")  and `tabQuotation Item`.supply_order_data = %s  and `tabQuotation`.is_multiple_quotation = 0 ''',j.name,as_dict=1)
					if sup_quo_q:
						for k in sup_quo_q:
							sqm_quo = sqm_quo + k["after_discount_cost"]
					else:
						sup_quo_q_2 = frappe.db.sql(''' select DISTINCT `tabQuotation`.after_discount_cost, `tabQuotation`.Workflow_state,`tabQuotation Item`.margin_amount from `tabQuotation` 
						left join `tabQuotation Item` on  `tabQuotation`.name = `tabQuotation Item`.parent
						where `tabQuotation`.Workflow_state in ("Quoted to Customer")  and `tabQuotation Item`.supply_order_data = %s  and `tabQuotation`.is_multiple_quotation = 0 ''',j.name,as_dict=1)
						if sup_quo_q_2:
							for k in sup_quo_q_2:
								sqm_quo = sqm_quo + k["after_discount_cost"]
						else:
							sup_quo_q_3 = frappe.db.sql(''' select DISTINCT `tabQuotation`.after_discount_cost, `tabQuotation`.Workflow_state,`tabQuotation Item`.margin_amount from `tabQuotation` 
							left join `tabQuotation Item` on  `tabQuotation`.name = `tabQuotation Item`.parent
							where `tabQuotation`.Workflow_state in ("Rejected by Customer")  and `tabQuotation Item`.supply_order_data = %s  and `tabQuotation`.is_multiple_quotation = 0 ''',j.name,as_dict=1)
							if sup_quo_q_3:
								for k in sup_quo_q_3:
									sqm_quo = sqm_quo + k["after_discount_cost"]
							else:
								sup_quo_q_4 = frappe.db.sql(''' select DISTINCT `tabQuotation`.after_discount_cost, `tabQuotation`.Workflow_state,`tabQuotation Item`.margin_amount from `tabQuotation` 
								left join `tabQuotation Item` on  `tabQuotation`.name = `tabQuotation Item`.parent
								where `tabQuotation`.Workflow_state in (Rejected")  and `tabQuotation Item`.supply_order_data = %s  and `tabQuotation`.is_multiple_quotation = 0 ''',j.name,as_dict=1)
								if sup_quo_q_4:
									for k in sup_quo_q_4:
										sqm_quo = sqm_quo + k["after_discount_cost"]
								else:
									sup_quo_q_5 = frappe.db.sql(''' select DISTINCT `tabQuotation`.after_discount_cost, `tabQuotation`.Workflow_state,`tabQuotation Item`.margin_amount from `tabQuotation` 
									left join `tabQuotation Item` on  `tabQuotation`.name = `tabQuotation Item`.parent
									where `tabQuotation`.Workflow_state in ("Approved By Management")  and `tabQuotation Item`.supply_order_data = %s  and `tabQuotation`.is_multiple_quotation = 0 ''',j.name,as_dict=1)
									if sup_quo_q_5:
										for k in sup_quo_q_5:
											sqm_quo = sqm_quo + k["after_discount_cost"]





				sqm_quo_2 = 0
				for j in sup_del:
					sup_quo_q_2 = frappe.db.sql(''' select DISTINCT `tabQuotation`.after_discount_cost, `tabQuotation`.Workflow_state,`tabQuotation Item`.margin_amount from `tabQuotation` 
					left join `tabQuotation Item` on  `tabQuotation`.name = `tabQuotation Item`.parent
					where `tabQuotation`.Workflow_state in ("Approved By Customer")  and `tabQuotation Item`.supply_order_data = %s  and `tabQuotation`.is_multiple_quotation = 1 ''',j.name,as_dict=1)
					if sup_quo_q_2:
						for k in sup_quo_q_2:
							sqm_quo_2 = sqm_quo_2 + k["margin_amount"]
					else:
						sup_quo_q_2_2 = frappe.db.sql(''' select DISTINCT `tabQuotation`.after_discount_cost, `tabQuotation`.Workflow_state,`tabQuotation Item`.margin_amount from `tabQuotation` 
						left join `tabQuotation Item` on  `tabQuotation`.name = `tabQuotation Item`.parent
						where `tabQuotation`.Workflow_state in ("Quoted to Customer")  and `tabQuotation Item`.supply_order_data = %s  and `tabQuotation`.is_multiple_quotation = 1 ''',j.name,as_dict=1)
						if sup_quo_q_2_2:
							for k in sup_quo_q_2_2:
								sqm_quo_2 = sqm_quo_2 + k["margin_amount"]
						else:
							sup_quo_q_2_3 = frappe.db.sql(''' select DISTINCT `tabQuotation`.after_discount_cost, `tabQuotation`.Workflow_state,`tabQuotation Item`.margin_amount from `tabQuotation` 
							left join `tabQuotation Item` on  `tabQuotation`.name = `tabQuotation Item`.parent
							where `tabQuotation`.Workflow_state in ("Rejected by Customer")  and `tabQuotation Item`.supply_order_data = %s  and `tabQuotation`.is_multiple_quotation = 1 ''',j.name,as_dict=1)
							if sup_quo_q_2_3:
								for k in sup_quo_q_2_3:
									sqm_quo_2 = sqm_quo_2 + k["margin_amount"]
							else:
								sup_quo_q_2_4 = frappe.db.sql(''' select DISTINCT `tabQuotation`.after_discount_cost, `tabQuotation`.Workflow_state,`tabQuotation Item`.margin_amount from `tabQuotation` 
								left join `tabQuotation Item` on  `tabQuotation`.name = `tabQuotation Item`.parent
								where `tabQuotation`.Workflow_state in ("Rejected")  and `tabQuotation Item`.supply_order_data = %s  and `tabQuotation`.is_multiple_quotation = 1 ''',j.name,as_dict=1)
								if sup_quo_q_2_4:
									for k in sup_quo_q_2_4:
										sqm_quo_2 = sqm_quo_2 + k["margin_amount"]
								else:
									sup_quo_q_2_5 = frappe.db.sql(''' DISTINCT select `tabQuotation`.after_discount_cost, `tabQuotation`.Workflow_state,`tabQuotation Item`.margin_amount from `tabQuotation` 
									left join `tabQuotation Item` on  `tabQuotation`.name = `tabQuotation Item`.parent
									where `tabQuotation`.Workflow_state in ("Approved By Management")  and `tabQuotation Item`.supply_order_data = %s  and `tabQuotation`.is_multiple_quotation = 1 ''',j.name,as_dict=1)
									if sup_quo_q_2_5:
										for k in sup_quo_q_2_5:
											sqm_quo_2 = sqm_quo_2 + k["margin_amount"]






				sqm_quo_total =sqm_quo + sqm_quo_2


				#Received
				sup_rec = frappe.get_all("Supply Order Data",{"sales_rep":i["sales_rep"],"status":"Received","posting_date": ["between", (Week_start,current_date)]})
			   
				sqm_rec = 0
				for j in sup_rec:
					sup_rec_q = frappe.db.sql(''' select DISTINCT `tabQuotation`.after_discount_cost, `tabQuotation`.Workflow_state,`tabQuotation Item`.margin_amount from `tabQuotation` 
					left join `tabQuotation Item` on  `tabQuotation`.name = `tabQuotation Item`.parent
					where `tabQuotation`.Workflow_state in ("Approved By Customer","Quoted to Customer","Rejected by Customer")  and `tabQuotation Item`.supply_order_data = %s  and `tabQuotation`.is_multiple_quotation = 0 ''',j.name,as_dict=1)
					
					for k in sup_rec_q:
						sqm_rec = sqm_rec + k["after_discount_cost"]
			

				sqm_rec_2 = 0
				for j in sup_rec:
					sup_rec_q_2 = frappe.db.sql(''' select DISTINCT `tabQuotation`.after_discount_cost, `tabQuotation`.Workflow_state,`tabQuotation Item`.margin_amount from `tabQuotation` 
					left join `tabQuotation Item` on  `tabQuotation`.name = `tabQuotation Item`.parent
					where `tabQuotation`.Workflow_state in ("Approved By Customer","Quoted to Customer","Rejected by Customer")  and `tabQuotation Item`.supply_order_data = %s  and `tabQuotation`.is_multiple_quotation = 1 ''',j.name,as_dict=1)
					
					for k in sup_rec_q_2:
						sqm_rec_2 = sqm_rec_2 + k["margin_amount"]
			
				sqm_rec_total =sqm_rec + sqm_rec_2



				# RSC	
				
				wds_rsc = frappe.get_all("Work Order Data",{"sales_rep":i["sales_rep"],"status":"RSC-Repaired and Shipped Client","posting_date": ["between", (Week_start,current_date)]},["*"])
				
				rsc_squot = 0
				for j in wds_rsc:
					rsc_q = frappe.db.sql(''' select DISTINCT `tabQuotation Item`.wod_no,`tabQuotation`.after_discount_cost, `tabQuotation`.Workflow_state,`tabQuotation Item`.margin_amount from `tabQuotation` 
					left join `tabQuotation Item` on  `tabQuotation`.name = `tabQuotation Item`.parent
					where `tabQuotation`.Workflow_state in ("Approved By Customer","Quoted to Customer")  and `tabQuotation Item`.wod_no = %s  and `tabQuotation`.is_multiple_quotation = 0 ''',j.name,as_dict=1)
					if rsc_q:
						for k in rsc_q:
							print(j.name)
							print( k["after_discount_cost"])
							rsc_squot = rsc_squot + k["after_discount_cost"]
					else:
						rsc_q_2 = frappe.db.sql(''' select DISTINCT `tabQuotation Item`.wod_no,`tabQuotation`.after_discount_cost, `tabQuotation`.Workflow_state,`tabQuotation Item`.margin_amount from `tabQuotation` 
						left join `tabQuotation Item` on  `tabQuotation`.name = `tabQuotation Item`.parent
						where `tabQuotation`.Workflow_state in ("Rejected by Customer")  and `tabQuotation Item`.wod_no = %s  and `tabQuotation`.is_multiple_quotation = 0 ''',j.name,as_dict=1)
						if rsc_q_2:
							for k in rsc_q_2:
								print(j.name)
								print( k["after_discount_cost"])
								rsc_squot = rsc_squot + k["after_discount_cost"]
						else:
							rsc_q_3 = frappe.db.sql(''' select DISTINCT `tabQuotation Item`.wod_no,`tabQuotation`.after_discount_cost, `tabQuotation`.Workflow_state,`tabQuotation Item`.margin_amount from `tabQuotation` 
							left join `tabQuotation Item` on  `tabQuotation`.name = `tabQuotation Item`.parent
							where `tabQuotation`.Workflow_state in ("Rejected")  and `tabQuotation Item`.wod_no = %s  and `tabQuotation`.is_multiple_quotation = 0 ''',j.name,as_dict=1)
							if rsc_q_3:
								for k in rsc_q_3:
									print(j.name)
									print( k["after_discount_cost"])
									rsc_squot = rsc_squot + k["after_discount_cost"]
						


				rsc_mquot = 0
				for j in wds_rsc:
					rsc_mq = frappe.db.sql(''' select DISTINCT `tabQuotation Item`.wod_no,`tabQuotation`.after_discount_cost, `tabQuotation`.Workflow_state,`tabQuotation Item`.margin_amount from `tabQuotation` 
					left join `tabQuotation Item` on  `tabQuotation`.name = `tabQuotation Item`.parent
					where `tabQuotation`.Workflow_state in ("Approved By Customer","Quoted to Customer")  and `tabQuotation Item`.wod_no = %s  and `tabQuotation`.is_multiple_quotation = 1 ''',j.name,as_dict=1)
					if rsc_mq:
						for k in rsc_mq:
							print(j.name)
							print(k["margin_amount"])
							rsc_mquot = rsc_mquot + k["margin_amount"]
					else:
						rsc_mq_2 = frappe.db.sql(''' select DISTINCT `tabQuotation Item`.wod_no,`tabQuotation`.after_discount_cost, `tabQuotation`.Workflow_state,`tabQuotation Item`.margin_amount from `tabQuotation` 
						left join `tabQuotation Item` on  `tabQuotation`.name = `tabQuotation Item`.parent
						where `tabQuotation`.Workflow_state in ("Rejected by Customer")  and `tabQuotation Item`.wod_no = %s  and `tabQuotation`.is_multiple_quotation = 1 ''',j.name,as_dict=1)
						if rsc_mq_2:
							for k in rsc_mq_2:
								print(j.name)
								print(k["margin_amount"])
								rsc_mquot = rsc_mquot + k["margin_amount"]
						else:
							rsc_mq_3 = frappe.db.sql(''' select DISTINCT `tabQuotation Item`.wod_no,`tabQuotation`.after_discount_cost, `tabQuotation`.Workflow_state,`tabQuotation Item`.margin_amount from `tabQuotation` 
							left join `tabQuotation Item` on  `tabQuotation`.name = `tabQuotation Item`.parent
							where `tabQuotation`.Workflow_state in ("Rejected")  and `tabQuotation Item`.wod_no = %s  and `tabQuotation`.is_multiple_quotation = 1 ''',j.name,as_dict=1)
							for k in rsc_mq_3:
								print(j.name)
								print(k["margin_amount"])
								rsc_mquot = rsc_mquot + k["margin_amount"]
			 
				#RS				
				wds_rs = frappe.get_all("Work Order Data",{"sales_rep":i["sales_rep"],"status":"RS-Repaired and Shipped","posting_date": ["between", (Week_start,current_date)]},["*"])
				
				rs_squot = 0
				for j in wds_rs:
					rs_q = frappe.db.sql(''' select DISTINCT `tabQuotation`.after_discount_cost, `tabQuotation`.Workflow_state,`tabQuotation Item`.margin_amount from `tabQuotation` 
					left join `tabQuotation Item` on  `tabQuotation`.name = `tabQuotation Item`.parent
					where `tabQuotation`.Workflow_state in ("Approved By Customer")  and `tabQuotation Item`.wod_no = %s  and `tabQuotation`.is_multiple_quotation = 0 ''',j.name,as_dict=1)
					if rs_q:
						for k in rs_q:
							rs_squot = rs_squot + k["after_discount_cost"]
					else:
						rs_q_2 = frappe.db.sql(''' select DISTINCT `tabQuotation`.after_discount_cost, `tabQuotation`.Workflow_state,`tabQuotation Item`.margin_amount from `tabQuotation` 
						left join `tabQuotation Item` on  `tabQuotation`.name = `tabQuotation Item`.parent
						where `tabQuotation`.Workflow_state in ("Quoted to Customer")  and `tabQuotation Item`.wod_no = %s  and `tabQuotation`.is_multiple_quotation = 0 ''',j.name,as_dict=1)
						if rs_q_2:
							for k in rs_q_2:
								rs_squot = rs_squot + k["after_discount_cost"]
						else:
							rs_q_3 = frappe.db.sql(''' select DISTINCT `tabQuotation`.after_discount_cost, `tabQuotation`.Workflow_state,`tabQuotation Item`.margin_amount from `tabQuotation` 
							left join `tabQuotation Item` on  `tabQuotation`.name = `tabQuotation Item`.parent
							where `tabQuotation`.Workflow_state in ("Rejected by Customer")  and `tabQuotation Item`.wod_no = %s  and `tabQuotation`.is_multiple_quotation = 0 ''',j.name,as_dict=1)
							if rs_q_3:
								for k in rs_q_3:
									rs_squot = rs_squot + k["after_discount_cost"]
							else:
								rs_q_4 = frappe.db.sql(''' select DISTINCT `tabQuotation`.after_discount_cost, `tabQuotation`.Workflow_state,`tabQuotation Item`.margin_amount from `tabQuotation` 
								left join `tabQuotation Item` on  `tabQuotation`.name = `tabQuotation Item`.parent
								where `tabQuotation`.Workflow_state in ("Rejected)  and `tabQuotation Item`.wod_no = %s  and `tabQuotation`.is_multiple_quotation = 0 ''',j.name,as_dict=1)
								if rs_q_4:
									for k in rs_q_4:
										rs_squot = rs_squot + k["after_discount_cost"]

								else:
									rs_q_5 = frappe.db.sql(''' select DISTINCT `tabQuotation`.after_discount_cost, `tabQuotation`.Workflow_state,`tabQuotation Item`.margin_amount from `tabQuotation` 
									left join `tabQuotation Item` on  `tabQuotation`.name = `tabQuotation Item`.parent
									where `tabQuotation`.Workflow_state in ("Approved By Management)  and `tabQuotation Item`.wod_no = %s  and `tabQuotation`.is_multiple_quotation = 0 ''',j.name,as_dict=1)
									if rs_q_5:
										for k in rs_q_5:
											rs_squot = rs_squot + k["after_discount_cost"]



				rs_mquot = 0
				for j in wds_rs:
					rs_mq = frappe.db.sql(''' select DISTINCT `tabQuotation`.after_discount_cost, `tabQuotation`.Workflow_state,`tabQuotation Item`.margin_amount from `tabQuotation` 
					left join `tabQuotation Item` on  `tabQuotation`.name = `tabQuotation Item`.parent
					where `tabQuotation`.Workflow_state in ("Approved By Customer")  and `tabQuotation Item`.wod_no = %s  and `tabQuotation`.is_multiple_quotation = 1 ''',j.name,as_dict=1)
					if rs_mq:
						for k in rs_mq:
							rs_mquot = rs_mquot + k["margin_amount"]
				
					else:
						rs_mq_2 = frappe.db.sql(''' select DISTINCT `tabQuotation`.after_discount_cost, `tabQuotation`.Workflow_state,`tabQuotation Item`.margin_amount from `tabQuotation` 
						left join `tabQuotation Item` on  `tabQuotation`.name = `tabQuotation Item`.parent
						where `tabQuotation`.Workflow_state in ("Quoted to Customer")  and `tabQuotation Item`.wod_no = %s  and `tabQuotation`.is_multiple_quotation = 1 ''',j.name,as_dict=1)
						if rs_mq_2:
							for k in rs_mq_2:
								rs_mquot = rs_mquot + k["margin_amount"]
						else:
							rs_mq_3 = frappe.db.sql(''' select DISTINCT `tabQuotation`.after_discount_cost, `tabQuotation`.Workflow_state,`tabQuotation Item`.margin_amount from `tabQuotation` 
							left join `tabQuotation Item` on  `tabQuotation`.name = `tabQuotation Item`.parent
							where `tabQuotation`.Workflow_state in ("Rejected by Customer")  and `tabQuotation Item`.wod_no = %s  and `tabQuotation`.is_multiple_quotation = 1 ''',j.name,as_dict=1)
							if rs_mq_3:
								for k in rs_mq_3:
									rs_mquot = rs_mquot + k["margin_amount"]
							else:
								rs_mq_4 = frappe.db.sql(''' select DISTINCT `tabQuotation`.after_discount_cost, `tabQuotation`.Workflow_state,`tabQuotation Item`.margin_amount from `tabQuotation` 
								left join `tabQuotation Item` on  `tabQuotation`.name = `tabQuotation Item`.parent
								where `tabQuotation`.Workflow_state in ("Rejected")  and `tabQuotation Item`.wod_no = %s  and `tabQuotation`.is_multiple_quotation = 1 ''',j.name,as_dict=1)
								if rs_mq_4:
									for k in rs_mq_4:
										rs_mquot = rs_mquot + k["margin_amount"]
								else:
									rs_mq_5 = frappe.db.sql(''' select DISTINCT `tabQuotation`.after_discount_cost, `tabQuotation`.Workflow_state,`tabQuotation Item`.margin_amount from `tabQuotation` 
									left join `tabQuotation Item` on  `tabQuotation`.name = `tabQuotation Item`.parent
									where `tabQuotation`.Workflow_state in ("Approved By Management)  and `tabQuotation Item`.wod_no = %s  and `tabQuotation`.is_multiple_quotation = 1 ''',j.name,as_dict=1)
									if rs_mq_5:
										for k in rs_mq_5:
											rs_mquot = rs_mquot + k["margin_amount"]
 

				#q-quoted
				wd_quot = 0
				wds_quot = frappe.get_all("Work Order Data",{"sales_rep":i["sales_rep"],"status":"Q-Quoted","posting_date": ["between", (Week_start,current_date)]},["*"])
				for j in wds_quot:
					total_q = frappe.db.sql(''' select sum(after_discount_cost) as t from `tabQuotation` 
					where Workflow_state in ("Approved By Customer")  and sales_rep = %s ''',i["sales_rep"],as_dict=1)
					if total_q:
						for k in total_q:
							wd_quot = wd_quot + k["after_discount_cost"]
					else:
						total_q_2 = frappe.db.sql(''' select sum(after_discount_cost) as t from `tabQuotation` 
						where Workflow_state in ("Quoted to Customer")  and sales_rep = %s ''',i["sales_rep"],as_dict=1)
						if total_q_2:
							for k in total_q_2:
								wd_quot = wd_quot + k["after_discount_cost"]
					   
				wd_mquot = 0
				for j in wds_quot:
					total_q_2 = frappe.db.sql(''' select sum(after_discount_cost) as t from `tabQuotation` 
					where Workflow_state in ("Approved By Customer")  and sales_rep = %s ''',i["sales_rep"],as_dict=1)
					if total_q_2:
						for k in total_q_2:
							wd_mquot = wd_mquot + k["after_discount_cost"]
					else:
						total_q_2_2 = frappe.db.sql(''' select sum(after_discount_cost) as t from `tabQuotation` 
						where Workflow_state in ("Quoted to Customer")  and sales_rep = %s ''',i["sales_rep"],as_dict=1)
						if total_q_2_2:
							for k in total_q_2_2:
								wd_mquot = wd_mquot + k["after_discount_cost"]

				wd_total_quot =   wd_quot + wd_mquot



				# if total_q[0]["t"] == None:
				#     total_q[0]["t"] = 0

				total_quotation = total_quotation + total_q[0]["t"] 

				#rsi
				rsi_old = 0
				# for i in wds:
				#     rsi_old = rsi_old + i.old_wo_q_amount
			  
				rsi_total = squot + mquot + rsi_old

				#rsc
				rsc_old = 0
				# for i in wds_rsc:
				#     rsc_old = rsc_old + i.old_wo_q_amount
			  
				rsc_total = rsc_squot + rsc_mquot + rsc_old
				print(rsc_total)
				#rs
				rs_old = 0
				# for i in wds_rs:
				#     rs_old = rs_old + i.old_wo_q_amount
			  
				rs_total = rs_squot + rs_mquot + rs_old

				total_rsi = total_rsi + rsi_total
				total_rsc = total_rsc + rsc_total
				total_rs = total_rs + rs_total 

				s_total_rsi = s_total_rsi + sqm_total
				s_total_rsc = s_total_rsc + sqm_del_total
				s_total_rs = s_total_rs + sqm_rec_total
				s_total_quotation =  s_total_quotation + sqm_quo_total

				
			   
				data += '<tr>'
				data += '<td style="border-color:#000000;"><center><b>%s</b></center></td>' %(i["sales_rep"])
				data += '<td style="border-color:#000000;"><center><b>%s</b></center></td>' %(rsi_total)
				data += '<td style="border-color:#000000;"><center><b>%s</b></center></td>' %(sqm_total)
				data += '<td style="border-color:#000000;"><center><b>%s</b></center></td>' %(round(rsc_total,2))
				data += '<td style="border-color:#000000;"><center><b>%s</b></center></td>' %(sqm_del_total)
				data += '<td style="border-color:#000000;"><center><b>%s</b></center></td>' %(rs_total)
				data += '<td style="border-color:#000000;"><center><b>%s</b></center></td>' %(sqm_rec_total)
				data += '<td style="border-color:#000000;"><center><b>%s</b></center></td>' %(total_q[0]["t"])
				data += '<td style="border-color:#000000;"><center><b>%s</b></center></td>'	%(sqm_quo_total)
				data += '</tr>'
	data += '<tr>'
	data += '<td style="border-color:#000000;"><center><b>Total</b></center></td>'
	data += '<td style="border-color:#000000;"><center><b>%s</b></center></td>' %(total_rsi)
	data += '<td style="border-color:#000000;"><center><b>%s</b></center></td>' %(s_total_rsi)
	data += '<td style="border-color:#000000;"><center><b>%s</b></center></td>' %(total_rsc)
	data += '<td style="border-color:#000000;"><center><b>%s</b></center></td>' %(s_total_rsi)
	data += '<td style="border-color:#000000;"><center><b>%s</b></center></td>' %(total_rs)
	data += '<td style="border-color:#000000;"><center><b>%s</b></center></td>' %(s_total_rs)
	data += '<td style="border-color:#000000;"><center><b>%s</b></center></td>'%(total_quotation )
	data += '<td style="border-color:#000000;"><center><b>%s</b></center></td>' %(s_total_quotation)
	data += '</tr>'

	data += '</table>'
	data += '</div>'
	
	# return data


@frappe.whitelist()
def check_wd():
	wd= frappe.db.sql(""" select  payment_entry_reference as payment_entry,payment_date as paid_date,name as wod_no,sales_rep,posting_date,remarks,expiry_date,customer,technician,status,department,returned_date,branch as branch_name,dn_no,dn_date,invoice_no,invoice_date,purchase_order_no as po_no  from `tabWork Order Data` where posting_date>= "%s" and posting_date <= "%s" """ %("2015-01-01","2024-10-12"),as_dict=1)
	for i in wd:
		if i.wod_no == "WOD-K24-13220":
			print(i.wod_no)

@frappe.whitelist()
def get_w():
	p = frappe.db.sql(""" select `tabPayment Entry Reference`.reference_name as pr from `tabPayment Entry` 
		left join `tabPayment Entry Reference` on `tabPayment Entry`.name = `tabPayment Entry Reference`.parent
		where  `tabPayment Entry Reference`.reference_doctype = "Sales Invoice" """ ,as_dict=1)
	
	for i in p:
		
		si = frappe.db.sql(""" select `tabSales Invoice Item`.wod_no as wd from `tabSales Invoice` 
		left join `tabSales Invoice Item` on `tabSales Invoice`.name = `tabSales Invoice Item`.parent
		where  `tabSales Invoice`.name = '%s' """ %(i.pr),as_dict=1)
		if si[0]["wd"]:
			status = frappe.get_value("Work Order Data",{"name":si[0]["wd"]},["status"])
			if not status == "P-Paid":
				print(si[0]["wd"])


@frappe.whitelist()
def purchase_report(date,company):
	ogdate = datetime.strptime(str(date),"%Y-%m-%d")

	# Format the date as a string in the desired format
	formatted_date = ogdate.strftime("%d-%m-%Y")
	# date = d.strftime("%Y-%m-%d")
	# original_date = datetime.strptime(str(d), "%Y-%m-%d")
	# original_date = date
	data= ""
	data += '<div class="table-container">'
	# data += '<h3><b><center>WORK ORDER<center><b><h3>'
	data += '<table class="table table-bordered">'
	data += '<tr>'
	data += '<td style="width:30%;border-color:#000000;"><img src = "/files/TSL Logo.png" align="left" width ="180"></td>'
	if company == "TSL COMPANY - Kuwait":  
		data += '<td style="width:30%;border-color:#000000;color:#055c9d;"><h3><center><b style="color:#055c9d;">TSL Company<br>Branch - Kuwait</b></center></h3></td>' 
	if company == "TSL COMPANY - UAE":  
		data += '<td style="width:40%;border-color:#000000;color:#055c9d;"><h3><center><b style="color:#055c9d;">TSL Company<br>Branch - UAE</b></center></h3></td>' 
	if company == "TSL COMPANY - Kuwait": 
		data += '<td style="width:30%;border-color:#000000;"><center><img src = "/files/kuwait flag.jpg" width ="100"></center></td>'
	if company == "TSL COMPANY - UAE":  
		data += '<td style="width:30%;border-color:#000000;"><center><img src = "/files/Flag_of_the_United_Arab_Emirates.svg.jpg" width ="100"></center></td>'
	data += '</tr>'
	data += '</table>'

	data += '<table class="table table-bordered">'
	
	data += '<tr>'
	data += '<td colspan = 2 style="border-right:hidden;text-align:left;border-color:#000000;padding:1px;font-size:15px;color:white;"><b>Purchase Report</b></td>'
	data += '<td colspan = 2 style="text-align:right;border-color:#000000;padding:1px;font-size:15px;color:white;"><b>Generation Date:%s</b></td>' %(formatted_date)
	data += '</tr>'

	data += '<tr>'
	data += '<td colspan = 4 style="border-color:#000000;padding:1px;font-size:15px;background-color:#00BFFF;color:white;"><center><b>WORK ORDER</b><center></td>'
	data += '</tr>'

	day = add_days(date,-1)
	
	# rnp = frappe.db.count("Work Order Data",{"status":"RNP-Return No Parts","posting_date": ["between", ["2015-07-23", date]]})

	rnp = frappe.db.sql(""" select count(DISTINCT `tabWork Order Data`.name) as ct from `tabWork Order Data` 
	left join `tabStatus Duration Details` on `tabWork Order Data`.name = `tabStatus Duration Details`.parent
	where  `tabStatus Duration Details`.status = "RNP-Return No Parts" and `tabWork Order Data`.company = '%s'
	and `tabStatus Duration Details`.date  LIKE "%s%%" """ %(company,date) ,as_dict=1)
	
	# ordered =  frappe.db.count("Work Order Data",{"status":"WP-Waiting Parts","posting_date": ["between", [date,date]]})
	ordered = frappe.db.sql(""" select count(DISTINCT `tabWork Order Data` .name) as wp from `tabWork Order Data` 
	left join `tabStatus Duration Details` on `tabWork Order Data`.name = `tabStatus Duration Details`.parent
	where  `tabStatus Duration Details`.status = "WP-Waiting Parts" and `tabWork Order Data`.company = '%s'
	and `tabStatus Duration Details`.date  LIKE "%s%%" """ %(company,date) ,as_dict=1)
	
	# ordered = frappe.db.count("Work Order Data",{"status":"Q-Quoted","posting_date": ["between", [date, date]]})
	quot = frappe.db.count("Work Order Data",{"status":"Parts Priced","posting_date": ["between", [date, date]]})
	rec =  frappe.db.count("Work Order Data",{"status":"TR-Technician Repair","posting_date": ["between", [date, date]]})
	# app =  frappe.db.count("Work Order Data",{"status":"A-Approved","posting_date": ["between", ["2015-07-23", date]]})

	not_qt =  frappe.db.count("Work Order Data",{"company":company,"old_wo_no":["is","not set"],"status":"SP-Searching Parts","posting_date": ["between", ["2015-07-23", date]]})

	wp =  frappe.db.count("Work Order Data",{"company":company,"status":"WP-Waiting Parts","posting_date": ["between", ["2015-07-23", date]]})
	np =  frappe.db.count("Work Order Data",{"company":company,"status":"RNP-Return No Parts","posting_date": ["between", [date, date]]})


	s_quoted = frappe.db.count("Supply Order Data",{"company":company,"status":"Quoted","posting_date": ["between", [date, date]]})
	s_ordered = frappe.db.count("Supply Order Data",{"company":company,"status":"Ordered","posting_date": ["between", [date, date]]})
	# s_received = frappe.db.count("Supply Order Data",{"status":"Received","posting_date": ["between", [date, date]]})
	s_approved = frappe.db.count("Supply Order Data",{"status":"Approved","posting_date": ["between", [date, date]]})
	s_not_quoted = frappe.db.count("Supply Order Data",{"company":company,"department":"Supply - TSL","docstatus":1,"company":company,"status":"Inquiry","posting_date": ["between", ["2015-07-23", date]]})

	s_not_found = frappe.db.count("Supply Order Data",{"department":"Supply - TSL","company":company,"status":"Not Found","posting_date": ["between", ["2015-07-23", date]]})

	waiting_so = frappe.db.count("Supply Order Data",{"department":"Supply - TSL","company":company,"status":"ordered","posting_date": ["between", ["2015-07-23", day]]})


	po = frappe.db.sql(""" select count(`tabPurchase Order Item`.work_order_data) as p from `tabPurchase Order`
		left join `tabPurchase Order Item` on `tabPurchase Order Item`.parent = `tabPurchase Order`.name
		where `tabPurchase Order`.transaction_date = '%s' and `tabPurchase Order`.docstatus != 2 and `tabPurchase Order`.per_received = %s and  `tabPurchase Order`.company = '%s' """ %("2024-09-14",0,company),as_dict =1)

	prw = frappe.db.sql(""" select count(`tabPurchase Receipt Item`.work_order_data) as pr from `tabPurchase Receipt`
		left join `tabPurchase Receipt Item` on `tabPurchase Receipt Item`.parent = `tabPurchase Receipt`.name
		where `tabPurchase Receipt`.posting_date = '%s' and `tabPurchase Receipt`.docstatus != 2  and `tabPurchase Receipt`.company = '%s' """ %(date,company),as_dict =1)

	# qc = frappe.db.sql(""" select count(`tabQuotation Item`.wod_no) as q from `tabQuotation`
	# 	left join `tabQuotation Item` on `tabQuotation Item`.parent = `tabQuotation`.name
	# 	where `tabQuotation`.transaction_date = '%s' and `tabQuotation`.workflow_state = "Approved By Customer" and  `tabQuotation`.company = "TSL COMPANY - Kuwait" """ %(date),as_dict =1)

	qc = frappe.db.sql(""" select count(DISTINCT `tabWork Order Data` .name) as q from `tabWork Order Data` 
	left join `tabStatus Duration Details` on `tabWork Order Data`.name = `tabStatus Duration Details`.parent
	where  `tabWork Order Data`.status = "A-Approved" and `tabWork Order Data`.company = '%s'
	and `tabStatus Duration Details`.date  LIKE "%s%%" """ %(company,date) ,as_dict=1)
		
	pre_eval =  frappe.get_all("Work Order Data",{"company":company,"status":"A-Approved","company":company,"old_wo_no":["is","not set"],"posting_date": ["between", ["2015-07-23", date]]})

	pre_wod_count = 0
	for i in pre_eval:
		
		eval_rep = frappe.db.exists("Evaluation Report",{"work_order_data":i.name})
		if not eval_rep:
			pre_q = frappe.db.sql(""" select `tabQuotation`.name from `tabQuotation`
			left join `tabQuotation Item` on `tabQuotation Item`.parent = `tabQuotation`.name
			where `tabQuotation`.workflow_state = "Approved By Customer" and `tabQuotation Item`.wod_no = "%s" """ %(i.name),as_dict =1)
			if pre_q:
				pre_wod_count = pre_wod_count + 1
	# frappe.errprint(pre_wod_count)



	pwod = 0
	if po:	
		pwod = po[0]["p"]

	prwod = 0
	if prw:	
		prwod = prw[0]["pr"]
	

	qcw = 0
	if qc:	
		qcw = qc[0]["q"]
	
	sq = frappe.db.count("Supplier Quotation",{"transaction_date":date,"work_order_data": ["is", "set"],"company":company})
	
	

	# data += '<tr>'
	# data += '<td colspan = 2 style="border-color:#000000;padding:1px;font-size:16px;background-color:#00BFFF;color:white;"><center><b>Achievments</b><center></td>'
	# data += '<td colspan = 2 style="border-color:#000000;padding:1px;font-size:16px;background-color:#00BFFF;color:white;"><center><b>Status</b><center></td>'
	# data += '</tr>'

	data += '<tr>'
	data += '<td style="border-color:#000000;padding:1px;font-size:14px;background-color:#004792;color:white;width:25%;text-align:center;font-weight:bold;color:#FFFFFF;">Current Status</td>'
	data += '<td style="border-color:#000000;padding:1px;font-size:14px;background-color:#004792;color:white;width:25%;text-align:center;font-weight:bold;color:#FFFFFF;">Count</td>'
	data += '<td style="border-color:#000000;padding:1px;font-size:14px;background-color:#004792;color:white;width:25%;text-align:center;font-weight:bold;color:#FFFFFF;">Status</td>'
	data += '<td style="border-color:#000000;padding:1px;font-size:14px;background-color:#004792;color:white;width:25%;text-align:center;font-weight:bold;color:#FFFFFF;">Count</td>'

	data += '</tr>'

	data += '<tr>'
	data += '<td style="border-color:#000000;padding:1px;font-size:14px;font-size:14px;background-color:#A7C7E7;"><center><b>Not Found</b><center></td>'
	data += '<td style="border-color:#000000;padding:1px;font-size:14px;font-size:14px;"><center><b>%s</b><center></td>' %(rnp[0]["ct"])
	data += '<td style="border-color:#000000;padding:1px;font-size:14px;font-size:14px;background-color:#A7C7E7;"><center><b>Approved</b><center></td>'
	data += '<td style="border-color:#000000;padding:1px;font-size:14px;font-size:14px;"><center><b>%s</b><center></td>' %(qcw)

	data += '</tr>'

	data += '<tr>'
	data += '<td style="border-color:#000000;padding:1px;font-size:14px;font-size:14px;background-color:#A7C7E7;"><center><b>Ordered</b><center></td>' 
	data += '<td style="border-color:#000000;padding:1px;font-size:14px;font-size:14px;"><center><b>%s</b><center></td>'%(ordered[0]["wp"] or 0)
	data += '<td style="border-color:#000000;padding:1px;font-size:14px;font-size:14px;background-color:#A7C7E7;"><center><b>Pre-Approved</b><center></td>'
	data += '<td style="border-color:#000000;padding:1px;font-size:14px;font-size:14px;"><center><b>%s</b><center></td>' %(pre_wod_count)

	
	data += '</tr>'

	data += '<tr>'
	data += '<td style="border-color:#000000;padding:1px;font-size:14px;font-size:14px;background-color:#A7C7E7;"><center><b>Quoted</b><center></td>'
	data += '<td style="border-color:#000000;padding:1px;font-size:14px;font-size:14px;"><center><b>%s</b><center></td>' %(sq)
	data += '<td style="border-color:#000000;padding:1px;font-size:14px;font-size:14px;background-color:#A7C7E7;"><center><b>Not Quoted</b><center></td>'
	data += '<td style="border-color:#000000;padding:1px;font-size:14px;font-size:14px;"><center><b>%s</b><center></td>' %(not_qt)

	
	data += '</tr>'

	prwod = frappe.db.sql(""" select count(DISTINCT `tabWork Order Data` .name) as ct from `tabWork Order Data` 
	left join `tabStatus Duration Details` on `tabWork Order Data`.name = `tabStatus Duration Details`.parent
	where  `tabStatus Duration Details`.status = "TR-Technician Repair" and `tabWork Order Data`.company = '%s'
	and `tabStatus Duration Details`.date  LIKE "%s%%" """ %(company,date) ,as_dict=1)
		

	data += '<tr>'
	data += '<td style="border-color:#000000;padding:1px;font-size:14px;font-size:14px;background-color:#A7C7E7;"><center><b>Received</b><center></td>'
	data += '<td style="border-color:#000000;padding:1px;font-size:14px;font-size:14px;"><center><b>%s</b><center></td>' %(prwod[0]["ct"] or 0)
	data += '<td style="border-color:#000000;padding:1px;font-size:14px;font-size:14px;background-color:#A7C7E7;"><center><b>Waiting PS</b><center></td>'
	data += '<td style="border-color:#000000;padding:1px;font-size:14px;font-size:14px;"><center><b>%s</b><center></td>' %(wp)

	data += '</tr>'
	data += '</table>'




	# data += '<h3><b><center>SUPPLY ORDER<center><b><h3>'
	data += '<table class="table table-bordered">'
	data += '<tr>'
	data += '<td colspan = 4 style="border-color:#000000;padding:1px;font-size:15px;background-color:#00BFFF;color:white;"><center><b>SUPPLY ORDER</b><center></td>'
	data += '</tr>'

	# data += '<tr>'
	# data += '<td colspan = 2 style="border-color:#000000;padding:1px;font-size:16px;background-color:#00BFFF;color:white;"><center><b>Achievments</b><center></td>'
	# data += '<td colspan = 2 style="border-color:#000000;padding:1px;font-size:16px;background-color:#00BFFF;color:white;"><center><b>Status</b><center></td>'
	# data += '</tr>'

	data += '<tr>'
	data += '<td style="border-color:#000000;padding:1px;font-size:14px;background-color:#004792;color:white;width:25%;text-align:center;font-weight:bold;color:#FFFFFF;">Current Status</td>'
	data += '<td style="border-color:#000000;padding:1px;font-size:14px;background-color:#004792;color:white;width:25%;text-align:center;font-weight:bold;color:#FFFFFF;">Count</td>'
	data += '<td style="border-color:#000000;padding:1px;font-size:14px;background-color:#004792;color:white;width:25%;text-align:center;font-weight:bold;color:#FFFFFF;">Status</td>'
	data += '<td style="border-color:#000000;padding:1px;font-size:14px;background-color:#004792;color:white;width:25%;text-align:center;font-weight:bold;color:#FFFFFF;">Count</td>'

	data += '</tr>'


	spo = frappe.db.sql(""" select count(`tabPurchase Order Item`.supply_order_data) as p from `tabPurchase Order`
		left join `tabPurchase Order Item` on `tabPurchase Order Item`.parent = `tabPurchase Order`.name
		where `tabPurchase Order`.transaction_date = '%s' and `tabPurchase Order`.docstatus != 2 and `tabPurchase Order`.per_received = %s and `tabPurchase Order`.cost_center != 'Supply Tender - TSL'
		and `tabPurchase Order`.company = '%s' """ %(date,0,company),as_dict =1)

	spwod = 0
	if spo:	
		spwod = spo[0]["p"]
	
	# sqs = frappe.db.count("Supplier Quotation",{"transaction_date":date,"supply_order_data": ["is", "set"],"company":"TSL COMPANY - Kuwait"})

	sqs = frappe.db.sql(""" select count(DISTINCT `tabSupply Order Data`.name) as ct from `tabSupply Order Data` 
	left join `tabStatus Duration Details` on `tabSupply Order Data`.name = `tabStatus Duration Details`.parent
	where  `tabStatus Duration Details`.status = "Parts Priced" and `tabSupply Order Data`.company = '%s' and `tabSupply Order Data`.department != 'Supply Tender - TSL'
	and `tabStatus Duration Details`.date  LIKE "%s%%" """ %(company,date) ,as_dict=1)
	

	# sqc = frappe.db.sql(""" select count(`tabQuotation Item`.supply_order_data) as qs from `tabQuotation`
	# 	left join `tabQuotation Item` on `tabQuotation Item`.parent = `tabQuotation`.name
	# 	where `tabQuotation`.transaction_date = '%s' and `tabQuotation`.workflow_state = "Approved By Customer" and  `tabQuotation`.company = "TSL COMPANY - Kuwait" """ %(date),as_dict =1)

	# sqcount = 0
	# if sqc:	
	# 	sqcount = sqc[0]["qs"]

	sqc = frappe.db.sql(""" select count(DISTINCT `tabSupply Order Data`.name) as ct from `tabSupply Order Data` 
	left join `tabStatus Duration Details` on `tabSupply Order Data`.name = `tabStatus Duration Details`.parent
	where  `tabSupply Order Data`.status = "Approved" and `tabSupply Order Data`.company = '%s' and `tabSupply Order Data`.department != 'Supply Tender - TSL'
	""" %(company) ,as_dict=1)
	

	s_not_found = frappe.db.sql(""" select count(DISTINCT `tabSupply Order Data`.name) as ct from `tabSupply Order Data` 
	left join `tabStatus Duration Details` on `tabSupply Order Data`.name = `tabStatus Duration Details`.parent
	where  `tabStatus Duration Details`.status = "Not Found" and `tabSupply Order Data`.company = '%s' and `tabSupply Order Data`.department != 'Supply Tender - TSL'
	and `tabStatus Duration Details`.date  LIKE "%s%%" """ %(company,date) ,as_dict=1)
	

	data += '<tr>'
	data += '<td style="border-color:#000000;padding:1px;font-size:14px;font-size:14px;background-color:#A7C7E7;"><center><b>Not Found</b><center></td>'
	data += '<td style="border-color:#000000;padding:1px;font-size:14px;font-size:14px;"><center><b>%s</b><center></td>' %(s_not_found[0]["ct"] or 0)
	data += '<td style="border-color:#000000;padding:1px;font-size:14px;font-size:14px;background-color:#A7C7E7;"><center><b>Approved</b><center></td>'
	data += '<td style="border-color:#000000;padding:1px;font-size:14px;font-size:14px;"><center><b>%s</b><center></td>' %(sqc[0]["ct"])

	data += '</tr>'

	data += '<tr>'
	data += '<td style="border-color:#000000;padding:1px;font-size:14px;background-color:#A7C7E7;"><center><b>Ordered</b><center></td>'
	data += '<td style="border-color:#000000;padding:1px;font-size:14px;"><center><b>%s</b><center></td>' %(spwod)
	data += '<td style="border-color:#000000;padding:1px;font-size:14px;background-color:#A7C7E7;"><center><b>Not Quoted</b><center></td>'
	data += '<td style="border-color:#000000;padding:1px;font-size:14px;"><center><b>%s</b><center></td>'  %(s_not_quoted)

	data += '</tr>'

	data += '<tr>'
	data += '<td style="border-color:#000000;padding:1px;font-size:14px;background-color:#A7C7E7;"><center><b>Quoted</b><center></td>'
	data += '<td style="border-color:#000000;padding:1px;font-size:14px;"><center><b>%s</b><center></td>' %(sqs[0]["ct"] or 0)
	data += '<td style="border-color:#000000;padding:1px;font-size:14px;background-color:#A7C7E7;"><center><b>Waiting SO</b><center></td>'
	data += '<td style="border-color:#000000;padding:1px;font-size:14px;"><center><b>%s</b><center></td>' %(waiting_so)

	data += '</tr>'

	s_received = frappe.db.sql(""" select count(DISTINCT `tabSupply Order Data`.name) as ct from `tabSupply Order Data` 
	left join `tabStatus Duration Details` on `tabSupply Order Data`.name = `tabStatus Duration Details`.parent
	where  `tabStatus Duration Details`.status = "Received" and `tabSupply Order Data`.company = '%s' and `tabSupply Order Data`.department != 'Supply Tender - TSL'
	and `tabStatus Duration Details`.date  LIKE "%s%%" """ %(company,date) ,as_dict=1)
	

	data += '<tr>'
	data += '<td style="border-color:#000000;padding:1px;font-size:14px;font-size:14px;background-color:#A7C7E7;"><center><b>Received</b><center></td>'
	data += '<td style="border-color:#000000;padding:1px;font-size:14px;font-size:14px;"><center><b>%s</b><center></td>' %(s_received[0]["ct"] or 0)
	data += '<td style="border-color:#000000;padding:1px;font-size:14px;font-size:14px;background-color:#A7C7E7;"><center><b></b><center></td>'
	data += '<td style="border-color:#000000;padding:1px;font-size:14px;font-size:14px;"><center><b></b><center></td>'

	data += '</tr>'
	
	data += '</table>'



	data += '</div>'
	return data


@frappe.whitelist()
def daily_lab_report(date,company):
	if company == "TSL COMPANY - Kuwait":
		d = datetime.now().date()
		ogdate = datetime.strptime(str(d),"%Y-%m-%d")
		ogdate_2 = datetime.strptime(str(date),"%Y-%m-%d")

		# Format the date as a string in the desired format
		formatted_date = ogdate.strftime("%d-%m-%Y")
		formatted_date_2 = ogdate_2.strftime("%d-%m-%Y")
		
		# date = d.strftime("%Y-%m-%d")
		# original_date = datetime.strptime(str(d), "%Y-%m-%d")
		original_date = date


		# Format datetime object to new date string format
		# date = datetime.strptime(str(original_date), "%Y-%m-%d")

		data = ""
		data += '<table class="table table-bordered">'
		data += '<tr>'
		data += '<td colspan = "8" align = "center" style="border-color:#000000;padding:1px;font-size:20px;background-color:#808080;color:white;"><b>Date :%s</b></td>' %(formatted_date_2)
		data += '</tr>'
		data += '<tr>'
		data += '<td colspan = 2 style="border-color:#000000;"><img src = "/files/TSL Logo.png" align="left" width ="220"></td>'
		data += '<td colspan = 4 style="border-color:#000000;color:#055c9d;"><h2><center><b style="color:#055c9d;">TSL Company<br>Branch - Kuwait</b></center></h2></td>'
		data += '<td colspan = 2 style="border-color:#000000;"><center><img src = "/files/kuwait flag.jpg" width ="130"></center></td>'
		
		data += '</tr>'

		
		data += '<tr>'
		# data += '<td style="border-color:#000000;width:15%;padding:1px;font-size:11px;font-size:11px;"><center><b></b><center></td>'
		data += '<td colspan = "2" style="line-height:0.8;border-color:#000000;padding:1px;font-size:18px;font-size:16px;background-color:#3333ff;color:white;"><center><b>Status</b><center></td>'
		data += '<td style="line-height:0.8;border-color:#000000;width:12%;padding:1px;font-size:12x;background-color:#FFA500;color:white;"><center><b>Sampat</b><center></td>'
		data += '<td style="line-height:0.8;border-color:#000000;width:12%;padding:1px;font-size:12px;background-color:#FFC000;color:white;"><center><b>Mari</b><center></td>'
		data += '<td style="line-height:0.8;border-color:#000000;width:12%;padding:1px;font-size:12px;background-color:#008000;color:white;"><center><b>ED</b><center></td>'
		data += '<td style="line-height:0.8;border-color:#000000;width:12%;padding:1px;font-size:12px;background-color:#0047AB;color:white;"><center><b>Aakib</b><center></td>'
		data += '<td style="line-height:0.8;border-color:#000000;width:12%;padding:1px;font-size:12px;background-color:#ae388b;color:white;"><center><b>Amir</b><center></td>'
		data += '<td style="line-height:0.8;border-color:#000000;width:12%;padding:1px;font-size:12px;background-color:#4682B4;color:white;"><center><b>Total</b><center></td>'
		
		data += '</tr>'

		tech = ["sampath@tsl-me.com","maari@tsl-me.com","eduardo@tsl-me.com","aakib@tsl-me.com","amir@tsl-me.com"]
		total_ne_s = 0
		total_ne_m = 0
		total_ne_e = 0
		total_ne_a = 0
		total_ne_am = 0

		lp_total_ne_s = 0
		lp_total_ne_m = 0
		lp_total_ne_e = 0
		lp_total_ne_a = 0
		lp_total_ne_am = 0

		total_ner_s = 0
		total_ner_m = 0
		total_ner_e = 0
		total_ner_a = 0
		total_ner_am = 0

		ue_s = 0
		ue_m = 0
		ue_e = 0
		ue_a = 0
		ue_am = 0

		utr_s = 0
		utr_m = 0
		utr_e = 0
		utr_a = 0
		utr_am = 0

		rs_s = 0
		rs_m = 0
		rs_e = 0
		rs_a = 0
		rs_am = 0

		w_s = 0
		w_m = 0
		w_e = 0
		w_a = 0
		w_am = 0

		rnr_s = 0
		rnr_m = 0
		rnr_e = 0
		rnr_a = 0
		rnr_am = 0

		rnf_s = 0
		rnf_m = 0
		rnf_e = 0
		rnf_a = 0
		rnf_am = 0

		com_s = 0
		com_m = 0
		com_e = 0
		com_a = 0
		com_am = 0

		# for t in tech:
		# 	ne = frappe.get_all("Work Order Data",{"technician":t},["*"])
		# 	for i in ne:
		# 		ev = frappe.db.exists("Evaluation Report",{"work_order_data":i.name,"date":original_date,"technician":t})
		# 		if ev:
		# 			# techni = frappe.get_value("Evaluation Report",{"name":ev},["technician"])
		# 			if t == "sampath@tsl-me.com":
		# 				total_ne_s = total_ne_s + 1
		# 			if t == "maari@tsl-me.com":
		# 				total_ne_m = total_ne_m + 1
		# 			if t == "eduardo@tsl-me.com":
		# 				total_ne_e = total_ne_e + 1
		# 			if t == "aakib@tsl-me.com":
		# 				total_ne_a = total_ne_a + 1
					
		
			# ne = frappe.get_all("Work Order Data",{"posting_date":original_date,"technician":t},["*"])
			# for i in ne:
			# 	sd = frappe.get_doc("Work Order Data",i.name)
			# 	for j in sd.status_duration_details:
			# 		timestamp = str(j.date) 
			# 		date_portion = timestamp[:10]
			# 		j.date= date_portion
			# 		if "NE-Need Evaluation"  == j.status and i.technician == "sampath@tsl-me.com" and str(j.date) :
			# 			total_ne_s = total_ne_s + 1
			# 		if "NE-Need Evaluation"  == j.status and i.technician == "maari@tsl-me.com" and str(j.date):
			# 			total_ne_m = total_ne_m + 1
			# 		if "NE-Need Evaluation"  == j.status and i.technician == "eduardo@tsl-me.com" and str(j.date) == original_date:
			# 			total_ne_e = total_ne_e + 1
			# 		if "NE-Need Evaluation"  == j.status and i.technician == "aakib@tsl-me.com" and str(j.date) == original_date:
			# 			total_ne_a = total_ne_a + 1
					
			# 		if "NER-Need Evaluation Return"  == j.status and i.technician == "sampath@tsl-me.com" and str(j.date) == original_date:
			# 			total_ner_s = total_ner_s + 1
			# 		if "NER-Need Evaluation Return"  == j.status and i.technician == "maari@tsl-me.com" and str(j.date) == original_date:
			# 			total_ner_m = total_ner_m + 1
			# 		if "NER-Need Evaluation Return"  == j.status and i.technician == "eduardo@tsl-me.com" and str(j.date) == original_date:
			# 			total_ner_e = total_ner_e + 1
			# 		if "NER-Need Evaluation Return"  == j.status and i.technician == "aakib@tsl-me.com" and str(j.date) == original_date:
			# 			total_ner_a = total_ner_a + 1
					
					
		
		# total_ne_s = frappe.db.count("Work Order Data",{"technician":"sampath@tsl-me.com","status":"NE-Need Evaluation","posting_date": ["BETWEEN", ["2016-01-01",date]]})
		# total_ne_m = frappe.db.count("Work Order Data",{"technician":"maari@tsl-me.com","status":"NE-Need Evaluation","posting_date": ["BETWEEN", ["2016-01-01",date]]})
		# total_ne_e = frappe.db.count("Work Order Data",{"technician":"eduardo@tsl-me.com","status":"NE-Need Evaluation","posting_date": ["BETWEEN", ["2016-01-01",date]]})
		# total_ne_a = frappe.db.count("Work Order Data",{"technician":"aakib@tsl-me.com","status":"NE-Need Evaluation","posting_date": ["BETWEEN", ["2016-01-01",date]]})
		
		for t in tech:
			emp = frappe.get_value("Employee",{"user_id":t})
			if emp:
				lp = frappe.db.sql(
					"""
					SELECT employee
					FROM `tabLeave Application Form` 
					WHERE employee = %s
					AND %s BETWEEN from_date AND to_date and docstatus = 1
					""",
					(emp,date),
					as_dict=True
				)
				if lp:
					if t == "sampath@tsl-me.com":
						lp_total_ne_s = "L"
					if t == "maari@tsl-me.com":
						lp_total_ne_m = "L"
					if t == "eduardo@tsl-me.com":
						lp_total_ne_e = "L"
					if t == "aakib@tsl-me.com":
						lp_total_ne_a = "L"
					if t == "amir@tsl-me.com":
						lp_total_ne_am = "L"
			
			
			ne = frappe.db.sql(""" select count(DISTINCT `tabWork Order Data`.name) as ct from `tabWork Order Data` 
					left join `tabStatus Duration Details` on `tabWork Order Data`.name = `tabStatus Duration Details`.parent
					where  `tabStatus Duration Details`.status = "UE-Under Evaluation"
					and `tabWork Order Data`.technician = "%s" and `tabStatus Duration Details`.date  LIKE "%s%%" """ %(t,original_date),as_dict=1)
				
			if t == "sampath@tsl-me.com":
				total_ne_s = ne[0]["ct"]
			if t == "maari@tsl-me.com":
				total_ne_m = ne[0]["ct"]
			if t == "eduardo@tsl-me.com":
				total_ne_e = ne[0]["ct"]
			if t == "aakib@tsl-me.com":
				total_ne_a = ne[0]["ct"]
			if t == "amir@tsl-me.com":
				total_ne_am = ne[0]["ct"]

		total_ner_s = frappe.db.count("Work Order Data",{"technician":"sampath@tsl-me.com","status":"NER-Need Evaluation Return","posting_date": ["BETWEEN", ["2016-01-01",date]]})
		total_ner_m = frappe.db.count("Work Order Data",{"technician":"maari@tsl-me.com","status":"NER-Need Evaluation Return","posting_date": ["BETWEEN", ["2016-01-01",date]]})
		total_ner_e = frappe.db.count("Work Order Data",{"technician":"eduardo@tsl-me.com","status":"NER-Need Evaluation Return","posting_date": ["BETWEEN", ["2016-01-01",date]]})
		total_ner_a = frappe.db.count("Work Order Data",{"technician":"aakib@tsl-me.com","status":"NER-Need Evaluation Return","posting_date": ["BETWEEN", ["2016-01-01",date]]})
		
		
		data += '<tr>'
		data += '<td style="line-height:0.8;border-bottom:hidden;border-color:#000000;width:10%;padding:1px;font-size:11px;font-size:11px;background-color:#A7C7E7;"><center><b></b><center></td>'
		data += '<td style="line-height:0.8;border-color:#000000;width:15%;padding:1px;font-size:11px;font-size:11px;background-color:#A7C7E7;"><center><b>UE</b><center></td>'
		
		if lp_total_ne_s == "L":
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>' %(lp_total_ne_s or 0)
			total_ne_s = 0
		else:
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>' %(total_ne_s or 0)

		if lp_total_ne_m == "L":
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>' %(lp_total_ne_m or 0)
			total_ne_m = 0
		else:
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>' %(total_ne_m or 0)
		
		if lp_total_ne_e == "L":
			total_ne_e = 0
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>' %(lp_total_ne_e or 0)
		else:
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>' %(total_ne_e or 0)
		
		if lp_total_ne_a == "L":
			total_ne_a = 0
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>'%(lp_total_ne_a or 0)
		else:
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>'%(total_ne_a or 0)

		if lp_total_ne_am == "L":
			total_ne_am = 0
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>'%(lp_total_ne_am or 0)
		else:
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>'%(total_ne_am or 0)

	
		data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>' %(total_ne_s + total_ne_m + total_ne_e + total_ne_a + total_ne_am)
		
		data += '</tr>'

		
		

		data += '<tr>'
		data += '<td style="line-height:0.8;border-bottom:hidden;border-color:#000000;width:10%;padding:1px;font-size:11px;font-size:11px;background-color:#A7C7E7;"><center><b>Started Work</b><center></td>'
		data += '<td style="line-height:0.8;border-color:#000000;width:15%;padding:1px;font-size:11px;font-size:11px;background-color:#A7C7E7;"><center><b>NER</b><center></td>'

		if lp_total_ne_s == "L":
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>' %(lp_total_ne_s or 0)
			total_ner_s = 0
		else:
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>' %(total_ner_s or 0)

		if lp_total_ne_m == "L":
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>' %(lp_total_ne_m or 0)
			total_ner_m = 0
		else:
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>' %(total_ner_m or 0)
		
		if lp_total_ne_e == "L":
			total_ner_e = 0
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>' %(lp_total_ne_e or 0)
		else:
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>' %(total_ner_e or 0)
		
		if lp_total_ne_a == "L":
			total_ner_a = 0
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>'%(lp_total_ne_a or 0)
		else:
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>'%(total_ner_a or 0)

		if lp_total_ne_am == "L":
			total_ner_am = 0
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>'%(lp_total_ne_am or 0)
		else:
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>'%(total_ner_am or 0)




		data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>' %(total_ner_s + total_ner_m + total_ner_e + total_ner_a + total_ner_am)
		data += '</tr>'

		data += '<tr>'
		data += '<td style="line-height:0.8;border-color:#000000;width:10%;padding:1px;font-size:11px;font-size:11px;background-color:#A7C7E7;"><center><b></b><center></td>'
		data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;background-color:#A7C7E7;"><center><b>Total</b><center></td>'
		data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;background-color:#A7C7E7;"><center><b>%s</b><center></td>' %(total_ne_s + total_ner_s)
		data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;background-color:#A7C7E7;"><center><b>%s</b><center></td>' %(total_ne_m + total_ner_m)
		data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;background-color:#A7C7E7;"><center><b>%s</b><center></td>' %(total_ne_e + total_ner_e)
		data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;background-color:#A7C7E7;"><center><b>%s</b><center></td>' %(total_ne_a + total_ner_a)
		data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;background-color:#A7C7E7;"><center><b>%s</b><center></td>' %(total_ne_am + total_ner_am)
		data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;background-color:#A7C7E7;"><center><b>%s</b><center></td>' %((total_ne_s + total_ne_m + total_ne_e + total_ne_a + total_ne_am)+(total_ner_s + total_ner_m + total_ner_e + total_ner_a + total_ner_am))
		data += '</tr>'


		data += '</table>'

		
		

		total_ue_s = frappe.db.count("Work Order Data",{"technician":"sampath@tsl-me.com","status":"UE-Under Evaluation","posting_date": ["BETWEEN", ["2016-01-01",date]]})
		total_ue_m = frappe.db.count("Work Order Data",{"technician":"maari@tsl-me.com","status":"UE-Under Evaluation","posting_date": ["BETWEEN", ["2016-01-01",date]]})
		total_ue_e = frappe.db.count("Work Order Data",{"technician":"eduardo@tsl-me.com","status":"UE-Under Evaluation","posting_date": ["BETWEEN", ["2016-01-01",date]]})
		total_ue_a = frappe.db.count("Work Order Data",{"technician":"aakib@tsl-me.com","status":"UE-Under Evaluation","posting_date": ["BETWEEN", ["2016-01-01",date]]})
		total_ue_am = frappe.db.count("Work Order Data",{"technician":"amir@tsl-me.com","status":"UE-Under Evaluation","posting_date": ["BETWEEN", ["2016-01-01",date]]})


		total_utr_s = frappe.db.count("Work Order Data",{"technician":"sampath@tsl-me.com","status":"UTR-Under Technician Repair","posting_date": ["BETWEEN", ["2016-01-01",date]]})
		total_utr_m = frappe.db.count("Work Order Data",{"technician":"maari@tsl-me.com","status":"UTR-Under Technician Repair","posting_date": ["BETWEEN", ["2016-01-01",date]]})
		total_utr_e = frappe.db.count("Work Order Data",{"technician":"eduardo@tsl-me.com","status":"UTR-Under Technician Repair","posting_date": ["BETWEEN", ["2016-01-01",date]]})
		total_utr_a = frappe.db.count("Work Order Data",{"technician":"aakib@tsl-me.com","status":"UTR-Under Technician Repair","posting_date": ["BETWEEN", ["2016-01-01",date]]})
		total_utr_am = frappe.db.count("Work Order Data",{"technician":"amir@tsl-me.com","status":"UTR-Under Technician Repair","posting_date": ["BETWEEN", ["2016-01-01",date]]})

		for t in tech:
			ue = frappe.db.sql(""" select count(DISTINCT `tabWork Order Data`.name) as ct from `tabWork Order Data` 
				left join `tabStatus Duration Details` on `tabWork Order Data`.name = `tabStatus Duration Details`.parent
				where  `tabStatus Duration Details`.status = "UE-Under Evaluation"
				and `tabWork Order Data`.technician = "%s" and `tabStatus Duration Details`.date  LIKE "%s%%" """ %(t,original_date) ,as_dict=1)
			
			if t == "sampath@tsl-me.com":
				ue_s = ue[0]["ct"]
			if t == "maari@tsl-me.com":
				ue_m = ue[0]["ct"]
			if t == "eduardo@tsl-me.com":
				ue_e = ue[0]["ct"]
			if t == "aakib@tsl-me.com":
				ue_a = ue[0]["ct"]
			if t == "amir@tsl-me.com":
				ue_am = ue[0]["ct"]

			utr = frappe.db.sql(""" select count(DISTINCT `tabWork Order Data`.name) as ct from `tabWork Order Data` 
				left join `tabStatus Duration Details` on `tabWork Order Data`.name = `tabStatus Duration Details`.parent
				where  `tabStatus Duration Details`.status = "UTR-Under Technician Repair"
				and `tabWork Order Data`.technician = "%s" and `tabStatus Duration Details`.date  LIKE "%s%%" """ %(t,original_date),as_dict=1)
			
			if t == "sampath@tsl-me.com":
				utr_s = utr[0]["ct"]
			if t == "maari@tsl-me.com":
				utr_m = utr[0]["ct"]
			if t == "eduardo@tsl-me.com":
				utr_e = utr[0]["ct"]
			if t == "aakib@tsl-me.com":
				utr_a = utr[0]["ct"]
			if t == "amir@tsl-me.com":
				utr_am = utr[0]["ct"]
			# ne = frappe.get_all("Work Order Data",{"technician":t},["*"])
			# for i in ne:
			# 	sd = frappe.get_doc("Work Order Data",i.name)
			# 	for j in sd.status_duration_details:
				
			# 		timestamp = str(j.date) 
			# 		date_portion = timestamp[:10]
					
			# 		j.date= date_portion
					
			# 		if "UE-Under Evaluation" == j.status and i.technician == "sampath@tsl-me.com" and str(j.date) == original_date:
			# 			ue_s = ue_s + 1
			# 		if "UE-Under Evaluation" == j.status and i.technician == "maari@tsl-me.com" and str(j.date) == original_date:
			# 			ue_m = ue_m + 1
					
			# 		if "UE-Under Evaluation" == j.status and i.technician == "eduardo@tsl-me.com" and str(j.date) == original_date:
			# 			ue_e = ue_e + 1
			# 		if "UE-Under Evaluation" == j.status and i.technician == "aakib@tsl-me.com" and str(j.date) == original_date:
			# 			ue_a = ue_a + 1
						
			# 		if "UTR-Under Technician Repair" == j.status and i.technician == "sampath@tsl-me.com" and str(j.date) == original_date:
			# 			utr_s = utr_s + 1
			# 		if "UTR-Under Technician Repair" == j.status and i.technician == "maari@tsl-me.com" and str(j.date) == original_date:
			# 			utr_m = utr_m + 1
					
			# 		if "UTR-Under Technician Repair" == j.status and i.technician == "eduardo@tsl-me.com" and str(j.date) == original_date:
			# 			utr_e = utr_e + 1
			# 		if "UTR-Under Technician Repair" == j.status and i.technician == "aakib@tsl-me.com" and str(j.date) == original_date:
			# 			utr_a = utr_a + 1
						

		
		data += '<table class="table table-bordered">'

		data += '<tr>'
		data += '<td colspan="2" style="border-left:none; border-top:none; border-right:none; border-color:#000000; padding:1px; font-size:11px; line-height:0.8;"><center><b></b></center></td>'
		data += '<td style="border-left:none; border-top:none; border-right:none; border-color:#000000; width:13%; padding:1px; font-size:11px; line-height:0.8;"><center><b></b></center></td>'
		data += '<td style="border-left:none; border-top:none; border-right:none; border-color:#000000; width:13%; padding:1px; font-size:11px; line-height:0.8;"><center><b></b></center></td>'
		data += '<td style="border-left:none; border-top:none; border-right:none; border-color:#000000; width:12%; padding:1px; font-size:11px; line-height:0.8;"><center><b></b></center></td>'
		data += '<td style="border-left:none; border-top:none; border-right:none; border-color:#000000; width:12%; padding:1px; font-size:11px; line-height:0.8;"><center><b></b></center></td>'
		data += '<td style="border-left:none; border-top:none; border-right:none; border-color:#000000; width:12%; padding:1px; font-size:11px; line-height:0.8;"><center><b></b></center></td>'
		data += '<td style="border-left:none; border-top:none; border-right:none; border-color:#000000; width:12%; padding:1px; font-size:11px; line-height:0.8;"><center><b></b></center></td>'
		data += '</tr>'

		data += '<tr>'
		data += '<td style="line-height:0.8;border-bottom:hidden;border-color:#000000;width:10%;padding:1px;font-size:11px;font-size:11px;background-color:#A7C7E7;"><center><b></b><center></td>'
		data += '<td style="line-height:0.8;border-color:#000000;width:12%;padding:1px;font-size:11px;font-size:11px;background-color:#A7C7E7;"><center><b>UE</b><center></td>' 
		
		
		if lp_total_ne_s == "L":
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>' %(lp_total_ne_s or 0)
			total_ue_s = 0
		else:
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>' %(total_ue_s or 0)

		if lp_total_ne_m == "L":
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>' %(lp_total_ne_m or 0)
			total_ue_m = 0
		else:
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>' %(total_ue_m or 0)
		
		if lp_total_ne_e == "L":
			total_ue_e = 0
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>' %(lp_total_ne_e or 0)
		else:
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>' %(total_ue_e or 0)
		
		if lp_total_ne_a == "L":
			total_ue_a = 0
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>'%(lp_total_ne_a or 0)
		else:
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>'%(total_ue_a or 0)

		if lp_total_ne_am == "L":
			total_ue_am = 0
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>'%(lp_total_ne_am or 0)
		else:
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>'%(total_ue_am or 0)

		
		data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>' %(total_ue_s + total_ue_m + total_ue_e + total_ue_a + total_ue_am)
		data += '</tr>'

		data += '<tr>'
		data += '<td style="line-height:0.8;border-bottom:hidden;border-color:#000000;width:10%;padding:1px;font-size:11px;font-size:11px;background-color:#A7C7E7;"><center><b>End of Work</b><center></td>'
		data += '<td style="line-height:0.8;border-color:#000000;width:15%;padding:1px;font-size:11px;font-size:11px;background-color:#A7C7E7;"><center><b>UTR</b><center></td>'
		
		if lp_total_ne_s == "L":
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>' %(lp_total_ne_s or 0)
			total_utr_s = 0
		else:
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>' %(total_utr_s or 0)

		if lp_total_ne_m == "L":
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>' %(lp_total_ne_m or 0)
			total_utr_m = 0
		else:
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>' %(total_utr_m or 0)
		
		if lp_total_ne_e == "L":
			total_utr_e = 0
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>' %(lp_total_ne_e or 0)
		else:
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>' %(total_utr_e or 0)
		
		if lp_total_ne_a == "L":
			total_utr_a = 0
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>'%(lp_total_ne_a or 0)
		else:
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>'%(total_utr_a or 0)

		if lp_total_ne_am == "L":
			total_utr_am = 0
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>'%(lp_total_ne_am or 0)
		else:
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>'%(total_utr_am or 0)


		data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>' %(total_utr_s +total_utr_m + total_utr_e + total_utr_a + total_utr_am)
		data += '</tr>'

		data += '<tr>'
		data += '<td style="line-height:0.8;border-bottom:hidden;border-color:#000000;width:10%;padding:1px;font-size:11px;background-color:#A7C7E7;"><center><b></b><center></td>'
		data += '<td style="line-height:0.8;border-color:#000000;width:15%;padding:1px;font-size:11px;font-size:11px;background-color:#A7C7E7;"><center><b>Site Visit</b><center></td>'
		
		if lp_total_ne_s == "L":
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>' %(lp_total_ne_s or 0)
			total_utr_s = 0
		else:
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>' %(0 or 0)

		if lp_total_ne_m == "L":
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>' %(lp_total_ne_m or 0)
			total_utr_m = 0
		else:
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>' %(0 or 0)
		
		if lp_total_ne_e == "L":
			total_utr_e = 0
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>' %(lp_total_ne_e or 0)
		else:
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>' %(0 or 0)
		
		if lp_total_ne_a == "L":
			total_utr_a = 0
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>'%(lp_total_ne_a or 0)
		else:
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>'%(0 or 0)

		if lp_total_ne_am == "L":
			total_utr_am = 0
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>'%(lp_total_ne_am or 0)
		else:
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>'%(0 or 0)

			
		data += '<td style="line-height:0.8;border-color:#000000;width:15%;padding:1px;font-size:11px;font-size:11px;"><center><b>0	</b><center></td>'
		
		data += '</tr>'

		data += '<tr>'
		data += '<td style="line-height:0.8;border-color:#000000;width:10%;padding:1px;font-size:11px;font-size:11px;background-color:#A7C7E7;"><center><b></b><center></td>'
		data += '<td style="line-height:0.8;border-color:#000000;width:15%;padding:1px;font-size:11px;font-size:11px;background-color:#A7C7E7;"><center><b>Total</b><center></td>'
		

		data += '<td style="line-height:0.8;border-color:#000000;;padding:1px;font-size:11px;font-size:11px;background-color:#A7C7E7;"><center><b>%s</b><center></td>' %(total_ue_s + total_utr_s)
		data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;background-color:#A7C7E7;"><center><b>%s</b><center></td>' %(total_ue_m + total_utr_m)
		data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;background-color:#A7C7E7;"><center><b>%s</b><center></td>' %(total_ue_e + total_utr_e)
		data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;background-color:#A7C7E7;"><center><b>%s</b><center></td>' %(total_ue_a + total_utr_a)
		data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;background-color:#A7C7E7;"><center><b>%s</b><center></td>' %(total_ue_am + total_utr_am)
		data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;background-color:#A7C7E7;"><center><b>%s</b><center></td>' %((total_ue_s + total_ue_m + total_ue_e + total_ue_a + total_ue_am)+(total_utr_s +  total_utr_m + total_utr_e + total_utr_a + total_utr_am))
		data += '</tr>'

		data += '</table>'

		data += '<table class="table table-bordered">'
		
		data += '<tr>'
		data += '<td colspan="2" style="border-left:none; border-top:none; border-right:none; border-color:#000000; padding:1px; font-size:11px; line-height:0.8;"><center><b></b></center></td>'
		data += '<td style="border-left:none; border-top:none; border-right:none; border-color:#000000; width:12%; padding:1px; font-size:11px; line-height:0.8;"><center><b></b></center></td>'
		data += '<td style="border-left:none; border-top:none; border-right:none; border-color:#000000; width:12%; padding:1px; font-size:11px; line-height:0.8;"><center><b></b></center></td>'
		data += '<td style="border-left:none; border-top:none; border-right:none; border-color:#000000; width:12%; padding:1px; font-size:11px; line-height:0.8;"><center><b></b></center></td>'
		data += '<td style="border-left:none; border-top:none; border-right:none; border-color:#000000; width:12%; padding:1px; font-size:11px; line-height:0.8;"><center><b></b></center></td>'
		data += '<td style="border-left:none; border-top:none; border-right:none; border-color:#000000; width:12%; padding:1px; font-size:11px; line-height:0.8;"><center><b></b></center></td>'
		data += '<td style="border-left:none; border-top:none; border-right:none; border-color:#000000; width:12%; padding:1px; font-size:11px; line-height:0.8;"><center><b></b></center></td>'
		data += '</tr>'

		for t in tech:
			s = frappe.db.sql(""" select count(DISTINCT `tabWork Order Data`.name) as ct from `tabWork Order Data` 
				left join `tabStatus Duration Details` on `tabWork Order Data`.name = `tabStatus Duration Details`.parent
				where  `tabStatus Duration Details`.status = "RS-Repaired and Shipped"
				and `tabWork Order Data`.technician = "%s" and `tabStatus Duration Details`.date  LIKE "%s%%" """ %(t,original_date) ,as_dict=1)
			
			if t == "sampath@tsl-me.com":
				rs_s = s[0]["ct"]
			if t == "maari@tsl-me.com":
				rs_m = s[0]["ct"]
			if t == "eduardo@tsl-me.com":
				rs_e = s[0]["ct"]
			if t == "aakib@tsl-me.com":
				rs_a = s[0]["ct"]
			if t == "amir@tsl-me.com":
				rs_am = s[0]["ct"]

			w = frappe.db.sql(""" select count(DISTINCT `tabWork Order Data`.name) as ct from `tabWork Order Data` 
				left join `tabStatus Duration Details` on `tabWork Order Data`.name = `tabStatus Duration Details`.parent
				where  `tabStatus Duration Details`.status = "W-Working" 
				and `tabWork Order Data`.technician = "%s" 
				and `tabStatus Duration Details`.date  LIKE "%s%%" """ %(t,original_date) ,as_dict=1)
			
			r = frappe.db.sql(""" select count(DISTINCT `tabWork Order Data` .name) as ct from `tabWork Order Data` 
				left join `tabStatus Duration Details` on `tabWork Order Data`.name = `tabStatus Duration Details`.parent
				where  `tabStatus Duration Details`.status = "RS-Repaired and Shipped" 
				and `tabWork Order Data`.technician = "%s" 
				and `tabStatus Duration Details`.date  LIKE "%s%%" """ %(t,original_date) ,as_dict=1)
			
			
			if t == "sampath@tsl-me.com":
				if not r:
					w_s = w[0]["ct"]
			if t == "maari@tsl-me.com":
				if not r:
					w_m = w[0]["ct"]
			if t == "eduardo@tsl-me.com":
				if not r:
					w_e = w[0]["ct"]
			if t == "aakib@tsl-me.com":
				if not r:
					w_a = w[0]["ct"]
			if t == "aakib@tsl-me.com":
				if not r:
					w_a = w[0]["ct"]
			if t == "amir@tsl-me.com":
				if not r:
					w_am = w[0]["ct"]
				
					


			rnr = frappe.db.sql(""" select count(DISTINCT `tabWork Order Data` .name) as ct from `tabWork Order Data` 
				left join `tabStatus Duration Details` on `tabWork Order Data`.name = `tabStatus Duration Details`.parent
				where  `tabStatus Duration Details`.status = "RNR-Return Not Repaired"
						and `tabWork Order Data`.technician = "%s" and `tabStatus Duration Details`.date  LIKE "%s%%" """ %(t,original_date) ,as_dict=1)
			
			if t == "sampath@tsl-me.com":
				rnr_s = rnr[0]["ct"]
			if t == "maari@tsl-me.com":
				rnr_m = rnr[0]["ct"]
			if t == "eduardo@tsl-me.com":
				rnr_e = rnr[0]["ct"]
			if t == "aakib@tsl-me.com":
				rnr_a = rnr[0]["ct"]
			if t == "amir@tsl-me.com":
				rnr_am = rnr[0]["ct"]

			rnf = frappe.db.sql(""" select count(DISTINCT `tabWork Order Data` .name) as ct from `tabWork Order Data` 
				left join `tabStatus Duration Details` on `tabWork Order Data`.name = `tabStatus Duration Details`.parent
				where  `tabStatus Duration Details`.status = "RNF-Return No Fault"
				and `tabWork Order Data`.technician = "%s" and `tabStatus Duration Details`.date  LIKE "%s%%" """ %(t,original_date) ,as_dict=1)
			
			if t == "sampath@tsl-me.com":
				rnf_s = rnf[0]["ct"]
			if t == "maari@tsl-me.com":
				rnf_m = rnf[0]["ct"]
			if t == "eduardo@tsl-me.com":
				rnf_e = rnf[0]["ct"]
			if t == "aakib@tsl-me.com":
				rnf_a = rnf[0]["ct"]
			if t == "amir@tsl-me.com":
				rnf_am = rnf[0]["ct"]

			com = frappe.db.sql(""" select count(DISTINCT `tabWork Order Data` .name) as ct from `tabWork Order Data` 
				left join `tabStatus Duration Details` on `tabWork Order Data`.name = `tabStatus Duration Details`.parent
				where  `tabStatus Duration Details`.status = "C-Comparison"
						and `tabWork Order Data`.technician = "%s" and `tabStatus Duration Details`.date  LIKE "%s%%" """ %(t,original_date) ,as_dict=1)
			
			if t == "sampath@tsl-me.com":
				com_s = com[0]["ct"]
			if t == "maari@tsl-me.com":
				com_m = com[0]["ct"]
			if t == "eduardo@tsl-me.com":
				com_e = com[0]["ct"]
			if t == "aakib@tsl-me.com":
				com_a = com[0]["ct"]
			if t == "amir@tsl-me.com":
				com_am = com[0]["ct"]


			# ne = frappe.get_all("Work Order Data",{"technician":t},["*"])
			# for i in ne:
			# 	sd = frappe.get_doc("Work Order Data",i.name)
			# 	for j in sd.status_duration_details:
				
			# 		timestamp = str(j.date) 
			# 		date_portion = timestamp[:10]
					
			# 		j.date= date_portion
					
			# 		if "RS-Repaired and Shipped"  == j.status and i.technician == "sampath@tsl-me.com" and str(j.date) == original_date:
			# 			rs_s = rs_s + 1
			# 		if "RS-Repaired and Shipped"  == j.status and i.technician == "maari@tsl-me.com" and str(j.date) == original_date:
			# 			rs_m = rs_m + 1
			# 		if "RS-Repaired and Shipped"  == j.status and i.technician == "eduardo@tsl-me.com" and str(j.date) == original_date:
			# 			rs_e = rs_e + 1
			# 		if "RS-Repaired and Shipped"  == j.status and i.technician == "aakib@tsl-me.com" and str(j.date) == original_date:
			# 			rs_a = rs_a + 1

			# 		s = frappe.db.sql(""" select count(DISTINCT `tabWork Order Data` .name) as ct from `tabWork Order Data` 
			# 			left join `tabStatus Duration Details` on `tabWork Order Data`.name = `tabStatus Duration Details`.parent
			# 			where  `tabStatus Duration Details`.status = "W-Working" 
			# 			and `tabWork Order Data`.technician = "%s" and `tabStatus Duration Details`.date LIKE "%s" """ %("sampath@tsl-me.com",original_date) ,as_dict=1)
			# 		
			# 		if "W-Working"  == j.status and i.technician == "sampath@tsl-me.com" and str(j.date) == original_date:
			# 			w_s = w_s + 1
						
			# 		if "W-Working"  == j.status and i.technician == "maari@tsl-me.com" and str(j.date) == original_date:
			# 			w_m = w_m + 1
						
			# 		if "W-Working"  == j.status and i.technician == "eduardo@tsl-me.com" and str(j.date) == original_date:
			# 			w_e = w_e + 1
			# 		if "W-Working"  == j.status and i.technician == "aakib@tsl-me.com" and str(j.date) == original_date:
			# 			w_a = w_a + 1

			# 		if "RNR-Return Not Repaired"  == j.status and i.technician == "sampath@tsl-me.com" and str(j.date) == original_date:
			# 			rnr_s = rnr_s + 1
			# 		if "RNR-Return Not Repaired"  == j.status and i.technician == "maari@tsl-me.com" and str(j.date) == original_date:
			# 			rnr_m = rnr_m + 1
			# 		if "RNR-Return Not Repaired"  == j.status and i.technician == "eduardo@tsl-me.com" and str(j.date) == original_date:
			# 			rnr_e = rnr_e + 1
			# 		if "RNR-Return Not Repaired"  == j.status and i.technician == "aakib@tsl-me.com" and str(j.date) == original_date:
			# 			rnr_a = rnr_a + 1
					
			# 		if "RNF-Return No Fault"  == j.status and i.technician == "sampath@tsl-me.com" and str(j.date) == original_date:
			# 			rnf_s = rnf_s + 1
			# 		if "RNF-Return No Fault"  == j.status and i.technician == "maari@tsl-me.com" and str(j.date) == original_date:
			# 			rnf_m = rnf_m + 1
			# 		if "RNF-Return No Fault"  == j.status and i.technician == "eduardo@tsl-me.com" and str(j.date) == original_date:
			# 			rnf_e = rnf_e + 1
			# 		if "RNF-Return No Fault"  == j.status and i.technician == "aakib@tsl-me.com" and str(j.date) == original_date:
			# 			rnf_a = rnf_a + 1

			# 		if "C-Comparison"  == j.status and i.technician == "sampath@tsl-me.com" and str(j.date) == original_date:
			# 			com_s = com_s + 1
			# 		if "C-Comparison"  == j.status and i.technician == "maari@tsl-me.com" and str(j.date) == original_date:
			# 			com_m = com_m + 1
			# 		if "C-Comparison"  == j.status and i.technician == "eduardo@tsl-me.com" and str(j.date) == original_date:
			# 			com_e = com_e + 1
			# 		if "C-Comparison"  == j.status and i.technician == "aakib@tsl-me.com" and str(j.date) == original_date:
			# 			com_a = com_a + 1
					
					
					
		
		data += '<tr>'
		data += '<td style="line-height:0.8;border-bottom:hidden;border-color:#000000;width:10%;padding:1px;font-size:11px;background-color:#A7C7E7;"><center><b></b><center></td>'
		data += '<td style="line-height:0.8;border-color:#000000;width:15%;padding:1px;font-size:11px;font-size:11px;background-color:#A7C7E7;"><center><b>RS</b><center></td>'
		
		if lp_total_ne_s == "L":
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>' %(lp_total_ne_s or 0)
			rs_s = 0
		else:
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>' %(rs_s or 0)

		if lp_total_ne_m == "L":
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>' %(lp_total_ne_m or 0)
			rs_m = 0
		else:
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>' %(rs_m or 0)
		
		if lp_total_ne_e == "L":
			rs_e = 0
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>' %(lp_total_ne_e or 0)
		else:
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>' %(rs_e or 0)
		
		if lp_total_ne_a == "L":
			rs_a = 0
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>'%(lp_total_ne_a or 0)
		else:
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>'%(rs_a or 0)

		if lp_total_ne_am == "L":
			rs_am = 0
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>'%(lp_total_ne_am or 0)
		else:
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>'%(rs_am or 0)

		
		data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>' %(rs_s + rs_m + rs_e + rs_a + rs_a)
		data += '</tr>'

		ev_s = 0
		ev_m = 0
		ev_e = 0
		ev_a = 0
		ev_am = 0

		# for t in tech:
		# 	ne = frappe.get_all("Work Order Data",{"date":original_date,"technician":t},["*"])
		# 	for i in ne:
		# 		ev = frappe.db.exists("Evaluation Report",{"work_order_data":i.name})
		# 		if ev:
		# 			techni = frappe.get_value("Evaluation Report",{"name":ev},["technician"])
		# 			if techni == "sampath@tsl-me.com":
		# 				ev_s = ev_s + 1
		# 			if techni == "maari@tsl-me.com":
		# 				ev_m = ev_m + 1
		# 			if techni == "eduardo@tsl-me.com":
		# 				ev_e = ev_e + 1
		# 			if techni == "aakib@tsl-me.com":
		# 				ev_a = ev_a + 1
		ev_s = frappe.db.count("Evaluation Report",{"date":original_date,"technician":"sampath@tsl-me.com"})
		ev_m = frappe.db.count("Evaluation Report",{"date":original_date,"technician":"maari@tsl-me.com"})
		ev_e = frappe.db.count("Evaluation Report",{"date":original_date,"technician":"eduardo@tsl-me.com"})
		ev_a = frappe.db.count("Evaluation Report",{"date":original_date,"technician":"aakib@tsl-me.com"})
		ev_am = frappe.db.count("Evaluation Report",{"date":original_date,"technician":"amir@tsl-me.com"})
					
		

		data += '<tr>'
		data += '<td style="line-height:0.8;border-bottom:hidden;border-color:#000000;width:10%;padding:1px;font-size:11px;background-color:#A7C7E7;"><center><b></b><center></td>'
		data += '<td style="line-height:0.8;border-color:#000000;width:15%;padding:1px;font-size:11px;background-color:#A7C7E7;"><center><b>PS</b><center></td>'
		
		
		if lp_total_ne_s == "L":
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>' %(lp_total_ne_s or 0)
			ev_s = 0
		else:
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>' %(ev_s or 0)

		if lp_total_ne_m == "L":
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>' %(lp_total_ne_m or 0)
			ev_m = 0
		else:
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>' %(ev_m or 0)
		
		if lp_total_ne_e == "L":
			ev_e = 0
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>' %(lp_total_ne_e or 0)
		else:
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>' %(ev_e or 0)
		
		if lp_total_ne_a == "L":
			ev_a = 0
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>'%(lp_total_ne_a or 0)
		else:
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>'%(ev_a or 0)

		if lp_total_ne_am == "L":
			ev_am = 0
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>'%(lp_total_ne_am or 0)
		else:
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>'%(ev_am or 0)

		
		
		
		data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>' %(ev_s + ev_m + ev_e + ev_a + ev_am)
		data += '</tr>'

		
		data += '<tr>'
		data += '<td style="line-height:0.8;border-bottom:hidden;border-color:#000000;width:10%;padding:1px;font-size:11px;background-color:#A7C7E7;"><center><b></b><center></td>'
		data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;background-color:#A7C7E7;"><center><b>W</b><center></td>'
		
		if lp_total_ne_s == "L":
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>' %(lp_total_ne_s or 0)
			w_s = 0
		else:
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>' %(w_s or 0)

		if lp_total_ne_m == "L":
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>' %(lp_total_ne_m or 0)
			w_m = 0
		else:
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>' %(w_m or 0)
		
		if lp_total_ne_e == "L":
			w_e = 0
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>' %(lp_total_ne_e or 0)
		else:
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>' %(w_e or 0)
		
		if lp_total_ne_a == "L":
			w_a = 0
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>'%(lp_total_ne_a or 0)
		else:
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>'%(w_a or 0)

		if lp_total_ne_am == "L":
			w_am = 0
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>'%(lp_total_ne_am or 0)
		else:
			frappe.errprint(w_am)
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>'%(w_am or 0)

		
		
		data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>' %(w_s + w_m + w_e + w_a + w_am)
		data += '</tr>'

		
		data += '<tr>'
		data += '<td style="line-height:0.8;border-bottom:hidden;border-color:#000000;width:10%;padding:1px;font-size:11px;background-color:#A7C7E7;"><center><b>Out - Flow</b><center></td>'
		data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;background-color:#A7C7E7;"><center><b>RNR</b><center></td>'


		if lp_total_ne_s == "L":
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>' %(lp_total_ne_s or 0)
			rnr_s = 0
		else:
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>' %(rnr_s or 0)

		if lp_total_ne_m == "L":
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>' %(lp_total_ne_m or 0)
			rnr_m = 0
		else:
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>' %(rnr_m or 0)
	
		if lp_total_ne_e == "L":
			rnr_e = 0
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>' %(lp_total_ne_e or 0)
		else:
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>' %(rnr_e or 0)
		
		if lp_total_ne_a == "L":
			rnr_a = 0
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>'%(lp_total_ne_a or 0)
		else:
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>'%(rnr_a or 0)

		if lp_total_ne_am == "L":
			rnr_am = 0
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>'%(lp_total_ne_am or 0)
		else:
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>'%(rnr_am or 0)

	

		data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>' %(rnr_s + rnr_m + rnr_e + rnr_a + rnr_am)
		data += '</tr>'

		# rnf_s = frappe.db.count("Work Order Data",{"posting_date":original_date,"technician":"sampath@tsl-me.com","status":"RNF-Return No Fault"})
		# rnf_m = frappe.db.count("Work Order Data",{"posting_date":original_date,"technician":"maari@tsl-me.com","status":"RNF-Return No Fault"})
		# rnf_e = frappe.db.count("Work Order Data",{"posting_date":original_date,"technician":"eduardo@tsl-me.com","status":"RNF-Return No Fault"})
		# rnf_a = frappe.db.count("Work Order Data",{"posting_date":original_date,"technician":"aakib@tsl-me.com","status":"RNF-Return No Fault"})


		data += '<tr>'
		data += '<td style="line-height:0.8;border-bottom:hidden;border-color:#000000;width:10%;padding:1px;font-size:11px;background-color:#A7C7E7;"><center><b></b><center></td>'
		data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;background-color:#A7C7E7;"><center><b>RNF</b><center></td>'

		if lp_total_ne_s == "L":
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>' %(lp_total_ne_s or 0)
			rnf_s = 0
		else:
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>' %(rnf_s or 0)

		if lp_total_ne_m == "L":
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>' %(lp_total_ne_m or 0)
			rnf_m = 0
		else:
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>' %(rnf_m or 0)
	
		if lp_total_ne_e == "L":
			rnf_e = 0
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>' %(lp_total_ne_e or 0)
		else:
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>' %(rnf_e or 0)
		
		if lp_total_ne_a == "L":
			rnf_a = 0
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>'%(lp_total_ne_a or 0)
		else:
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>'%(rnf_a or 0)

		if lp_total_ne_am == "L":
			rnf_am = 0
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>'%(lp_total_ne_am or 0)
		else:
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>'%(rnf_am or 0)



		data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>' %(rnf_a + rnf_s + rnf_e + rnf_m + rnf_am) 
		data += '</tr>'

		# com_s = frappe.db.count("Work Order Data",{"posting_date":original_date,"technician":"sampath@tsl-me.com","status":"C-Comparison"})
		# com_m = frappe.db.count("Work Order Data",{"posting_date":original_date,"technician":"maari@tsl-me.com","status":"C-Comparison"})
		# com_e = frappe.db.count("Work Order Data",{"posting_date":original_date,"technician":"eduardo@tsl-me.com","status":"C-Comparison"})
		# com_a = frappe.db.count("Work Order Data",{"posting_date":original_date,"technician":"aakib@tsl-me.com","status":"C-Comparison"})


		data += '<tr>'
		data += '<td style="line-height:0.8;border-bottom:hidden;border-color:#000000;width:10%;padding:1px;font-size:11px;background-color:#A7C7E7;"><center><b></b><center></td>'
		data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;background-color:#A7C7E7;"><center><b>COMP</b><center></td>'

		if lp_total_ne_s == "L":
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>' %(lp_total_ne_s or 0)
			com_s = 0
		else:
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>' %(com_s or 0)

		if lp_total_ne_m == "L":
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>' %(lp_total_ne_m or 0)
			com_m = 0
		else:
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>' %(com_m or 0)
	
		if lp_total_ne_e == "L":
			com_e = 0
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>' %(lp_total_ne_e or 0)
		else:
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>' %(com_e or 0)
		
		if lp_total_ne_a == "L":
			com_a = 0
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>'%(lp_total_ne_a or 0)
		else:
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>'%(com_a or 0)

		if lp_total_ne_am == "L":
			com_am = 0
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>'%(lp_total_ne_am or 0)
		else:
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>'%(com_am or 0)


		data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>' %(com_s + com_m + com_e + com_a)
		data += '</tr>'

		total_out_s = (rs_s or 0) + (ev_s or 0) + (w_s or 0) + (rnr_s or 0) + (rnf_s or 0) + (com_s or 0)
		total_out_m = (rs_m or 0) + (ev_m or 0) + (w_m or 0) + (rnr_m or 0) + (rnf_m or 0) + (com_m or 0)
		total_out_e = (rs_e or 0) + (ev_e or 0) + (w_e or 0) + (rnr_e or 0) + (rnf_e or 0) + (com_e or 0)
		total_out_a = (rs_a or 0) + (ev_a or 0) + (w_a or 0) + (rnr_a or 0) + (rnf_a or 0) + (com_a or 0)
		total_out_am = (rs_am or 0) + (ev_am or 0) + (w_am or 0) + (rnr_am or 0) + (rnf_am or 0) + (com_am or 0)

		sum_total_out = (rs_s + rs_m + rs_e + rs_a +  rs_am) + (ev_s + ev_m + ev_e + ev_a + ev_am) + (w_s + w_m + w_e + w_a + w_am) + (rnr_s + rnr_m + rnr_e + rnr_a + rnr_am) + (rnf_a + rnf_s + rnf_e + rnf_m + rnf_am)

		data += '<tr>'
		data += '<td style="line-height:0.8;border-color:#000000;width:10%;padding:1px;font-size:11px;background-color:#A7C7E7;"><center><b></b><center></td>'
		data += '<td style="line-height:0.8;border-color:#000000;width:15%;padding:1px;font-size:11px;background-color:#A7C7E7;"><center><b>Total</b><center></td>'
		data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;background-color:#A7C7E7;"><center><b>%s</b><center></td>' %(total_out_s or 0)
		data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;background-color:#A7C7E7;"><center><b>%s</b><center></td>' %(total_out_m or 0)
		data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;background-color:#A7C7E7;"><center><b>%s</b><center></td>' %(total_out_e or 0)
		data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;background-color:#A7C7E7;"><center><b>%s</b><center></td>' %(total_out_a or 0)
		data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;background-color:#A7C7E7;"><center><b>%s</b><center></td>' %(total_out_am or 0)
		data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;background-color:#A7C7E7;"><center><b>%s</b><center></td>' %(sum_total_out or 0)
		data += '</tr>'

		data += '</table>'


		data += '<table class="table table-bordered">'

		data += '<tr>'
		data += '<td colspan="2" style="border-left:none; border-top:none; border-right:none; border-color:#000000; padding:1px; font-size:11px; line-height:0.8;"><center><b></b></center></td>'
		data += '<td style="border-left:none; border-top:none; border-right:none; border-color:#000000; width:12%; padding:1px; font-size:11px; line-height:0.8;"><center><b></b></center></td>'
		data += '<td style="border-left:none; border-top:none; border-right:none; border-color:#000000; width:12%; padding:1px; font-size:11px; line-height:0.8;"><center><b></b></center></td>'
		data += '<td style="border-left:none; border-top:none; border-right:none; border-color:#000000; width:12%; padding:1px; font-size:11px; line-height:0.8;"><center><b></b></center></td>'
		data += '<td style="border-left:none; border-top:none; border-right:none; border-color:#000000; width:12%; padding:1px; font-size:11px; line-height:0.8;"><center><b></b></center></td>'
		data += '<td style="border-left:none; border-top:none; border-right:none; border-color:#000000; width:12%; padding:1px; font-size:11px; line-height:0.8;"><center><b></b></center></td>'
		data += '<td style="border-left:none; border-top:none; border-right:none; border-color:#000000; width:12%; padding:1px; font-size:11px; line-height:0.8;"><center><b></b></center></td>'
		data += '</tr>'
		
		p_total_ne_s = 0
		p_total_ne_m = 0
		p_total_ne_e = 0
		p_total_ne_a = 0

		p_total_ner_s = 0
		p_total_ner_m = 0
		p_total_ner_e = 0
		p_total_ner_a = 0
		dat = add_days(today(), -1)
		

		p_total_ne_s = frappe.db.count("Work Order Data",{"technician":"sampath@tsl-me.com","status":"NE-Need Evaluation","posting_date": ["BETWEEN", ["2016-01-01", dat]]})
		p_total_ne_m = frappe.db.count("Work Order Data",{"technician":"maari@tsl-me.com","status":"NE-Need Evaluation","posting_date": ["BETWEEN", ["2016-01-01", dat]]})
		p_total_ne_e = frappe.db.count("Work Order Data",{"technician":"eduardo@tsl-me.com","status":"NE-Need Evaluation","posting_date": ["BETWEEN", ["2016-01-01",dat]]})
		p_total_ne_a = frappe.db.count("Work Order Data",{"technician":"aakib@tsl-me.com","status":"NE-Need Evaluation","posting_date": ["BETWEEN", ["2016-01-01",dat]]})
		p_total_ne_am = frappe.db.count("Work Order Data",{"technician":"amir@tsl-me.com","status":"NE-Need Evaluation","posting_date": ["BETWEEN", ["2016-01-01",dat]]})


		p_total_ner_s = frappe.db.count("Work Order Data",{"technician":"sampath@tsl-me.com","status":"NER-Need Evaluation Return","posting_date": ["BETWEEN", ["2016-01-01", dat]]})
		p_total_ner_m = frappe.db.count("Work Order Data",{"technician":"maari@tsl-me.com","status":"NER-Need Evaluation Return","posting_date": ["BETWEEN", ["2016-01-01", dat]]})
		p_total_ner_e = frappe.db.count("Work Order Data",{"technician":"eduardo@tsl-me.com","status":"NER-Need Evaluation Return","posting_date": ["BETWEEN", ["2016-01-01",dat]]})
		p_total_ner_a = frappe.db.count("Work Order Data",{"technician":"aakib@tsl-me.com","status":"NER-Need Evaluation Return","posting_date": ["BETWEEN", ["2016-01-01",dat]]})
		p_total_ner_am = frappe.db.count("Work Order Data",{"technician":"amir@tsl-me.com","status":"NER-Need Evaluation Return","posting_date": ["BETWEEN", ["2016-01-01",dat]]})
		
		
		nes = total_ne_s- p_total_ne_s
		
		nem = total_ne_m - p_total_ne_m
		nee = total_ne_e - p_total_ne_e
	
		nea = total_ne_a - p_total_ne_a		
		neam = total_ne_am - p_total_ne_am		
					
		ners = total_ner_s- p_total_ner_s
		
		nerm = total_ner_m - p_total_ner_m
		nere = total_ner_e - p_total_ner_e
	
		nera = total_ner_a - p_total_ner_a		
		neram = total_ner_am - p_total_ner_am		
					

		
		data += '<tr>'
		data += '<td style="line-height:0.8;border-bottom:hidden;border-color:#000000;padding:1px;font-size:11px;background-color:#A7C7E7;"></td>'
		data += '<td style="line-height:0.8;border-color:#000000;width:15%;padding:1px;font-size:11px;background-color:#A7C7E7;"><center><b>NE</b><center></td>'

		if lp_total_ne_s == "L":
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>' %(lp_total_ne_s or 0)
			p_total_ne_s = 0
		else:
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>' %(p_total_ne_s or 0)

		if lp_total_ne_m == "L":
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>' %(lp_total_ne_m or 0)
			p_total_ne_m = 0
		else:
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>' %(p_total_ne_m or 0)
		
		if lp_total_ne_e == "L":
			p_total_ne_e = 0
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>' %(lp_total_ne_e or 0)
		else:
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>' %(p_total_ne_e or 0)
		
		if lp_total_ne_a == "L":
			p_total_ne_a = 0
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;"><center><b>%s</b><center></td>'%(lp_total_ne_a or 0)
		else:
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;"><center><b>%s</b><center></td>'%(p_total_ne_a or 0)
		
		if lp_total_ne_am == "L":
			p_total_ne_am = 0
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;"><center><b>%s</b><center></td>'%(lp_total_ne_am or 0)
		else:
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;"><center><b>%s</b><center></td>'%(p_total_ne_am or 0)
	

		data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>'%(p_total_ne_s + p_total_ne_m + p_total_ne_e + p_total_ne_a + p_total_ne_am)
		data += '</tr>'

		data += '<tr>'
		data += '<td style="line-height:0.8;border-bottom:hidden;border-color:#000000;width:10%;padding:1px;font-size:11px;background-color:#A7C7E7;"><center><b>Pending Work</b><center></td>'
		data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;background-color:#A7C7E7;"><center><b>NER</b><center></td>'

		if lp_total_ne_s == "L":
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>' %(lp_total_ne_s or 0)
			p_total_ner_s = 0
		else:
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>' %(p_total_ner_s or 0)

		if lp_total_ne_m == "L":
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>' %(lp_total_ne_m or 0)
			p_total_ner_m = 0
		else:
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>' %(p_total_ner_m or 0)
		
		if lp_total_ne_e == "L":
			p_total_ner_e = 0
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>' %(lp_total_ne_e or 0)
		else:
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>' %(p_total_ner_e or 0)
		
		if lp_total_ne_a == "L":
			p_total_ner_a = 0
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>'%(lp_total_ne_a or 0)
		else:
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>'%(p_total_ner_a or 0)

		if lp_total_ne_am == "L":
			p_total_ner_am = 0
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>'%(lp_total_ne_am or 0)
		else:
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>'%(p_total_ner_am or 0)
	
		data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>' %(p_total_ner_s + p_total_ner_m + p_total_ner_e + p_total_ner_a + p_total_ner_am)
		data += '</tr>'

				
		tr_s = frappe.db.count("Work Order Data",{"technician":"sampath@tsl-me.com","status":"TR-Technician Repair"})
		tr_m = frappe.db.count("Work Order Data",{"technician":"maari@tsl-me.com","status":"TR-Technician Repair"})
		tr_e = frappe.db.count("Work Order Data",{"technician":"eduardo@tsl-me.com","status":"TR-Technician Repair"})
		tr_a = frappe.db.count("Work Order Data",{"technician":"aakib@tsl-me.com","status":"TR-Technician Repair"})
		tr_am = frappe.db.count("Work Order Data",{"technician":"amir@tsl-me.com","status":"TR-Technician Repair"})

		data += '<tr>'
		data += '<td style="line-height:0.8;border-bottom:hidden;border-color:#000000;width:10%;padding:1px;font-size:11px;background-color:#A7C7E7;"><center><b></b><center></td>'
		data += '<td style="line-height:0.8;border-color:#000000;width:15%;padding:1px;font-size:11px;font-size:11px;background-color:#A7C7E7;"><center><b>TR</b><center></td>'
		
		if lp_total_ne_s == "L":
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>' %(lp_total_ne_s or 0)
			tr_s = 0
		else:
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>' %(tr_s or 0)

		if lp_total_ne_m == "L":
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>' %(lp_total_ne_m or 0)
			tr_m = 0
		else:
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>' %(tr_m or 0)
		
		if lp_total_ne_e == "L":
			tr_e = 0
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>' %(lp_total_ne_e or 0)
		else:
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>' %(tr_e or 0)
		
		if lp_total_ne_a == "L":
			tr_a = 0
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>'%(lp_total_ne_a or 0)
		else:
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>'%(tr_a or 0)


		if lp_total_ne_am == "L":
			tr_am = 0
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>'%(lp_total_ne_am or 0)
		else:
			data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>'%(tr_am or 0)


		data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;"><center><b>%s</b><center></td>' %(tr_s + tr_m + tr_e + tr_a + tr_am)
		data += '</tr>'

		# t_p_s = nes + ners + tr_s
		# t_p_m = nem + nerm + tr_m
		# t_p_e = nee + nere + tr_e
		# t_p_a = nea + nera + tr_a

		data += '<tr>'
		data += '<td style="line-height:0.8;border-color:#000000;width:10%;padding:1px;font-size:11px;background-color:#A7C7E7;"><center><b></b><center></td>'
		data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;background-color:#A7C7E7;"><center><b>Total</b><center></td>'
		data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;background-color:#A7C7E7;"><center><b>%s</b><center></td>'  %(p_total_ne_s + p_total_ner_s + tr_s)
		data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;background-color:#A7C7E7;"><center><b>%s</b><center></td>'  %(p_total_ne_m + p_total_ner_m + tr_m)
		data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;background-color:#A7C7E7;"><center><b>%s</b><center></td>'  %(p_total_ne_e + p_total_ner_e + tr_e)
		data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;background-color:#A7C7E7;"><center><b>%s</b><center></td>'  %(p_total_ne_a + p_total_ner_a + tr_a)
		data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;background-color:#A7C7E7;"><center><b>%s</b><center></td>'  %(p_total_ne_am + p_total_ner_am + tr_am)
		
		data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;font-size:11px;background-color:#A7C7E7;"><center><b>%s</b><center></td>'  %((p_total_ne_s + p_total_ne_m + p_total_ne_e + p_total_ne_a + p_total_ne_am)+(p_total_ner_s + p_total_ner_m + p_total_ner_e + p_total_ner_a + p_total_ner_am)+(tr_s + tr_m + tr_e + tr_a + tr_am))
		data += '</tr>'

		data += '</table>'
		return data
	




@frappe.whitelist()
def daily_lab_report_uae(date,company):
	d = datetime.now().date()
	ogdate = datetime.strptime(str(d),"%Y-%m-%d")
	ogdate_2 = datetime.strptime(str(date),"%Y-%m-%d")

	# Format the date as a string in the desired format
	formatted_date = ogdate.strftime("%d-%m-%Y")
	formatted_date_2 = ogdate_2.strftime("%d-%m-%Y")
	
	# date = d.strftime("%Y-%m-%d")
	# original_date = datetime.strptime(str(d), "%Y-%m-%d")
	original_date = date


	# Format datetime object to new date string format
	# date = datetime.strptime(str(original_date), "%Y-%m-%d")
	# total_ne = 0

	data = ""
	data += '<table class="table table-bordered">'
	data += '<tr>'
	data += '<td colspan = "7" align = "center" style="border-color:#000000;padding:1px;font-size:15px;background-color:#808080;color:white;"><b>Date :%s</b></td>' %(formatted_date_2)
	data += '</tr>'
	data += '<tr>'
	data += '<td style="border-color:#000000;width:30%"><img src = "/files/TSL Logo.png" align="left" width ="150"></td>'
	data += '<td style="border-color:#000000;width:35%;color:#055c9d;font-weight:bold;text-align:center;font-size:14px;">TSL Company<br>Branch - UAE</td>'
	data += '<td style="border-color:#000000;width:30%""><center><img src = "/files/Flag_of_the_United_Arab_Emirates.svg.jpg" width ="100"></center></td>'
	tech = frappe.get_all("Employee",{"company":company,"status": "Active","designation":["in",["Technician","Senior Technician"]]},["first_name"])
	
	data += '</tr>'
	data += '</table">'


	data += '<table class="table table-bordered">'
	data += '<tr>'
	# data += '<td style="border-color:#000000;width:15%;padding:1px;font-size:15px;font-size:15px;"><center><b></b><center></td>'
	data += '<td style="width:25%;border-color:#000000;padding:1px;font-size:18px;font-size:12px;background-color:#4682B4;color:white;"><center><b>Status</b><center></td>'
	technician = frappe.get_all("Employee",{"company":company,"status": "Active","designation":["in",["Technician","Senior Technician"]]},["user_id"])
	for i in technician:
		# data += '<td style="border-color:#000000;width:15%;padding:1px;font-size:18px;font-size:12px;line-height:0.8;background-color:#FFA500;;color:white;"><center><b>Rajesh</b><center></td>'
		# data += '<td style="border-color:#000000;width:15%;padding:1px;font-size:18px;font-size:12px;line-height:0.8;background-color:#FFC000;color:white;"><center><b>Glyen</b><center></td>'
		# data += '<td style="border-color:#000000;width:15%;padding:1px;font-size:18px;font-size:12px;line-height:0.8;background-color:#008000;color:white;"><center><b>Marwin</b><center></td>'
		if i.user_id == "ahmedmaster75@gmail.com":
			data += '<td colspan="1" style="border-color:#000000;padding:1px;font-size:18px;font-size:12px;line-height:0.8;background-color:#0047AB;color:white;"><center><b>Ahmed</b><center></td>'
		if i.user_id == "03glyen06mariano83@gmail.com":
			data += '<td colspan="1" style="border-color:#000000;padding:1px;font-size:18px;font-size:12px;line-height:0.8;background-color:#FFC000;color:white;"><center><b>Glyen</b><center></td>'
		if i.user_id == "rajesh-uae@tsl-me.com":
			data += '<td colspan="1" style="border-color:#000000;padding:1px;font-size:18px;font-size:12px;line-height:0.8;background-color:#FFA500;color:white;"><center><b>Rajesh</b><center></td>'
		if i.user_id == "marwin-uae@tsl-me.com":
			data += '<td colspan="1" style="border-color:#000000;padding:1px;font-size:18px;font-size:12px;line-height:0.8;background-color:#008000;color:white;"><center><b>Marwin</b><center></td>'
	
	data += '<td style="border-color:#000000;width:15%;padding:1px;font-size:18px;font-size:12px;line-height:0.8;background-color:#4682B4;color:white;"><center><b>Total</b><center></td>'
	data += '</tr>'
	data += '</table>'

	technician = frappe.get_all("Employee",{"company":company,"status": "Active","designation":["in",["Technician","Senior Technician"]]},["user_id"])
	data += '<table class="table table-bordered">'
	data += '<tr>'
	data += '<td style="line-height:0.8;border-bottom:hidden;border-color:#000000;width:10%;padding:1px;font-size:15px;font-size:15px;background-color:#A7C7E7;"><center><b></b><center></td>'
	data += '<td style="line-height:0.8;border-color:#000000;width:15%;padding:1px;font-size:15px;font-size:11px;background-color:#A7C7E7;"><center><b>UE</b><center></td>'
	total_ne = 0
	for t in technician:
		# emp = frappe.get_value("Employee",{"user_id":t})
		# if emp:
		# 	lp = frappe.db.sql(
		# 		"""
		# 		SELECT employee
		# 		FROM `tabLeave Application Form` 
		# 		WHERE employee = %s
		# 		AND %s BETWEEN from_date AND to_date and docstatus = 1
		# 		""",
		# 		(emp,date),
		# 		as_dict=True
		# 	)
		# 	if lp:
		# 		if t == "sampath@tsl-me.com":
		# 			lp_total_ne_s = "L"
		# 		if t == "maari@tsl-me.com":
		# 			lp_total_ne_m = "L"
		# 		if t == "eduardo@tsl-me.com":
		# 			lp_total_ne_e = "L"
		# 		if t == "aakib@tsl-me.com":
		# 			lp_total_ne_a = "L"
		
		ne = frappe.db.sql(""" select count(DISTINCT `tabWork Order Data`.name) as ct from `tabWork Order Data` 
				left join `tabStatus Duration Details` on `tabWork Order Data`.name = `tabStatus Duration Details`.parent
				where  `tabStatus Duration Details`.status = "UE-Under Evaluation"
				and `tabWork Order Data`.technician = "%s" and `tabStatus Duration Details`.date  LIKE "%s%%" """ %(t.user_id,original_date),as_dict=1)
		total_ne = total_ne + ne[0]["ct"]
		data += '<td style="border-color:#000000;padding:1px;font-size:15px;font-size:11px;line-height:0.8;"><center><b>%s</b><center></td>' %(ne[0]["ct"] or 0)

	data += '<td style="border-color:#000000;padding:1px;font-size:15px;font-size:11px;line-height:0.8;"><center><b>%s</b><center></td>' %(total_ne or 0)

		
	data += '</tr>'
	

	data += '<tr>'
	data += '<td style="border-bottom:hidden;border-color:#000000;width:10%;line-height:0.8;padding:1px;font-size:11px;font-size:11px;background-color:#A7C7E7;"><center><b>Started Work</b><center></td>'
	data += '<td style="border-color:#000000;width:15%;padding:1px;line-height:0.8;font-size:15px;font-size:11px;background-color:#A7C7E7;"><center><b>NER</b><center></td>'
	
	total_ner = 0
	for t in technician:
		ner = frappe.db.count("Work Order Data",{"technician":t.user_id,"status":"NER-Need Evaluation Return","posting_date": ["BETWEEN", ["2016-01-01",date]]})
		data += '<td style="border-color:#000000;padding:1px;font-size:15px;font-size:11px;line-height:0.8;"><center><b>%s</b><center></td>' %(ner or 0)

	total_ner = total_ner + ner

	data += '<td style="border-color:#000000;padding:1px;font-size:15px;font-size:11px;line-height:0.8;"><center><b>%s</b><center></td>' %(total_ner or 0)

	data += '<tr>'
	data += '<td style="border-color:#000000;width:10%;padding:1px;font-size:15px;font-size:15px;line-height:0.8;background-color:#A7C7E7;"><center><b></b><center></td>'
	data += '<td style="border-color:#000000;padding:1px;font-size:15px;font-size:11px;background-color:#A7C7E7;line-height:0.8;"><center><b>Total</b><center></td>'
	
	sum_ne_ner = 0

	for t in technician:
		ne = frappe.db.sql(""" select count(DISTINCT `tabWork Order Data`.name) as ct from `tabWork Order Data` 
		left join `tabStatus Duration Details` on `tabWork Order Data`.name = `tabStatus Duration Details`.parent
		where  `tabStatus Duration Details`.status = "UE-Under Evaluation"
		and `tabWork Order Data`.technician = "%s" and `tabStatus Duration Details`.date  LIKE "%s%%" """ %(t.user_id,original_date),as_dict=1)
		ner = frappe.db.count("Work Order Data",{"technician":t.user_id,"status":"NER-Need Evaluation Return","posting_date": ["BETWEEN", ["2016-01-01",date]]})
		
		sum_ne_ner = (sum_ne_ner + ner+ne[0]["ct"])
		data += '<td style="border-color:#000000;padding:1px;font-size:15px;line-height:0.8;font-size:11px;background-color:#A7C7E7;"><center><b>%s</b><center></td>' %(ner + ne[0]["ct"])
	
	data += '<td style="border-color:#000000;padding:1px;font-size:11px;line-height:0.8;background-color:#A7C7E7;"><center><b>%s</b><center></td>' %(sum_ne_ner or 0)

	data += '</tr>'

	data += '</table>'

	data += '<table class="table table-bordered">'
	data += '<tr>'
	data += '<td style="border-bottom:hidden;border-color:#000000;width:10%;line-height:0.8;padding:1px;font-size:11px;background-color:#A7C7E7;"><center><b></b><center></td>'
	data += '<td style="border-color:#000000;width:15%;padding:1px;line-height:0.8;font-size:11px;background-color:#A7C7E7;"><center><b>UE</b><center></td>' 
	
	total_ue = 0
	for t in technician:
		ue = frappe.db.sql(""" select count(DISTINCT `tabWork Order Data`.name) as ct from `tabWork Order Data` 
			left join `tabStatus Duration Details` on `tabWork Order Data`.name = `tabStatus Duration Details`.parent
			where  `tabStatus Duration Details`.status = "UE-Under Evaluation"
			and `tabWork Order Data`.technician = "%s" and `tabStatus Duration Details`.date  LIKE "%s%%" """ %(t.user_id,original_date) ,as_dict=1)
		
		data += '<td style="border-color:#000000;line-height:0.8;padding:1px;font-size:15px;font-size:11px;"><center><b>%s</b><center></td>' %(ue[0]["ct"] or 0)
		total_ue = total_ue + ue[0]["ct"]
	data += '<td style="border-color:#000000;padding:1px;line-height:0.8;font-size:15px;font-size:11px;"><center><b>%s</b><center></td>' %(total_ue or 0)

	data += '</tr>'
	data += '<tr>'
	data += '<td style="line-height:0.8;border-bottom:hidden;border-color:#000000;width:10%;padding:1px;font-size:15px;font-size:11px;background-color:#A7C7E7;"><center><b>End of Work</b><center></td>'
	data += '<td style="line-height:0.8;border-color:#000000;width:15%;padding:1px;font-size:15px;font-size:11px;background-color:#A7C7E7;"><center><b>UTR</b><center></td>'
	total_utr = 0 
	for t in technician:

		utr = frappe.db.sql(""" select count(DISTINCT `tabWork Order Data`.name) as ct from `tabWork Order Data` 
			left join `tabStatus Duration Details` on `tabWork Order Data`.name = `tabStatus Duration Details`.parent
			where  `tabStatus Duration Details`.status = "UTR-Under Technician Repair"
			and `tabWork Order Data`.technician = "%s" and `tabStatus Duration Details`.date  LIKE "%s%%" """ %(t.user_id,original_date),as_dict=1)
		
		data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:15px;font-size:11px;"><center><b>%s</b><center></td>' %(utr[0]["ct"] or 0)
		total_utr = total_utr + utr[0]["ct"]
	data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:15px;font-size:11px;"><center><b>%s</b><center></td>' %(total_utr or 0)

	data += '<tr>'

	data += '<tr>'
	data += '<td style="line-height:0.8;border-bottom:hidden;border-top:hidden;border-color:#000000;width:10%;padding:1px;font-size:11px;background-color:#A7C7E7;"><center><b></b><center></td>'
	data += '<td style="line-height:0.8;border-color:#000000;width:15%;padding:1px;font-size:15px;font-size:11px;background-color:#A7C7E7;"><center><b>Site Visit</b><center></td>'
	for t in technician:
		data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:15px;font-size:11px;"><center><b>%s</b><center></td>' %(0)
	data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:15px;font-size:11px;"><center><b>%s</b><center></td>' %(0)


	data += '<tr>'
	data += '<td style="line-height:0.8;border-color:#000000;width:10%;padding:1px;font-size:15px;font-size:11px;background-color:#A7C7E7;"><center><b></b><center></td>'
	data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:15px;font-size:11px;background-color:#A7C7E7;"><center><b>Total</b><center></td>'

	sum_ue_utr = 0
	for t in technician:
		ue = frappe.db.sql(""" select count(DISTINCT `tabWork Order Data`.name) as ct from `tabWork Order Data` 
			left join `tabStatus Duration Details` on `tabWork Order Data`.name = `tabStatus Duration Details`.parent
			where  `tabStatus Duration Details`.status = "UE-Under Evaluation"
			and `tabWork Order Data`.technician = "%s" and `tabStatus Duration Details`.date  LIKE "%s%%" """ %(t.user_id,original_date) ,as_dict=1)
		
	
		utr = frappe.db.sql(""" select count(DISTINCT `tabWork Order Data`.name) as ct from `tabWork Order Data` 
			left join `tabStatus Duration Details` on `tabWork Order Data`.name = `tabStatus Duration Details`.parent
			where  `tabStatus Duration Details`.status = "UTR-Under Technician Repair"
			and `tabWork Order Data`.technician = "%s" and `tabStatus Duration Details`.date  LIKE "%s%%" """ %(t.user_id,original_date),as_dict=1)
		sum_ue_utr = (sum_ue_utr + ue[0]["ct"]+utr[0]["ct"])
		data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:15px;font-size:11px;background-color:#A7C7E7;"><center><b>%s</b><center></td>' %(ue[0]["ct"] + utr[0]["ct"])
	
	data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:15px;font-size:11px;background-color:#A7C7E7;"><center><b>%s</b><center></td>' %(sum_ue_utr or 0)

	data += '</table>'

	
	data += '<table class="table table-bordered">'
	
	data += '<tr>'
	data += '<td colspan="2" style="border-left:none; border-top:none; border-right:none; border-color:#000000; padding:1px; font-size:11px; line-height:0.8;"><center><b></b></center></td>'
	data += '<td style="border-left:none; border-top:none; border-right:none; border-color:#000000; width:15%; padding:1px; font-size:11px; line-height:0.8;"><center><b></b></center></td>'
	data += '<td style="border-left:none; border-top:none; border-right:none; border-color:#000000; width:15%; padding:1px; font-size:11px; line-height:0.8;"><center><b></b></center></td>'
	data += '<td style="border-left:none; border-top:none; border-right:none; border-color:#000000; width:15%; padding:1px; font-size:11px; line-height:0.8;"><center><b></b></center></td>'
	data += '<td style="border-left:none; border-top:none; border-right:none; border-color:#000000; width:15%; padding:1px; font-size:11px; line-height:0.8;"><center><b></b></center></td>'
	data += '<td style="border-left:none; border-top:none; border-right:none; border-color:#000000; width:15%; padding:1px; font-size:11px; line-height:0.8;"><center><b></b></center></td>'
	data += '</tr>'
	
	#RS
	data += '<tr>'
	data += '<td style="line-height:0.8;border-bottom:hidden;border-color:#000000;width:10%;padding:1px;font-size:11px;background-color:#A7C7E7;"><center><b></b><center></td>'
	data += '<td style="line-height:0.8;border-color:#000000;width:15%;padding:1px;font-size:15px;font-size:11px;background-color:#A7C7E7;"><center><b>RS</b><center></td>'
	
	total_rs = 0
	for t in technician:
		rs = frappe.db.sql(""" select count(DISTINCT `tabWork Order Data`.name) as ct from `tabWork Order Data` 
			left join `tabStatus Duration Details` on `tabWork Order Data`.name = `tabStatus Duration Details`.parent
			where  `tabStatus Duration Details`.status = "RS-Repaired and Shipped"
			and `tabWork Order Data`.technician = "%s" and `tabStatus Duration Details`.date  LIKE "%s%%" """ %(t.user_id,original_date) ,as_dict=1)
		data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:15px;font-size:11px;"><center><b>%s</b><center></td>' %(rs[0]["ct"] or 0)
		total_rs = total_rs + rs[0]["ct"]
	data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:15px;font-size:11px;"><center><b>%s</b><center></td>' %(total_rs)
	#PS
	data += '<tr>'
	data += '<td style="line-height:0.8;border-bottom:hidden;border-color:#000000;width:10%;padding:1px;font-size:11px;background-color:#A7C7E7;"><center><b></b><center></td>'
	data += '<td style="line-height:0.8;border-color:#000000;width:15%;padding:1px;font-size:15px;font-size:11px;background-color:#A7C7E7;"><center><b>PS</b><center></td>'
	
	total_ps = 0
	for t in technician:
		ps= frappe.db.count("Evaluation Report",{"date":original_date,"technician":t.user_id})

		data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:15px;font-size:11px;"><center><b>%s</b><center></td>' %(ps or 0)
		total_ps = total_ps + ps
	data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:15px;font-size:11px;"><center><b>%s</b><center></td>' %(total_ps)

	#Working
	data += '<tr>'
	data += '<td style="line-height:0.8;border-bottom:hidden;border-color:#000000;width:10%;padding:1px;font-size:11px;background-color:#A7C7E7;"><center><b></b><center></td>'
	data += '<td style="line-height:0.8;border-color:#000000;width:15%;padding:1px;font-size:15px;font-size:11px;background-color:#A7C7E7;"><center><b>W</b><center></td>'
	
	total_w = 0
	for t in technician:
		w = frappe.db.sql(""" select count(DISTINCT `tabWork Order Data`.name) as ct from `tabWork Order Data` 
			left join `tabStatus Duration Details` on `tabWork Order Data`.name = `tabStatus Duration Details`.parent
			where  `tabStatus Duration Details`.status = "W-Working" 
			and `tabWork Order Data`.technician = "%s" and `tabStatus Duration Details`.date  LIKE "%s%%" """ %(t.user_id,original_date) ,as_dict=1)
		
		rs = frappe.db.sql(""" select count(DISTINCT `tabWork Order Data`.name) as ct from `tabWork Order Data` 
			left join `tabStatus Duration Details` on `tabWork Order Data`.name = `tabStatus Duration Details`.parent
			where  `tabStatus Duration Details`.status = "RS-Repaired and Shipped"
			and `tabWork Order Data`.technician = "%s" and `tabStatus Duration Details`.date  LIKE "%s%%" """ %(t.user_id,original_date) ,as_dict=1)
		# if rs:
			# data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:15px;font-size:11px;"><center><b>%s</b><center></td>' %(w[0]["ct"] - rs[0]["ct"]  or 0)
		# else:
		data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:15px;font-size:11px;"><center><b>%s</b><center></td>' %(w[0]["ct"] or 0)
		total_w = total_w + w[0]["ct"]
	data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:15px;font-size:11px;"><center><b>%s</b><center></td>' %(total_w)

	
	#RNR
	data += '<tr>'
	data += '<td style="line-height:0.8;border-bottom:hidden;border-color:#000000;width:10%;padding:1px;font-size:11px;background-color:#A7C7E7;"><center><b>Out - Flow</b><center></td>'
	data += '<td style="line-height:0.8;border-color:#000000;width:15%;padding:1px;font-size:15px;font-size:11px;background-color:#A7C7E7;"><center><b>RNR</b><center></td>'
	
	total_rnr = 0
	for t in technician:
		rnr = frappe.db.sql(""" select count(DISTINCT `tabWork Order Data` .name) as ct from `tabWork Order Data` 
		left join `tabStatus Duration Details` on `tabWork Order Data`.name = `tabStatus Duration Details`.parent
		where  `tabStatus Duration Details`.status = "RNR-Return Not Repaired"
				and `tabWork Order Data`.technician = "%s" and `tabStatus Duration Details`.date  LIKE "%s%%" """ %(t.user_id,original_date) ,as_dict=1)

		data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:15px;font-size:11px;"><center><b>%s</b><center></td>' %(rnr[0]["ct"] or 0)
		total_rnr = total_rnr + rnr[0]["ct"]
	data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:15px;font-size:11px;"><center><b>%s</b><center></td>' %(total_rnr)

	#RNF
	data += '<tr>'
	data += '<td style="line-height:0.8;border-bottom:hidden;border-color:#000000;width:10%;padding:1px;font-size:11px;background-color:#A7C7E7;"><center><b></b><center></td>'
	data += '<td style="line-height:0.8;border-color:#000000;width:15%;padding:1px;font-size:15px;font-size:11px;background-color:#A7C7E7;"><center><b>RNF</b><center></td>'
	
	total_rnf = 0
	for t in technician:
		rnf = frappe.db.sql(""" select count(DISTINCT `tabWork Order Data` .name) as ct from `tabWork Order Data` 
			left join `tabStatus Duration Details` on `tabWork Order Data`.name = `tabStatus Duration Details`.parent
			where  `tabStatus Duration Details`.status = "RNF-Return No Fault"
			and `tabWork Order Data`.technician = "%s" and `tabStatus Duration Details`.date  LIKE "%s%%" """ %(t.user_id,original_date) ,as_dict=1)
		
	
		data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:15px;font-size:11px;"><center><b>%s</b><center></td>' %(rnf[0]["ct"] or 0)
		total_rnf = total_rnf + rnf[0]["ct"]
	data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:15px;font-size:11px;"><center><b>%s</b><center></td>' %(total_rnf)

	#Com
	data += '<tr>'
	data += '<td style="line-height:0.8;border-bottom:hidden;border-color:#000000;width:10%;padding:1px;font-size:11px;background-color:#A7C7E7;"><center><b></b><center></td>'
	data += '<td style="line-height:0.8;border-color:#000000;width:15%;padding:1px;font-size:15px;font-size:11px;background-color:#A7C7E7;"><center><b>COM</b><center></td>'
	
	total_com = 0
	for t in technician:
		com = frappe.db.sql(""" select count(DISTINCT `tabWork Order Data` .name) as ct from `tabWork Order Data` 
		left join `tabStatus Duration Details` on `tabWork Order Data`.name = `tabStatus Duration Details`.parent
		where  `tabStatus Duration Details`.status = "C-Comparison"
		and `tabWork Order Data`.technician = "%s" and `tabStatus Duration Details`.date  LIKE "%s%%" """ %(t,original_date) ,as_dict=1)
	
	
		data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:15px;font-size:11px;"><center><b>%s</b><center></td>' %(com[0]["ct"] or 0)
		total_com = total_com + com[0]["ct"]
	data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:15px;font-size:11px;"><center><b>%s</b><center></td>' %(total_com)

	data += '<tr>'
	data += '<td style="line-height:0.8;border-color:#000000;width:10%;padding:1px;font-size:15px;font-size:11px;background-color:#A7C7E7;"><center><b></b><center></td>'
	data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:15px;font-size:11px;background-color:#A7C7E7;"><center><b>Total</b><center></td>'

	total_outflow = 0
	for t in technician:
		rs = frappe.db.sql(""" select count(DISTINCT `tabWork Order Data`.name) as ct from `tabWork Order Data` 
			left join `tabStatus Duration Details` on `tabWork Order Data`.name = `tabStatus Duration Details`.parent
			where  `tabStatus Duration Details`.status = "RS-Repaired and Shipped"
			and `tabWork Order Data`.technician = "%s" and `tabStatus Duration Details`.date  LIKE "%s%%" """ %(t.user_id,original_date) ,as_dict=1)

		ps= frappe.db.count("Evaluation Report",{"date":original_date,"technician":t.user_id})

		w = frappe.db.sql(""" select count(DISTINCT `tabWork Order Data`.name) as ct from `tabWork Order Data` 
			left join `tabStatus Duration Details` on `tabWork Order Data`.name = `tabStatus Duration Details`.parent
			where  `tabStatus Duration Details`.status = "W-Working" 
			and `tabWork Order Data`.technician = "%s" and `tabStatus Duration Details`.date  LIKE "%s%%" """ %(t.user_id,original_date) ,as_dict=1)
		
		rnr = frappe.db.sql(""" select count(DISTINCT `tabWork Order Data` .name) as ct from `tabWork Order Data` 
		left join `tabStatus Duration Details` on `tabWork Order Data`.name = `tabStatus Duration Details`.parent
		where  `tabStatus Duration Details`.status = "RNR-Return Not Repaired"
				and `tabWork Order Data`.technician = "%s" and `tabStatus Duration Details`.date  LIKE "%s%%" """ %(t.user_id,original_date) ,as_dict=1)

		rnf = frappe.db.sql(""" select count(DISTINCT `tabWork Order Data` .name) as ct from `tabWork Order Data` 
			left join `tabStatus Duration Details` on `tabWork Order Data`.name = `tabStatus Duration Details`.parent
			where  `tabStatus Duration Details`.status = "RNF-Return No Fault"
			and `tabWork Order Data`.technician = "%s" and `tabStatus Duration Details`.date  LIKE "%s%%" """ %(t.user_id,original_date) ,as_dict=1)

		com = frappe.db.sql(""" select count(DISTINCT `tabWork Order Data` .name) as ct from `tabWork Order Data` 
		left join `tabStatus Duration Details` on `tabWork Order Data`.name = `tabStatus Duration Details`.parent
		where  `tabStatus Duration Details`.status = "C-Comparison"
		and `tabWork Order Data`.technician = "%s" and `tabStatus Duration Details`.date  LIKE "%s%%" """ %(t,original_date) ,as_dict=1)
	
		data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;background-color:#A7C7E7;"><center><b>%s</b><center></td>' %(rs[0]["ct"] + ps + w[0]["ct"] + rnr[0]["ct"] + rnf[0]["ct"] + com[0]["ct"])
		total_outflow =  total_outflow + (rs[0]["ct"] + ps + w[0]["ct"] + rnr[0]["ct"] + rnf[0]["ct"] + com[0]["ct"])
	data += '<td style="line-height:0.8;line-height:0.8;border-color:#000000;padding:1px;font-size:11px;background-color:#A7C7E7;"><center><b>%s</b><center></td>' %(total_outflow or 0)

	#Pending Works
	data += '<table class="table table-bordered">'

	data += '<tr>'
	data += '<td colspan="2" style="border-left:none; border-top:none; border-right:none; border-color:#000000; padding:1px; font-size:11px; line-height:0.8;"><center><b></b></center></td>'
	data += '<td style="border-left:none; border-top:none; border-right:none; border-color:#000000; width:15%; padding:1px; font-size:11px; line-height:0.8;"><center><b></b></center></td>'
	data += '<td style="border-left:none; border-top:none; border-right:none; border-color:#000000; width:15%; padding:1px; font-size:11px; line-height:0.8;"><center><b></b></center></td>'
	data += '<td style="border-left:none; border-top:none; border-right:none; border-color:#000000; width:15%; padding:1px; font-size:11px; line-height:0.8;"><center><b></b></center></td>'
	data += '<td style="border-left:none; border-top:none; border-right:none; border-color:#000000; width:15%; padding:1px; font-size:11px; line-height:0.8;"><center><b></b></center></td>'
	data += '<td style="border-left:none; border-top:none; border-right:none; border-color:#000000; width:15%; padding:1px; font-size:11px; line-height:0.8;"><center><b></b></center></td>'
	data += '</tr>'
	
	dat = add_days(today(), -1)
				
	
	data += '<tr>'
	data += '<td style="line-height:0.8;border-bottom:hidden;border-color:#000000;padding:1px;font-size:11px;background-color:#A7C7E7;"><center><b></b><center></td>'
	data += '<td style="line-height:0.8;border-color:#000000;width:15%;padding:1px;font-size:11px;background-color:#A7C7E7;"><center><b>NE</b><center></td>'
	
	total_pne = 0
	for t in technician:
		p_ne = frappe.db.count("Work Order Data",{"technician":t.user_id,"status":"NE-Need Evaluation","posting_date": ["BETWEEN", ["2016-01-01", dat]]})

		data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:15px;font-size:11px;"><center><b>%s</b><center></td>' %(p_ne or 0)
		total_pne = total_pne + p_ne
	data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:15px;font-size:11px;"><center><b>%s</b><center></td>' %(total_pne)
	data += '</tr>'

	data += '<tr>'
	data += '<td style="line-height:0.8;border-bottom:hidden;border-color:#000000;padding:1px;font-size:11px;background-color:#A7C7E7;"><center><b>Pending</b><center></td>'
	data += '<td style="line-height:0.8;border-color:#000000;width:15%;padding:1px;font-size:11px;background-color:#A7C7E7;"><center><b>NER</b><center></td>'

	total_pner = 0
	for t in technician:
		p_ner = frappe.db.count("Work Order Data",{"technician":t.user_id,"status":"NER-Need Evaluation Return","posting_date": ["BETWEEN", ["2016-01-01", dat]]})

		data += '<td style="border-color:#000000;padding:1px;font-size:15px;font-size:11px;"><center><b>%s</b><center></td>' %(p_ner or 0)
		total_pner = total_pner + p_ner
	
	data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:15px;font-size:11px;"><center><b>%s</b><center></td>' %(total_pner)
	data += '</tr>'

	data += '<tr>'
	data += '<td style="line-height:0.8;border-bottom:hidden;border-color:#000000;padding:1px;font-size:11px;background-color:#A7C7E7;"><center><b></b><center></td>'
	data += '<td style="line-height:0.8;border-color:#000000;width:15%;padding:1px;font-size:11px;background-color:#A7C7E7;"><center><b>TR</b><center></td>'
	
	total_ptr = 0
	for t in technician:
		p_tr = frappe.db.count("Work Order Data",{"technician":t.user_id,"status":"TR-Technician Repair","posting_date": ["BETWEEN", ["2016-01-01", dat]]})

		data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:15px;font-size:11px;"><center><b>%s</b><center></td>' %(p_tr or 0)
		total_ptr = total_ptr + p_tr
	
	data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:15px;font-size:11px;"><center><b>%s</b><center></td>' %(total_ptr)

		
	sum_pending_works = 0

	data += '<tr>'
	data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;background-color:#A7C7E7;"><center><b></b><center></td>'
	data += '<td style="line-height:0.8;border-color:#000000;width:15%;padding:1px;font-size:11px;background-color:#A7C7E7;"><center><b>Total</b><center></td>'
	
	for t in technician:
		p_ne = frappe.db.count("Work Order Data",{"technician":t.user_id,"status":"NE-Need Evaluation","posting_date": ["BETWEEN", ["2016-01-01", dat]]})

		p_ner = frappe.db.count("Work Order Data",{"technician":t.user_id,"status":"NER-Need Evaluation Return","posting_date": ["BETWEEN", ["2016-01-01", dat]]})

		p_tr = frappe.db.count("Work Order Data",{"technician":t.user_id,"status":"TR-Technician Repair","posting_date": ["BETWEEN", ["2016-01-01", dat]]})
		sum_pending_works =  sum_pending_works + (p_ne + p_ner + p_tr)
		data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;background-color:#A7C7E7;"><center><b>%s</b><center></td>'  %(p_ne + p_ner + p_tr)
	data += '<td style="line-height:0.8;border-color:#000000;padding:1px;font-size:11px;background-color:#A7C7E7;"><center><b>%s</b><center></td>'  %(sum_pending_works)
	data += '</tr>'


	data += '</table>'

	return data



@frappe.whitelist()
def weekly_lab_report(company):
	# d = datetime.now().date()
	# ogdate = datetime.strptime(str(d),"%Y-%m-%d")

	# # Format the date as a string in the desired format
	# formatted_date = ogdate.strftime("%d-%m-%Y")
	
	# from_date = add_days(d,-7)
	# br = ""
	# if company == "TSL COMPANY - Kuwait":
	#     br = "Kuwait"
	# if company == "TSL COMPANY - UAE":
	#     br = "UAE"
		

	# data= ""
	# data += '<div class="table-container">'
	# data += '<table class="table table-bordered">'
	# data += '<tr>'
	# data += '<td colspan = 1 style="border-color:#000000;"><img src = "/files/TSL Logo.png" align="left" width ="150"></td>'
	# data += '<td colspan = 2 style="border-color:#000000;"><h2><center><b>TSL Company <br> Branch - %s</b></center></h2></td>' %(br)
	# data += '<td colspan = 1 style="border-color:#000000;"><center><img src = "/files/kuwait flag.jpg" width ="100"></center></td>'
	
	# data += '<tr>'
	# data += '<td colspan = 4 style="border-color:#000000;padding:1px;font-size:20px;background-color:#808080;color:white;"><b>%s</b></td>' %(formatted_date)
	# data += '</tr>'

	# data += '<tr>'
	# data += '<td style="border-color:#000000;padding:1px;font-size:16px;background-color:#3333ff;color:white;width:25%;"><center><b>Status</b><center></td>'
	# data += '<td style="border-color:#000000;padding:1px;font-size:16px;background-color:#3333ff;color:white;width:25%;"><center><b>WOD Count</b><center></td>'
	# data += '<td style="border-color:#000000;padding:1px;font-size:16px;background-color:#3333ff;color:white;width:25%;"><center><b>More than a Week</b><center></td>'
	# data += '<td style="border-color:#000000;padding:1px;font-size:16px;background-color:#3333ff;color:white;width:25%;"><center><b>Remarks</b><center></td>'

	# data += '</tr>'

	# ne_1 =  frappe.db.count("Work Order Data",{"status":"NE-Need Evaluation","company":company,"old_wo_no":["is","not set"]})
	# ner_1 =  frappe.db.count("Work Order Data",{"status":"NER-Need Evaluation Return","company":company,"old_wo_no":["is","not set"]})
	# ue_1 =  frappe.db.count("Work Order Data",{"status":"UE-Under Evaluation","company":company,"old_wo_no":["is","not set"]})
	# utr_1 =  frappe.db.count("Work Order Data",{"status":"UTR-Under Technician Repair","company":company,"old_wo_no":["is","not set"]})
	# tr_1 =  frappe.db.count("Work Order Data",{"status":"TR-Technician Repair","company":company,"old_wo_no":["is","not set"]})
	# sp_1 =  frappe.db.count("Work Order Data",{"status":"SP-Searching Parts","company":company,"old_wo_no":["is","not set"]})
	# wp_1 =  frappe.db.count("Work Order Data",{"status":"WP-Waiting Parts","company":company,"old_wo_no":["is","not set"]})

	# ne = frappe.db.sql(""" select count(`tabWork Order Data` .name) as wd from `tabWork Order Data` 
	# where  `tabWork Order Data`.posting_date < '%s' and `tabWork Order Data`.status = "NE-Need Evaluation" and `tabWork Order Data`.company = '%s' and`tabWork Order Data`.old_wo_no IS NULL """ %(from_date,company) ,as_dict=1)

	# ner = frappe.db.sql(""" select count(`tabWork Order Data` .name) as wd from `tabWork Order Data` 
	# where  `tabWork Order Data`.posting_date < '%s' and `tabWork Order Data`.status = "NER-Need Evaluation Return" and `tabWork Order Data`.company = '%s' and`tabWork Order Data`.old_wo_no IS NULL """ %(from_date,company) ,as_dict=1)
	
	# ue = frappe.db.sql(""" select count(`tabWork Order Data` .name) as wd from `tabWork Order Data` 
	# where  `tabWork Order Data`.posting_date < '%s' and `tabWork Order Data`.status = "UE-Under Evaluation" and `tabWork Order Data`.company = '%s' and`tabWork Order Data`.old_wo_no IS NULL """ %(from_date,company) ,as_dict=1)
	
	# utr = frappe.db.sql(""" select count(`tabWork Order Data` .name) as wd from `tabWork Order Data` 
	# where  `tabWork Order Data`.posting_date < '%s' and `tabWork Order Data`.status = "UTR-Under Technician Repair" and `tabWork Order Data`.company = '%s' and`tabWork Order Data`.old_wo_no IS NULL """ %(from_date,company) ,as_dict=1)
	
	# tr = frappe.db.sql(""" select count(`tabWork Order Data` .name) as wd from `tabWork Order Data` 
	# where  `tabWork Order Data`.posting_date < '%s' and `tabWork Order Data`.status = "TR-Technician Repair" and `tabWork Order Data`.company = '%s' and`tabWork Order Data`.old_wo_no IS NULL """ %(from_date,company) ,as_dict=1)
	
	# sp = frappe.db.sql(""" select count(`tabWork Order Data` .name) as wd from `tabWork Order Data` 
	# where  `tabWork Order Data`.posting_date < '%s' and `tabWork Order Data`.status = "SP-Searching Parts"and `tabWork Order Data`.company = '%s' and`tabWork Order Data`.old_wo_no IS NULL  """ %(from_date,company) ,as_dict=1)
	
	# wp = frappe.db.sql(""" select count(`tabWork Order Data` .name) as wd from `tabWork Order Data` 
	# where  `tabWork Order Data`.posting_date < '%s' and `tabWork Order Data`.status = "WP-Waiting Parts" and `tabWork Order Data`.company = '%s' and`tabWork Order Data`.old_wo_no IS NULL """ %(from_date,company) ,as_dict=1)
	
	# # frappe.errprint(ner[0]["wd"])

	# data += '<tr>'
	# data += '<td style="border-color:#000000;padding:1px;font-size:16px;width:25%;"><center><b>NE</b><center></td>'
	# data += '<td style="border-color:#000000;padding:1px;font-size:16px;"><center><b>%s</b><center></td>' %(ne_1)
	# data += '<td style="border-color:#000000;padding:1px;font-size:16px;"><center><b>%s</b><center></td>' %(ne[0]["wd"])
	# data += '<td style="border-color:#000000;padding:1px;font-size:16px;width:25%;"><center><b></b><center></td>'
	# data += '</tr>'

	
	# data += '<tr>'
	# data += '<td style="border-color:#000000;padding:1px;font-size:16px;width:25%;"><center><b>NER</b><center></td>'
	# data += '<td style="border-color:#000000;padding:1px;font-size:16px;"><center><b>%s</b><center></td>' %(ner_1)
	# data += '<td style="border-color:#000000;padding:1px;font-size:16px;"><center><b>%s</b><center></td>' %(ner[0]["wd"])
	# data += '<td style="border-color:#000000;padding:1px;font-size:16px;width:25%;"><center><b></b><center></td>'
	# data += '</tr>'
	
	# data += '<tr>'
	# data += '<td style="border-color:#000000;padding:1px;font-size:16px;width:25%;"><center><b>UE</b><center></td>'
	# data += '<td style="border-color:#000000;padding:1px;font-size:16px;"><center><b>%s</b><center></td>' %(ue_1)
	# data += '<td style="border-color:#000000;padding:1px;font-size:16px;"><center><b>%s</b><center></td>' %(ue[0]["wd"])
	# data += '<td style="border-color:#000000;padding:1px;font-size:16px;width:25%;"><center><b></b><center></td>'

	# data += '</tr>'
		
	# data += '<tr>'
	# data += '<td style="border-color:#000000;padding:1px;font-size:16px;width:25%;"><center><b>UTR</b><center></td>'
	# data += '<td style="border-color:#000000;padding:1px;font-size:16px;"><center><b>%s</b><center></td>' %(utr_1)
	# data += '<td style="border-color:#000000;padding:1px;font-size:16px;"><center><b>%s</b><center></td>' %(utr[0]["wd"])
	# data += '<td style="border-color:#000000;padding:1px;font-size:16px;width:25%;"><center><b></b><center></td>'

	# data += '</tr>'

	# data += '<tr>'
	# data += '<td style="border-color:#000000;padding:1px;font-size:16px;width:25%;"><center><b>TR</b><center></td>'
	# data += '<td style="border-color:#000000;padding:1px;font-size:16px;"><center><b>%s</b><center></td>' %(tr_1)
	# data += '<td style="border-color:#000000;padding:1px;font-size:16px;"><center><b>%s</b><center></td>' %(tr[0]["wd"])
	# data += '<td style="border-color:#000000;padding:1px;font-size:16px;"><center><b></b><center></td>'

	# data += '</tr>'
		
	# data += '<tr>'
	# data += '<td style="border-color:#000000;padding:1px;font-size:16px;width:25%;"><center><b>SP</b><center></td>'
	# data += '<td style="border-color:#000000;padding:1px;font-size:16px;"><center><b>%s</b><center></td>' %(sp_1)
	# data += '<td style="border-color:#000000;padding:1px;font-size:16px;"><center><b>%s</b><center></td>' %(sp[0]["wd"])
	# data += '<td style="border-color:#000000;padding:1px;font-size:16px;width:25%;"><center><b></b><center></td>'

	# data += '</tr>'
	
	# data += '<tr>'
	# data += '<td style="border-color:#000000;padding:1px;font-size:16px;width:25%;"><center><b>WP</b><center></td>'
	# data += '<td style="border-color:#000000;padding:1px;font-size:16px;"><center><b>%s</b><center></td>' %(wp_1)
	# data += '<td style="border-color:#000000;padding:1px;font-size:16px;"><center><b>%s</b><center></td>' %(wp[0]["wd"])
	# data += '<td style="border-color:#000000;padding:1px;font-size:16px;width:25%;"><center><b></b><center></td>'

	# data += '</tr>'
	
	# data += '<tr>'
	# data += '<td style="border-color:#000000;padding:1px;font-size:16px;width:25%;background-color:#3333ff;color:white;"><center><b>Total</b><center></td>'
	# data += '<td style="border-color:#000000;padding:1px;font-size:16px;background-color:#3333ff;color:white;"><center><b>%s</b><center></td>' %(ne_1 + ner_1 + ue_1 + utr_1 + sp_1 + tr_1 + wp_1)
	# data += '<td style="border-color:#000000;padding:1px;font-size:16px;background-color:#3333ff;color:white;"><center><b>%s</b><center></td>' %(ne[0]["wd"] + ner[0]["wd"] + ue[0]["wd"] + utr[0]["wd"] + tr[0]["wd"] + sp[0]["wd"] + wp[0]["wd"])
	# data += '<td style="border-color:#000000;padding:1px;font-size:16px;background-color:#3333ff;color:white;width:25%;"><center><b></b><center></td>'

	# data += '</tr>'
	# data += '</table>'



	# data += '</div>'
	d = datetime.now().date()
	ogdate = datetime.strptime(str(d),"%Y-%m-%d")

	# Format the date as a string in the desired format
	formatted_date = ogdate.strftime("%d-%m-%Y")
	
	from_date = add_days(d,-7)
	br = ""
	if company == "TSL COMPANY - Kuwait":
		br = "Kuwait"
	if company == "TSL COMPANY - UAE":
		br = "UAE"
		

	data= ""
	data += '<div class="table-container">'
	data += '<table class="table table-bordered">'
	data += '<tr>'
	data += '<td colspan = 1 align = center style="border-color:#000000;"><img src = "/files/TSL Logo.png" align="left" width ="250"></td>'
	data += '<td colspan = 2 style="border-color:#000000;"><h2><center><b style="color:#055c9d;" >TSL Company <br> Branch - %s</b></center></h2></td>' %(br)
	
	if company == "TSL COMPANY - Kuwait":
		data += '<td colspan = 1 style="border-color:#000000;"><center><img src = "/files/kuwait flag.jpg" width ="140"></center></td>'
	if company == "TSL COMPANY - UAE":
		data += '<td colspan = 1 style="border-color:#000000;"><center><img src = "/files/Flag_of_the_United_Arab_Emirates.svg.jpg" width ="140"></center></td>'
		
	data += '<tr>'
	data += '<td colspan = 4 style="border-color:#000000;padding:1px;font-size:20px;background-color:#0e86d4;color:white;"><b style = "color:white;" >%s</b></td>' %(formatted_date)
	data += '</tr>'

	data += '<tr>'
	data += '<td style="border-color:#000000;padding:1px;font-size:16px;background-color:#0e86d4;color:white;width:25%;"><center><b style = "color:white;" >Status</b><center></td>'
	data += '<td style="border-color:#000000;padding:1px;font-size:16px;background-color:#0e86d4;color:white;width:25%;"><center><b style = "color:white;" >WOD Count</b><center></td>'
	data += '<td style="border-color:#000000;padding:1px;font-size:16px;background-color:#0e86d4;color:white;width:25%;"><center><b style = "color:white;" >More than a Week</b><center></td>'
	data += '<td style="border-color:#000000;padding:1px;font-size:16px;background-color:#0e86d4;color:white;width:25%;"><center><b style = "color:white;" >Remarks</b><center></td>'

	data += '</tr>'

	ne_1 =  frappe.db.count("Work Order Data",{"status":"NE-Need Evaluation","company":company,"old_wo_no":["is","not set"]})
	ner_1 =  frappe.db.count("Work Order Data",{"status":"NER-Need Evaluation Return","company":company,"old_wo_no":["is","not set"]})
	ue_1 =  frappe.db.count("Work Order Data",{"status":"UE-Under Evaluation","company":company,"old_wo_no":["is","not set"]})
	utr_1 =  frappe.db.count("Work Order Data",{"status":"UTR-Under Technician Repair","company":company,"old_wo_no":["is","not set"]})
	tr_1 =  frappe.db.count("Work Order Data",{"status":"TR-Technician Repair","company":company,"old_wo_no":["is","not set"]})
	sp_1 =  frappe.db.count("Work Order Data",{"status":"SP-Searching Parts","company":company,"old_wo_no":["is","not set"]})
	wp_1 =  frappe.db.count("Work Order Data",{"status":"WP-Waiting Parts","company":company,"old_wo_no":["is","not set"]})

	# ne = frappe.db.sql(""" select count(`tabWork Order Data` .name) as wd from `tabWork Order Data` 
	# where  `tabWork Order Data`.posting_date < '%s' and `tabWork Order Data`.status = "NE-Need Evaluation" and `tabWork Order Data`.company = '%s' and`tabWork Order Data`.old_wo_no IS NULL """ %(from_date,self.company) ,as_dict=1)

	ne = frappe.db.sql("""
	SELECT COUNT(DISTINCT `tabWork Order Data`.name) AS wd
	FROM `tabWork Order Data`
	LEFT JOIN `tabStatus Duration Details` ON `tabWork Order Data`.name = `tabStatus Duration Details`.parent
	WHERE `tabStatus Duration Details`.status = "NE-Need Evaluation"
	AND `tabWork Order Data`.status = "NE-Need Evaluation"
	AND `tabWork Order Data`.company = '%s' and`tabWork Order Data`.old_wo_no IS NULL
	AND DATE(`tabStatus Duration Details`.date) < '%s'
	""" % (company,from_date), as_dict=1)


	# ner = frappe.db.sql(""" select count(`tabWork Order Data` .name) as wd from `tabWork Order Data` 
	# where  `tabWork Order Data`.status_cap_date < '%s' and `tabWork Order Data`.status = "NER-Need Evaluation Return" and `tabWork Order Data`.company = '%s' and`tabWork Order Data`.old_wo_no IS NULL """ %(from_date,self.company) ,as_dict=1)
	
	ner = frappe.db.sql("""
	SELECT COUNT(DISTINCT `tabWork Order Data`.name) AS wd
	FROM `tabWork Order Data`
	LEFT JOIN `tabStatus Duration Details` ON `tabWork Order Data`.name = `tabStatus Duration Details`.parent
	WHERE `tabStatus Duration Details`.status = "NER-Need Evaluation Return"
	AND `tabWork Order Data`.status = "NER-Need Evaluation Return"
	AND `tabWork Order Data`.company = '%s' and`tabWork Order Data`.old_wo_no IS NULL
	and DATE(`tabStatus Duration Details`.date) BETWEEN '%s' AND '%s';
	""" % (company,from_date,d), as_dict=1)


	# ue = frappe.db.sql(""" select count(`tabWork Order Data` .name) as wd from `tabWork Order Data` 
	# where  `tabWork Order Data`.posting_date < '%s' and `tabWork Order Data`.status = "UE-Under Evaluation" and `tabWork Order Data`.company = '%s' and`tabWork Order Data`.old_wo_no IS NULL """ %(from_date,self.company) ,as_dict=1)
	
	ue = frappe.db.sql("""
	SELECT COUNT(DISTINCT `tabWork Order Data`.name) AS wd
	FROM `tabWork Order Data`
	LEFT JOIN `tabStatus Duration Details` ON `tabWork Order Data`.name = `tabStatus Duration Details`.parent
	WHERE `tabStatus Duration Details`.status = 'UE-Under Evaluation'
	AND `tabWork Order Data`.status = 'UE-Under Evaluation'
	AND `tabWork Order Data`.company = '%s' and`tabWork Order Data`.old_wo_no IS NULL
	AND DATE(`tabStatus Duration Details`.date) < '%s'
	""" % (company,from_date), as_dict=1)


	utr = frappe.db.sql("""
	SELECT COUNT(DISTINCT `tabWork Order Data`.name) AS wd
	FROM `tabWork Order Data`
	LEFT JOIN `tabStatus Duration Details` ON `tabWork Order Data`.name = `tabStatus Duration Details`.parent
	WHERE `tabStatus Duration Details`.status = "UTR-Under Technician Repair"
	AND `tabWork Order Data`.status = "UTR-Under Technician Repair"
	AND `tabWork Order Data`.company = '%s' and`tabWork Order Data`.old_wo_no IS NULL
	AND DATE(`tabStatus Duration Details`.date) < '%s' AND DATE(`tabStatus Duration Details`.date) > '%s'
	""" % (company,from_date,from_date), as_dict=1)

	# utr = frappe.db.sql(""" select count(`tabWork Order Data` .name) as wd from `tabWork Order Data` 
	# where  `tabWork Order Data`.posting_date < '%s' and `tabWork Order Data`.status = "UTR-Under Technician Repair" and `tabWork Order Data`.company = '%s' and`tabWork Order Data`.old_wo_no IS NULL """ %(from_date,self.company) ,as_dict=1)
	
	tr = frappe.db.sql("""
	SELECT COUNT(DISTINCT `tabWork Order Data`.name) AS wd
	FROM `tabWork Order Data`
	LEFT JOIN `tabStatus Duration Details` ON `tabWork Order Data`.name = `tabStatus Duration Details`.parent
	WHERE `tabStatus Duration Details`.status = "TR-Technician Repair"
	AND `tabWork Order Data`.status = "TR-Technician Repair"
	AND `tabWork Order Data`.company = '%s' and`tabWork Order Data`.old_wo_no IS NULL
	AND DATE(`tabStatus Duration Details`.date) < '%s'
	""" % (company,from_date), as_dict=1)

	# tr = frappe.db.sql(""" select count(`tabWork Order Data` .name) as wd from `tabWork Order Data` 
	# where  `tabWork Order Data`.posting_date < '%s' and `tabWork Order Data`.status = "TR-Technician Repair" and `tabWork Order Data`.company = '%s' and`tabWork Order Data`.old_wo_no IS NULL """ %(from_date,self.company) ,as_dict=1)
	
	sp = frappe.db.sql("""
	SELECT COUNT(DISTINCT `tabWork Order Data`.name) AS wd
	FROM `tabWork Order Data`
	LEFT JOIN `tabStatus Duration Details` ON `tabWork Order Data`.name = `tabStatus Duration Details`.parent
	WHERE `tabStatus Duration Details`.status = "SP-Searching Parts"
	AND `tabWork Order Data`.status = "SP-Searching Parts"
	AND `tabWork Order Data`.company = '%s' and`tabWork Order Data`.old_wo_no IS NULL
	AND DATE(`tabStatus Duration Details`.date) < '%s'
	""" % (company,from_date), as_dict=1)

	# sp = frappe.db.sql(""" select count(`tabWork Order Data` .name) as wd from `tabWork Order Data` 
	# where  `tabWork Order Data`.posting_date < '%s' and `tabWork Order Data`.status = "SP-Searching Parts" and `tabWork Order Data`.company = '%s' and`tabWork Order Data`.old_wo_no IS NULL  """ %(from_date,self.company) ,as_dict=1)
	
	# wp = frappe.db.sql(""" select count(`tabWork Order Data` .name) as wd from `tabWork Order Data` 
	# where  `tabWork Order Data`.posting_date < '%s' and `tabWork Order Data`.status = "WP-Waiting Parts" and `tabWork Order Data`.company = '%s' and`tabWork Order Data`.old_wo_no IS NULL """ %(from_date,self.company) ,as_dict=1)
	
	wp = frappe.db.sql("""
	SELECT COUNT(DISTINCT `tabWork Order Data`.name) AS wd
	FROM `tabWork Order Data`
	LEFT JOIN `tabStatus Duration Details` ON `tabWork Order Data`.name = `tabStatus Duration Details`.parent
	WHERE `tabStatus Duration Details`.status = "WP-Waiting Parts"
	AND `tabWork Order Data`.status = "WP-Waiting Parts"
	AND `tabWork Order Data`.company = '%s' and`tabWork Order Data`.old_wo_no IS NULL
	AND DATE(`tabStatus Duration Details`.date) < '%s'
	""" % (company,from_date), as_dict=1)

	# frappe.errprint(ner[0]["wd"])
	ner_3 = ner_1-ner[0]["wd"]

	data += '<tr>'
	data += '<td style="border-color:#000000;padding:1px;font-size:16px;width:25%;"><center><b>NE</b><center></td>'
	data += '<td style="border-color:#000000;padding:1px;font-size:16px;"><center><b>%s</b><center></td>' %(ne_1)
	data += '<td style="border-color:#000000;padding:1px;font-size:16px;"><center><b>%s</b><center></td>' %(ne[0]["wd"])
	data += '<td style="border-color:#000000;padding:1px;font-size:16px;width:25%;"><center><b></b><center></td>'
	data += '</tr>'

	
	data += '<tr>'
	data += '<td style="border-color:#000000;padding:1px;font-size:16px;width:25%;"><center><b>NER</b><center></td>'
	data += '<td style="border-color:#000000;padding:1px;font-size:16px;"><center><b>%s</b><center></td>' %(ner_1)
	data += '<td style="border-color:#000000;padding:1px;font-size:16px;"><center><b>%s</b><center></td>' %(ner_3)
	data += '<td style="border-color:#000000;padding:1px;font-size:16px;width:25%;"><center><b></b><center></td>'
	data += '</tr>'
	
	data += '<tr>'
	data += '<td style="border-color:#000000;padding:1px;font-size:16px;width:25%;"><center><b>UE</b><center></td>'
	data += '<td style="border-color:#000000;padding:1px;font-size:16px;"><center><b>%s</b><center></td>' %(ue_1)
	data += '<td style="border-color:#000000;padding:1px;font-size:16px;"><center><b>%s</b><center></td>' %(ue[0]["wd"])
	data += '<td style="border-color:#000000;padding:1px;font-size:16px;width:25%;"><center><b></b><center></td>'

	data += '</tr>'
		
	data += '<tr>'
	data += '<td style="border-color:#000000;padding:1px;font-size:16px;width:25%;"><center><b>UTR</b><center></td>'
	data += '<td style="border-color:#000000;padding:1px;font-size:16px;"><center><b>%s</b><center></td>' %(utr_1)
	data += '<td style="border-color:#000000;padding:1px;font-size:16px;"><center><b>%s</b><center></td>' %(utr[0]["wd"])
	data += '<td style="border-color:#000000;padding:1px;font-size:16px;width:25%;"><center><b></b><center></td>'

	data += '</tr>'

	data += '<tr>'
	data += '<td style="border-color:#000000;padding:1px;font-size:16px;width:25%;"><center><b>TR</b><center></td>'
	data += '<td style="border-color:#000000;padding:1px;font-size:16px;"><center><b>%s</b><center></td>' %(tr_1)
	data += '<td style="border-color:#000000;padding:1px;font-size:16px;"><center><b>%s</b><center></td>' %(tr[0]["wd"])
	data += '<td style="border-color:#000000;padding:1px;font-size:16px;"><center><b></b><center></td>'

	data += '</tr>'
		
	data += '<tr>'
	data += '<td style="border-color:#000000;padding:1px;font-size:16px;width:25%;"><center><b>SP</b><center></td>'
	data += '<td style="border-color:#000000;padding:1px;font-size:16px;"><center><b>%s</b><center></td>' %(sp_1)
	data += '<td style="border-color:#000000;padding:1px;font-size:16px;"><center><b>%s</b><center></td>' %(sp[0]["wd"])
	data += '<td style="border-color:#000000;padding:1px;font-size:16px;width:25%;"><center><b></b><center></td>'

	data += '</tr>'
	
	data += '<tr>'
	data += '<td style="border-color:#000000;padding:1px;font-size:16px;width:25%;"><center><b>WP</b><center></td>'
	data += '<td style="border-color:#000000;padding:1px;font-size:16px;"><center><b>%s</b><center></td>' %(wp_1)
	data += '<td style="border-color:#000000;padding:1px;font-size:16px;"><center><b>%s</b><center></td>' %(wp[0]["wd"])
	data += '<td style="border-color:#000000;padding:1px;font-size:16px;width:25%;"><center><b></b><center></td>'

	data += '</tr>'
	
	data += '<tr>'
	data += '<td style="border-color:#000000;padding:1px;font-size:16px;width:25%;background-color:#0e86d4;color:white;"><center><b style = "color:white;">Total</b><center></td>'
	data += '<td style="border-color:#000000;padding:1px;font-size:16px;background-color:#0e86d4;color:white;"><center><b style = "color:white;">%s</b><center></td>' %(ne_1 + ner_1 + ue_1 + utr_1 + sp_1 + tr_1 + wp_1)
	data += '<td style="border-color:#000000;padding:1px;font-size:16px;background-color:#0e86d4;color:white;"><center><b style = "color:white;">%s</b><center></td>' %(ne[0]["wd"] + ner[0]["wd"] + ue[0]["wd"] + utr[0]["wd"] + tr[0]["wd"] + sp[0]["wd"] + wp[0]["wd"])
	data += '<td style="border-color:#000000;padding:1px;font-size:16px;background-color:#0e86d4;color:white;width:25%;"><center><b></b><center></td>'

	data += '</tr>'
	data += '</table>'



	data += '</div>'
	return data


@frappe.whitelist()
def invoice_request(qu):

	wd = frappe.db.sql(""" select  `tabQuotation Item`.wod_no from `tabQuotation` 
			left join `tabQuotation Item` on `tabQuotation Item`.parent = `tabQuotation`.name
			where `tabQuotation`.name = '%s' """ %(qu),as_dict = 1)


	sd = frappe.db.sql(""" select  `tabQuotation Item`.supply_order_data from `tabQuotation` 
			left join `tabQuotation Item` on `tabQuotation Item`.parent = `tabQuotation`.name
			where `tabQuotation`.name = '%s' """ %(qu),as_dict = 1)
			
	
	return wd

@frappe.whitelist()
def invoice_request_2(qu):
	sd = frappe.db.sql(""" select  `tabQuotation Item`.supply_order_data from `tabQuotation` 
			left join `tabQuotation Item` on `tabQuotation Item`.parent = `tabQuotation`.name
			where `tabQuotation`.name = '%s' """ %(qu),as_dict = 1)
			
	
	return sd

@frappe.whitelist()
def get_item_details(md):
	it = frappe.db.exists("Item",{"model":md})
	if it:
		item = frappe.get_value("Item",it,["name","item_number","description","stock_uom"])
		# frappe.errprint(it)
		return item


@frappe.whitelist()
def update_eval():
	wo = frappe.db.sql(""" UPDATE `tabEvaluation Report` SET technician_id = 11 WHERE  technician = "marwin-uae@tsl-me.com" ; """)

@frappe.whitelist()
def fill_employee_details(self):
	self.set("employees", [])
	employees = self.get_emp_list()
	if not employees:
		error_msg = _(
			"No employees found for the mentioned criteria:<br>Company: {0}<br> Currency: {1}<br>Payroll Payable Account: {2}"
		).format(
			frappe.bold(self.company),
			frappe.bold(self.currency),
			frappe.bold(self.payroll_payable_account),
		)
		if self.branch:
			error_msg += "<br>" + _("Branch: {0}").format(frappe.bold(self.branch))
		if self.department:
			error_msg += "<br>" + _("Department: {0}").format(frappe.bold(self.department))
		if self.designation:
			error_msg += "<br>" + _("Designation: {0}").format(frappe.bold(self.designation))
		if self.start_date:
			error_msg += "<br>" + _("Start date: {0}").format(frappe.bold(self.start_date))
		if self.end_date:
			error_msg += "<br>" + _("End date: {0}").format(frappe.bold(self.end_date))
		frappe.throw(error_msg, title=_("No employees found"))

	for d in employees:
		start_date = datetime.strptime(str(self.start_date), '%Y-%m-%d')
		end_date = datetime.strptime(str(self.end_date), '%Y-%m-%d')
		leave_salary = frappe.db.sql(""" select name from `tabLeave Salary` 
							WHERE docstatus = 1 
							AND employee = '%s' 
							AND month(from_date) = '%s' 
							AND year(from_date) = '%s'
							ORDER BY creation DESC
							LIMIT 1
						"""%(d['employee'],start_date.month,start_date.year))
		fnf = frappe.db.get_value("Full and Final Settlement",{
			"employee":d["employee"],
			"last_day_of_work": ["between", (self.start_date, self.end_date)],
			"docstatus":1
		},['name'])

		# loan = frappe.db.sql("""
		# 	SELECT l.name 
		# 	FROM `tabLoan` l
		# 	JOIN `tabLoan Pause Details` lpd ON lpd.parent = l.name
		# 	WHERE l.docstatus = 1 
		# 	AND l.applicant = '%s' 
		# 	AND MONTH(lpd.pause_from) = '%s'
		# 	AND YEAR(lpd.pause_upto) = '%s'
		# 	ORDER BY l.creation DESC
		# 	LIMIT 1
		# """ % (d['employee'], start_date.month, start_date.year))
		if leave_salary or fnf:
			continue

		self.append("employees", d)

	self.number_of_employees = len(self.employees)
	return self.get_employees_with_unmarked_attendance()

@frappe.whitelist()
def check_ss():
	sup_quoted= frappe.get_all("Supply Order Data",{"sales_rep":"Maaz","status":"Quoted","posting_date": ["between", ("2015-01-01","2024-11-24")]})
	sp_qted =0
	sp_qted_2 =0
	for j in sup_quoted:
		sup_qamt_qted = frappe.db.sql(''' select `tabQuotation`.name as q_name,
			`tabQuotation`.default_discount_percentage as dis,
			`tabQuotation`.approval_date as a_date,
			`tabQuotation`.is_multiple_quotation as is_m,
			`tabQuotation`.after_discount_cost as adc,
			`tabQuotation Item`.unit_price as up,
			`tabQuotation Item`.qty as qty,
			`tabQuotation Item`.margin_amount as ma 
			from `tabQuotation` 
			left join `tabQuotation Item` on  `tabQuotation`.name = `tabQuotation Item`.parent
			where  `tabQuotation`.Workflow_state in ("Approved By Customer") and 
			`tabQuotation Item`.supply_order_data = '%s' and 
			`tabQuotation`.quotation_type in ("Customer Quotation - Supply","Revised Quotation - Supply") 
			and transaction_date between '%s' and '%s' ''' %(j.name,"2015-01-01","2024-11-24"),as_dict=1)

		if sup_qamt_qted:
			if sup_qamt_qted[0]["is_m"] == 1:
				for k in sup_qamt_qted:
					per = (k.up * k.dis)/100
					q_amt = (k.up - per) * k.qty
				
					# print(j.name)
					print(q_amt)
					sp_qted = sp_qted + q_amt

			else:
				q_amt = sup_qamt_qted[0]["adc"]
				# print(round(q_amt))
				# print(j.name)
				print(q_amt)
				sp_qted = sp_qted + q_amt
			# print(sp_qted)
		else:
			sup_qamt_qted_2 = frappe.db.sql(''' select `tabQuotation`.name as q_name,
			`tabQuotation`.default_discount_percentage as dis,
			`tabQuotation`.approval_date as a_date,
			`tabQuotation`.is_multiple_quotation as is_m,
			`tabQuotation`.after_discount_cost as adc,
			`tabQuotation`.Workflow_state,
			`tabQuotation Item`.qty as qty,
			`tabQuotation Item`.unit_price as up,
			`tabQuotation Item`.margin_amount as ma from `tabQuotation` 
			left join `tabQuotation Item` on  `tabQuotation`.name = `tabQuotation Item`.parent
			where `tabQuotation`.Workflow_state in ("Quoted to Customer") and
			`tabQuotation`.quotation_type in ("Customer Quotation - Supply","Revised Quotation - Supply")
			and `tabQuotation Item`.supply_order_data = '%s' 
			and transaction_date between '%s' and '%s' ''' %(j.name,"2015-01-01","2024-11-24") ,as_dict=1)

			if sup_qamt_qted_2:
				if sup_qamt_qted_2[0]["is_m"] == 1:
					for k in sup_qamt_qted_2:
						per = (k.up * k.dis)/100
						q_amt_2 = (k.up - per)  * k.qty
					
						# print(j.name)
						print(q_amt_2)
						sp_qted_2 = sp_qted_2 + q_amt_2

				else:
					q_amt_2 = sup_qamt_qted_2[0]["adc"]
					
					# print(j.name)
					print(q_amt_2)
					sp_qted_2 = sp_qted_2 + q_amt_2
				
	print(sp_qted_2)


@frappe.whitelist()
def check_si():
	wd = "WOD-K24-13440"
	ev = frappe.get_doc("Evaluation Report",{"work_order_data":wd})
	for i in ev.items:
		frappe.errprint(i.part)
	# d = datetime.now().date()
	# ogdate = datetime.strptime(str(d),"%Y-%m-%d")

	# # Format the date as a string in the desired format
	# formatted_date = ogdate.strftime("%d-%m-%Y")
	
	# from_date = add_days(d,-7)

	# ue = frappe.db.sql("""
	# SELECT DISTINCT `tabWork Order Data`.name AS wd
	# FROM `tabWork Order Data`
	# LEFT JOIN `tabStatus Duration Details` ON `tabWork Order Data`.name = `tabStatus Duration Details`.parent
	# WHERE `tabStatus Duration Details`.status = 'UTR-Under Technician Repair'
	# AND `tabWork Order Data`.status = 'UTR-Under Technician Repair'
	# AND `tabWork Order Data`.company = '%s' and`tabWork Order Data`.old_wo_no IS NULL
	# AND DATE(`tabStatus Duration Details`.date) < '%s'
	# """ % ("TSL COMPANY - Kuwait",from_date), as_dict=1)
	# # print(ue)
	# count = 0
	# for i in ue:
	#     s = frappe.db.sql("""
	#     SELECT `tabStatus Duration Details`.date as date
	#     FROM `tabWork Order Data`
	#     LEFT JOIN `tabStatus Duration Details` ON `tabWork Order Data`.name = `tabStatus Duration Details`.parent
	#     WHERE `tabWork Order Data`.name = '%s' and `tabStatus Duration Details`.status = 'UTR-Under Technician Repair' """ % (i.wd), as_dict=1)
	#     # print(d)
	#     # print(i.wd)
	#     for k in s:
	#         if k.date.date() < from_date and not k.date.date() < from_date:
	#             print(k.date.date())
	#         # else:
	#         #     count = count + 
	# print(count)

	# si = frappe.db.sql(""" select DISTINCT `tabSales Invoice Item`.supply_order_data as so,
	# `tabSales Invoice`.name as n
	# from `tabSales Invoice` 
	# left join `tabSales Invoice Item` on `tabSales Invoice`.name = `tabSales Invoice Item`.parent 
	# where `tabSales Invoice`.status = 'Paid' and `tabSales Invoice`.department = "Supply - TSL" """ ,as_dict =1)
	# for i in si:
	#     if i.so:
	#         print(i.so)
	#     c = frappe.get_value("Supply Order Data",i.so,["status"])
		
	#     if c in ["Invoiced","Delivered and Invoiced"]:
	#         print(i.so)
	#         frappe.db.set_value("Supply Order Data",i.so,"Status","Paid")
			# frappe.db.set_value("Supply Order Data",i.so,"Workflow_state","Paid")
   

@frappe.whitelist()
def get_q_amt_for_po(wo):
	q = frappe.db.sql(''' select `tabQuotation`.name as q_name,
	`tabQuotation`.wod_no as wod,                      
	`tabQuotation`.default_discount_percentage as dis,
	`tabQuotation`.approval_date as a_date,
	`tabQuotation`.is_multiple_quotation as is_m,
	`tabQuotation`.after_discount_cost as adc,
	`tabQuotation`.Workflow_state,
	`tabQuotation Item`.unit_price as up,
	`tabQuotation Item`.qty as qty,
	`tabQuotation Item`.item_code as ic,
	`tabQuotation Item`.margin_amount as mar,
	`tabQuotation Item`.margin_amount as ma from `tabQuotation` 
	left join `tabQuotation Item` on  `tabQuotation`.name = `tabQuotation Item`.parent
	where `tabQuotation`.Workflow_state in ("Approved By Customer") and
	`tabQuotation`.quotation_type in ("Customer Quotation - Repair","Revised Quotation - Repair")
	and `tabQuotation Item`.wod_no = '%s' ''' %(wo) ,as_dict=1)
	data = ''
	data += '<table class="table table-bordered">'
	data += '<tr>'
	data += '<td style="text-align:center;width:20%;font-size:14px;background-color:#3333ff;color:white;"><b>Work Order Data</b></td>'
	data += '<td style="text-align:center;width:20%;font-size:14px;background-color:#3333ff;color:white;"><b>Quotation</b></td>'
	data += '<td style="text-align:center;width:20%;font-size:14px;background-color:#3333ff;color:white;"><b>SKU</b></td>'
	data += '<td style="text-align:center;width:20%;font-size:14px;background-color:#3333ff;color:white;"><b>Quantity</b></td>'
	data += '<td style="text-align:center;width:20%;font-size:14px;background-color:#3333ff;color:white;"><b>Price</b></td>'
	data+= '</tr>'
	if q:
		for i in q:
			data += '<tr>'
			data += '<td style="text-align:center;font-size:12px;"><b>%s</b></td>' %(wo)
			data += '<td style="text-align:center;font-size:12px;"><b>%s</b></td>' %(i.q_name)
			data += '<td style="text-align:center;font-size:12px;"><b>%s</b></td>' %(i.ic)
			data += '<td style="text-align:center;font-size:12px;"><b>%s</b></td>' %(i.qty)
			if not i.is_m:
				data += '<td style="text-align:center;font-size:12px;"><b>%s</b></td>' %(i.adc)
			else:
				per = (i.up * i.dis)/100
				quot_amt = i.up - per
				data += '<td style="text-align:center;font-size:12px;"><b>%s</b></td>' %(round(quot_amt))
			data+= '</tr>'

	data += '</table>'
	return data
	
# override - Total Working Days in Salary Slip
def get_working_days_details(self, joining_date=None, relieving_date=None, lwp=None, for_preview=0):
	include_holidays_in_total_working_days = frappe.db.get_value("Company Wise Payroll Days",{"parent":"Payroll Settings","company":self.company}, "include_holidays")

	payroll_based_on = frappe.db.get_value("Payroll Settings", None, "payroll_based_on")
	
	working_days = frappe.db.get_value("Company Wise Payroll Days",{"company": self.company},"total_working_days")

	if not (joining_date and relieving_date):
		joining_date, relieving_date = self.get_joining_and_relieving_dates()

	if for_preview:
		self.total_working_days = working_days
		self.payment_days = working_days
		return

	holidays = self.get_holidays_for_employee(self.start_date, self.end_date)
	working_days_list = [add_days(getdate(self.start_date), days=day) for day in range(0, working_days)]

	if not cint(include_holidays_in_total_working_days):
		working_days_list = [i for i in working_days_list if i not in holidays]
		working_days = frappe.db.get_value("Company Wise Payroll Days",{"company": self.company},"total_working_days")

		if working_days < 0:
			frappe.throw(_("There are more holidays than working days this month."))

	if not payroll_based_on:
		frappe.throw(_("Please set Payroll based on in Payroll settings"))

	if payroll_based_on == "Attendance":
		actual_lwp, absent = self.calculate_lwp_ppl_and_absent_days_based_on_attendance(
			holidays, relieving_date
		)
		self.absent_days = absent
	else:
		actual_lwp = self.calculate_lwp_or_ppl_based_on_leave_application(
			holidays, working_days_list, relieving_date
		)

	if not lwp:
		lwp = actual_lwp
	elif lwp != actual_lwp:
		frappe.msgprint(
			_("Leave Without Pay does not match with approved {} records").format(payroll_based_on)
		)

	self.leave_without_pay = lwp
	self.total_working_days = working_days

	payment_days = self.get_payment_days(
		joining_date, relieving_date, include_holidays_in_total_working_days
	)
	if payment_days > working_days:
		payment_days = working_days

	if flt(payment_days) > flt(lwp):
		self.payment_days = flt(payment_days) - flt(lwp)

		if payroll_based_on == "Attendance":
			self.payment_days -= flt(absent)

		consider_unmarked_attendance_as = (
			frappe.db.get_value("Payroll Settings", None, "consider_unmarked_attendance_as") or "Present"
		)

		if payroll_based_on == "Attendance" and consider_unmarked_attendance_as == "Absent":
			unmarked_days = self.get_unmarked_days(include_holidays_in_total_working_days, holidays)
			self.absent_days += unmarked_days  # will be treated as absent
			self.payment_days -= unmarked_days
	else:
		self.payment_days = 0


@frappe.whitelist()
def wo_cu_list():
	three_months_ago = add_months(today(), -3)
	print (three_months_ago)
	msg1="""<div><style>.sh-src a{text-decoration:none!important;}</style></div> <br> <table cellpadding="0" cellspacing="0" border="0" class="sh-src" style="margin: 0px; border-collapse: collapse;"><tr><td style="padding: 0px 1px 0px 0px;"><table cellpadding="0" cellspacing="0" border="0" style="margin: 0px; border-collapse: collapse;"><tr><td align="center" style="padding: 0px 18px 0px 0px; vertical-align: top;"><table cellpadding="0" cellspacing="0" border="0" style="margin: 0px; border-collapse: collapse;"><tr><td style="padding: 0px 1px 13px 0px;"><p style="margin: 1px;"><img src="https://signaturehound.com/api/v1/file/4hvclm3gxkcjh" alt="" title="Profile Picture" width="100" height="100" class="" style="display: block; border: 0px; max-width: 100px;"></p></td></tr></table> <table cellpadding="0" cellspacing="0" border="0" style="margin: 0px; border-collapse: collapse;"><tr><td style="padding: 0px 1px 0px 0px;"><p style="margin: 1px;"><a href="https://tsl-me.com/" target="_blank"><img src="https://signaturehound.com/api/v1/file/3lwizllxn10tyt" alt="" title="Logo" width="150" height="50" style="display: block; border: 0px; max-width: 150px;"></a></p></td></tr></table></td> <td width="5" style="padding: 1px 0px 0px;"></td> <td style="padding: 0px 1px 0px 0px; vertical-align: top;"><table cellpadding="0" cellspacing="0" border="0" style="margin: 0px; border-collapse: collapse;"><tr><td style="padding: 0px 1px 10px 0px; border-bottom: 2px solid rgb(0,92,163); font-family: Arial, sans-serif; font-size: 13px; line-height: 15px; white-space: nowrap;"><p style="font-family: Arial, sans-serif; font-size: 13px; line-height: 15px; font-weight: 700; color: rgb(0,92,163); white-space: nowrap; margin: 1px;">Aswin Baby
					</p> <p style="font-family: Arial, sans-serif; font-size: 13px; line-height: 15px; white-space: nowrap; color: rgb(136,136,136); margin: 1px;">Customer Support Officer</p> <!----> <p style="font-family: Arial, sans-serif; font-size: 13px; line-height: 15px; white-space: nowrap; color: rgb(136,136,136); margin: 1px;">
					  TSL Group | Kuwait</p> <!----></td></tr> <tr><td style="padding: 10px 1px 10px 0px; border-bottom: 2px solid rgb(0,92,163);"><table cellpadding="0" cellspacing="0" border="0" style="margin: 0px; border-collapse: collapse;"><tr><td valign="middle" style="padding: 1px 5px 1px 0px; vertical-align: middle;"><p style="margin: 1px;"><img src="https://signaturehound.com/api/v1/png/email/round-outlined/0088cc.png" alt="" width="22" height="22" style="display: block; border: 0px; width: 22px; height: 22px;"></p></td> <td style="font-family: Arial, sans-serif; font-size: 13px; line-height: 15px; white-space: nowrap; color: rgb(136,136,135) !important; padding: 1px 0px; vertical-align: middle;"><p style="margin: 1px;"><a href="mailto:info@tsl-me.com" target="_blank" style="font-family: Arial, sans-serif; font-size: 13px; line-height: 15px; white-space: nowrap; color: rgb(136,136,136); text-decoration: none !important;"><span style="font-family: Arial, sans-serif; font-size: 13px; line-height: 15px; white-space: nowrap; color: rgb(136,136,136); text-decoration: none !important;">info@tsl-me.com</span></a></p></td></tr> <tr><td valign="middle" style="padding: 1px 5px 1px 0px; vertical-align: middle;"><p style="margin: 1px;"><img src="https://signaturehound.com/api/v1/png/direct/round-outlined/0088cc.png" alt="" width="22" height="22" style="display: block; border: 0px; width: 22px; height: 22px;"></p></td> <td style="font-family: Arial, sans-serif; font-size: 13px; line-height: 15px; white-space: nowrap; color: rgb(136,136,135) !important; padding: 1px 0px; vertical-align: middle;"><p style="margin: 1px;"><a href="tel:+96524741313" target="_blank" style="font-family: Arial, sans-serif; font-size: 13px; line-height: 15px; white-space: nowrap; color: rgb(136,136,136); text-decoration: none !important;"><span style="font-family: Arial, sans-serif; font-size: 13px; line-height: 15px; white-space: nowrap; color: rgb(136,136,136); text-decoration: none !important;">+965 24741313</span></a></p></td></tr><tr><td valign="middle" style="padding: 1px 5px 1px 0px; vertical-align: middle;"><p style="margin: 1px;"><img src="https://signaturehound.com/api/v1/png/fax/round-outlined/0088cc.png" alt="" width="22" height="22" style="display: block; border: 0px; width: 22px; height: 22px;"></p></td> <td style="font-family: Arial, sans-serif; font-size: 13px; line-height: 15px; white-space: nowrap; color: rgb(136,136,135) !important; padding: 1px 0px; vertical-align: middle;"><p style="margin: 1px;"><a href="tel:+96524741311" target="_blank" style="font-family: Arial, sans-serif; font-size: 13px; line-height: 15px; white-space: nowrap; color: rgb(136,136,136); text-decoration: none !important;"><span style="font-family: Arial, sans-serif; font-size: 13px; line-height: 15px; white-space: nowrap; color: rgb(136,136,136); text-decoration: none !important;">+965 24741311</span></a></p></td></tr> <tr><td valign="top" style="padding: 1px 5px 1px 0px; vertical-align: top;"><p style="margin: 1px;"><img src="https://signaturehound.com/api/v1/png/map/round-outlined/0088cc.png" alt="" width="22" height="22" style="display: block; border: 0px; width: 22px; height: 22px;"></p></td> <td style="font-family: Arial, sans-serif; font-size: 13px; line-height: 15px; white-space: nowrap; color: rgb(136,136,135) !important; padding: 1px 0px; vertical-align: middle;"><p style="margin: 1px;"><a href="https://www.google.com/maps/place/TSL+For+Industrial+Electronics+Repairing+%26+Supply+-+Kuwait/@29.3082683,47.9352691,18z/data=!4m6!3m5!1s0x3fcf9a9068440231:0x7321607de759fdc1!8m2!3d29.3082309!4d47.9365244!16s%2Fg%2F11c0rlsd_t?entry=ttu" target="_blank" style="font-family: Arial, sans-serif; font-size: 13px; line-height: 15px; white-space: nowrap; color: rgb(136,136,136); text-decoration: none !important;"><span style="font-family: Arial, sans-serif; font-size: 13px; line-height: 15px; white-space: nowrap; color: rgb(136,136,136); text-decoration: none !important;">Bldg: 1473, Unit: 13, Street: 24, Block: 1,<br>Al Rai Industrial Area, Kuwait</span></a></p></td></tr> <tr><td valign="middle" style="padding: 1px 5px 1px 0px; vertical-align: middle;"><p style="margin: 1px;"><img src="https://signaturehound.com/api/v1/png/website/round-outlined/0088cc.png" alt="" width="22" height="22" style="display: block; border: 0px; width: 22px; height: 22px;"></p></td> <td style="font-family: Arial, sans-serif; font-size: 13px; line-height: 15px; white-space: nowrap; color: rgb(0,92,162) !important; font-weight: 700; padding: 1px 0px; vertical-align: middle;"><p style="margin: 1px;"><a href="https://tsl-me.com/" target="_blank" style="font-family: Arial, sans-serif; font-size: 13px; line-height: 15px; white-space: nowrap; color: rgb(0,92,163); font-weight: 700; text-decoration: none !important;"><span style="font-family: Arial, sans-serif; font-size: 13px; line-height: 15px; white-space: nowrap; color: rgb(0,92,163); font-weight: 700; text-decoration: none !important;">tsl-me.com</span></a></p></td></tr></table></td></tr> <tr><td style="padding: 0px 1px 0px 0px;"><table cellpadding="0" cellspacing="0" border="0" style="margin: 0px; border-collapse: collapse;"><tr><td width="27" style="font-size: 0px; line-height: 0px; padding: 13px 1px 0px 0px;"><p style="margin: 1px;"><a href="https://www.linkedin.com/company/tsl-me/mycompany/" target="_blank"><img src="https://signaturehound.com/api/v1/png/linkedin/round/0088cc.png" alt="" width="27" height="27" style="display: block; border: 0px; width: 27px; height: 27px;"></a></p></td> <td width="3" style="padding: 0px 0px 1px;"></td><td width="27" style="font-size: 0px; line-height: 0px; padding: 13px 1px 0px 0px;"><p style="margin: 1px;"><a href="https://x.com/tsl_mecompany?s=11&amp;t=Zxza0-9Q_18nsDCddfTQPw" target="_blank"><img src="https://signaturehound.com/api/v1/png/x/round/0088cc.png" alt="" width="27" height="27" style="display: block; border: 0px; width: 27px; height: 27px;"></a></p></td> <td width="3" style="padding: 0px 0px 1px;"></td><td width="27" style="font-size: 0px; line-height: 0px; padding: 13px 1px 0px 0px;"><p style="margin: 1px;"><a href="https://www.instagram.com/tslcom/?igshid=MzRlODBiNWFlZA%3D%3D" target="_blank"><img src="https://signaturehound.com/api/v1/png/instagram/round/0088cc.png" alt="" width="27" height="27" style="display: block; border: 0px; width: 27px; height: 27px;"></a></p></td> <td width="3" style="padding: 0px 0px 1px;"></td><td width="27" style="font-size: 0px; line-height: 0px; padding: 13px 1px 0px 0px;"><p style="margin: 1px;"><a href="https://www.facebook.com/people/TSL-Industrial-Electronics-Services/61550277093129/" target="_blank"><img src="https://signaturehound.com/api/v1/png/facebook/round/0088cc.png" alt="" width="27" height="27" style="display: block; border: 0px; width: 27px; height: 27px;"></a></p></td> <td width="3" style="padding: 0px 0px 1px;"></td><td width="27" style="font-size: 0px; line-height: 0px; padding: 13px 1px 0px 0px;"><p style="margin: 1px;"><a href="https://www.youtube.com/@TSLELECTRONICSSERVICES" target="_blank"><img src="https://signaturehound.com/api/v1/png/youtube/round/0088cc.png" alt="" width="27" height="27" style="display: block; border: 0px; width: 27px; height: 27px;"></a></p></td> <td width="3" style="padding: 0px 0px 1px;"></td></tr></table></td></tr></table></td></tr></table></td></tr> <!----> <tr>
					  <td style="padding: 0px 1px 0px 0px;"><table cellpadding="0" cellspacing="0" border="0" style="max-width: 600px; margin: 0px; border-collapse: collapse;"><tr><td style="padding: 15px 1px 0px 0px; font-family: Arial, sans-serif; font-size: 10px; line-height: 12px; color: rgb(136,136,136);"><p style="font-family: Arial, sans-serif; font-size: 10px; line-height: 12px; color: rgb(136,136,136); margin: 1px;">The content of this email is confidential and intended for the recipient specified in message only. It is strictly forbidden to share any part of this message with any third party, without a written consent of the sender. If you received this message by mistake, please reply to this message and follow with its deletion, so that we can ensure such a mistake does not occur in the future.</p></td></tr></table></td></tr> <tr><td style="padding: 0px 1px 0px 0px;"><table cellpadding="0" cellspacing="0" border="0" style="margin: 0px; border-collapse: collapse;"><tr><td valign="middle" style="padding: 15px 4px 1px 0px; vertical-align: middle;"><p style="margin: 1px;"></p></td></tr></table></td></tr> <!----></table>"""
	
	msg = """Dear Team,<br><br>Kindly follow-up with these customers who where Inactive since 3 months<br><br>Thanks & Regards,<br><br>"""		
	
	inactive_customers = frappe.db.sql("""
		SELECT 
			wo.customer,
			wo.sales_rep,
			wo.company,
			COUNT(wo.name) AS total_work_orders
		FROM 
			`tabWork Order Data` wo
		
		WHERE 
			wo.docstatus = 1
			AND wo.sales_rep IS NOT NULL
			AND wo.customer NOT IN (
				SELECT DISTINCT customer
				FROM `tabWork Order Data`
				WHERE received_date < %s AND docstatus = 1
			)
		GROUP BY 
			wo.sales_rep, wo.customer
		ORDER BY 
			wo.sales_rep, wo.customer
	""", (three_months_ago), as_dict=True)
	# data= ""
	# data += '<table class="table table-bordered">'
	# data += '<tr>'
	# data += '<td style="padding:2px;font-size:14px;font-size:12px;background-color:#3da8de;color:white;"><center><b>Customers</b><center></td>'
	# data += '<td style="padding:2px;font-size:14px;font-size:12px;background-color:#3da8de;color:white;"><center><b>No of Units Received</b><center></td>'
	# data += '</tr>'
	# for ic in inactive_customers:
	# 	if ic.company == 'TSL COMPANY - Kuwait' and ic.sales_rep =="Maaz":
		
	# 		data +='<tr>'
	# 		data += '<td style="text-align:left" ><b>%s</b></td>'%(ic.customer)
	# 		data += '<td style="text-align:left"><b>%s</b></td>'%(ic.total_work_orders)
	# 		data += '</tr>'	
	
				
	# data += '</table>'
	# frappe.sendmail(
	# 	recipients= ['maaz@tsl-me.com'],
	# 	cc=['omar@tsl-me.com','yousuf@tsl-me.com'],
	# 	subject="Inactive Customers for Past 2 months",
	# 	message = msg+msg1+data)

	# data= ""
	# data += '<table class="table table-bordered">'
	# data += '<tr>'
	# data += '<td style="padding:2px;font-size:14px;font-size:12px;background-color:#3da8de;color:white;"><center><b>Customers</b><center></td>'
	# data += '<td style="padding:2px;font-size:14px;font-size:12px;background-color:#3da8de;color:white;"><center><b>No of Units Received</b><center></td>'
	# data += '</tr>'
	# for icv in inactive_customers:
		
	# 	if icv.company == 'TSL COMPANY - Kuwait' and icv.sales_rep =="Vazeem":
		
	# 		data +='<tr>'
	# 		data += '<td style="text-align:left" ><b>%s</b></td>'%(icv.customer)
	# 		data += '<td style="text-align:left"><b>%s</b></td>'%(icv.total_work_orders)
	# 		data += '</tr>'	
	

	# data += '</table>'
	# frappe.sendmail(
	# 	recipients= ['vazeem@tsl-me.com'],
	# 	cc=['omar@tsl-me.com','yousuf@tsl-me.com'],
	# 	subject="Inactive Customers for Past 2 months",
	# 	message = msg+msg1+data)

	data= ""
	data += '<table class="table table-bordered">'
	data += '<tr>'
	data += '<td style="padding:2px;font-size:14px;font-size:12px;background-color:#3da8de;color:white;"><center><b>Customers</b><center></td>'
	data += '<td style="padding:2px;font-size:14px;font-size:12px;background-color:#3da8de;color:white;"><center><b>No of Units Received</b><center></td>'
	data += '<td style="padding:2px;font-size:14px;font-size:12px;background-color:#3da8de;color:white;"><center><b>Sales Person</b><center></td>'
	data += '</tr>'
	for icv in inactive_customers:
		
		if icv.company == 'TSL COMPANY - Kuwait' and icv.sales_rep in ("Maaz"):
		
			data +='<tr>'
			data += '<td style="text-align:left" ><b>%s</b></td>'%(icv.customer)
			data += '<td style="text-align:left"><b>%s</b></td>'%(icv.total_work_orders)
			data += '<td style="text-align:left"><b>%s</b></td>'%(icv.sales_rep)
			data += '</tr>'	
	

	data += '</table>'
	frappe.sendmail(
		recipients= ['yousuf@tsl-me.com'],
		cc=['omar@tsl-me.com'],
		subject="Inactive Customers for Past 3 months",
		message = msg+msg1+data)

@frappe.whitelist()
def gratuity_amount(employee,date):
	filters = {'employee': employee,'date': date}
	from tsl.tsl.report.gratuity.gratuity import execute

	result = execute(filters)
	records = result[1]
	if records:
		gratuity = {
			'termination_days': records[0][7],
			'termination_amount': records[0][8], 
			'resignation_days': records[0][9],
			'resignation_amount': records[0][10]
		}
		return gratuity
	else:
		gratuity = {
			'termination_days': 0,
			'termination_amount': 0, 
			'resignation_days': 0,
			'resignation_amount': 0
		}
		return gratuity


@frappe.whitelist()
def calculate_leave_payment_amount(company,total_working_days,basic,leave_balance):
	working_days = frappe.db.get_value("Company Wise Payroll Days",{"company": company},"total_working_days")
	return float(basic)/float(30) * float(leave_balance)

@frappe.whitelist()
def submit_dubai_wo():
	s = frappe.db.sql(""" UPDATE `tabWork Order Data`
		SET `docstatus` = 1
		WHERE `docstatus` = 0
		AND `company` = 'TSL COMPANY - UAE';
		""")
		
def get_payroll_period_days(start_date, end_date, employee, company=None):
	if not company:
		company = frappe.db.get_value("Employee", employee, "company")
	payroll_period = frappe.db.sql(
		"""
		select name, start_date, end_date
		from `tabPayroll Period`
		where
			company=%(company)s
			and %(start_date)s between start_date and end_date
			and %(end_date)s between start_date and end_date
	""",
		{"company": company, "start_date": start_date, "end_date": end_date},
	)

	if len(payroll_period) > 0:
		actual_no_of_days = date_diff(getdate(payroll_period[0][2]), getdate(payroll_period[0][1])) + 1
		working_days = actual_no_of_days
		if not cint(frappe.db.get_value("Company Wise Payroll Days",{"parent":"Payroll Settings","company":company}, "include_holidays")):
			holidays = get_holiday_dates_for_employee(
				employee, getdate(payroll_period[0][1]), getdate(payroll_period[0][2])
			)
			working_days -= len(holidays)
		return payroll_period[0][0], working_days, actual_no_of_days
	return False, False, False


#to allocate leaves on date of joining every year
from frappe.utils import getdate, add_years, today, add_days
@frappe.whitelist()
def allocate_leave_on_anniversary():
	today = datetime.now()
	current_year = today.year
	current_month = today.month
	current_day = today.day
	emp_list = frappe.db.get_all("Employee",{"status":"Active"},['name', 'date_of_joining','company'])
	
	for emp in emp_list:
		joining_date = getdate(emp['date_of_joining'])
		
		if joining_date.month == current_month and joining_date.day == current_day:
			allocate_sick_leaves(emp['name'],emp['company'],joining_date,today)

def allocate_sick_leaves(employee, company,joining_date,today):
	nex_ann = joining_date.replace(year = today.year + 1)
	day_before_last_day = add_days(nex_ann,-1)
	leave_type = [
		{'type': "Sick Leave 25%", 'days': frappe.db.get_value("Leave Allocation Table", {"company": company}, "sick_leave_25") or 0},
		{'type': "Sick Leave 50%", 'days': frappe.db.get_value("Leave Allocation Table", {"company": company}, "sick_leave_50") or 0},
		{'type': "Sick Leave 75%", 'days': frappe.db.get_value("Leave Allocation Table", {"company": company}, "sick_leave_75") or 0},
		{'type': "Sick Leave 100%", 'days': frappe.db.get_value("Leave Allocation Table", {"company": company}, "sick_leave_100") or 0}
	]
	for i in leave_type:
		alc = frappe.db.exists("Leave Allocation", {"employee": employee, "leave_type": i['type'], "to_date": day_before_last_day})
		
		if not alc:
			if i['days'] > 0:
				al = frappe.new_doc("Leave Allocation")
				al.employee = employee
				al.leave_type = i['type']
				al.from_date = today  
				al.to_date = day_before_last_day
				al.new_leaves_allocated = i['days']
				al.total_leaves_allocated = i['days']
				al.save(ignore_permissions=True)
				al.submit()

# @frappe.whitelist()
# def customer_rep():
# 	w = frappe.get_all("Work Order Data",{"company":"TSL COMPANY - UAE"},["*"])
# 	for i in w:
# 		c = frappe.db.exists("Customer",{"name":i.customer})
# 		if c:
# 			print(i.customer)
# 			cs = frappe.db.exists("Contact Details",{"parent":i.customer})
# 			if cs:
# 				cus = frappe.get_doc("Contact Details",{"parent":i.customer})
			
# 				print(cus.name1)
# 				# frappe.db.set_value("Work Order Data")
# 				frappe.db.sql(
# 					"""
# 					UPDATE `tabWork Order Data`
# 					SET customer_rep = %s
# 					WHERE customer = %s;
# 					""",
# 					(cus.name1,i.customer)
# 				)

@frappe.whitelist()
def crt_wo(import_file):
	from datetime import datetime
	filepath = get_file(import_file)
	data = read_csv_content(filepath[1])
	count = 0
	a_date = ''
	p_date = ''
	d_date = ''
	Pd_date = ''
	app_date = ''
	inv_date = ''

	for i in data[1:20]:
		count = count + 1
		print(count)
		wo = frappe.new_doc("Work Order Data")
		cus = frappe.db.exists("Customer",{"name": i[3]})
		if not cus:
			c = frappe.new_doc("Customer")
			c.customer_name = i[3]
			c.territory = i[4]
			c.customer_type = "Company"
			c.company_group = "Dubai"
			c.save(ignore_permissions = 1)
			
		sp = frappe.db.exists("Sales Person",{"name": i[1]})
		if not sp:
			s = frappe.new_doc("Sales Person")
			s.sales_person_name= i[1]
			s.parent_sales_person = "Sales Team"
			s.save(ignore_permissions =1)
		
		# Input date string
		if i[2]:
			date_ad = str(i[2])
			day, month, year = date_ad.split("-")
			if month == "Jan":
				month = "1"
			if month == "Feb":
				month = "2"
			if month == "Mar":
				month = "3"
			if month == "Apr":
				month = "4"
			if month == "May":
				month = "5"
			if month == "Jun":
				month = "6"
			if month == "Jul":
				month = "7"
			if month == "Aug":
				month = "8"
			if month == "Sep":
				month = "9"
			if month == "Oct":
				month = "10"
			if month == "Nov":
				month = "11"
			if month == "Dec":
				month = "12"
			
			if year == "17":
				year = "2017" 
			if year == "18":
				year = "2018" 
			if year == "19":
				year = "2019" 
			if year == "20":
				year = "2020" 
			if year == "21":
				year = "2021" 
			if year == "22":
				year = "2022" 
			if year == "23":
				year = "2023" 
			if year == "24":
				year = "2024" 


			result = "-".join([year,month,day])
			p_date = result
		else:
			p_date = ''
		
		
		if i[16]:
			date_ad = str(i[16])
			day, month, year = date_ad.split("-")
			if month == "Jan":
				month = "1"
			if month == "Feb":
				month = "2"
			if month == "Mar":
				month = "3"
			if month == "Apr":
				month = "4"
			if month == "May":
				month = "5"
			if month == "Jun":
				month = "6"
			if month == "Jul":
				month = "7"
			if month == "Aug":
				month = "8"
			if month == "Sep":
				month = "9"
			if month == "Oct":
				month = "10"
			if month == "Nov":
				month = "11"
			if month == "Dec":
				month = "12"
			
			if year == "17":
				year = "2017" 
			if year == "18":
				year = "2018" 
			if year == "19":
				year = "2019" 
			if year == "20":
				year = "2020" 
			if year == "21":
				year = "2021" 
			if year == "22":
				year = "2022" 
			if year == "23":
				year = "2023" 
			if year == "24":
				year = "2024" 


			result = "-".join([year,month,day])
			app_date = result
		else:
			app_date = ''
		
		



		if i[15]:
			date_ad = str(i[15])
			day, month, year = date_ad.split("-")
			if month == "Jan":
				month = "1"
			if month == "Feb":
				month = "2"
			if month == "Mar":
				month = "3"
			if month == "Apr":
				month = "4"
			if month == "May":
				month = "5"
			if month == "Jun":
				month = "6"
			if month == "Jul":
				month = "7"
			if month == "Aug":
				month = "8"
			if month == "Sep":
				month = "9"
			if month == "Oct":
				month = "10"
			if month == "Nov":
				month = "11"
			if month == "Dec":
				month = "12"
			
			if year == "17":
				year = "2017" 
			if year == "18":
				year = "2018" 
			if year == "19":
				year = "2019" 
			if year == "20":
				year = "2020" 
			if year == "21":
				year = "2021" 
			if year == "22":
				year = "2022" 
			if year == "23":
				year = "2023" 
			if year == "24":
				year = "2024" 


			result = "-".join([year,month,day])
			a_date = result
		else:
			a_date = ''
		
			

		if i[19]:
			date_ad = str(i[19])
			day, month, year = date_ad.split("-")
			if month == "Jan":
				month = "1"
			if month == "Feb":
				month = "2"
			if month == "Mar":
				month = "3"
			if month == "Apr":
				month = "4"
			if month == "May":
				month = "5"
			if month == "Jun":
				month = "6"
			if month == "Jul":
				month = "7"
			if month == "Aug":
				month = "8"
			if month == "Sep":
				month = "9"
			if month == "Oct":
				month = "10"
			if month == "Nov":
				month = "11"
			if month == "Dec":
				month = "12"
			
			if year == "17":
				year = "2017" 
			if year == "18":
				year = "2018" 
			if year == "19":
				year = "2019" 
			if year == "20":
				year = "2020" 
			if year == "21":
				year = "2021" 
			if year == "22":
				year = "2022" 
			if year == "23":
				year = "2023" 
			if year == "24":
				year = "2024" 


			result = "-".join([year,month,day])
			d_date = result
		else:
			d_date = ''
		
		

		if i[22]:
			date_ad = str(i[22])
			day, month, year = date_ad.split("-")
			if month == "Jan":
				month = "1"
			if month == "Feb":
				month = "2"
			if month == "Mar":
				month = "3"
			if month == "Apr":
				month = "4"
			if month == "May":
				month = "5"
			if month == "Jun":
				month = "6"
			if month == "Jul":
				month = "7"
			if month == "Aug":
				month = "8"
			if month == "Sep":
				month = "9"
			if month == "Oct":
				month = "10"
			if month == "Nov":
				month = "11"
			if month == "Dec":
				month = "12"
			
			if year == "17":
				year = "2017" 
			if year == "18":
				year = "2018" 
			if year == "19":
				year = "2019" 
			if year == "20":
				year = "2020" 
			if year == "21":
				year = "2021" 
			if year == "22":
				year = "2022" 
			if year == "23":
				year = "2023" 
			if year == "24":
				year = "2024" 


			result = "-".join([year,month,day])
			pd_date = result
		else:
			pd_date = ''
		
		if i[20]:
			date_ad = str(i[20])
			day, month, year = date_ad.split("-")
			if month == "Jan":
				month = "1"
			if month == "Feb":
				month = "2"
			if month == "Mar":
				month = "3"
			if month == "Apr":
				month = "4"
			if month == "May":
				month = "5"
			if month == "Jun":
				month = "6"
			if month == "Jul":
				month = "7"
			if month == "Aug":
				month = "8"
			if month == "Sep":
				month = "9"
			if month == "Oct":
				month = "10"
			if month == "Nov":
				month = "11"
			if month == "Dec":
				month = "12"
			
			if year == "17":
				year = "2017" 
			if year == "18":
				year = "2018" 
			if year == "19":
				year = "2019" 
			if year == "20":
				year = "2020" 
			if year == "21":
				year = "2021" 
			if year == "22":
				year = "2022" 
			if year == "23":
				year = "2023" 
			if year == "24":
				year = "2024" 


			result = "-".join([year,month,day])
			inv_date = result
		else:
			inv_date = ''
		
		# print(i[0])
		# print(p_date)
		# print(a_date)
		# print(d_date)
		# print(pd_date)
			
		
		if i[14] == "P" or i[6] == "p":
			i[14] = 'P-Paid'
		if i[14] == "RNRC":
			i[14] = 'RNRC-Return Not Repaired Client'
		if i[14] == "A":
			i[14] = 'A-Approved'
		if i[14] == "C":
			i[14] = 'C-Comparison'
		if i[14] == "CC":
			i[14] = 'CC-Comparison Client'
		if i[14] == "EP" or i[6] == "ED":
			i[14] = 'EP-Extra Parts'
		if i[14] == "NE":
			i[14] = 'NE-Need Evaluation'
		if i[14] == "NER":
			i[14] = 'NER-Need Evaluation Return'
		if i[14] == "Q":
			i[14] = 'Q-Quoted'
		if i[14] == "RNA":
			i[14] = 'RNA-Return Not Approved'
		if i[14] == "RNAC":
			i[14] = 'RNAC-Return Not Approved Client'
		if i[14] == "RNF":
			i[14] = 'RNF-Return No Fault'
		if i[14] == "RNFC":
			i[14] = 'RNFC-Return No Fault Client'
		if i[14] == "RNP":
			i[14] = 'RNP-Return No Parts'
		if i[14] == "RNPC":
			i[14] = 'RNPC-Return No Parts Client'
		if i[14] == "RNR":
			i[14] = 'RNR-Return Not Repaired'
		if i[14] == "RNRC":
			i[14] = 'RNRC-Return Not Repaired Client'
		if i[14] == "RS":
			i[14] = 'RS-Repaired and Shipped'
		if i[14] == "RSC":
			i[14] = 'RSC-Repaired and Shipped Client'
		if i[14] == "RSI":
			i[14] = 'RSI-Repaired and Shipped Invoiced'
		if i[14] == "SP":
			i[14] = 'SP-Searching Parts'
		if i[14] == "TR":
			i[14] ='TR-Technician Repair'
		if i[14] == "UE":
			i[14] = 'UE-Under Evaluation'
		if i[14] == "UTR":
			i[14] = 'UTR-Under Technician Repair'
		if i[14] == "W":
			i[14] = "W-Working"
		if i[14] == "WP":
			i[14] = 'WP-Waiting Parts'
		if i[14] == "CT":
			i[14] = 'CT-Customer Testing'

		wo.naming_series = "WOD-DU-O-.YY.-"
		wo.customer = i[3]
		wo.old_wo_no = i[0]
		wo.status = i[14]
		wo.sales_rep = i[1]
		wo.posting_date = p_date
		wo.old_wo_q_amount = i[12]
		wo.quoted_date = a_date
		wo.payment_date = pd_date
		wo.invoice_date = inv_date
		wo.delivery = d_date
		wo.quotation_approved_date =  app_date
		wo.no_power = 1
		wo.old_wo_vat = i[11]
		wo.old_wo_total_amt = i[13]
		wo.append("material_list", {
				
			'item_code': "001300",
			'model_no': "Old",
			"mfg":"Old mfg",
			"type": "Old",
			"item_name":"001300",
			"quantity" :1
				
		})

		wo.save(ignore_permissions = 1)


# @frappe.whitelist()
# def crt_sod_old(import_file):
# 	from datetime import datetime
# 	filepath = get_file(import_file)
# 	data = read_csv_content(filepath[1])
# 	count = 0
# 	a_date = ''
# 	p_date = ''
# 	d_date = ''
# 	pd_date = ''
# 	app_date = ''
# 	inv_date = ''

# 	for i in data[1:]:
# 		# print(i[16])
# 		count = count + 1
# 		print(count)
# 		wo = frappe.new_doc("Supply Order Data")
# 		cus = frappe.db.exists("Customer",{"name": i[3]})
# 		if not cus:
# 			c = frappe.new_doc("Customer")
# 			c.customer_name = i[3]
# 			c.territory = i[4]
# 			c.customer_type = "Company"
# 			c.company_group = "Dubai"
# 			c.save(ignore_permissions = 1)
			
# 		# sp = frappe.db.exists("Sales Person",{"name": i[1]})
# 		# if not sp:
# 		# 	s = frappe.new_doc("Sales Person")
# 		# 	s.sales_person_name= i[2]
# 		# 	s.parent_sales_person = "Sales Team"
# 		# 	s.save(ignore_permissions =1)
		
# 		# # # # Input date string
# 		if i[1]:
# 			date_ad = str(i[1])
# 			day, month, year = date_ad.split("-" or "/")
# 			if month == "Jan":
# 				month = "1"
# 			if month == "Feb":
# 				month = "2"
# 			if month == "Mar":
# 				month = "3"
# 			if month == "Apr":
# 				month = "4"
# 			if month == "May":
# 				month = "5"
# 			if month == "Jun":
# 				month = "6"
# 			if month == "Jul":
# 				month = "7"
# 			if month == "Aug":
# 				month = "8"
# 			if month == "Sep":
# 				month = "9"
# 			if month == "Oct":
# 				month = "10"
# 			if month == "Nov":
# 				month = "11"
# 			if month == "Dec":
# 				month = "12"
			
			
# 			if year == "17" or year == "2017":
# 				year = "2017" 
# 			if year == "18" or year == "2018" :
# 				year = "2018" 
# 			if year == "19" or year == "2019" :
# 				year = "2019" 
# 			if year == "20" or year == "2020" :
# 				year = "2020" 
# 			if year == "21" or year == "2021" :
# 				year = "2021" 
# 			if year == "22" or year == "2012" :
# 				year = "2022" 
# 			if year == "23" or year == "2023" :
# 				year = "2023" 
# 			if year == "24" or year == "2024" :
# 				year = "2024" 


# 			result = "-".join([year,month,day])
# 			p_date = result
# 		else:
# 			p_date = ''
		
		
# 		if i[21]:
# 			date_ad = str(i[21])
# 			day, month, year = date_ad.split("-" or "/")
# 			if month == "Jan":
# 				month = "1"
# 			if month == "Feb":
# 				month = "2"
# 			if month == "Mar":
# 				month = "3"
# 			if month == "Apr":
# 				month = "4"
# 			if month == "May":
# 				month = "5"
# 			if month == "Jun":
# 				month = "6"
# 			if month == "Jul":
# 				month = "7"
# 			if month == "Aug":
# 				month = "8"
# 			if month == "Sep":
# 				month = "9"
# 			if month == "Oct":
# 				month = "10"
# 			if month == "Nov":
# 				month = "11"
# 			if month == "Dec":
# 				month = "12"
			
			
# 			if year == "17" or year == "2017":
# 				year = "2017" 
# 			if year == "18" or year == "2018" :
# 				year = "2018" 
# 			if year == "19" or year == "2019" :
# 				year = "2019" 
# 			if year == "20" or year == "2020" :
# 				year = "2020" 
# 			if year == "21" or year == "2021" :
# 				year = "2021" 
# 			if year == "22" or year == "2012" :
# 				year = "2022" 
# 			if year == "23" or year == "2023" :
# 				year = "2023" 
# 			if year == "24" or year == "2024" :
# 				year = "2024" 

# 			result = "-".join([year,month,day])
# 			app_date = result
# 		else:
# 			app_date = ''
		
		



# 		if i[20]:
# 			date_ad = str(i[20])
# 			day, month, year = date_ad.split("-" or "/")
# 			if month == "Jan":
# 				month = "1"
# 			if month == "Feb":
# 				month = "2"
# 			if month == "Mar":
# 				month = "3"
# 			if month == "Apr":
# 				month = "4"
# 			if month == "May":
# 				month = "5"
# 			if month == "Jun":
# 				month = "6"
# 			if month == "Jul":
# 				month = "7"
# 			if month == "Aug":
# 				month = "8"
# 			if month == "Sep":
# 				month = "9"
# 			if month == "Oct":
# 				month = "10"
# 			if month == "Nov":
# 				month = "11"
# 			if month == "Dec":
# 				month = "12"
			
			
# 			if year == "17" or year == "2017":
# 				year = "2017" 
# 			if year == "18" or year == "2018" :
# 				year = "2018" 
# 			if year == "19" or year == "2019" :
# 				year = "2019" 
# 			if year == "20" or year == "2020" :
# 				year = "2020" 
# 			if year == "21" or year == "2021" :
# 				year = "2021" 
# 			if year == "22" or year == "2012" :
# 				year = "2022" 
# 			if year == "23" or year == "2023" :
# 				year = "2023" 
# 			if year == "24" or year == "2024" :
# 				year = "2024" 


# 			result = "-".join([year,month,day])
# 			a_date = result
# 		else:
# 			a_date = ''
		
			

# 		if i[23]:
# 			date_ad = str(i[23])
# 			day, month, year = date_ad.split("-" or "/")
# 			if month == "Jan":
# 				month = "1"
# 			if month == "Feb":
# 				month = "2"
# 			if month == "Mar":
# 				month = "3"
# 			if month == "Apr":
# 				month = "4"
# 			if month == "May":
# 				month = "5"
# 			if month == "Jun":
# 				month = "6"
# 			if month == "Jul":
# 				month = "7"
# 			if month == "Aug":
# 				month = "8"
# 			if month == "Sep":
# 				month = "9"
# 			if month == "Oct":
# 				month = "10"
# 			if month == "Nov":
# 				month = "11"
# 			if month == "Dec":
# 				month = "12"
			
			
# 			if year == "17" or year == "2017":
# 				year = "2017" 
# 			if year == "18" or year == "2018" :
# 				year = "2018" 
# 			if year == "19" or year == "2019" :
# 				year = "2019" 
# 			if year == "20" or year == "2020" :
# 				year = "2020" 
# 			if year == "21" or year == "2021" :
# 				year = "2021" 
# 			if year == "22" or year == "2012" :
# 				year = "2022" 
# 			if year == "23" or year == "2023" :
# 				year = "2023" 
# 			if year == "24" or year == "2024" :
# 				year = "2024" 


# 			result = "-".join([year,month,day])
# 			d_date = result
# 		else:
# 			d_date = ''
		
		

# 		if i[26]:
# 			date_ad = str(i[26])
# 			day, month, year = date_ad.split("-" or "/")
# 			if month == "Jan":
# 				month = "1"
# 			if month == "Feb":
# 				month = "2"
# 			if month == "Mar":
# 				month = "3"
# 			if month == "Apr":
# 				month = "4"
# 			if month == "May":
# 				month = "5"
# 			if month == "Jun":
# 				month = "6"
# 			if month == "Jul":
# 				month = "7"
# 			if month == "Aug":
# 				month = "8"
# 			if month == "Sep":
# 				month = "9"
# 			if month == "Oct":
# 				month = "10"
# 			if month == "Nov":
# 				month = "11"
# 			if month == "Dec":
# 				month = "12"
			
			
# 			if year == "17" or year == "2017":
# 				year = "2017" 
# 			if year == "18" or year == "2018" :
# 				year = "2018" 
# 			if year == "19" or year == "2019" :
# 				year = "2019" 
# 			if year == "20" or year == "2020" :
# 				year = "2020" 
# 			if year == "21" or year == "2021" :
# 				year = "2021" 
# 			if year == "22" or year == "2012" :
# 				year = "2022" 
# 			if year == "23" or year == "2023" :
# 				year = "2023" 
# 			if year == "24" or year == "2024" :
# 				year = "2024" 


# 			result = "-".join([year,month,day])
# 			pd_date = result
# 		else:
# 			pd_date = ''
		
# 		if i[24]:
# 			date_ad = str(i[20])
# 			day, month, year = date_ad.split("-")
# 			if month == "Jan":
# 				month = "1"
# 			if month == "Feb":
# 				month = "2"
# 			if month == "Mar":
# 				month = "3"
# 			if month == "Apr":
# 				month = "4"
# 			if month == "May":
# 				month = "5"
# 			if month == "Jun":
# 				month = "6"
# 			if month == "Jul":
# 				month = "7"
# 			if month == "Aug":
# 				month = "8"
# 			if month == "Sep":
# 				month = "9"
# 			if month == "Oct":
# 				month = "10"
# 			if month == "Nov":
# 				month = "11"
# 			if month == "Dec":
# 				month = "12"
			
# 			if year == "17" or year == "2017":
# 				year = "2017" 
# 			if year == "18" or year == "2018" :
# 				year = "2018" 
# 			if year == "19" or year == "2019" :
# 				year = "2019" 
# 			if year == "20" or year == "2020" :
# 				year = "2020" 
# 			if year == "21" or year == "2021" :
# 				year = "2021" 
# 			if year == "22" or year == "2012" :
# 				year = "2022" 
# 			if year == "23" or year == "2023" :
# 				year = "2023" 
# 			if year == "24" or year == "2024" :
# 				year = "2024" 


# 			result = "-".join([year,month,day])
# 			inv_date = result
# 		else:
# 			inv_date = ''	
		
# 		if i[19] == "PAID":
# 			i[19] = 'Paid'
# 		if i[19] == "APPROVED":
# 			i[19] = 'Approved'
# 		if i[19] == "NOT APPROVED":
# 			i[19] = 'Approved'
# 		if i[19] == "QUOTED":
# 			i[19] = 'Quoted'
# 		if i[19] == "INVOICED":
# 			i[19] = 'Invoiced'
		
# 		text = str(i[16])

# 		# Extract the currency (letters)
# 		currency = ''.join(filter(str.isalpha, text))

# 		# Extract the numeric value (digits, commas, and dots)
# 		numeric_value = ''.join(filter(lambda x: x.isdigit() or x in [',', '.'], text))
# 		# print(currency)
		

# 		text2 = str(i[17])

# 		# Extract the currency (letters)
# 		currency2 = ''.join(filter(str.isalpha, text2))

# 		# Extract the numeric value (digits, commas, and dots)
# 		numeric_value2 = ''.join(filter(lambda x: x.isdigit() or x in [',', '.'], text2))
# 		# print(currency)

# 		text3 = str(i[18])

# 		# Extract the currency (letters)
# 		currency3 = ''.join(filter(str.isalpha, text3))

# 		# Extract the numeric value (digits, commas, and dots)
# 		numeric_value3 = ''.join(filter(lambda x: x.isdigit() or x in [',', '.'], text3))
# 		# print(currency)
	
		
# 		wo.naming_series = "SOD-DU-O-.YY.-"
# 		wo.customer = i[3]
# 		wo.old_sod_no = i[0]
# 		wo.status = i[19]
# 		wo.sales_rep = i[2]
# 		wo.company = "TSL COMPANY - UAE"
# 		wo.department = "Supply - TSL"
# 		wo.posting_date = p_date
# 		wo.received_date = p_date
# 		wo.so_old_quoted_amt = numeric_value
# 		wo.so_old_vat = numeric_value2
# 		wo.so_old_total_amt = numeric_value3
# 		wo.quoted_date = a_date
# 		wo.payment_date = pd_date
# 		wo.po_number = i[17]
# 		wo.invoice_date = inv_date
# 		wo.remarks = i[27]
# 		wo.delivery = d_date
# 		wo.quotation_approved_date =  app_date
		
		
# 		wo.so_currency_old = currency
# 		wo.append("material_list", {
# 			'item_code': "001300",
# 			'model': "Old",
# 			"mfg":"Old mfg",
# 			"type": "Old",
# 			"item_name":"001300",
# 			"quantity" :1
				
# 		})

# 		wo.save(ignore_permissions = 1)

# @frappe.whitelist()
# def sales_check():
# 	wo = frappe.db.sql(""" select name from `tabSales Person` where company = '%s' """ %("TSL COMPANY - UAE") ,as_dict =1)
# 	for i in wo:
# 		print(i)


# @frappe.whitelist()
# def leave_allo():
# 	la = frappe.db.sql("""select employee,new_leaves_allocated from `tabLeave Allocation` where leave_type = 'Annual Leave'""",as_dict=1)
# 	for l in la:
# 		exist = l.new_leaves_allocated
# 		new_le = 0.567
# 		tot_leave = exist - new_le
# 		frappe.db.sql("""update `tabLeave Allocation` set new_leaves_allocated = %s , total_leaves_allocated = %s """,(tot_leave,tot_leave))
# 		print(tot_leave)


@frappe.whitelist()
def remove_loan_repayment(rows_to_remove):
	rows_to_remove = json.loads(rows_to_remove)
	for rows in rows_to_remove:
		doc = frappe.get_doc("Repayment Schedule",rows['name'])
		doc.cancel()
		doc.delete()

@frappe.whitelist()
def get_bin(company,item,user):
	branch = frappe.get_value("Employee",{"user_id":user},"branch")
	s = frappe.db.exists("Bin Details",{"parent":item})
	if s:
		k = frappe.get_doc("Item",{"name":item})
		if k:
			for d in k.bin_details:
				if d.company and d.company == company and d.branch == branch:
					frappe.errprint(d.bin)
					return d.bin
	

def send_mail_to_hr_on_rejoining(doc,method):
	message = doc.doctype +" - "+ doc.name +" is Created"
	if doc.company == "TSL COMPANY - KSA":
		frappe.sendmail(
			recipients='hr@tsl-me.com',
			cc = ["yousuf@tsl-me.com","admin@tsl-me.com"],
			sender="info@tsl-me.com",
			subject = message,
			message = message
		)
	else:
		frappe.sendmail(
			recipients='hr@tsl-me.com',
			cc = ["yousuf@tsl-me.com"],
			sender="info@tsl-me.com",
			subject = message,
			message = message
		)

def send_mail_to_hr(doc,method):
	message = doc.doctype +" - "+ doc.name +" is Submitted"
	if doc.company == "TSL COMPANY - KSA":
		frappe.sendmail(
			recipients='hr@tsl-me.com',
			cc = ["yousuf@tsl-me.com","admin@tsl-me.com"],
			sender="info@tsl-me.com",
			subject = message,
			message = message
		)
	else:
		frappe.sendmail(
			recipients='hr@tsl-me.com',
			cc = ["yousuf@tsl-me.com"],
			sender="info@tsl-me.com",
			subject = message,
			message = message
		)


@frappe.whitelist()
def update_wod(import_file):
	from datetime import datetime
	filepath = get_file(import_file)
	data = read_csv_content(filepath[1])
	count = 0
	for i in data[1:]:
		count = count + 1
		print(count)
		# print(i[0])
		if i[14] == "P" or i[6] == "p":
			i[14] = 'P-Paid'
		if i[14] == "RNRC":
			i[14] = 'RNRC-Return Not Repaired Client'
		if i[14] == "A":
			i[14] = 'A-Approved'
		if i[14] == "C":
			i[14] = 'C-Comparison'
		if i[14] == "CC":
			i[14] = 'CC-Comparison Client'
		if i[14] == "EP" or i[6] == "ED":
			i[14] = 'EP-Extra Parts'
		if i[14] == "NE":
			i[14] = 'NE-Need Evaluation'
		if i[14] == "NER":
			i[14] = 'NER-Need Evaluation Return'
		if i[14] == "Q":
			i[14] = 'Q-Quoted'
		if i[14] == "RNA":
			i[14] = 'RNA-Return Not Approved'
		if i[14] == "RNAC":
			i[14] = 'RNAC-Return Not Approved Client'
		if i[14] == "RNF":
			i[14] = 'RNF-Return No Fault'
		if i[14] == "RNFC":
			i[14] = 'RNFC-Return No Fault Client'
		if i[14] == "RNP":
			i[14] = 'RNP-Return No Parts'
		if i[14] == "RNPC":
			i[14] = 'RNPC-Return No Parts Client'
		if i[14] == "RNR":
			i[14] = 'RNR-Return Not Repaired'
		if i[14] == "RNRC":
			i[14] = 'RNRC-Return Not Repaired Client'
		if i[14] == "RS":
			i[14] = 'RS-Repaired and Shipped'
		if i[14] == "RSC":
			i[14] = 'RSC-Repaired and Shipped Client'
		if i[14] == "RSI":
			i[14] = 'RSI-Repaired and Shipped Invoiced'
		if i[14] == "SP":
			i[14] = 'SP-Searching Parts'
		if i[14] == "TR":
			i[14] ='TR-Technician Repair'
		if i[14] == "UE":
			i[14] = 'UE-Under Evaluation'
		if i[14] == "UTR":
			i[14] = 'UTR-Under Technician Repair'
		if i[14] == "W":
			i[14] = "W-Working"
		if i[14] == "WP":
			i[14] = 'WP-Waiting Parts'
		if i[14] == "CT":
			i[14] = 'CT-Customer Testing'
		w = frappe.db.exists("Work Order Data",{"old_wo_no":i[0],"company":"TSL COMPANY - UAE"})\
		
		

		if w:
			# if i[12]:
			# 	i[12] = i[12].replace(",","")
			# if i[13]:
			# 	i[13] = i[13].replace(",","")
			# if i[11]:
			# 	i[11] = i[11].replace(",","")
			wo = frappe.db.sql("""
			UPDATE `tabWork Order Data` 
			SET 
				old_wo_q_amount = %s, 
				old_wo_vat = %s, 
				old_wo_total_amt = %s,
				status = %s 
			WHERE name = %s;
			""", (i[12] or 0, i[11] or 0, i[13] or 0,i[14],w))

@frappe.whitelist()
def update_sod(import_file):
	from datetime import datetime
	filepath = get_file(import_file)
	data = read_csv_content(filepath[1])
	count = 0
	for i in data[1:]:
		count = count + 1
		print(count)
		
		w = frappe.db.exists("Supply Order Data",{"old_sod_no":i[0],"company":"TSL COMPANY - UAE"})
		if w:
			# Extract the currency (letters)
			text = str(i[16])
			currency = ''.join(filter(str.isalpha, text))
			

			# Extract the numeric value (digits, commas, and dots)
			numeric_value = ''.join(filter(lambda x: x.isdigit() or x in [',', '.'], text))
			# print(currency)
			

			text2 = str(i[17])

			# Extract the currency (letters)
			currency2 = ''.join(filter(str.isalpha, text2))

			# Extract the numeric value (digits, commas, and dots)
			numeric_value2 = ''.join(filter(lambda x: x.isdigit() or x in [',', '.'], text2))
			# print(currency)

			text3 = str(i[18])

			# Extract the currency (letters)
			currency3 = ''.join(filter(str.isalpha, text3))

			# Extract the numeric value (digits, commas, and dots)
			numeric_value3 = ''.join(filter(lambda x: x.isdigit() or x in [',', '.'], text3))
			# print(currency)
	
			print(i[0])
			
			if currency == "AED":
				print(currency)
			else:
				currency = "USD"
				print("USD")

			if i[19] == "PAID":
				i[19] = 'Paid'
			if i[19] == "APPROVED":
				i[19] = 'Approved'
			if i[19] == "NOT APPROVED":
				i[19] = 'Approved'
			if i[19] == "QUOTED":
				i[19] = 'Quoted'
			if i[19] == "INVOICED":
				i[19] = 'Invoiced'
				
			wo = frappe.db.sql("""
			UPDATE `tabSupply Order Data` 
			SET 
				so_old_quoted_amt = %s, 
				so_old_vat = %s, 
				so_old_total_amt = %s,
				status = %s ,
				so_currency_old = %s
					  
			WHERE name = %s;
		""", (numeric_value or 0, numeric_value2 or 0, numeric_value3 or 0,i[19],currency,w))


# @frappe.whitelist()
# def crt_pr(import_file):
# 	from datetime import datetime
# 	filepath = get_file(import_file)
# 	data = read_csv_content(filepath[1])
	# count = 0
	# a_date = ''
	# p_date = ''
	# d_date = ''
	# pd_date = ''
	# app_date = ''
	# inv_date = ''

	# for i in data[1:]:
	# 	# print(i[0])
	# 	w = frappe.db.exists("Work Order Data",{"company":"TSL COMPANY - UAE","old_wo_no":i[0]})
	# 	if w:
	# 		tech = ""
	# 		tech_id = ""
	# 		# if i[10] == "Rajesh" or i[10] == "RAjesh":
	# 		# 	tech = "rajesh-uae@tsl-me.com"
	# 		# 	t = frappe.get_value("Technician ID",{"technician":tech},"name")
	# 		# 	tech_id = t
	# 		if i[10] == "Bala" or i[10] == "BALA":
	# 			tech = "bala@tsl-me.com"
	# 			t = frappe.get_value("Technician ID",{"technician":tech},"name")
	# 			tech_id = t
	# 			print(i[10])
	# 			frappe.set_value("Work Order Data",w,"technician",tech)
	# 			frappe.set_value("Work Order Data",w,"tech_id",tech_id)
			# if i[10] == "Glyen" or i[10] == "glyen"  :
			# 	tech = "03glyen06mariano83@gmail.com"
			# 	t = frappe.get_value("Technician ID",{"technician":tech},"name")
			# 	tech_id = t
			# if i[10] == "Marwin" or "MARWIN":
			# 	tech = "marwin-uae@tsl-me.com"
			# 	t = frappe.get_value("Technician ID",{"technician":tech},"name")
			# 	tech_id = t
			# if i[10] == "Ahmed":
			# 	tech = "ahmedmaster75@gmail.com"
			# 	t = frappe.get_value("Technician ID",{"technician":tech},"name")
			# 	tech_id = t
			# if i[10] == "Sami":
			# 	tech = "sami@tsl-me.com"
			# 	t = frappe.get_value("Technician ID",{"technician":tech},"name")
			# 	tech_id = t
			
			# print(tech_id)

		# # count = count + 1
		# print(i[3])
		# wo = frappe.new_doc("Work Order Data")
		# cus = frappe.db.exists("Customer",{"name": i[3]})
		# if not cus:
		# 	c = frappe.new_doc("Customer")
		# 	c.customer_name = i[3]
		# 	c.territory = "DUBAI"
		# 	c.customer_type = "Company"
		# 	# c.company_group = "Dubai"
		# 	c.save(ignore_permissions = 1)
			
		
		# if i[2]:
		# 	date_ad = str(i[2])
		# 	day, month, year = date_ad.split("/")
		# 	if month == "Jan":
		# 		month = "1"
		# 	if month == "Feb":
		# 		month = "2"
		# 	if month == "Mar":
		# 		month = "3"
		# 	if month == "Apr":
		# 		month = "4"
		# 	if month == "May":
		# 		month = "5"
		# 	if month == "Jun":
		# 		month = "6"
		# 	if month == "Jul":
		# 		month = "7"
		# 	if month == "Aug":
		# 		month = "8"
		# 	if month == "Sep":
		# 		month = "9"
		# 	if month == "Oct":
		# 		month = "10"
		# 	if month == "Nov":
		# 		month = "11"
		# 	if month == "Dec":
		# 		month = "12"
			
			
		# 	if year == "17" or year == "2017":
		# 		year = "2017" 
		# 	if year == "18" or year == "2018" :
		# 		year = "2018" 
		# 	if year == "19" or year == "2019" :
		# 		year = "2019" 
		# 	if year == "20" or year == "2020" :
		# 		year = "2020" 
		# 	if year == "21" or year == "2021" :
		# 		year = "2021" 
		# 	if year == "22" or year == "2012" :
		# 		year = "2022" 
		# 	if year == "23" or year == "2023" :
		# 		year = "2023" 
		# 	if year == "24" or year == "2024" :
		# 		year = "2024" 


		# 	result = "-".join([year,month,day])
		# 	p_date = result
		# else:
		# 	p_date = ''
		
		
		# if i[14]:
		# 	date_ad = str(i[14])
		# 	day, month, year = date_ad.split("/")
		# 	if month == "Jan":
		# 		month = "1"
		# 	if month == "Feb":
		# 		month = "2"
		# 	if month == "Mar":
		# 		month = "3"
		# 	if month == "Apr":
		# 		month = "4"
		# 	if month == "May":
		# 		month = "5"
		# 	if month == "Jun":
		# 		month = "6"
		# 	if month == "Jul":
		# 		month = "7"
		# 	if month == "Aug":
		# 		month = "8"
		# 	if month == "Sep":
		# 		month = "9"
		# 	if month == "Oct":
		# 		month = "10"
		# 	if month == "Nov":
		# 		month = "11"
		# 	if month == "Dec":
		# 		month = "12"
			
			
		# 	if year == "17" or year == "2017":
		# 		year = "2017" 
		# 	if year == "18" or year == "2018" :
		# 		year = "2018" 
		# 	if year == "19" or year == "2019" :
		# 		year = "2019" 
		# 	if year == "20" or year == "2020" :
		# 		year = "2020" 
		# 	if year == "21" or year == "2021" :
		# 		year = "2021" 
		# 	if year == "22" or year == "2012" :
		# 		year = "2022" 
		# 	if year == "23" or year == "2023" :
		# 		year = "2023" 
		# 	if year == "24" or year == "2024" :
		# 		year = "2024" 

		# 	result = "-".join([year,month,day])
		# 	app_date = result
		# else:
		# 	app_date = ''
		
		



		# if i[13]:
		# 	date_ad = str(i[13])
		# 	day, month, year = date_ad.split("/")
		# 	if month == "Jan":
		# 		month = "1"
		# 	if month == "Feb":
		# 		month = "2"
		# 	if month == "Mar":
		# 		month = "3"
		# 	if month == "Apr":
		# 		month = "4"
		# 	if month == "May":
		# 		month = "5"
		# 	if month == "Jun":
		# 		month = "6"
		# 	if month == "Jul":
		# 		month = "7"
		# 	if month == "Aug":
		# 		month = "8"
		# 	if month == "Sep":
		# 		month = "9"
		# 	if month == "Oct":
		# 		month = "10"
		# 	if month == "Nov":
		# 		month = "11"
		# 	if month == "Dec":
		# 		month = "12"
			
			
		# 	if year == "17" or year == "2017":
		# 		year = "2017" 
		# 	if year == "18" or year == "2018" :
		# 		year = "2018" 
		# 	if year == "19" or year == "2019" :
		# 		year = "2019" 
		# 	if year == "20" or year == "2020" :
		# 		year = "2020" 
		# 	if year == "21" or year == "2021" :
		# 		year = "2021" 
		# 	if year == "22" or year == "2012" :
		# 		year = "2022" 
		# 	if year == "23" or year == "2023" :
		# 		year = "2023" 
		# 	if year == "24" or year == "2024" :
		# 		year = "2024" 


		# 	result = "-".join([year,month,day])
		# 	a_date = result
		# else:
		# 	a_date = ''
		
			

		# if i[16]:
		# 	date_ad = str(i[16])
		# 	day, month, year = date_ad.split("/")
		# 	if month == "Jan":
		# 		month = "1"
		# 	if month == "Feb":
		# 		month = "2"
		# 	if month == "Mar":
		# 		month = "3"
		# 	if month == "Apr":
		# 		month = "4"
		# 	if month == "May":
		# 		month = "5"
		# 	if month == "Jun":
		# 		month = "6"
		# 	if month == "Jul":
		# 		month = "7"
		# 	if month == "Aug":
		# 		month = "8"
		# 	if month == "Sep":
		# 		month = "9"
		# 	if month == "Oct":
		# 		month = "10"
		# 	if month == "Nov":
		# 		month = "11"
		# 	if month == "Dec":
		# 		month = "12"
			
			
		# 	if year == "17" or year == "2017":
		# 		year = "2017" 
		# 	if year == "18" or year == "2018" :
		# 		year = "2018" 
		# 	if year == "19" or year == "2019" :
		# 		year = "2019" 
		# 	if year == "20" or year == "2020" :
		# 		year = "2020" 
		# 	if year == "21" or year == "2021" :
		# 		year = "2021" 
		# 	if year == "22" or year == "2012" :
		# 		year = "2022" 
		# 	if year == "23" or year == "2023" :
		# 		year = "2023" 
		# 	if year == "24" or year == "2024" :
		# 		year = "2024" 


		# 	result = "-".join([year,month,day])
		# 	d_date = result
		# else:
		# 	d_date = ''
		
		

		# if i[19]:
		# 	date_ad = str(i[19])
		# 	day, month, year = date_ad.split("-" or "/")
		# 	if month == "Jan":
		# 		month = "1"
		# 	if month == "Feb":
		# 		month = "2"
		# 	if month == "Mar":
		# 		month = "3"
		# 	if month == "Apr":
		# 		month = "4"
		# 	if month == "May":
		# 		month = "5"
		# 	if month == "Jun":
		# 		month = "6"
		# 	if month == "Jul":
		# 		month = "7"
		# 	if month == "Aug":
		# 		month = "8"
		# 	if month == "Sep":
		# 		month = "9"
		# 	if month == "Oct":
		# 		month = "10"
		# 	if month == "Nov":
		# 		month = "11"
		# 	if month == "Dec":
		# 		month = "12"
			
			
		# 	if year == "17" or year == "2017":
		# 		year = "2017" 
		# 	if year == "18" or year == "2018" :
		# 		year = "2018" 
		# 	if year == "19" or year == "2019" :
		# 		year = "2019" 
		# 	if year == "20" or year == "2020" :
		# 		year = "2020" 
		# 	if year == "21" or year == "2021" :
		# 		year = "2021" 
		# 	if year == "22" or year == "2012" :
		# 		year = "2022" 
		# 	if year == "23" or year == "2023" :
		# 		year = "2023" 
		# 	if year == "24" or year == "2024" :
		# 		year = "2024" 


		# 	result = "-".join([year,month,day])
		# 	pd_date = result
		# else:
		# 	pd_date = ''
		
		# if i[17]:
		# 	date_ad = str(i[17])
		# 	day, month, year = date_ad.split("/")
		# 	if month == "Jan":
		# 		month = "1"
		# 	if month == "Feb":
		# 		month = "2"
		# 	if month == "Mar":
		# 		month = "3"
		# 	if month == "Apr":
		# 		month = "4"
		# 	if month == "May":
		# 		month = "5"
		# 	if month == "Jun":
		# 		month = "6"
		# 	if month == "Jul":
		# 		month = "7"
		# 	if month == "Aug":
		# 		month = "8"
		# 	if month == "Sep":
		# 		month = "9"
		# 	if month == "Oct":
		# 		month = "10"
		# 	if month == "Nov":
		# 		month = "11"
		# 	if month == "Dec":
		# 		month = "12"
			
		# 	if year == "17" or year == "2017":
		# 		year = "2017" 
		# 	if year == "18" or year == "2018" :
		# 		year = "2018" 
		# 	if year == "19" or year == "2019" :
		# 		year = "2019" 
		# 	if year == "20" or year == "2020" :
		# 		year = "2020" 
		# 	if year == "21" or year == "2021" :
		# 		year = "2021" 
		# 	if year == "22" or year == "2012" :
		# 		year = "2022" 
		# 	if year == "23" or year == "2023" :
		# 		year = "2023" 
		# 	if year == "24" or year == "2024" :
		# 		year = "2024" 


		# 	result = "-".join([year,month,day])
		# 	inv_date = result
		# else:
		# 	inv_date = ''	
		
		# if i[12] == "P":
		# 	i[12] = 'P-Paid'
		# if i[12] == "RSI":
		# 	i[12] = 'RSI-Repaired and Shipped Invoiced'
		# if i[12] == "Q":
		# 	i[12] = 'Q-Quoted'
		# if i[12] == "A":
		# 	i[12] = 'A-Approved'
		
		
		# text = str(i[9])

		# # Extract the currency (letters)
		# currency = ''.join(filter(str.isalpha, text))

		# # Extract the numeric value (digits, commas, and dots)
		# numeric_value = ''.join(filter(lambda x: x.isdigit() or x in [',', '.'], text))
		# # print(currency)
		

		# text2 = str(i[10])

		# # Extract the currency (letters)
		# currency2 = ''.join(filter(str.isalpha, text2))

		# # Extract the numeric value (digits, commas, and dots)
		# numeric_value2 = ''.join(filter(lambda x: x.isdigit() or x in [',', '.'], text2))
		# # print(currency)

		# text3 = str(i[11])

		# # Extract the currency (letters)
		# currency3 = ''.join(filter(str.isalpha, text3))

		# # Extract the numeric value (digits, commas, and dots)
		# numeric_value3 = ''.join(filter(lambda x: x.isdigit() or x in [',', '.'], text3))
		# # print(currency)
	
		
		# wo.naming_series = "WOD-DU.YY.-P."
		# wo.customer = i[3]
		# wo.old_wod_no = i[0]
		# wo.status = i[12]
		# wo.sales_rep = i[1]
		# wo.company = "TSL COMPANY - UAE"
		# # wo.department = "Supply - TSL"
		# wo.posting_date = p_date
		# wo.received_date = p_date
		# wo.old_wo_q_amount = numeric_value
		# wo.old_wo_vat = numeric_value2
		# wo.old_wo_total_amt = numeric_value3
		# wo.quoted_date = a_date
		# wo.payment_date = pd_date
		# wo.po_number = i[17]
		# wo.invoice_date = inv_date
		# wo.remarks = i[20]
		# wo.delivery = d_date
		# wo.po_no = i[15]
		# wo.quotation_approved_date =  app_date
		
		
		# # wo.so_currency_old = currency
		# wo.append("material_list", {
		# 	'item_code': "001300",
		# 	'model': "Old",
		# 	"mfg":"Old mfg",
		# 	"type": "Old",
		# 	"item_name":"001300",
		# 	"quantity" :1
				
		# })

		# wo.save(ignore_permissions = 1)

@frappe.whitelist()
def update_serial_no(sn,wo):
	st = frappe.get_value("Stock Entry",{"work_order_data":wo},"name")
	frappe.errprint(st)
	
	serial = frappe.get_value("Serial No",{"purchase_document_no":st},"name")
	frappe.db.set_value("Serial No",serial,"name",sn)
	
	frappe.db.sql("""
	UPDATE `tabStock Entry Detail` AS sii
	JOIN `tabStock Entry` AS si ON sii.parent = si.name
	SET sii.serial_no = %s
	WHERE si.work_order_data = %s
	""", (sn,wo))

@frappe.whitelist()
def set_tech(company,branch):
	# company = filters.get("company")
	# sales_list = []
	sales = frappe.get_all("Employee",{"company":company,"branch":branch,"technician":1},["user_id"])
	# for i in sales:
	# 	sales_list.append(i.user_id)
	# frappe.errprint(sales_list)
	return sales
	
	
# from frappe.permissions import (
# 	add_user_permission,
# 	get_doc_permissions,
# 	has_permission,
# 	remove_user_permission,
# 	set_user_permission_if_allowed,
# 	get_role_permissions
# )
def set_user_permission_if_allowed(doctype, name, user, with_message=False,hide_descendants = 0):
	if get_role_permissions(frappe.get_meta(doctype), user).set_user_permissions != 1:
		add_user_permission(doctype, name, user,hide_descendants=hide_descendants)

def update_user_permissions(self):
	if not self.create_user_permission:
		return
	if not has_permission("User Permission", ptype="write", raise_exception=False):
		return

	employee_user_permission_exists = frappe.db.exists(
		"User Permission", {"allow": "Employee", "for_value": self.name, "user": self.user_id}
	)

	if employee_user_permission_exists:
		return

	add_user_permission("Employee", self.name, self.user_id,hide_descendants = 1)
	add_user_permission("Branch", self.branch, self.user_id,hide_descendants = 1)
	set_user_permission_if_allowed("Company", self.company, self.user_id,hide_descendants = 1)


@frappe.whitelist()

def test_email():
# 	make(
# 		sender = "info@tsl-me.com",
# 		recipients="yousuf@tsl-me.com",
# 		content= "TEST EMAIL",
# 		subject="TEST",
# 		send_mail=True,
# 		cc="mohamedyousufesi46@gmail.com",
# 
	# )
	sq = frappe.db.sql("""
		UPDATE `tabSupply Order Data` 
		SET branch = 'Dubai - TSL' 
		WHERE company = 'TSL COMPANY - UAE'
	""", as_dict=1)

@frappe.whitelist()
def get_item_number(item_details,company):
	item_details = json.loads(item_details)
	data = ''
	if item_details:
		data = ''
		data += '<h5><center><b>PART DETAILS</b></center></h5>'
		data += '<table class="table table-bordered">'
		data += '<tr>'
		data += '<td style="width:07%;padding:1px;font-size:14px;font-size:12px;background-color:#3333ff;color:white;"><center><b>ITEM NUMBER</b><center></td>'
		data += '<td style="width:07%;padding:1px;font-size:14px;font-size:12px;background-color:#3333ff;color:white;"><center><b>SKU</b><center></td>'
		data += '</tr>'
		for j in item_details:
			data += '<tr>'
			data += '<td style="font-size:12px;"><center><b>%s</b><center></td>' %(j["item_number"])
			s = ""
			sku = frappe.get_value("Item",{"item_number":j["item_number"]},["name"])
			if sku:
				s = sku
			else:
				s = "NA"
			data += '<td style="font-size:12px;"><center><b>%s</b><center></td>' %(s)
			data += '</tr>'

		data += '</table>'
		return data 



import frappe
from frappe.utils.pdf import get_pdf

@frappe.whitelist()
def download_custom_pdf(doctype, name, print_format="Standard", no_letterhead=0):
	doc = frappe.get_doc(doctype, name)
	html = frappe.get_print(doctype, name, print_format, doc=doc, no_letterhead=no_letterhead)
	pdf = get_pdf(html)

	# Clean filename (remove spaces, special chars)
	customer = (doc.customer_name or "").replace(" ", "_").replace("/", "_")
	filename = f"{name}_{customer}.pdf"

	frappe.local.response.filename = filename
	frappe.local.response.filecontent = pdf
	frappe.local.response.type = "download"


@frappe.whitelist()
def create_replacement_item(customer,wod,items):
	ev = frappe.db.exists("Evaluation Report",{"work_order_data":wod})
	if ev:
		evaluation = frappe.get_doc("Evaluation Report",ev)
		items = json.loads(items)
		for i in items:
			evaluation.append("replacement_unit",{
				"item":i["item_code"],
				"model":i["model_no"],
				"manufacturer":i["mfg"],
				"type":i["type"],
				"serial_no":"",
				"description":i["item_name"]
			  

			})
		evaluation.save(ignore_permissions = True)

		wd = frappe.new_doc("Replacement Unit")
		wd.name = wod

		
		doclist = get_mapped_doc("Work Order Data",wod, {
		"Work Order Data": {
			"doctype": "Work Order Data",	
		},
		},wd)
		
		for i in doclist.get('material_list'):		
			i.serial_no = ""
		wd.status = "Inquiry"
		wd.status_duration_details = ""
		wd.save()

		w = frappe.get_doc("Work Order Data",wod)
		st = frappe.new_doc("Stock Entry")
		st.company = w.company
		st.stock_entry_type = "Material Issue"
		for i in w.material_list:
			st.append("items",{
			"item_code":i.item_code,
			"qty":i.quantity,
			"s_warehouse":"Repair - Kuwait - TSL",
			"uom":"Nos",
			"stock_uom":"Nos",
			"serial_no":i.serial_no,
			'conversion_factor':1,
					
		})
		st.save(ignore_permissions = True)


# @frappe.whitelist()
# def get_sales_person():
	# s = frappe.db.sql(""" select distinct sales_rep from `tabSales Invoice` """)
	# for i in s:
	# 	print(i[0])
	# k = frappe.db.sql(""" UPDATE `tabQuotation`
 	# 		SET workflow_state = "Approved By Management"
 	# 		WHERE name = "SUP-QTN-INT-DU25-00174" """)


# @frappe.whitelist()
# def set_sales_person():
# 	k = frappe.db.sql(""" UPDATE `tabQuotation`
# 	SET sales_rep = "yehia@tsl-me.com"
# 	WHERE custom_sales_person = 'Ahmed Yahia' """)
# 	s = frappe.db.sql(""" select distinct sales_rep from `tabSales Invoice` """)
# 	for i in s:
		
# 		user = frappe.get_value("Sales Person",{"custom_user":i[0]},["name"])
# 		if user:
# 			print(user)
	



