# Copyright (c) 2024, Tsl and contributors
# For license information, please see license.txt

import frappe
from frappe.model.document import Document

class ReportDashboard(Document):
	@frappe.whitelist()
	def get_wod(self):
		return "Yesss"



import io
import frappe
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side

@frappe.whitelist()
def download_work_order_with_operations_excel():
	args = frappe.local.form_dict
	from_date = args.from_date
	to_date = args.to_date
	company = args.company

	work_orders = frappe.get_all(
		'Work Order Data',
		filters={
			"posting_date": ("between", (from_date, to_date)),
			"company": company
		},
		fields=['name', 'status', 'posting_date'],
		order_by = 'posting_date asc'
	)
	end_row = 0
	# Create workbook
	wb = Workbook()
	ws = wb.active
	ws.title = "Work Order Report"

	# Add header
	headers = ['Work Order ID', 'Date', 'Current Status', 'Status', 'Duration (in Days)']
	ws.append(headers)
	header_font = Font(bold=True)
	header_alignment = Alignment(horizontal="center", vertical="center")

	for col_num, _ in enumerate(headers, start=1):
		cell = ws.cell(row=1, column=col_num)
		cell.font = header_font
		cell.alignment = header_alignment

	current_row = 2  # Header is in row 1

	for work_order in work_orders:
		operations = frappe.get_all(
			"Status Duration Details",
			filters={'parenttype': "Work Order Data", 'parent': work_order.name},
			fields=['status', 'duration'],
			order_by = 'idx asc'
		)

		start_row = current_row

		if not operations:
			ws.append([work_order.name, frappe.utils.format_date(work_order.posting_date), work_order.status, '', ''])
			current_row += 1
		else:
			for operation in operations:
				try:
					if operation.duration:
						duration = operation.duration
						hours = int(duration.split('hrs')[0].strip())
						minutes = int(duration.split('hrs')[1].split('min')[0].strip())
						total_minutes = (hours * 60) + minutes
						total_days = round(total_minutes / (24 * 60), 2)
					else:
						total_days = 0
				except Exception:
					total_days = 0

				ws.append([
					work_order.name,
					frappe.utils.format_date(work_order.posting_date),
					work_order.status,
					operation.status,
					total_days
				])
				current_row += 1

		end_row = current_row - 1

		# Merge cells for Work Order ID and Current Status if there are multiple rows
		if end_row > start_row:
			ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
			ws.merge_cells(start_row=start_row, start_column=2, end_row=end_row, end_column=2)
			ws.merge_cells(start_row=start_row, start_column=3, end_row=end_row, end_column=3)

			# Center align merged cells
			ws.cell(row=start_row, column=1).alignment = Alignment(horizontal="center", vertical="center")
			ws.cell(row=start_row, column=2).alignment = Alignment(horizontal="center", vertical="center")
			ws.cell(row=start_row, column=3).alignment = Alignment(horizontal="center", vertical="center")



	thin_border = Border(
		left=Side(style='thin'),
		right=Side(style='thin'),
		top=Side(style='thin'),
		bottom=Side(style='thin')
	)

	max_row = end_row
	max_col = ws.max_column

	for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
		for cell in row:
			cell.border = thin_border

	# Set column widths
	column_widths = [15, 20, 30, 30, 19]
	for i, width in enumerate(column_widths, start=1):
		col_letter = ws.cell(row=1, column=i).column_letter
		ws.column_dimensions[col_letter].width = width

	# Save to BytesIO
	output = io.BytesIO()
	wb.save(output)
	output.seek(0)

	frappe.local.response.filename = "Work Order Data.xlsx"
	frappe.local.response.filecontent = output.getvalue()
	frappe.local.response.type = "binary"
	frappe.local.response.status_code = 200
