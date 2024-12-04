import frappe
from frappe.utils import cstr, add_days, date_diff, getdate, format_date
from frappe import _, bold
from frappe.utils.csvutils import UnicodeWriter, read_csv_content
from frappe.utils.data import format_date
from frappe.utils.file_manager import get_file
from frappe.model.document import Document
from frappe.utils.background_jobs import enqueue

from datetime import date, timedelta, datetime
import openpyxl
from openpyxl import Workbook
import re
from frappe import _
import frappe
from frappe.model.document import Document
from datetime import date, timedelta, datetime,time
from frappe.utils import (getdate, cint, add_months, date_diff, add_days,
    nowdate, get_datetime_str, cstr, get_datetime, now_datetime, format_datetime,today, format_date)
import math
from frappe.utils import add_months, cint, flt, getdate, time_diff_in_hours,time_diff_in_seconds
import locale


import openpyxl
import xlrd
import re
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
from six import BytesIO, string_types
import openpyxl.styles as styles
from frappe.utils import flt, fmt_money


@frappe.whitelist()
def download():
    filename = 'Work Order Data Time Log'
    test = build_xlsx_response(filename)
    
        
def make_xlsx(data, sheet_name=None, wb=None, column_widths=None):
    args = frappe.local.form_dict
    column_widths = column_widths or []
    if wb is None:
        wb = openpyxl.Workbook()
         
    ws = wb.create_sheet(sheet_name, 0)
    # ws.column_dimensions['A'].width = 20
    # ws.column_dimensions['B'].width = 20
    # ws.column_dimensions['C'].width = 20 
    # ws.column_dimensions['D'].width = 50
    # ws.append(["Customer Ledger Summary Report"," "," "," ","",""])
    # from_date_str = args.from_date
    # from_date_obj = datetime.strptime(from_date_str, "%Y-%m-%d")
    # formatted_from_date = from_date_obj.strftime("%d-%m-%Y")
    # to_date_str = args.to_date
    # to_date_obj = datetime.strptime(to_date_str, "%Y-%m-%d")
    # formatted_to_date = to_date_obj.strftime("%d-%m-%Y")
    # header3=["For the Period From: "+ formatted_from_date +" "+"to"+" "+formatted_to_date]
    # ws.append(header3 + [""] * 5)
    # ws.append([''])
    # ws.append(["Customer Name",args.customer,"","Currency",args.currency,""])
    # ws.append(['Date','Voucher No','Voucher Type','Narration',"Debit","Credit"])
    ws.append(['Work Order','Customer','Date','Status','Technician','SKU','MFG','Type','Serial No','Description'
    # 'NE-Need Evaluation',
    # 'AP-Available Parts',
    # 'SP-Searching Parts',
    # 'Q-Quoted',
    # 'CT-Customer Testing',
    # 'RNP-Return No Parts',
    # 'RNPC-Return No Parts Client',
    # 'A-Approved',
    # 'EP-Extra Parts',
    # 'RNA-Return Not Approved',
    # 'RNAC-Return Not Approved Client',
    # 'TR-Technician Repair',
    # 'RS-Repaired and Shipped',
    # 'RSC-Repaired and Shipped Client',
    # 'WP-Waiting Parts',
    # 'RNR-Return Not Repaired',
    # 'RNRC-Return Not Repaired Client',
    # 'W-Working',
    # 'P-Paid',
    # 'C-Comparison',
    # 'CC-Comparison Client',
    # 'NER-Need Evaluation Return',
    # 'UE-Under Evaluation',
    # 'IQ-Internally Quoted',
    # 'UTR-Under Technician Repair',
    # 'RNF-Return No Fault',
    # 'RNFC-Return No Fault Client',
    # 'Parts Priced',
    # 'Pending Internal Approval',
    # 'RSI-Repaired and Shipped Invoiced',
               ])
    wd = frappe.get_all("Work Order Data",{"company":"TSL COMPANY - Kuwait","posting_date": ["between", (args.from_date,args.to_date)]},["*"])
    for i in wd:
        ml = frappe.get_doc("Material List",{"parent":i.name})
        row = [i.name,i.customer,i.posting_date,i.status,i.technician,ml.item_code,ml.mfg,ml.type,ml.serial_no or '',ml.item_name]
        ws.append(row)
        s = frappe.db.sql(""" select date,status,duration from `tabStatus Duration Details` where parent = '%s' ORDER BY status """ %(i.name),as_dict=1)
        
        ws.append(["<b>Status</b>","Date and Time","Duration"])
        for a in s:
            row_1 = [a.status,a.date,a.duration]
            ws.append(row_1)
            
        ws.append([""])
        
    
             
    # data1= get_data(args)
    # for row in data1:
    #     ws.append(row)
    # align_center = Alignment(horizontal='center',vertical='center')
    # align_right = Alignment(horizontal='right',vertical='bottom')
    for header in ws.iter_rows(min_row=1 , max_row=1, min_col=1, max_col=10):
        for cell in header:
            cell.font = Font(bold=True)
    
     # Make every cell bold for Status, Date and Time, and Duration columns
    # for col in [1,2,3,4, 5, 6,7,8,9,10]:  # Assuming Status is in column D, Date and Time in column E, and Duration in column F
    #     for row in range(2, len(wd) * 3):  # Adjusted for the number of rows
    #         ws.cell(row=row, column=col).font = Font(bold=True)
    # for header in ws.iter_rows(min_row=3 , max_row=3, min_col=1, max_col=2):
    #     for cell in header:
    #         cell.font = Font(bold=True)
    #         cell.alignment = align_center
    # for header in ws.iter_rows(min_row=6 , max_row=6, min_col=1, max_col=6):
    #     for cell in header:
    #         cell.font = Font(bold=True)
    #         cell.alignment = align_right
    # for header in ws.iter_rows(min_row=len(get_data(args))+5 , max_row=len(get_data(args))+5, min_col=1, max_col=6):
    #     for cell in header:
    #         cell.font = Font(bold=True)
    #         cell.alignment = align_center
    # for header in ws.iter_rows(min_row=len(get_data(args))+6 , max_row=len(get_data(args))+6, min_col=1, max_col=6):
    #     for cell in header:
    #         cell.font = Font(bold=True)
    #         cell.alignment = align_right
    # for header in ws.iter_rows(min_row=5 , max_row=5, min_col=1, max_col=6):
    #     for cell in header:
    #         cell.font = Font(bold=True)
    #         cell.fill = PatternFill(fgColor='D3D3D3', fill_type = "solid")
    border_thin = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin'))
    header_range = ws['A1':ws.cell(row=len(wd)+1, column=10).coordinate]
    for row in header_range:
        for cell in row:
            cell.border = border_thin
    # ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6 )
    # ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6 )
    # ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=6 )
    # ws.merge_cells(start_row=4, start_column=2, end_row=4, end_column=3 )
    # ws.merge_cells(start_row=4, start_column=5, end_row=4, end_column=6 )
    # ws.merge_cells(start_row=6, start_column=1, end_row=6, end_column=6 )
    # ws.merge_cells(start_row=len(get_data(args))+5, start_column=2, end_row=len(get_data(args))+5, end_column=6)
    # ws.merge_cells(start_row=len(get_data(args))+6, start_column=1, end_row=len(get_data(args))+6, end_column=4 )
    

    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    return xlsx_file

def get_data(args):
    row2=[]
    row3=[]
    data=[]
    credit=0
    debit=0
    currency = frappe.db.get_all("GL Entry",{'against':args.customer},['name','posting_date','account_currency','voucher_type','remarks','credit_in_account_currency','debit_in_account_currency'])
    for cur in currency:
        row1=[]
        credit +=cur.credit_in_account_currency
        debit +=cur.debit_in_account_currency
        row1+=[cur.posting_date.strftime("%d-%m-%Y"),cur.name,cur.voucher_type,cur.remarks,fmt_money(cur.debit_in_account_currency),fmt_money(cur.credit_in_account_currency)]
        data.append(row1)
    count = frappe.db.count("GL Entry",{'against':args.customer})
    row2+=["No. of Entries:",count,"","","",""]
    row3+=["Total","","","",debit,credit]
    data.append(row2)
    data.append(row3)
    return data
    
    

def build_xlsx_response(filename):
    xlsx_file = make_xlsx(filename)
    frappe.response['filename'] = filename + '.xlsx'
    frappe.response['filecontent'] = xlsx_file.getvalue()
    frappe.response['type'] = 'binary'